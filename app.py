
import streamlit as st
import pandas as pd
import numpy as np
import random, re, os, glob, unicodedata, time, io, gc, json
import requests
from difflib import SequenceMatcher
from itertools import combinations

st.set_page_config(page_title="Tennis IA v23.25.8 Fallback Lectura", page_icon="🎾", layout="wide")

APP_VERSION = "v23.49.15-ta-jsfrag-html-parser"
QUALITY_ENGINE_VERSION = "v23.39.5-wta-over17-rankgap80-2026-05-28"

# =========================================================
# 🔒 OVER ENGINE LOCK
# Regla de proyecto: desde esta base NO se toca ninguna fórmula,
# umbral, guard, simulación ni exportación que afecte a mercados Over
# salvo orden explícita del usuario.
# Este flag es informativo y no modifica ningún cálculo.
# =========================================================
OVER_ENGINE_LOCKED = True
OVER_ENGINE_LOCK_NOTE = "Over blindado: no tocar matemáticas ni lógica de Over sin orden explícita."

# v23.21: WTA Over17 Export Fix + Watchlist Tight + Strict Surname Fix.
# v23.39.1: WTA Reason Fix - corrige motivos ML vs gana al menos 1 set sin tocar cálculos.
# v23.39.5: WTA Over17 RankGap80 - baja el corte WTA Over 17.5 a >=80% con RankGap 21-50, sin tocar motor Over.

# =========================================================
# TENNIS IA v15
# Stable core + Hard Compression + Smart Tie-break + Analyzer
# =========================================================

def normalizar_texto(x):
    if pd.isna(x): return ""
    t = unicodedata.normalize("NFKD", str(x)).encode("ascii", "ignore").decode("ascii")
    return t.replace("\xa0", " ").strip()

def limpiar(x):
    t = re.sub(r"\[.*?\]|\(.*?\)", "", normalizar_texto(x))
    return re.sub(r"[^A-Z0-9]", "", t.upper())

def tokens(x):
    t = re.sub(r"\[.*?\]|\(.*?\)", "", normalizar_texto(x))
    return re.findall(r"[A-Z]+", t.upper())

def similitud_nombre(a, b):
    ac, bc = limpiar(a), limpiar(b)
    if not ac or not bc: return 0.0
    if ac == bc: return 1.0
    ta, tb = tokens(a), tokens(b)
    sa, sb = set(ta), set(tb)
    score = len(sa & sb) / len(sa | sb) if sa and sb else 0.0

    # "Tiafoe F." vs "Frances Tiafoe"
    if len(ta) >= 2 and len(tb) >= 2:
        surname, initial = ta[0], ta[1][0]
        if surname in tb and any(x != surname and x.startswith(initial) for x in tb):
            score = max(score, 0.94)
    if len(tb) >= 2 and len(ta) >= 2:
        surname, initial = tb[0], tb[1][0]
        if surname in ta and any(x != surname and x.startswith(initial) for x in ta):
            score = max(score, 0.94)

    return max(score, SequenceMatcher(None, ac, bc).ratio())


# =========================================================
# v23.47.2 TA ALIAS MANAGER
# Evita que el mismo jugador aparezca como faltante por variantes:
#   Y. Bu -> Bu Yunchaokete
#   T. Kokkinakis -> Thanasi Kokkinakis
# No toca el motor Over; solo normaliza nombres antes de buscar perfiles.
# =========================================================

TA_ALIASES_PATH = os.path.join("datos", "ta_aliases.json")


def _ta_alias_ensure_dir():
    try:
        os.makedirs(os.path.dirname(TA_ALIASES_PATH), exist_ok=True)
    except Exception:
        pass


def _ta_alias_key(name):
    return limpiar(name)


@st.cache_data(show_spinner=False)
def _ta_alias_load_cached(cache_version=APP_VERSION):
    try:
        if not os.path.exists(TA_ALIASES_PATH):
            return {}
        with open(TA_ALIASES_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {}
        out = {}
        for k, v in data.items():
            kk = _ta_alias_key(k)
            vv = normalizar_texto(v)
            if kk and vv:
                out[kk] = vv
        return out
    except Exception:
        return {}


def _ta_alias_load():
    return dict(_ta_alias_load_cached())


def _ta_alias_save(alias_map):
    _ta_alias_ensure_dir()
    try:
        clean = {}
        for k, v in (alias_map or {}).items():
            kk = _ta_alias_key(k)
            vv = normalizar_texto(v)
            if kk and vv:
                clean[kk] = vv
        with open(TA_ALIASES_PATH, "w", encoding="utf-8") as f:
            json.dump(clean, f, ensure_ascii=False, indent=2, sort_keys=True)
        try:
            st.cache_data.clear()
        except Exception:
            pass
        return True, f"{len(clean)} alias guardados"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _ta_alias_resolve(name):
    """Devuelve el nombre canónico si existe alias; si no, devuelve el original."""
    raw = normalizar_texto(name)
    if not raw:
        return raw
    aliases = _ta_alias_load_cached()
    seen = set()
    cur = raw
    # Permite alias encadenados sin bucles.
    for _ in range(4):
        k = _ta_alias_key(cur)
        if not k or k in seen or k not in aliases:
            break
        seen.add(k)
        cur = aliases[k]
    return normalizar_texto(cur) or raw


def _ta_alias_add(alias_name, canonical_name):
    alias_name = normalizar_texto(alias_name)
    canonical_name = normalizar_texto(canonical_name)
    if not alias_name or not canonical_name:
        return False, "alias o nombre correcto vacío"
    if _ta_alias_key(alias_name) == _ta_alias_key(canonical_name):
        return False, "el alias y el nombre correcto son iguales"
    m = _ta_alias_load()
    m[_ta_alias_key(alias_name)] = canonical_name
    return _ta_alias_save(m)


def _ta_alias_debug_rows(names):
    aliases = _ta_alias_load_cached()
    rows = []
    for n in names or []:
        k = _ta_alias_key(n)
        canon = aliases.get(k, "")
        rows.append({
            "Jugador detectado": n,
            "Alias guardado": canon,
            "Nombre usado por la app": canon or n,
            "Estado alias": "✅ Alias" if canon else "—"
        })
    return rows

def buscar_columna(df, posibles):
    cols = {limpiar(c): c for c in df.columns}
    for p in posibles:
        if limpiar(p) in cols:
            return cols[limpiar(p)]
    return None

def leer_float(v, default):
    try:
        if pd.isna(v): return default
        s = str(v).replace(",", ".").replace("%", "").strip()
        if s == "": return default
        return float(s)
    except:
        return default

def leer_porcentaje(v, default):
    n = leer_float(v, default)
    if n is None: return None
    if n > 1: n /= 100
    return n

def elo_prob(e1, e2):
    return 1 / (1 + 10 ** ((e2 - e1) / 400))

def nivel(p):
    if p >= 0.72: return "🔥 Alta"
    if p >= 0.60: return "✅ Media-alta"
    if p >= 0.53: return "⚖️ Ajustada"
    return "⚠️ Baja"

def perfil_saque(ace):
    if ace >= 0.16: return "elite_server"
    if ace >= 0.12: return "big_server"
    if ace >= 0.08: return "good_server"
    return "normal"

def perfil_legible(p):
    return {
        "elite_server": "🚀 Elite server",
        "big_server": "🔥 Big server",
        "good_server": "✅ Buen sacador",
        "normal": "Normal"
    }.get(p, "Normal")

def calibrar_probabilidad(p, surface):
    if p >= 0.50:
        fav, sign = p, 1
    else:
        fav, sign = 1 - p, -1

    if fav < 0.58: shrink = 0.02
    elif fav < 0.65: shrink = 0.05
    elif fav < 0.72: shrink = 0.08
    elif fav < 0.80: shrink = 0.12
    else: shrink = 0.16

    if surface == "Clay": shrink += 0.02
    elif surface == "Hard": shrink += 0.01

    cal = 0.50 + (fav - 0.50) * (1 - shrink)
    cal = np.clip(cal, 0.50, 0.88)
    return cal if sign == 1 else 1 - cal

def edge_calibracion(raw, cal):
    d = abs(raw - cal)
    if d < 0.025: return "🟢 estable"
    if d < 0.055: return "🟡 algo inflada"
    return "🔴 muy inflada"

def rutas(circuito):
    base = f"datos/{circuito.lower()}"
    return {
        "base": base,
        "elo": f"{base}/{circuito.lower()}_elo.xlsx",
        "serve": f"{base}/{circuito.lower()}_serve.xlsx",
        "return": f"{base}/{circuito.lower()}_return.xlsx",
        "break": f"{base}/{circuito.lower()}_break.xlsx",
        "historicos": f"{base}/historicos"
    }

def ruta_stats_superficie(circuito, tipo, surface):
    """
    v15 Surface Adaptive:
    Intenta leer primero:
      datos/atp/atp_serve_hard.xlsx
      datos/atp/atp_return_clay.xlsx
      datos/atp/atp_break_grass.xlsx
    Si no existe, usa el archivo general:
      datos/atp/atp_serve.xlsx
    """
    base = f"datos/{circuito.lower()}"
    surf = str(surface).lower()
    specific = f"{base}/{circuito.lower()}_{tipo}_{surf}.xlsx"
    general = f"{base}/{circuito.lower()}_{tipo}.xlsx"
    return specific if os.path.exists(specific) else general

def merge_surface_stats(general_stats, surface_stats):
    """
    Combina general + superficie.
    La superficie pisa al general cuando trae dato.
    """
    out = general_stats.copy()
    for k, v in surface_stats.items():
        if v is not None:
            out[k] = v
    return out

def stats_filename_label(circuito, tipo, surface):
    path = ruta_stats_superficie(circuito, tipo, surface)
    return os.path.basename(path) if path else "N/A"

def stats_default_por_elo(elo_surface, rank=999, surface="Clay", circuito="ATP"):
    if elo_surface >= 1800: hold, ace, ret, brk = 0.825, 0.075, 0.300, 0.300
    elif elo_surface >= 1700: hold, ace, ret, brk = 0.810, 0.070, 0.285, 0.285
    elif elo_surface >= 1600: hold, ace, ret, brk = 0.795, 0.065, 0.270, 0.270
    elif elo_surface >= 1500: hold, ace, ret, brk = 0.780, 0.055, 0.250, 0.250
    else: hold, ace, ret, brk = 0.760, 0.050, 0.225, 0.225

    if rank <= 50: hold += 0.006; ret += 0.012; brk += 0.012
    elif rank <= 100: hold += 0.003; ret += 0.006; brk += 0.006
    elif rank > 180: hold -= 0.006; ret -= 0.006; brk -= 0.006

    if surface == "Clay": hold -= 0.010; ret += 0.025; brk += 0.025; ace -= 0.005
    elif surface == "Grass": hold += 0.012; ret -= 0.012; brk -= 0.012; ace += 0.006

    if circuito == "WTA": hold -= 0.035; ret += 0.030; brk += 0.030; ace -= 0.015

    ace = np.clip(ace, 0.035, 0.090)
    return {
        "found_stats": False, "raw_name_stats": "NO ENCONTRADO",
        "hold": np.clip(hold, 0.72, 0.84), "ace": ace, "df": 0.035,
        "1in": 0.62, "1w": 0.70, "2w": 0.50,
        "rpw": np.clip(ret, 0.18, 0.38),
        "break_pct": np.clip(brk, 0.15, 0.42),
        "bp_conv": 0.38, "bp_saved": 0.58,
        "serve_profile": perfil_saque(ace), "match_type": "default_elo"
    }

def merge_stats(base, extra):
    out = base.copy()
    for k, v in extra.items():
        if v is not None:
            out[k] = v
    return out

def leer_archivo_stats(path, tipo):
    stats = {}
    if not os.path.exists(path): return stats

    df = pd.read_excel(path)
    col_player = buscar_columna(df, ["Player", "Name", "Jugador"])
    if col_player is None: return stats

    if tipo == "serve":
        col_ace = buscar_columna(df, ["Ace%", "Ace"])
        col_df = buscar_columna(df, ["DF%", "DF"])
        col_1in = buscar_columna(df, ["1stIn", "1st In"])
        col_1w = buscar_columna(df, ["1st%", "1st Won"])
        col_2w = buscar_columna(df, ["2nd%", "2nd Won"])
        col_hold = buscar_columna(df, ["Hld%", "Hold%", "Hold"])
        for _, row in df.iterrows():
            nombre = normalizar_texto(row.get(col_player, ""))
            nid = limpiar(nombre)
            if not nid: continue
            ace = leer_porcentaje(row.get(col_ace), 0.05)
            stats[nid] = {
                "found_stats": True, "raw_name_stats": nombre,
                "ace": ace,
                "df": leer_porcentaje(row.get(col_df), 0.035),
                "1in": leer_porcentaje(row.get(col_1in), 0.62),
                "1w": leer_porcentaje(row.get(col_1w), 0.70),
                "2w": leer_porcentaje(row.get(col_2w), 0.50),
                "hold": leer_porcentaje(row.get(col_hold), 0.78),
                "serve_profile": perfil_saque(ace),
                "match_type": "serve"
            }

    elif tipo == "return":
        col_rpw = buscar_columna(df, ["RPW", "RPW%", "Return Points Won", "ReturnPointsWon"])
        col_brk = buscar_columna(df, ["Brk%", "Break%", "Break"])
        col_v1 = buscar_columna(df, ["v1st%", "v1st", "vs1st"])
        col_v2 = buscar_columna(df, ["v2nd%", "v2nd", "vs2nd"])
        for _, row in df.iterrows():
            nombre = normalizar_texto(row.get(col_player, ""))
            nid = limpiar(nombre)
            if not nid: continue
            stats[nid] = {
                "found_return": True, "raw_name_return": nombre,
                "rpw": leer_porcentaje(row.get(col_rpw), None) if col_rpw else None,
                "break_pct": leer_porcentaje(row.get(col_brk), None) if col_brk else None,
                "v1st": leer_porcentaje(row.get(col_v1), None) if col_v1 else None,
                "v2nd": leer_porcentaje(row.get(col_v2), None) if col_v2 else None
            }

    elif tipo == "break":
        col_bpconv = buscar_columna(df, ["BPConv%", "BP Conv%", "Break Points Converted"])
        col_bpsaved = buscar_columna(df, ["BPSvd%", "BP Saved%", "Break Points Saved"])
        col_brkset = buscar_columna(df, ["Breaks/Set", "Brk/Set", "Breaks Set"])
        for _, row in df.iterrows():
            nombre = normalizar_texto(row.get(col_player, ""))
            nid = limpiar(nombre)
            if not nid: continue
            stats[nid] = {
                "found_break": True, "raw_name_break": nombre,
                "bp_conv": leer_porcentaje(row.get(col_bpconv), None) if col_bpconv else None,
                "bp_saved": leer_porcentaje(row.get(col_bpsaved), None) if col_bpsaved else None,
                "breaks_set": leer_float(row.get(col_brkset), None) if col_brkset else None
            }

    return stats

def buscar_stats(nombre, stats_map):
    nombre_alias = _ta_alias_resolve(nombre)
    nid = limpiar(nombre_alias)
    if nid in stats_map: return stats_map[nid]
    mejor, best = None, 0
    for sid, data in stats_map.items():
        score = max(similitud_nombre(nombre_alias, sid), similitud_nombre(nombre, sid), similitud_nombre(nombre_alias, data.get("raw_name_stats", sid)), similitud_nombre(nombre, data.get("raw_name_stats", sid)))
        if score > best:
            best, mejor = score, data
    if mejor is not None and best >= 0.72:
        out = mejor.copy()
        out["match_type"] = "aproximado"
        return out
    return None


# =========================================================
# v23.39 WTA REAL DATA UPGRADE
# Usa datos/wta/wta_players.csv + datos/wta/historicos/wta_matches_2026.csv
# como capa de datos WTA. No cambia fórmulas ni motor Over: solo mejora
# perfiles, ranking, muestra por superficie y estadísticas base WTA.
# =========================================================

def _wta_real_match_paths():
    base = os.path.join("datos", "wta")
    patterns = [
        os.path.join(base, "historicos", "wta_matches_2026.csv"),
        os.path.join(base, "wta_matches_2026.csv"),
        os.path.join(base, "matches_2026.csv"),
        os.path.join(base, "historicos", "*wta*matches*.csv"),
        os.path.join(base, "historicos", "*matches*2026*.csv"),
    ]
    found = []
    for pat in patterns:
        if any(ch in pat for ch in "*?["):
            found.extend(glob.glob(pat))
        elif os.path.exists(pat):
            found.append(pat)
    out = []
    seen = set()
    for x in found:
        nx = os.path.normpath(x)
        if nx not in seen and os.path.exists(nx):
            seen.add(nx)
            out.append(nx)
    return out


def _parse_score_games(score):
    txt = normalizar_texto(score).upper()
    if not txt or any(x in txt for x in ["W/O", "WALKOVER", "RET", "DEF", "ABD", "DEFAULT"]):
        return {"valid": False, "games": None, "sets": None, "set3": False}
    pairs = re.findall(r"(\d+)\s*-\s*(\d+)", txt)
    if not pairs:
        return {"valid": False, "games": None, "sets": None, "set3": False}
    games = 0
    sets = 0
    for a, b in pairs:
        try:
            ga, gb = int(a), int(b)
        except Exception:
            continue
        if ga > 20 or gb > 20:
            continue
        games += ga + gb
        sets += 1
    if sets <= 0:
        return {"valid": False, "games": None, "sets": None, "set3": False}
    return {"valid": True, "games": games, "sets": sets, "set3": sets >= 3}


def _safe_div(num, den, default=None):
    try:
        num = float(num)
        den = float(den)
        if den <= 0 or pd.isna(num) or pd.isna(den):
            return default
        return num / den
    except Exception:
        return default


def _wta_empty_bucket(name):
    return {
        "name": normalizar_texto(name),
        "matches_total": 0,
        "wins": 0,
        "losses": 0,
        "rank_values": [],
        "matches_surface": {"Hard": 0, "Clay": 0, "Grass": 0},
        "wins_surface": {"Hard": 0, "Clay": 0, "Grass": 0},
        "games": [],
        "games_surface": {"Hard": [], "Clay": [], "Grass": []},
        "set3": [],
        "set3_surface": {"Hard": [], "Clay": [], "Grass": []},
        "serve_pts": 0.0, "aces": 0.0, "dfs": 0.0, "first_in": 0.0,
        "first_won": 0.0, "second_won": 0.0, "second_pts": 0.0,
        "sv_gms": 0.0, "bp_saved": 0.0, "bp_faced": 0.0,
        "ret_pts": 0.0, "ret_won": 0.0, "return_games": 0.0, "breaks_won": 0.0,
        "serve_surface": {sf: {"serve_pts":0.0,"aces":0.0,"dfs":0.0,"first_in":0.0,"first_won":0.0,"second_won":0.0,"second_pts":0.0,"sv_gms":0.0,"bp_saved":0.0,"bp_faced":0.0,"ret_pts":0.0,"ret_won":0.0,"return_games":0.0,"breaks_won":0.0} for sf in ["Hard","Clay","Grass"]},
    }


def _wta_add_stats(bucket, surface, prefix, opp_prefix, row):
    def v(col):
        try:
            return leer_float(row.get(col), 0.0)
        except Exception:
            return 0.0

    svpt = v(f"{prefix}_svpt")
    first_in = v(f"{prefix}_1stIn")
    second_pts = max(0.0, svpt - first_in)
    serve_updates = {
        "serve_pts": svpt,
        "aces": v(f"{prefix}_ace"),
        "dfs": v(f"{prefix}_df"),
        "first_in": first_in,
        "first_won": v(f"{prefix}_1stWon"),
        "second_won": v(f"{prefix}_2ndWon"),
        "second_pts": second_pts,
        "sv_gms": v(f"{prefix}_SvGms"),
        "bp_saved": v(f"{prefix}_bpSaved"),
        "bp_faced": v(f"{prefix}_bpFaced"),
    }

    opp_svpt = v(f"{opp_prefix}_svpt")
    opp_first_won = v(f"{opp_prefix}_1stWon")
    opp_second_won = v(f"{opp_prefix}_2ndWon")
    opp_svgms = v(f"{opp_prefix}_SvGms")
    opp_bpsaved = v(f"{opp_prefix}_bpSaved")
    opp_bpfaced = v(f"{opp_prefix}_bpFaced")
    return_updates = {
        "ret_pts": opp_svpt,
        "ret_won": max(0.0, opp_svpt - opp_first_won - opp_second_won),
        "return_games": opp_svgms,
        "breaks_won": max(0.0, opp_bpfaced - opp_bpsaved),
    }

    for k, val in {**serve_updates, **return_updates}.items():
        if val and not pd.isna(val):
            bucket[k] += float(val)
            if surface in bucket["serve_surface"]:
                bucket["serve_surface"][surface][k] += float(val)


def _wta_bucket_to_profile(bucket):
    def build_stats(src, fallback_surface="Hard"):
        serve_pts = src.get("serve_pts", 0.0)
        sv_gms = src.get("sv_gms", 0.0)
        bp_faced = src.get("bp_faced", 0.0)
        bp_saved = src.get("bp_saved", 0.0)
        return_games = src.get("return_games", 0.0)
        breaks_won = src.get("breaks_won", 0.0)
        hold = 1.0 - _safe_div(max(0.0, bp_faced - bp_saved), sv_gms, 0.22)
        break_pct = _safe_div(breaks_won, return_games, None)
        stats = {
            "found_stats": bool(serve_pts >= 80 or sv_gms >= 8),
            "raw_name_stats": bucket["name"],
            "hold": float(np.clip(hold if hold is not None else 0.64, 0.46, 0.82)),
            "ace": float(np.clip(_safe_div(src.get("aces", 0.0), serve_pts, 0.035), 0.005, 0.14)),
            "df": float(np.clip(_safe_div(src.get("dfs", 0.0), serve_pts, 0.050), 0.005, 0.16)),
            "1in": float(np.clip(_safe_div(src.get("first_in", 0.0), serve_pts, 0.62), 0.45, 0.82)),
            "1w": float(np.clip(_safe_div(src.get("first_won", 0.0), src.get("first_in", 0.0), 0.64), 0.45, 0.82)),
            "2w": float(np.clip(_safe_div(src.get("second_won", 0.0), src.get("second_pts", 0.0), 0.46), 0.28, 0.66)),
            "rpw": float(np.clip(_safe_div(src.get("ret_won", 0.0), src.get("ret_pts", 0.0), 0.38), 0.18, 0.52)),
            "break_pct": float(np.clip(break_pct if break_pct is not None else 0.34, 0.12, 0.56)),
            "bp_conv": float(np.clip(break_pct if break_pct is not None else 0.38, 0.15, 0.62)),
            "bp_saved": float(np.clip(_safe_div(bp_saved, bp_faced, 0.55), 0.25, 0.82)),
            "serve_profile": "normal",
            "match_type": "wta_real_2026",
        }
        stats["serve_profile"] = perfil_saque(stats.get("ace", 0.035))
        return stats

    total = int(bucket["matches_total"])
    if total <= 0:
        return None
    rank_vals = [x for x in bucket["rank_values"] if x and not pd.isna(x) and float(x) > 0]
    rank = int(np.nanmedian(rank_vals)) if rank_vals else 999
    win_rate = bucket["wins"] / total if total else 0.50
    games_avg = float(np.nanmean(bucket["games"])) if bucket["games"] else 19.5
    set3_rate = float(np.nanmean(bucket["set3"])) if bucket["set3"] else 0.30

    stats_general = build_stats(bucket)
    stats_by_surface = {}
    for sf in ["Hard", "Clay", "Grass"]:
        src = bucket["serve_surface"].get(sf, {})
        if bucket["matches_surface"].get(sf, 0) >= 3:
            stats_by_surface[sf] = build_stats(src, sf)
        else:
            stats_by_surface[sf] = stats_general.copy()
            stats_by_surface[sf]["match_type"] = "wta_real_2026_general_fallback"

    stability = {}
    confidence = {}
    for sf in ["Hard", "Clay", "Grass"]:
        n_sf = int(bucket["matches_surface"].get(sf, 0))
        sample_score = np.sqrt(n_sf / (n_sf + 10)) if n_sf > 0 else 0.0
        total_score = np.sqrt(total / (total + 20))
        stab = (sample_score * 0.75) + (total_score * 0.25)
        conf = 0.42 + 0.58 * stab
        if n_sf < 3:
            conf -= 0.18
        elif n_sf < 6:
            conf -= 0.09
        confidence[sf] = float(np.clip(conf, 0.30, 0.96))
        stability[sf] = float(np.clip(stab, 0.05, 1.00))

    return {
        "Player": bucket["name"],
        "Rank": rank,
        "Stats": stats_general,
        "StatsBySurface": stats_by_surface,
        "Quality": {
            "matches_total": total,
            "matches_surface": {k:int(v) for k, v in bucket["matches_surface"].items()},
            "level_counts": {"tour": total, "challenger": 0, "itf": 0, "qualy": 0, "unknown": 0},
            "tour_quality": 0.92,
            "stability": stability,
            "confidence": confidence,
            "raw_names": [bucket["name"]],
            "aliases": sorted(list(variantes_nombre_quality(bucket["name"]))),
            "source_files": _wta_real_match_paths(),
            "matched_name": bucket["name"],
            "match_score": 1.0,
        },
        "WTARealProfile": {
            "active": True,
            "matches_total": total,
            "rank": rank,
            "win_rate_2026": float(win_rate),
            "games_avg": float(games_avg),
            "set3_rate": float(set3_rate),
            "over17_rate": float(np.mean([g > 17.5 for g in bucket["games"]])) if bucket["games"] else None,
            "over18_rate": float(np.mean([g > 18.5 for g in bucket["games"]])) if bucket["games"] else None,
        }
    }


@st.cache_data(show_spinner=False)
def cargar_wta_real_profiles(cache_version=APP_VERSION):
    paths = _wta_real_match_paths()
    if not paths:
        return {}

    buckets = {}
    for path in paths:
        try:
            df = pd.read_csv(path, low_memory=False)
        except Exception:
            continue
        if df.empty:
            continue
        for _, row in df.iterrows():
            surface = normalizar_superficie_hist(row.get("surface", "Hard"), default="Hard")
            if surface not in ["Hard", "Clay", "Grass"]:
                surface = "Hard"
            score_info = _parse_score_games(row.get("score", ""))
            winner = normalizar_texto(row.get("winner_name", ""))
            loser = normalizar_texto(row.get("loser_name", ""))
            if not winner or not loser:
                continue
            for name, won, rank_col, prefix, opp_prefix in [
                (winner, True, "winner_rank", "w", "l"),
                (loser, False, "loser_rank", "l", "w"),
            ]:
                key = limpiar(name)
                if not key:
                    continue
                b = buckets.setdefault(key, _wta_empty_bucket(name))
                b["matches_total"] += 1
                b["wins"] += int(bool(won))
                b["losses"] += int(not won)
                b["matches_surface"][surface] = b["matches_surface"].get(surface, 0) + 1
                if won:
                    b["wins_surface"][surface] = b["wins_surface"].get(surface, 0) + 1
                rank_v = leer_float(row.get(rank_col), None)
                if rank_v is not None and not pd.isna(rank_v):
                    b["rank_values"].append(rank_v)
                if score_info.get("valid"):
                    b["games"].append(float(score_info["games"]))
                    b["games_surface"][surface].append(float(score_info["games"]))
                    b["set3"].append(1.0 if score_info.get("set3") else 0.0)
                    b["set3_surface"][surface].append(1.0 if score_info.get("set3") else 0.0)
                _wta_add_stats(b, surface, prefix, opp_prefix, row)

    profiles = {}
    for key, bucket in buckets.items():
        prof = _wta_bucket_to_profile(bucket)
        if prof:
            profiles[key] = prof
    return profiles


def _wta_rank_to_elo(rank, win_rate=0.50, surface_boost=0.0):
    try:
        r = float(rank)
    except Exception:
        r = 999.0
    if r <= 0 or pd.isna(r):
        r = 999.0
    # Escala conservadora para WTA. Evita ratings extremos en jugadoras nuevas.
    elo = 1890.0 - 2.15 * min(max(r - 1.0, 0.0), 260.0)
    elo += (float(win_rate) - 0.50) * 90.0
    elo += surface_boost
    return float(np.clip(elo, 1320, 1915))


def aplicar_wta_real_data_upgrade(players):
    profiles = cargar_wta_real_profiles()
    if not profiles:
        return players

    out = dict(players or {})
    existing_clean = {limpiar(k): k for k in out.keys()}

    # 1) Enriquecer jugadoras ya existentes en wta_elo.xlsx.
    for clean_name, app_name in list(existing_clean.items()):
        prof = profiles.get(clean_name)
        if prof is None:
            # Fuzzy controlado para diferencias de formato.
            best_key, best_score = None, 0.0
            for pk, pd_prof in profiles.items():
                sc = similitud_nombre(app_name, pd_prof.get("Player", pk))
                if sc > best_score:
                    best_key, best_score = pk, sc
            if best_key and best_score >= 0.92:
                prof = profiles.get(best_key)
        if prof is None:
            continue

        d = out[app_name].copy()
        d["Rank"] = int(min(d.get("Rank", 999), prof.get("Rank", 999)))
        d["Stats"] = merge_surface_stats(d.get("Stats", {}), prof.get("Stats", {}))
        ss = dict(d.get("StatsBySurface", {}))
        for sf in ["Hard", "Clay", "Grass"]:
            current = ss.get(sf, d.get("Stats", {}))
            ss[sf] = merge_surface_stats(current, prof.get("StatsBySurface", {}).get(sf, {}))
            ss[sf]["match_type"] = prof.get("StatsBySurface", {}).get(sf, {}).get("match_type", "wta_real_2026")
        d["StatsBySurface"] = ss
        d["Quality"] = prof.get("Quality", d.get("Quality", {}))
        d["WTARealProfile"] = prof.get("WTARealProfile", {})
        d["FuenteDB"] = "WTA real 2026"
        out[app_name] = d

    # 2) Añadir jugadoras que no existen en el Excel WTA pero sí en los partidos 2026.
    existing_clean = {limpiar(k) for k in out.keys()}
    for key, prof in profiles.items():
        if key in existing_clean:
            continue
        name = prof.get("Player", key)
        rank = prof.get("Rank", 999)
        wta_meta = prof.get("WTARealProfile", {}) or {}
        wr = wta_meta.get("win_rate_2026", 0.50) or 0.50
        surf_counts = prof.get("Quality", {}).get("matches_surface", {}) or {}
        hard_boost = 8 if surf_counts.get("Hard", 0) >= surf_counts.get("Clay", 0) else 0
        clay_boost = 8 if surf_counts.get("Clay", 0) > surf_counts.get("Hard", 0) else 0
        base = _wta_rank_to_elo(rank, wr, 0)
        out[name] = {
            "Player": name,
            "Rank": int(rank),
            "Hard": _wta_rank_to_elo(rank, wr, hard_boost),
            "Clay": _wta_rank_to_elo(rank, wr, clay_boost),
            "Grass": base,
            "Stats": prof.get("Stats", stats_default_por_elo(base, rank, "Hard", "WTA")),
            "StatsBySurface": prof.get("StatsBySurface", {}),
            "Fatigue": {"matches_14d": 0, "matches_7d": 0, "b2b": False, "fatigue_score": 0.0},
            "Quality": prof.get("Quality", {}),
            "WTARealProfile": wta_meta,
            "FuenteDB": "WTA real 2026 fallback"
        }
        existing_clean.add(key)

    return out

# =========================================================
# v22.3 MATCH COUNT FIX + TOUR QUALITY ENGINE
# =========================================================

def detectar_nivel_torneo(row):
    """
    v22.17 Tour Quality + Challenger CSV Fix.

    Soporta dos familias de históricos:
    1) Excel tipo tennis-data ATP Tour con columnas Series/Tournament/Round.
    2) CSV Jeff Sackmann / tennis-data qual_chall con columnas:
       tourney_level, tourney_name, round, winner_name, loser_name.

    Reglas:
    - tourney_level = C => Challenger.
    - round Q1/Q2/Q3/QR... o winner_entry/loser_entry=Q => Qualy, salvo Challenger explícito.
    - tourney_level = G/M/A/F/D => Tour si NO es qualy.
    - SourceFile con qual_chall no convierte por sí solo; solo ayuda al diagnóstico.
    - Si no hay evidencia clara, Unknown.
    """
    def val(*names):
        for c in names:
            try:
                if c in row.index:
                    v = normalizar_texto(row.get(c, ""))
                    if v != "" and v.lower() != "nan":
                        return v
            except Exception:
                pass
        return ""

    # Columnas estándar Jeff Sackmann / tennis-data CSV.
    tl_raw = val("MC_Level", "tourney_level", "Tourney Level", "level", "Level", "Series")
    round_raw = val("MC_Round", "round", "Round", "ronda")
    entry_raw = " ".join([
        val("winner_entry", "Winner Entry", "w_entry"),
        val("loser_entry", "Loser Entry", "l_entry")
    ])
    tourney_raw = val("MC_Tournament", "tourney_name", "Tournament", "Event", "Location")
    source_raw = val("SourceFile", "source_file")

    tl = limpiar(tl_raw)
    rnd = limpiar(round_raw)
    entry = limpiar(entry_raw)

    # 1) Challenger explícito por tourney_level C. Esto es lo más importante para tus CSV qual_chall.
    if tl in ["C", "CH", "CHALLENGER", "ATPCHALLENGER"]:
        return "challenger"

    # 2) Qualy explícita. En CSV qual_chall los partidos de previa ATP suelen venir como tourney_level=A + round=Q1/Q2.
    if rnd in ["Q", "Q1", "Q2", "Q3", "Q4", "QR", "QR1", "QR2", "QR3"]:
        return "qualy"
    if entry in ["Q", "WCQ"] or "QUAL" in entry:
        return "qualy"

    # 3) Tour explícito por código o texto.
    if tl in ["G", "M", "A", "F", "D", "ATPM", "ATP", "ATP250", "ATP500", "MASTERS1000", "GRANDSLAM"]:
        return "tour"

    def _txt(cols):
        return " ".join(str(row.get(c, "")) for c in cols if c in getattr(row, 'index', []))

    level_txt = _txt(["MC_Level", "Series", "Level", "Tour", "Category", "Circuit", "Event Type", "tourney_level"])
    event_txt = _txt(["MC_Tournament", "MC_Round", "Tournament", "Event", "Location", "Round", "Comment", "tourney_name", "round"])
    source_txt = _txt(["SourceFile", "source_file"])

    all_txt = f" {level_txt} {event_txt} {source_txt} ".lower()
    all_clean = " " + re.sub(r"[^a-z0-9]+", " ", all_txt) + " "
    level_clean = " " + re.sub(r"[^a-z0-9]+", " ", level_txt.lower()) + " "
    event_clean = " " + re.sub(r"[^a-z0-9]+", " ", event_txt.lower()) + " "

    # Qualy por texto.
    qualy_tokens = [
        " qual ", " qualifying ", " qualification ", " qualies ",
        " q1 ", " q2 ", " q3 ", " qr1 ", " qr2 ", " qr3 ",
        " qualifying draw "
    ]
    if any(x in all_clean for x in qualy_tokens):
        return "qualy"

    challenger_tokens = [
        " challenger ", " atp challenger ", " wta 125 ", " wta125 ",
        " ch50 ", " ch 50 ", " ch75 ", " ch 75 ", " ch100 ", " ch 100 ",
        " ch125 ", " ch 125 ", " ch175 ", " ch 175 "
    ]
    if any(x in all_clean for x in challenger_tokens):
        return "challenger"

    itf_tokens = [
        " itf ", " futures ", " future ", " m15 ", " m25 ", " m35 ",
        " w15 ", " w25 ", " w35 ", " w50 ", " w75 ", " w100 "
    ]
    if any(x in all_clean for x in itf_tokens):
        return "itf"

    tour_tokens = [
        " grand slam ", " grandslam ", " gs ",
        " masters 1000 ", " atp masters ", " wta 1000 ", " wta1000 ", " atp1000 ", " atp 1000 ",
        " atp 500 ", " atp500 ", " wta 500 ", " wta500 ",
        " atp 250 ", " atp250 ", " wta 250 ", " wta250 ",
        " atp finals ", " wta finals ", " main tour ", " tour level ",
        " olympics ", " united cup ", " davis cup ", " billie jean king cup "
    ]
    if any(x in level_clean for x in tour_tokens):
        return "tour"

    major_tour_events = [
        " australian open ", " roland garros ", " french open ", " wimbledon ", " us open ",
        " monte carlo ", " indian wells ", " miami ", " madrid ", " rome ", " roma ",
        " canada masters ", " cincinnati ", " shanghai ", " paris masters ",
        " queens ", " halle ", " basel ", " vienna ", " barcelona ", " doha ",
        " dubai ", " acapulco ", " rotterdam ", " beijing ", " tokyo "
    ]
    if any(x in event_clean for x in major_tour_events):
        return "tour"

    return "unknown"

def normalizar_superficie_hist(v, default="Hard"):
    """Normaliza superficies históricas: C/Clay/Red Clay, H/Hard, G/Grass."""
    raw = normalizar_texto(v).strip()
    key = limpiar(raw)
    if key in ["C", "CL", "CLAY", "REDCLAY", "RCLAY", "TIERRA", "TIERRABATIDA"] or "CLAY" in key:
        return "Clay"
    if key in ["G", "GR", "GRASS", "LAWN", "HIERBA"] or "GRASS" in key:
        return "Grass"
    if key in ["H", "HD", "HARD", "HARDO", "INDOORHARD", "OUTDOORHARD"] or "HARD" in key:
        return "Hard"
    if key in ["CARPET", "CARPETA"]:
        return "Hard"
    return default




def detectar_superficie_quality(row, col_surface=None):
    """Detecta superficie desde columna Surface o, si falta, desde Tournament/SourceFile."""
    primary = row.get(col_surface, "") if col_surface else ""
    surf = normalizar_superficie_hist(primary, default="")
    if surf in ["Hard", "Clay", "Grass"]:
        return surf

    txt = " ".join(str(row.get(c, "")) for c in [
        "MC_Surface", "Surface", "surface", "Superficie", "Court", "Court Surface", "MC_Tournament", "Tournament", "Event", "SourceFile", "Location"
    ])
    key = limpiar(txt)
    if any(x in key for x in ["CLAY", "REDCLAY", "TIERRA", "RG", "ROLANDGARROS", "MONTECARLO", "MADRID", "ROMA", "ROME"]):
        return "Clay"
    if any(x in key for x in ["GRASS", "HIERBA", "WIMBLEDON", "HALLE", "QUEENS", "NEWPORT"]):
        return "Grass"
    return "Hard"

def apellido_inicial_key(nombre):
    """Clave canónica tipo tennis-data: Matteo Arnaldi <-> Arnaldi M.
    Devuelve ARNALDI_M. Evita que el Match Count dependa del orden exacto.
    """
    toks = tokens(nombre)
    if len(toks) < 2:
        return ""

    # Si viene como "Bublik A." o "Arnaldi M.": apellido + inicial.
    if len(toks[-1]) == 1:
        surname = toks[0]
        initial = toks[-1][0]
    else:
        # Si viene como "Matteo Arnaldi": último token = apellido, primero = inicial.
        surname = toks[-1]
        initial = toks[0][0]

    surname = limpiar(surname)
    initial = limpiar(initial)
    return f"{surname}_{initial}" if surname and initial else ""


def surname_key(nombre):
    toks = tokens(nombre)
    if not toks:
        return ""
    if len(toks) >= 2 and len(toks[-1]) == 1:
        return limpiar(toks[0])
    return limpiar(toks[-1])


def variantes_nombre_quality(nombre):
    """
    Genera claves de búsqueda robustas para históricos:
    - Matteo Arnaldi
    - Arnaldi Matteo
    - Arnaldi M / M Arnaldi
    - clave canónica ARNALDI_M
    - MATTEOARNALDI limpio
    """
    raw = normalizar_texto(nombre)
    toks = tokens(raw)
    out = set()
    clean = limpiar(raw)
    if clean:
        out.add(clean)

    ai = apellido_inicial_key(raw)
    if ai:
        out.add(ai)
        out.add(limpiar(ai))

    sk = surname_key(raw)
    if sk:
        out.add(sk)

    if len(toks) >= 2:
        # Formato normal completo
        first, last = toks[0], toks[-1]
        out.add(limpiar(f"{last} {first}"))
        out.add(limpiar(f"{last} {first[0]}"))
        out.add(limpiar(f"{first[0]} {last}"))
        out.add(limpiar(f"{last}{first[0]}"))
        out.add(limpiar(f"{first[0]}{last}"))

        # Si el histórico viene abreviado tipo "Arnaldi M.", toks[0]=apellido, toks[1]=inicial.
        if len(toks[-1]) == 1:
            surname, initial = toks[0], toks[-1][0]
            out.add(limpiar(f"{surname} {initial}"))
            out.add(limpiar(f"{initial} {surname}"))
            out.add(limpiar(f"{surname}{initial}"))
            out.add(limpiar(f"{initial}{surname}"))

    return {x for x in out if x}


def fusionar_quality_rows(items):
    """Fusiona varias identidades históricas que pertenecen al mismo jugador."""
    rows = []
    raw_names = set()
    aliases = set()
    source_files = set()
    for item in items:
        rows.extend(item.get("_rows", []))
        raw_names.update(item.get("raw_names", []))
        aliases.update(item.get("aliases", []))
        source_files.update(item.get("source_files", []))

    total = len(rows)
    if total <= 0:
        return None

    surface_counts = {sf: sum(1 for r in rows if r["surface"] == sf) for sf in ["Hard", "Clay", "Grass"]}
    level_counts = {lv: sum(1 for r in rows if r["level"] == lv) for lv in ["tour", "challenger", "itf", "qualy", "unknown"]}

    tour_weighted = (
        level_counts["tour"] * 1.00 +
        level_counts["challenger"] * 0.62 +
        level_counts["qualy"] * 0.50 +
        level_counts["itf"] * 0.32 +
        level_counts["unknown"] * 0.42
    ) / total

    out = {
        "matches_total": total,
        "matches_surface": surface_counts,
        "level_counts": level_counts,
        "tour_quality": float(np.clip(tour_weighted, 0.30, 1.00)),
        "raw_names": sorted(raw_names)[:12],
        "aliases": sorted(aliases),
        "source_files": sorted(source_files)[:8],
        "_rows": rows,
    }

    stability = {}
    confidence = {}
    for sf in ["Hard", "Clay", "Grass"]:
        n_sf = surface_counts.get(sf, 0)
        sample_score = np.sqrt(n_sf / (n_sf + 14)) if n_sf > 0 else 0.0
        total_score = np.sqrt(total / (total + 28))
        stab = (sample_score * 0.72) + (total_score * 0.28)
        conf = 0.38 + 0.62 * stab * out["tour_quality"]

        if n_sf < 5:
            conf -= 0.18
        elif n_sf < 10:
            conf -= 0.09
        elif n_sf < 18:
            conf -= 0.04

        if level_counts["tour"] == 0 and level_counts["challenger"] >= 1:
            conf -= 0.07
        if level_counts["itf"] >= max(3, total * 0.45):
            conf -= 0.10
        if level_counts["unknown"] >= max(4, total * 0.50):
            conf -= 0.08

        stability[sf] = float(np.clip(stab, 0.05, 1.00))
        confidence[sf] = float(np.clip(conf, 0.28, 1.00))

    out["stability"] = stability
    out["confidence"] = confidence
    return out


@st.cache_data(show_spinner=False)
def crear_quality_map(circuito, cache_version=QUALITY_ENGINE_VERSION):
    """
    v22.3 Match Count Fix.
    Cuenta partidos desde históricos aunque los nombres vengan en formatos distintos.
    Devuelve también raw_names/source_files para debug visual.
    """
    hist = cargar_historicos(circuito, cache_version=cache_version)
    quality = {}

    if hist.empty:
        return quality

    df = hist.copy()
    if "Comment" in df.columns:
        # v22.17: filtro por Comment sin borrar filas CSV Challenger/Qualy.
        # Al concatenar Excel Tour + CSV qual_chall, las filas CSV suelen tener Comment vacío/NaN.
        # Antes, como los Excel sí tenían Comment=Completed, se aplicaba df[completed_mask]
        # y se eliminaban TODOS los CSV. Ahora solo filtramos filas que realmente traen Comment.
        comment_txt = df["Comment"].apply(normalizar_texto).replace({"nan": "", "NaN": "", "None": ""})
        has_comment = comment_txt.str.strip().ne("")
        completed_mask = comment_txt.str.contains("Completed", case=False, na=False)
        if has_comment.any() and completed_mask.sum() > 0:
            df = df[(~has_comment) | completed_mask]

    col_winner = buscar_columna(df, ["MC_Winner", "Winner", "Ganador", "Player1", "Player 1", "WName", "winner_name", "winner_name_clean", "Jugador1", "Home", "P1"])
    col_loser = buscar_columna(df, ["MC_Loser", "Loser", "Perdedor", "Player2", "Player 2", "LName", "loser_name", "loser_name_clean", "Jugador2", "Away", "P2"])
    col_surface = buscar_columna(df, ["MC_Surface", "Surface", "surface", "Superficie", "Court Surface", "Court", "surface_name"])

    if col_winner is None or col_loser is None:
        return quality

    player_buckets = {}

    for _, row in df.iterrows():
        level = detectar_nivel_torneo(row)
        surface = detectar_superficie_quality(row, col_surface)
        source = normalizar_texto(row.get("SourceFile", ""))

        for player in [normalizar_texto(row.get(col_winner, "")), normalizar_texto(row.get(col_loser, ""))]:
            if not player:
                continue
            pid = limpiar(player)
            if not pid:
                continue
            bucket = player_buckets.setdefault(pid, {
                "_rows": [], "raw_names": set(), "aliases": set(), "source_files": set()
            })
            bucket["_rows"].append({"surface": surface, "level": level})
            bucket["raw_names"].add(player)
            bucket["aliases"].update(variantes_nombre_quality(player))
            ai = apellido_inicial_key(player)
            if ai:
                bucket["aliases"].add(ai)
                bucket["aliases"].add(limpiar(ai))
            if source:
                bucket["source_files"].add(source)

    # Primera pasada: identidad exacta por PID histórico.
    alias_usage = {}
    for pid, data in player_buckets.items():
        for alias in data.get("aliases", []):
            alias_usage.setdefault(str(alias), set()).add(pid)

    for pid, data in player_buckets.items():
        q = fusionar_quality_rows([data])
        if q:
            quality[pid] = q
            # Índice espejo por clave apellido_inicial.
            # Esto permite casar Matteo Arnaldi con Arnaldi M. sin depender de fuzzy matching.
            for alias in data.get("aliases", []):
                alias = str(alias)
                # Guardamos alias fuertes siempre y alias simples solo si no chocan con otro jugador.
                if "_" in alias or len(alias_usage.get(alias, set())) == 1:
                    quality.setdefault(alias, q)
                    quality.setdefault(limpiar(alias), q)

    # Meta interna para debug/cache.
    raw_player_count = len(player_buckets)
    quality["__meta__"] = {
        "version": QUALITY_ENGINE_VERSION,
        "raw_player_count": raw_player_count,
        "quality_keys": len([k for k in quality.keys() if not str(k).startswith("__")]),
        "sample_names": sorted([next(iter(v.get("raw_names", [k])), k) if isinstance(v.get("raw_names", []), list) and v.get("raw_names", []) else str(k) for k, v in list(quality.items())[:12] if not str(k).startswith("__")])[:8],
    }

    return quality



def closest_quality_candidates(nombre, quality_map, limit=8):
    """
    v22.6 Name Radar Pro.
    Devuelve candidatos aunque el jugador no exista, y da prioridad a:
    - clave apellido_inicial exacta: ARNALDI_M
    - mismo apellido
    - fuzzy score normal
    """
    candidates = []
    user_ai = apellido_inicial_key(nombre)
    user_surname = surname_key(nombre)

    seen_objs = set()
    for qid, data in quality_map.items():
        if str(qid).startswith("__"):
            continue
        # Evita duplicados por índices espejo.
        obj_id = id(data)
        if obj_id in seen_objs:
            continue
        seen_objs.add(obj_id)

        labels = [qid] + list(data.get("raw_names", [])) + list(data.get("aliases", []))
        best_score = 0.0
        best_label = qid
        reason = "fuzzy"

        for cand in labels:
            cand_ai = apellido_inicial_key(cand)
            cand_surname = surname_key(cand)
            sc = similitud_nombre(nombre, cand)

            if user_ai and cand_ai and user_ai == cand_ai:
                sc = max(sc, 0.995)
                reason = "apellido+inicial"
            elif user_surname and cand_surname and user_surname == cand_surname:
                sc = max(sc, 0.86)
                reason = "apellido"
            elif user_surname and cand_surname and user_surname != cand_surname:
                # v22.21: radar estricto, no elevar candidatos con otro apellido por compartir nombre común.
                sc = min(sc, 0.89)

            t_user = tokens(nombre)
            t_cand = tokens(cand)
            if len(t_user) >= 2 and len(t_cand) >= 1:
                user_first = t_user[0]
                user_last = t_user[-1]
                if user_last in t_cand and any(x.startswith(user_first[0]) for x in t_cand if x != user_last):
                    sc = max(sc, 0.94)
                    reason = "apellido+inicial"
                if user_last in t_cand and user_first in t_cand:
                    sc = max(sc, 0.98)
                    reason = "nombre completo"

            if sc > best_score:
                best_score = sc
                best_label = cand

        if best_score >= 0.18 or (user_surname and user_surname in " ".join(map(str, labels))):
            candidates.append({
                "name": best_label,
                "score": float(best_score),
                "reason": reason,
                "matches_total": int(data.get("matches_total", 0)),
                "surface": data.get("matches_surface", {}),
                "source_files": data.get("source_files", [])[:3],
            })
    candidates.sort(key=lambda x: (x["score"], x["matches_total"]), reverse=True)
    return candidates[:limit]



def nombre_score_quality_direct(objetivo, candidato):
    """
    v22.21 Name Match Strict.
    Evita falsos positivos tipo:
    - Martin Landaluce -> Alvaro Lopez San Martin

    Reglas:
    - Exacto limpio: 1.00
    - Apellido + inicial exactos: 0.995
    - Mismo apellido + inicial compatible: 0.96
    - Mismo apellido sin inicial clara: 0.84
    - Fuzzy sin mismo apellido queda capado a 0.89, aunque SequenceMatcher dé alto.
    """
    if not objetivo or not candidato:
        return 0.0, "empty"

    n1, n2 = limpiar(objetivo), limpiar(candidato)
    if n1 and n1 == n2:
        return 1.0, "exacto"

    ai1, ai2 = apellido_inicial_key(objetivo), apellido_inicial_key(candidato)
    if ai1 and ai2 and ai1 == ai2:
        return 0.995, "apellido+inicial"

    s1, s2 = surname_key(objetivo), surname_key(candidato)
    t1, t2 = tokens(objetivo), tokens(candidato)

    if s1 and s2 and s1 == s2:
        initials1 = {x[0] for x in t1 if x and limpiar(x) != s1}
        initials2 = {x[0] for x in t2 if x and limpiar(x) != s2}
        if initials1 and initials2 and initials1 & initials2:
            return 0.96, "apellido+inicial"
        return 0.84, "apellido"

    # Si el apellido objetivo no aparece como token del candidato, NO permitimos match fuerte.
    # Esto bloquea falsos positivos por nombres comunes: Martin, Alejandro, Rafael, etc.
    sc = float(similitud_nombre(objetivo, candidato))
    if s1 and s1 not in {limpiar(x) for x in t2}:
        sc = min(sc, 0.89)

    return sc, "fuzzy"


@st.cache_data(show_spinner=False)
def buscar_quality_directo_historicos(nombre, circuito, cache_version=QUALITY_ENGINE_VERSION):
    """
    v22.9 Direct Match Count + Tour Quality Fix.
    No depende del quality_map guardado en cargar_datos/cache.
    Lee históricos cargados y cuenta el jugador en Winner/Loser al vuelo.
    """
    hist = cargar_historicos(circuito, cache_version=cache_version)
    meta = {
        "version": QUALITY_ENGINE_VERSION,
        "mode": "direct_historicos_cached",
        "hist_rows": int(len(hist)) if hist is not None else 0,
        "raw_player_count": 0,
        "quality_keys": 0,
        "sample_names": [],
    }

    fallback = {
        "matches_total": 0,
        "matches_surface": {"Hard": 0, "Clay": 0, "Grass": 0},
        "level_counts": {"tour": 0, "challenger": 0, "itf": 0, "qualy": 0, "unknown": 0},
        "tour_quality": 0.45,
        "stability": {"Hard": 0.05, "Clay": 0.05, "Grass": 0.05},
        "confidence": {"Hard": 0.32, "Clay": 0.32, "Grass": 0.32},
        "raw_names": [],
        "aliases": [],
        "source_files": [],
        "matched_name": "NO ENCONTRADO",
        "match_score": 0.0,
        "quality_map_size": 0,
        "quality_meta": meta,
        "candidate_matches": [],
        "direct_engine": True,
    }

    if hist is None or hist.empty:
        return fallback

    df = hist.copy()
    if "Comment" in df.columns:
        # v22.18: conservar CSV Challenger/Qualy aunque no traigan Comment.
        comment_txt = df["Comment"].apply(normalizar_texto).replace({"nan": "", "NaN": "", "None": ""})
        has_comment = comment_txt.str.strip().ne("")
        completed_mask = comment_txt.str.contains("Completed", case=False, na=False)
        if has_comment.any() and completed_mask.sum() > 0:
            df = df[(~has_comment) | completed_mask]

    col_winner = buscar_columna(df, ["MC_Winner", "Winner", "Ganador", "Player1", "Player 1", "WName", "winner_name", "winner_name_clean", "Jugador1", "Home", "P1"])
    col_loser = buscar_columna(df, ["MC_Loser", "Loser", "Perdedor", "Player2", "Player 2", "LName", "loser_name", "loser_name_clean", "Jugador2", "Away", "P2"])
    col_surface = buscar_columna(df, ["MC_Surface", "Surface", "surface", "Superficie", "Court Surface", "Court", "surface_name"])

    if col_winner is None or col_loser is None:
        return fallback

    rows = []
    raw_names = set()
    aliases = set()
    source_files = set()
    candidate_best = {}
    unique_players = set()
    best_score = 0.0
    best_name = "NO ENCONTRADO"

    # Umbral algo flexible: apellido+inicial entra seguro; fuzzy puro exige más.
    ACCEPT_SCORE = 0.92

    for _, row in df.iterrows():
        for col in [col_winner, col_loser]:
            cand = normalizar_texto(row.get(col, ""))
            if not cand:
                continue
            unique_players.add(cand)
            sc, reason = nombre_score_quality_direct(nombre, cand)

            # Guardamos radar de candidatos por nombre histórico.
            old = candidate_best.get(cand)
            if old is None or sc > old["score"]:
                candidate_best[cand] = {
                    "name": cand,
                    "score": float(sc),
                    "reason": reason,
                    "matches_total": 0,
                    "surface": {"Hard": 0, "Clay": 0, "Grass": 0},
                    "source_files": [],
                }

            if sc > best_score:
                best_score = float(sc)
                best_name = cand

            if sc >= ACCEPT_SCORE:
                surface = detectar_superficie_quality(row, col_surface)
                level = detectar_nivel_torneo(row)
                rows.append({"surface": surface, "level": level})
                raw_names.add(cand)
                aliases.update(variantes_nombre_quality(cand))
                ai = apellido_inicial_key(cand)
                if ai:
                    aliases.add(ai)
                    aliases.add(limpiar(ai))
                src = normalizar_texto(row.get("SourceFile", ""))
                if src:
                    source_files.add(src)

                cb = candidate_best[cand]
                cb["matches_total"] += 1
                cb["surface"][surface] = cb["surface"].get(surface, 0) + 1
                if src and src not in cb["source_files"]:
                    cb["source_files"].append(src)

    meta["raw_player_count"] = int(len(unique_players))
    meta["quality_keys"] = int(len(unique_players))
    meta["sample_names"] = sorted(list(unique_players))[:8]
    fallback["quality_meta"] = meta
    fallback["quality_map_size"] = int(len(unique_players))

    radar = sorted(candidate_best.values(), key=lambda x: (x["score"], x["matches_total"]), reverse=True)[:8]
    fallback["candidate_matches"] = radar
    if best_score > 0:
        fallback["matched_name"] = best_name
        fallback["match_score"] = float(best_score)

    if not rows:
        return fallback

    q = fusionar_quality_rows([{
        "_rows": rows,
        "raw_names": raw_names,
        "aliases": aliases,
        "source_files": source_files,
    }])
    if not q:
        return fallback

    q["matched_name"] = sorted(raw_names)[0] if raw_names else best_name
    q["match_score"] = float(best_score)
    q["quality_map_size"] = int(len(unique_players))
    q["quality_meta"] = meta
    q["candidate_matches"] = radar
    q["direct_engine"] = True
    return q

def buscar_quality(nombre, quality_map):
    """
    v22.3: busca primero por alias fuerte y después por similitud contra nombres reales.
    """
    nid = limpiar(nombre)
    variants = variantes_nombre_quality(nombre)
    ai_key = apellido_inicial_key(nombre)

    # Exacto o clave canónica apellido_inicial.
    for key in [nid, ai_key, limpiar(ai_key)]:
        if key and key in quality_map:
            out = quality_map[key].copy()
            raw = out.get("raw_names", [])
            out["matched_name"] = raw[0] if raw else key
            out["match_score"] = 1.0
            return out

    alias_hits = []
    for qid, data in quality_map.items():
        aliases = set(data.get("aliases", [])) | {qid}
        if variants & aliases:
            alias_hits.append((qid, data))

    if alias_hits:
        fused = fusionar_quality_rows([x[1] for x in alias_hits])
        if fused:
            fused["matched_name"] = ", ".join(sorted({x[0] for x in alias_hits})[:3])
            fused["match_score"] = 0.96
            return fused

    mejor, best = None, 0.0
    best_label = ""
    for qid, data in quality_map.items():
        candidates = [qid] + data.get("raw_names", []) + data.get("aliases", [])
        for cand in candidates:
            score = similitud_nombre(nombre, cand)
            # Refuerzo apellido+inicial para formatos tipo ARNALDI M.
            t_user = tokens(nombre)
            t_cand = tokens(cand)
            if len(t_user) >= 2 and len(t_cand) >= 1:
                user_last = t_user[-1]
                user_first_initial = t_user[0][0]
                cand_txt = " ".join(t_cand)
                if user_last in t_cand and user_first_initial in cand_txt:
                    score = max(score, 0.94)
            if score > best:
                best = score
                mejor = data
                best_label = cand

    if mejor is not None and best >= 0.72:
        out = mejor.copy()
        out["matched_name"] = best_label
        out["match_score"] = float(best)
        return out

    # No encontrado: devolvemos radar de candidatos para depurar el histórico.
    return {
        "matches_total": 0,
        "matches_surface": {"Hard": 0, "Clay": 0, "Grass": 0},
        "level_counts": {"tour": 0, "challenger": 0, "itf": 0, "qualy": 0, "unknown": 0},
        "tour_quality": 0.45,
        "stability": {"Hard": 0.05, "Clay": 0.05, "Grass": 0.05},
        "confidence": {"Hard": 0.32, "Clay": 0.32, "Grass": 0.32},
        "raw_names": [],
        "aliases": [],
        "source_files": [],
        "matched_name": "NO ENCONTRADO",
        "match_score": 0.0,
        "quality_map_size": len([k for k in quality_map.keys() if not str(k).startswith("__")]),
        "quality_meta": quality_map.get("__meta__", {}),
        "candidate_matches": closest_quality_candidates(nombre, quality_map, limit=8),
    }



# =========================================================
# v23.42.0 AUTO PROFILE FINDER + CACHE
# Objetivo: evitar copiar/pegar perfiles manuales cada día.
# - Detecta jugadores del plan global / tanda actual.
# - Busca primero en caché local.
# - Opcional: descarga ranking Elo TennisAbstract y guarda coincidencias.
# - Integra la caché en cargar_datos() sin tocar motor Over.
# =========================================================

AUTO_PROFILE_VERSION = "v23.42.0-auto-profile-finder"
AUTO_PROFILE_DIR = os.path.join("datos", "auto_profiles")


def _auto_profile_ensure_dir_v23420():
    try:
        os.makedirs(AUTO_PROFILE_DIR, exist_ok=True)
    except Exception:
        pass


def _auto_profile_circuit_key_v23420(circuito):
    c = str(circuito or "ATP").upper().strip()
    if c == "CHALLENGER" or c == "CHALLENGER_ATP":
        return "CHALLENGER"
    if c == "WTA":
        return "WTA"
    return "ATP"


def _auto_profile_path_v23420(circuito):
    _auto_profile_ensure_dir_v23420()
    c = _auto_profile_circuit_key_v23420(circuito).lower()
    return os.path.join(AUTO_PROFILE_DIR, f"{c}_profiles.csv")


def _auto_profile_read_cache_v23420(circuito):
    path = _auto_profile_path_v23420(circuito)
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df = pd.read_csv(path)
        if "Player" not in df.columns:
            return pd.DataFrame()
        return df
    except Exception:
        return pd.DataFrame()


def _auto_profile_write_cache_v23420(circuito, rows):
    if rows is None:
        return 0
    new = pd.DataFrame(rows)
    if new.empty or "Player" not in new.columns:
        return 0
    path = _auto_profile_path_v23420(circuito)
    old = _auto_profile_read_cache_v23420(circuito)
    df = pd.concat([old, new], ignore_index=True) if not old.empty else new
    df["_clean"] = df["Player"].apply(limpiar)
    df = df[df["_clean"].astype(str).str.len() > 0].copy()
    if "SourceScore" not in df.columns:
        df["SourceScore"] = 0.0
    df["SourceScore"] = pd.to_numeric(df["SourceScore"], errors="coerce").fillna(0.0)
    df = df.sort_values(["SourceScore"], ascending=False).drop_duplicates("_clean", keep="first")
    df = df.drop(columns=["_clean"], errors="ignore")
    try:
        df.to_csv(path, index=False)
        return len(new)
    except Exception:
        return 0


def _auto_profile_default_quality_v23420(name=""):
    return {
        "matches_total": 0,
        "matches_surface": {"Hard": 0, "Clay": 0, "Grass": 0},
        "level_counts": {"tour": 0, "challenger": 0, "itf": 0, "qualy": 0, "unknown": 0},
        "tour_quality": 0.45,
        "stability": {"Hard": 0.05, "Clay": 0.05, "Grass": 0.05},
        "confidence": {"Hard": 0.32, "Clay": 0.32, "Grass": 0.32},
        "raw_names": [name] if name else [],
        "aliases": list(variantes_nombre_quality(name)) if name else [],
        "source_files": ["auto_profile_cache"],
        "matched_name": name or "AUTO PROFILE",
        "match_score": 1.0 if name else 0.0,
        "quality_map_size": 0,
        "quality_meta": {"version": AUTO_PROFILE_VERSION, "mode": "auto_profile_cache"},
        "candidate_matches": [],
        "auto_profile": True,
    }


def aplicar_auto_profile_cache_v23420(players, circuito):
    """Añade perfiles auto-completados desde datos/auto_profiles/*.csv.
    No pisa jugadores ya encontrados; solo rellena faltantes.
    """
    try:
        df = _auto_profile_read_cache_v23420(circuito)
        if df.empty:
            return players
        out = dict(players or {})
        existing = {limpiar(k) for k in out.keys()}
        circ = str(circuito or "ATP").upper().strip()
        circ_for_stats = "ATP" if circ in ["CHALLENGER", "CHALLENGER_ATP"] else circ
        for _, row in df.iterrows():
            name = normalizar_texto(row.get("Player", ""))
            if not name or limpiar(name) in existing:
                continue
            rank = int(leer_float(row.get("Rank", row.get("EloRank", 999)), 999))
            elo = leer_float(row.get("Elo", 1500), 1500)
            hard = leer_float(row.get("Hard", row.get("hElo", elo)), elo)
            clay = leer_float(row.get("Clay", row.get("cElo", elo)), elo)
            grass = leer_float(row.get("Grass", row.get("gElo", elo)), elo)
            stats_general = stats_default_por_elo(clay, rank, "Clay", circ_for_stats)
            stats_general["found_stats"] = True
            stats_general["raw_name_stats"] = name
            stats_general["match_type"] = "auto_profile_cache"
            stats_by_surface = {
                "Hard": stats_default_por_elo(hard, rank, "Hard", circ_for_stats),
                "Clay": stats_default_por_elo(clay, rank, "Clay", circ_for_stats),
                "Grass": stats_default_por_elo(grass, rank, "Grass", circ_for_stats),
            }
            for sf in stats_by_surface:
                stats_by_surface[sf]["found_stats"] = True
                stats_by_surface[sf]["raw_name_stats"] = name
                stats_by_surface[sf]["match_type"] = f"auto_profile_{sf.lower()}"
            out[name] = {
                "Player": name,
                "Rank": rank,
                "Hard": hard,
                "Clay": clay,
                "Grass": grass,
                "Stats": stats_general,
                "StatsBySurface": stats_by_surface,
                "Fatigue": {"recent_matches": 0, "fatigue_score": 0.0, "days_since_last": None, "surface_fatigue": 0.0},
                "Quality": _auto_profile_default_quality_v23420(name),
                "FuenteDB": "AutoProfileCache",
            }
            existing.add(limpiar(name))
        return out
    except Exception:
        return players


def _auto_profile_ta_report_url_v23420(circuito):
    c = _auto_profile_circuit_key_v23420(circuito)
    if c == "WTA":
        return "https://tennisabstract.com/reports/wta_elo_ratings.html"
    return "https://tennisabstract.com/reports/atp_elo_ratings.html"


def _auto_profile_abs_url_v23431(href):
    href = str(href or "").strip()
    if not href:
        return ""
    if href.startswith("http"):
        return href
    if href.startswith("/"):
        return "https://www.tennisabstract.com" + href
    return "https://www.tennisabstract.com/" + href.lstrip("./")


def _auto_profile_parse_ta_elo_html_v23420(html_text):
    """Parser del leaderboard Elo de TennisAbstract con enlace a ficha cuando está disponible.
    Mantiene compatibilidad con el parser anterior y añade PlayerUrl.
    """
    try:
        import html as _html
        raw = str(html_text or "")
        rows = []
        pat = re.compile(r"^\s*(\d+)\s+(.+?)\s+(\d{1,2}\.\d)\s+(\d{3,4}(?:\.\d)?)\s+(\d+)\s+(\d{3,4}(?:\.\d)?)\s+(\d+)\s+(\d{3,4}(?:\.\d)?)\s+(\d+)\s+(\d{3,4}(?:\.\d)?)\b")

        # Preferimos filas HTML para poder recuperar el enlace player.cgi.
        html_rows = re.findall(r"<tr[^>]*>(.*?)</tr>", raw, flags=re.I | re.S)
        iterable = html_rows if html_rows else raw.splitlines()
        for row_html in iterable:
            href = ""
            mh = re.search(r'href=["\']([^"\']*player\.cgi[^"\']*)["\']', row_html, flags=re.I)
            if mh:
                href = _auto_profile_abs_url_v23431(mh.group(1))

            line = re.sub(r"<br\s*/?>", "\n", row_html, flags=re.I)
            line = re.sub(r"<[^>]+>", " ", line)
            line = _html.unescape(line)
            line = re.sub(r"\s+", " ", line).strip()
            if not line or "Updated weekly" in line or "Elo Rank Player" in line:
                continue
            m = pat.search(line)
            if not m:
                continue
            elo_rank, player, age, elo, h_rank, helo, c_rank, celo, g_rank, gelo = m.groups()
            player = normalizar_texto(player)
            if not player or len(player) < 3:
                continue
            rows.append({
                "Player": player,
                "EloRank": int(float(elo_rank)),
                "Rank": int(float(elo_rank)),
                "Elo": float(elo),
                "Hard": float(helo),
                "Clay": float(celo),
                "Grass": float(gelo),
                "PlayerUrl": href,
                "Source": "TennisAbstract Elo leaderboard",
                "Updated": time.strftime("%Y-%m-%d %H:%M:%S"),
                "SourceScore": 0.90,
            })
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()




def _ta_match_urls_from_elo_leaderboard_v23493(name, circuito="ATP", min_score=0.90):
    """Busca la URL real de TennisAbstract en el leaderboard Elo.
    Esto corrige casos donde la URL directa falla por iniciales, acentos o variantes.
    No toca el motor Over: solo resuelve la ficha TA.
    """
    out = []
    try:
        # ATP por defecto; Challenger también vive en el ranking ATP de TA.
        circuits = [circuito] if str(circuito).upper() in ["ATP", "WTA"] else ["ATP"]
        if "ATP" not in [str(c).upper() for c in circuits]:
            circuits.append("ATP")
        for c in circuits:
            url = _auto_profile_ta_report_url_v23420(c)
            html_text, status = _ta_match_fetch_url_v23432(url, timeout_sec=18)
            if not html_text:
                continue
            df = _auto_profile_parse_ta_elo_html_v23420(html_text)
            if df is None or df.empty or "PlayerUrl" not in df.columns:
                continue
            best = []
            target = _ta_alias_resolve(name)
            for _, r in df.iterrows():
                player = normalizar_texto(r.get("Player", ""))
                pu = normalizar_texto(r.get("PlayerUrl", ""))
                if not player or not pu:
                    continue
                sc = max(similitud_nombre(target, player), similitud_nombre(name, player))
                # Refuerzo apellido+inicial para nombres abreviados tipo N. Kyrgios.
                try:
                    ai1 = apellido_inicial_key(target)
                    ai2 = apellido_inicial_key(player)
                    if ai1 and ai2 and ai1 == ai2:
                        sc = max(sc, 0.99)
                except Exception:
                    pass
                if sc >= min_score:
                    best.append((sc, int(float(r.get("EloRank", 9999) or 9999)), pu))
            best.sort(key=lambda x: (-x[0], x[1]))
            for _, __, u in best[:3]:
                if u and u not in out:
                    out.append(u)
    except Exception:
        pass
    return out

def _auto_profile_direct_urls_v23431(name):
    """Candidatos de URL para fichas TennisAbstract. Se usan solo si la tabla Elo no trae enlace."""
    toks = tokens(name)
    if not toks:
        return []
    if len(toks) >= 2:
        # Si viene "Apellido I" intentamos apellido + inicial, y también formato normal si existe.
        first = toks[0]
        last = toks[-1]
        if len(last) == 1 and len(toks) >= 2:
            # Mejor no inventar demasiado con iniciales abreviadas; probablemente no baste para URL directa.
            base1 = "".join(toks)
            base2 = "-".join(toks)
        else:
            base1 = "".join([t.title() for t in toks])
            base2 = "-".join([t.title() for t in toks])
    else:
        base1 = toks[0].title(); base2 = toks[0].title()
    out = []
    for b in [base1, base2]:
        if b and b not in out:
            out.append(f"https://www.tennisabstract.com/cgi-bin/player.cgi?p={b}")
    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def _auto_profile_fetch_url_v23431(url):
    try:
        resp = requests.get(str(url), timeout=12, headers={"User-Agent": "Mozilla/5.0 TennisIA/AutoProfile"})
        if resp.status_code != 200:
            return "", f"HTTP {resp.status_code}"
        return resp.text, "OK"
    except Exception as e:
        return "", f"{type(e).__name__}: {e}"


def _auto_profile_parse_player_page_hint_v23431(html_text, name):
    """Extrae señales básicas de la ficha. No inventa stats: solo confirma que la ficha parece del jugador.
    Si encuentra números Elo en la ficha, los usa; si no, devuelve vacío.
    """
    try:
        import html as _html
        raw = str(html_text or "")
        txt = re.sub(r"<br\s*/?>", "\n", raw, flags=re.I)
        txt = re.sub(r"<[^>]+>", " ", txt)
        txt = _html.unescape(txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        if similitud_nombre(name, txt[:300]) < 0.15 and limpiar(name) not in limpiar(txt[:800]):
            return {}

        def find_num(label_patterns):
            for lp in label_patterns:
                m = re.search(lp + r"\D{0,40}(\d{3,4}(?:\.\d)?)", txt, flags=re.I)
                if m:
                    val = float(m.group(1))
                    if 900 <= val <= 2600:
                        return val
            return None

        elo = find_num([r"\bElo\b", r"Current\s+Elo"])
        hard = find_num([r"Hard\s+Elo", r"hElo", r"Hard"])
        clay = find_num([r"Clay\s+Elo", r"cElo", r"Clay"])
        grass = find_num([r"Grass\s+Elo", r"gElo", r"Grass"])
        out = {"FichaOK": True}
        if elo: out["Elo"] = elo
        if hard: out["Hard"] = hard
        if clay: out["Clay"] = clay
        if grass: out["Grass"] = grass
        return out
    except Exception:
        return {}


def _auto_profile_enrich_with_player_pages_v23431(found_rows, requested_missing=None, max_pages=25):
    """Abre fichas de TennisAbstract de forma controlada y añade confirmación/enlace.
    Para filas sin URL intenta URL directa. No cambia motor; solo mejora caché.
    """
    enriched = []
    for i, row in enumerate(found_rows or []):
        if i >= int(max_pages):
            enriched.append(row); continue
        r = dict(row)
        req = r.get("RequestedName") or r.get("Player") or ""
        urls = []
        if r.get("PlayerUrl"):
            urls.append(r.get("PlayerUrl"))
        urls.extend([u for u in _auto_profile_direct_urls_v23431(r.get("Player", req)) if u not in urls])
        ok = False
        for url in urls[:3]:
            html_text, status = _auto_profile_fetch_url_v23431(url)
            if not html_text:
                continue
            hint = _auto_profile_parse_player_page_hint_v23431(html_text, r.get("Player", req))
            if hint.get("FichaOK"):
                for k in ["Elo", "Hard", "Clay", "Grass"]:
                    if k in hint and (pd.isna(r.get(k, None)) or not r.get(k, None) or float(r.get(k, 0)) <= 0):
                        r[k] = hint[k]
                r["PlayerUrl"] = url
                r["FichaTA"] = "OK"
                r["Source"] = str(r.get("Source", "TennisAbstract")) + " + ficha"
                r["SourceScore"] = max(float(r.get("SourceScore", 0.90) or 0.90), 0.94)
                ok = True
                break
        if not ok:
            r["FichaTA"] = r.get("FichaTA", "No confirmada")
        enriched.append(r)
    return enriched


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def _auto_profile_download_ta_elo_v23420(circuito):
    url = _auto_profile_ta_report_url_v23420(circuito)
    try:
        resp = requests.get(url, timeout=12, headers={"User-Agent": "Mozilla/5.0 TennisIA/AutoProfile"})
        if resp.status_code != 200:
            return pd.DataFrame(), f"HTTP {resp.status_code}"
        df = _auto_profile_parse_ta_elo_html_v23420(resp.text)
        if df.empty:
            return pd.DataFrame(), "No se pudo parsear la tabla Elo"
        return df, "OK"
    except Exception as e:
        return pd.DataFrame(), f"{type(e).__name__}: {e}"


def _auto_profile_match_names_v23420(missing_names, ta_df, min_score=0.86):
    if ta_df is None or ta_df.empty or not missing_names:
        return [], []
    found, not_found = [], []
    ta_rows = ta_df.to_dict("records")
    for name in missing_names:
        search_name = _ta_alias_resolve(name)
        best = None
        best_score = 0.0
        for row in ta_rows:
            cand = row.get("Player", "")
            score = similitud_nombre(search_name, cand)
            ai1, ai2 = apellido_inicial_key(search_name, cand)
            if ai1 and ai2 and ai1 == ai2:
                score = max(score, 0.98)
            if score > best_score:
                best_score = score
                best = row
        if best is not None and best_score >= min_score:
            r = dict(best)
            r["RequestedName"] = name
            if search_name != name:
                r["AliasUsado"] = search_name
            r["Player"] = best.get("Player", search_name or name)
            r["MatchScore"] = round(float(best_score), 3)
            r["SourceScore"] = max(float(r.get("SourceScore", 0.90) or 0.90), float(best_score))
            found.append(r)
        else:
            not_found.append({"Jugador": name, "Mejor score": round(float(best_score), 3), "Candidato": best.get("Player", "") if best else ""})
    return found, not_found


def _auto_profile_extract_players_from_partido_v23420(partido):
    txt = normalizar_texto(partido)
    if not txt:
        return []
    for sep in [" vs ", " VS ", " v ", " V ", " - "]:
        if sep in txt:
            parts = [p.strip() for p in txt.split(sep) if p.strip()]
            return parts[:2]
    return []


def _auto_profile_players_from_df_v23420(df):
    if df is None or df.empty or "Partido" not in df.columns:
        return []
    names = []
    for partido in df["Partido"].dropna().astype(str).tolist():
        names.extend(_auto_profile_extract_players_from_partido_v23420(partido))
    seen, out = set(), []
    for n in names:
        ck = limpiar(n)
        if ck and ck not in seen:
            seen.add(ck); out.append(n)
    return out


def _auto_profile_known_clean_v23420(circuito):
    known = set()
    try:
        db = cargar_datos(circuito)
        known.update(limpiar(k) for k in db.keys())
    except Exception:
        pass
    try:
        cache = _auto_profile_read_cache_v23420(circuito)
        if not cache.empty:
            known.update(cache["Player"].dropna().astype(str).apply(limpiar).tolist())
    except Exception:
        pass
    # v23.47.2: si Y. Bu -> Bu Yunchaokete y Bu ya existe, Y. Bu también cuenta como OK.
    try:
        aliases = _ta_alias_load_cached()
        known_now = set(known)
        for alias_key, canonical in aliases.items():
            if limpiar(canonical) in known_now:
                known.add(alias_key)
    except Exception:
        pass
    return {x for x in known if x}


def render_auto_profile_finder_v23420(current_df=None, global_df=None):
    st.markdown("#### 🔎 Auto perfiles faltantes")
    st.caption("Detecta jugadores del plan/tanda y busca perfiles en TennisAbstract en bloque. No toca el motor Over; solo evita meter fichas a mano.")

    c1, c2, c3 = st.columns([1, 1, 1.4])
    with c1:
        circuito_busqueda = st.selectbox("Circuito para buscar perfiles", ["ATP", "CHALLENGER", "WTA"], index=0, key="auto_profile_circuit_v23420")
    with c2:
        min_score = st.slider("Coincidencia mínima", 0.80, 0.98, 0.86, 0.01, key="auto_profile_score_v23420")
    with c3:
        modo_ficha = st.selectbox("Modo", ["Elo rápido", "Elo + abrir ficha TA"], index=1, key="auto_profile_mode_v23431")

    names = []
    names.extend(_auto_profile_players_from_df_v23420(global_df))
    names.extend(_auto_profile_players_from_df_v23420(current_df))
    seen, all_names = set(), []
    for n in names:
        ck = limpiar(n)
        if ck and ck not in seen:
            seen.add(ck); all_names.append(n)

    if not all_names:
        st.warning("No he detectado jugadores en picks actuales/acumulados todavía.")
        return

    known = _auto_profile_known_clean_v23420(circuito_busqueda)
    missing = [n for n in all_names if limpiar(n) not in known]
    ok_count = len(all_names) - len(missing)
    m1, m2, m3 = st.columns(3)
    m1.metric("Jugadores detectados", len(all_names))
    m2.metric("Perfil ya OK", ok_count)
    m3.metric("Fichas a buscar", len(missing))

    with st.expander("Ver fichas que la app intentará buscar", expanded=len(missing) > 0):
        st.dataframe(pd.DataFrame(_ta_alias_debug_rows(missing)), width='stretch', hide_index=True)

    # v23.47.2: gestor manual de alias para casos tipo Y. Bu -> Bu Yunchaokete.
    if missing:
        with st.expander("🔗 Corregir alias manualmente", expanded=False):
            st.caption("Usa esto cuando la app pide una ficha que ya tienes guardada con otro nombre. Ejemplo: Y. Bu → Bu Yunchaokete.")
            c_alias1, c_alias2 = st.columns([1, 1.3])
            with c_alias1:
                alias_src = st.selectbox("Jugador detectado", missing, key="ta_alias_src_v23472")
            with c_alias2:
                alias_dst = st.text_input("Nombre correcto guardado / TennisAbstract", value=_ta_alias_resolve(alias_src) if _ta_alias_resolve(alias_src) != alias_src else "", key="ta_alias_dst_v23472", placeholder="Ej: Bu Yunchaokete")
            if st.button("💾 Guardar alias", key="ta_alias_save_v23472", use_container_width=True):
                ok_alias, msg_alias = _ta_alias_add(alias_src, alias_dst)
                if ok_alias:
                    st.success(f"Alias guardado: {alias_src} → {alias_dst}. Pulsa Rerun o reanaliza para que cuente como perfil OK.")
                    try:
                        st.rerun()
                    except Exception:
                        pass
                else:
                    st.error(msg_alias)

            current_aliases = _ta_alias_load_cached()
            if current_aliases:
                st.caption(f"Alias activos: {len(current_aliases)}")
                sample = pd.DataFrame([{"Alias": k, "Nombre correcto": v} for k, v in list(current_aliases.items())[:80]])
                st.dataframe(sample, width='stretch', hide_index=True)

    if not missing:
        st.success("No veo perfiles faltantes para este circuito en la tanda/plan actual.")
        return

    st.info("Pulsa el botón y la app intentará buscar esos jugadores en TennisAbstract en bloque. Si encuentra Elo/ficha, lo guarda en caché para no pedírtelo mañana.")

    if st.button("🔎 Buscar fichas faltantes en TennisAbstract y guardar caché", key="auto_profile_search_v23420", use_container_width=True):
        with st.spinner("Buscando fichas en TennisAbstract..."):
            ta_df, status = _auto_profile_download_ta_elo_v23420(circuito_busqueda)
            if ta_df.empty:
                st.error(f"No se pudo descargar/leer TennisAbstract Elo: {status}")
                return
            found, not_found = _auto_profile_match_names_v23420(missing, ta_df, min_score=float(min_score))

            if modo_ficha == "Elo + abrir ficha TA" and found:
                found = _auto_profile_enrich_with_player_pages_v23431(found, missing, max_pages=min(25, len(found)))

            saved = _auto_profile_write_cache_v23420(circuito_busqueda, found)
            if saved > 0:
                try:
                    st.cache_data.clear()
                except Exception:
                    pass
                st.success(f"Guardados {saved} perfiles/fichas en caché. Reejecuta el análisis o pulsa Rerun para que entren en el modelo.")
            else:
                st.warning("No se guardó ningún perfil nuevo.")
            if found:
                show_found = pd.DataFrame(found)
                cols = ["RequestedName", "AliasUsado", "Player", "MatchScore", "Elo", "Hard", "Clay", "Grass", "Rank", "FichaTA", "PlayerUrl", "Source"]
                st.markdown("##### ✅ Fichas encontradas")
                st.dataframe(show_found[[c for c in cols if c in show_found.columns]], width='stretch', hide_index=True)
            if not_found:
                st.markdown("##### ⚠️ No encontrados con suficiente seguridad")
                st.dataframe(pd.DataFrame(not_found), width='stretch', hide_index=True)

    with st.expander("Cómo usarlo", expanded=False):
        st.markdown("""
1. Analiza la lista de partidos.
2. Entra aquí y pulsa **Buscar fichas faltantes**.
3. Si guarda perfiles, pulsa **Rerun** o vuelve a analizar.
4. Los jugadores encontrados quedarán guardados en `datos/auto_profiles/`.

Si alguno no aparece, la app seguirá usando fallback prudente y te lo dejará como perfil débil, pero ya no tendrás que copiar 20 fichas una por una.
""")


@st.cache_data
def cargar_datos(circuito, cache_version=QUALITY_ENGINE_VERSION):
    r = rutas(circuito)
    fatigue_map = crear_fatigue_map(circuito)
    quality_map = crear_quality_map(circuito)

    # =========================
    # Mapas generales
    # =========================
    serve_general = leer_archivo_stats(r["serve"], "serve")
    return_general = leer_archivo_stats(r["return"], "return")
    break_general = leer_archivo_stats(r["break"], "break")

    def construir_stats_map(serve_map, return_map, break_map):
        stats_map = {}
        all_ids = set(serve_map) | set(return_map) | set(break_map)

        for nid in all_ids:
            base = {
                "found_stats": False, "raw_name_stats": "NO ENCONTRADO",
                "hold": 0.78, "ace": 0.05, "df": 0.035,
                "1in": 0.62, "1w": 0.70, "2w": 0.50,
                "rpw": None, "break_pct": None, "bp_conv": None, "bp_saved": None,
                "serve_profile": "normal", "match_type": "stats"
            }
            if nid in serve_map:
                base = merge_stats(base, serve_map[nid])
            if nid in return_map:
                base = merge_stats(base, return_map[nid])
            if nid in break_map:
                base = merge_stats(base, break_map[nid])
            base["serve_profile"] = perfil_saque(base.get("ace", 0.05))
            stats_map[nid] = base

        return stats_map

    stats_general_map = construir_stats_map(serve_general, return_general, break_general)

    # =========================
    # Mapas por superficie
    # =========================
    stats_surface_maps = {}

    for surface in ["Hard", "Clay", "Grass"]:
        serve_surface = leer_archivo_stats(ruta_stats_superficie(circuito, "serve", surface), "serve")
        return_surface = leer_archivo_stats(ruta_stats_superficie(circuito, "return", surface), "return")
        break_surface = leer_archivo_stats(ruta_stats_superficie(circuito, "break", surface), "break")

        surface_map_raw = construir_stats_map(serve_surface, return_surface, break_surface)

        # Combina cada jugador con general + superficie
        combined = {}
        all_ids = set(stats_general_map) | set(surface_map_raw)

        for nid in all_ids:
            if nid in stats_general_map and nid in surface_map_raw:
                combined[nid] = merge_surface_stats(stats_general_map[nid], surface_map_raw[nid])
                combined[nid]["match_type"] = f"surface_{surface.lower()}"
            elif nid in surface_map_raw:
                combined[nid] = surface_map_raw[nid]
                combined[nid]["match_type"] = f"surface_{surface.lower()}"
            else:
                combined[nid] = stats_general_map[nid].copy()
                combined[nid]["match_type"] = "general_fallback"

            combined[nid]["serve_profile"] = perfil_saque(combined[nid].get("ace", 0.05))

        stats_surface_maps[surface] = combined

    # v22.29 safety: ensure fatigue_map exists even after cache/merge issues
    if 'fatigue_map' not in locals() or fatigue_map is None:
        fatigue_map = {}

    players = {}

    if not os.path.exists(r["elo"]):
        if str(circuito).upper().strip() == "WTA":
            players = aplicar_wta_real_data_upgrade(players)
        players = aplicar_auto_profile_cache_v23420(players, circuito)
        return players

    df = pd.read_excel(r["elo"])

    col_player = buscar_columna(df, ["Player", "Name", "Jugador"])
    col_rank = buscar_columna(df, ["ATP Rank", "WTA Rank", "Rank"])
    col_elo = buscar_columna(df, ["Elo"])
    col_hard = buscar_columna(df, ["hElo", "Hard Elo"])
    col_clay = buscar_columna(df, ["cElo", "Clay Elo"])
    col_grass = buscar_columna(df, ["gElo", "Grass Elo"])

    if col_player is None:
        return players

    for _, row in df.iterrows():
        nombre = normalizar_texto(row.get(col_player, ""))
        if nombre == "":
            continue

        rank = int(leer_float(row.get(col_rank), 999))
        elo_general = leer_float(row.get(col_elo), 1500)

        hard = leer_float(row.get(col_hard), elo_general)
        clay = leer_float(row.get(col_clay), elo_general)
        grass = leer_float(row.get(col_grass), elo_general)

        stats_general = buscar_stats(nombre, stats_general_map)
        if stats_general is None:
            stats_general = stats_default_por_elo(clay, rank, "Clay", circuito)

        stats_by_surface = {}
        for surface, elo_surface in [("Hard", hard), ("Clay", clay), ("Grass", grass)]:
            s = buscar_stats(nombre, stats_surface_maps.get(surface, {}))
            if s is None:
                s = stats_default_por_elo(elo_surface, rank, surface, circuito)
            stats_by_surface[surface] = s

        players[nombre] = {
            "Player": nombre,
            "Rank": rank,
            "Hard": hard,
            "Clay": clay,
            "Grass": grass,
            "Stats": stats_general,
            "StatsBySurface": stats_by_surface,
            "Fatigue": buscar_fatigue(nombre, fatigue_map),
            "Quality": buscar_quality(nombre, quality_map)
        }


    if str(circuito).upper().strip() == "WTA":
        players = aplicar_wta_real_data_upgrade(players)

    players = aplicar_auto_profile_cache_v23420(players, circuito)

    return players


# =========================================================
# v23.26 CHALLENGER DATA ENGINE
# =========================================================
# Si el usuario trabaja en modo ATP, la app carga ATP + Challenger.
# - ATP conserva prioridad para jugadores que ya existen en atp_elo.xlsx.
# - Challenger añade todos los jugadores que faltan desde datos/challenger/.
# Esto evita que partidos Challenger/Qualy caigan como "No encontrados" cuando
# el jugador no está en la base ATP principal.

@st.cache_data(show_spinner=False)
def cargar_datos_app(circuito, cache_version=QUALITY_ENGINE_VERSION):
    circuito = str(circuito).upper().strip()

    if circuito != "ATP":
        return cargar_datos(circuito, cache_version=cache_version)

    atp_db = cargar_datos("ATP", cache_version=cache_version)
    challenger_dir = os.path.join("datos", "challenger")
    challenger_elo = os.path.join(challenger_dir, "challenger_elo.xlsx")

    if not os.path.exists(challenger_elo):
        return atp_db

    try:
        challenger_db = cargar_datos("challenger", cache_version=cache_version)
    except Exception:
        challenger_db = {}

    if not challenger_db:
        return atp_db

    merged = dict(atp_db)
    existing_clean = {limpiar(k) for k in merged.keys()}

    for name, data in challenger_db.items():
        ck = limpiar(name)
        if not ck or ck in existing_clean:
            continue
        try:
            data = data.copy()
            data["FuenteDB"] = "Challenger"
        except Exception:
            pass
        merged[name] = data
        existing_clean.add(ck)

    return merged


def circuito_lookup_para_match(m, circuito_ui):
    """Devuelve el circuito de datos para fallback/históricos.
    En modo ATP, los bloques Challenger/ITF/Qualy usan datos/challenger.
    El motor de simulación sigue tratándolos como ATP para no romper guards ATP/WTA.
    """
    src = str((m or {}).get("circuito_detectado", "")).upper().strip()
    ui = str(circuito_ui).upper().strip()
    if ui == "CHALLENGER":
        return "challenger"
    if ui == "ATP" and (src in {"CHALLENGER_ATP", "ITF_ATP"} or _es_entorno_challenger_match(m, circuito_ui)):
        return "challenger"
    return circuito_ui


def circuito_sim_para_lookup(lookup_circuit, circuito_ui):
    lk = str(lookup_circuit).upper().strip()
    if lk == "CHALLENGER":
        return "ATP"
    return circuito_ui


def get_stats_surface(player_data, surface):
    return player_data.get("StatsBySurface", {}).get(surface, player_data.get("Stats", {}))


def calc_hold(stats, elo_diff, surface, circuito):
    base = stats.get("hold", 0.78)
    ace = stats.get("ace", 0.05)
    profile = stats.get("serve_profile", "normal")
    surface_adj = {"Hard": -0.005, "Clay": -0.105, "Grass": 0.015} if circuito == "ATP" else {"Hard": -0.010, "Clay": -0.085, "Grass": 0.010}
    elo_adj = np.clip(elo_diff / 3900, -0.040, 0.040)

    serve_bonus = 0
    if profile == "good_server": serve_bonus = 0.006
    elif profile == "big_server": serve_bonus = 0.012
    elif profile == "elite_server": serve_bonus = 0.018

    bp_saved = stats.get("bp_saved")
    clutch_bonus = np.clip((bp_saved - 0.58) * 0.05, -0.010, 0.012) if bp_saved is not None else 0
    hold = base + surface_adj[surface] + elo_adj + serve_bonus + ace * 0.01 + clutch_bonus
    return np.clip(hold, 0.46, 0.84)

def calc_return_strength(stats, elo_surface, surface):
    if stats.get("rpw") is not None:
        ret = stats["rpw"]
    elif stats.get("break_pct") is not None:
        ret = stats["break_pct"]
    else:
        ret = 1.0 - stats.get("hold", 0.78) + ((elo_surface - 1500) / 4200)

    if stats.get("break_pct") is not None:
        ret = (ret * 0.65) + (stats["break_pct"] * 0.35)

    if stats.get("bp_conv") is not None:
        ret += np.clip((stats["bp_conv"] - 0.38) * 0.05, -0.010, 0.012)

    if surface == "Clay": ret += 0.020
    elif surface == "Grass": ret -= 0.010
    return np.clip(ret, 0.12, 0.40)

def aplicar_return_pressure(h1, h2, r1, r2, surface):
    weight = 0.19 if surface == "Clay" else 0.10 if surface == "Hard" else 0.09
    return np.clip(h1 - r2 * weight, 0.44, 0.84), np.clip(h2 - r1 * weight, 0.44, 0.84)

def calcular_match_volatility(e1, e2, surface, fav_raw_est=None):
    gap = abs(e1 - e2)
    vol = 0.040
    if gap > 300: vol += 0.018
    elif gap > 200: vol += 0.010
    elif gap > 120: vol += 0.005
    if surface == "Clay": vol += 0.010
    elif surface == "Grass": vol -= 0.004

    # v14: Hard Compression
    if surface == "Hard" and fav_raw_est is not None:
        if fav_raw_est >= 0.72: vol -= 0.010
        elif fav_raw_est >= 0.66: vol -= 0.006

    return float(np.clip(vol, 0.020, 0.075))

def hard_compression(surface, fav, h1, h2):
    noise_mult, pressure_add, shift_mult = 1.0, 0.0, 1.0
    if surface == "Hard" and fav >= 0.72:
        noise_mult, pressure_add, shift_mult = 0.88, 0.008, 0.86
    elif surface == "Hard" and fav >= 0.66:
        noise_mult, pressure_add, shift_mult = 0.93, 0.004, 0.92

    if surface == "Hard" and h1 > 0.76 and h2 > 0.76:
        noise_mult = min(1.0, noise_mult + 0.05)
        shift_mult = min(1.0, shift_mult + 0.05)

    return noise_mult, pressure_add, shift_mult


def calcular_tiebreak_boost(stats1, stats2, hold1, hold2, surface, p1_big, p2_big):
    """
    v16 Tie-break Intelligence Engine.
    Estima el entorno de tie-break con:
    - hold combinado
    - ace%
    - 1st serve won
    - 2nd serve won
    - superficie
    - big server profile
    """
    ace1 = stats1.get("ace", 0.05)
    ace2 = stats2.get("ace", 0.05)

    first_won1 = stats1.get("1w", 0.70)
    first_won2 = stats2.get("1w", 0.70)

    second_won1 = stats1.get("2w", 0.50)
    second_won2 = stats2.get("2w", 0.50)

    hold_avg = (hold1 + hold2) / 2
    ace_avg = (ace1 + ace2) / 2
    first_avg = (first_won1 + first_won2) / 2
    second_avg = (second_won1 + second_won2) / 2

    boost = 0.0

    # Hold alto de ambos = más 6-6
    if hold1 > 0.78 and hold2 > 0.78:
        boost += 0.070
    elif hold1 > 0.76 and hold2 > 0.76:
        boost += 0.050
    elif hold1 > 0.74 and hold2 > 0.74:
        boost += 0.030
    elif hold_avg < 0.66:
        boost -= 0.025

    # Aces y primeros servicios potentes
    if ace_avg >= 0.12:
        boost += 0.050
    elif ace_avg >= 0.09:
        boost += 0.030
    elif ace_avg >= 0.07:
        boost += 0.015

    if first_avg >= 0.76:
        boost += 0.030
    elif first_avg >= 0.72:
        boost += 0.015

    if second_avg >= 0.55:
        boost += 0.015

    # Perfil de sacadores
    if p1_big and p2_big:
        boost += 0.060
    elif p1_big or p2_big:
        boost += 0.030

    # Superficie
    if surface == "Grass":
        boost += 0.055
    elif surface == "Hard":
        boost += 0.035
    elif surface == "Clay":
        boost -= 0.015

    return float(np.clip(boost, -0.040, 0.160))


def pressure_collapse_params(stats1, stats2, surface):
    """
    v16 Pressure Collapse.
    Ajusta ligeramente los games finales según BP saved / BP conversion.
    No cambia mucho el ML; afecta más al cierre de sets y tie-breaks.
    """
    bp_saved1 = stats1.get("bp_saved", None)
    bp_saved2 = stats2.get("bp_saved", None)
    bp_conv1 = stats1.get("bp_conv", None)
    bp_conv2 = stats2.get("bp_conv", None)

    pressure1 = 0.0
    pressure2 = 0.0

    if bp_saved1 is not None:
        pressure1 += np.clip((bp_saved1 - 0.58) * 0.08, -0.012, 0.014)
    if bp_saved2 is not None:
        pressure2 += np.clip((bp_saved2 - 0.58) * 0.08, -0.012, 0.014)

    if bp_conv1 is not None:
        pressure1 += np.clip((bp_conv1 - 0.38) * 0.04, -0.008, 0.010)
    if bp_conv2 is not None:
        pressure2 += np.clip((bp_conv2 - 0.38) * 0.04, -0.008, 0.010)

    if surface == "Clay":
        pressure1 *= 1.10
        pressure2 *= 1.10

    return float(np.clip(pressure1, -0.018, 0.020)), float(np.clip(pressure2, -0.018, 0.020))



def sim_set(hold1, hold2, surface, shift, p1_big, p2_big, fav_raw_est=0.5, stats1=None, stats2=None):
    g1 = g2 = 0
    tb = False
    server = random.choice([1, 2])

    stats1 = stats1 or {}
    stats2 = stats2 or {}
    tb_boost = calcular_tiebreak_boost(stats1, stats2, hold1, hold2, surface, p1_big, p2_big)
    pressure_skill1, pressure_skill2 = pressure_collapse_params(stats1, stats2, surface)

    if surface == "Clay":
        noise, pressure = 0.068, -0.085
    elif surface == "Hard":
        noise, pressure = 0.038, -0.030
    else:
        noise, pressure = 0.028, -0.018

    nm, pa, sm = hard_compression(surface, fav_raw_est, hold1, hold2)
    noise *= nm
    pressure += pa
    shift *= sm

    while True:
        extra = pressure if g1 >= 4 and g2 >= 4 else 0
        # En games finales, BP saved/BP converted modula el cierre.
        clutch1 = pressure_skill1 if g1 >= 4 and g2 >= 4 else 0.0
        clutch2 = pressure_skill2 if g1 >= 4 and g2 >= 4 else 0.0

        h1 = np.clip(np.random.normal(hold1 + shift, noise) + extra + clutch1, 0.30, 0.93)
        h2 = np.clip(np.random.normal(hold2 - shift, noise) + extra + clutch2, 0.30, 0.93)

        if server == 1:
            if random.random() < h1: g1 += 1
            else: g2 += 1
        else:
            if random.random() < h2: g2 += 1
            else: g1 += 1

        server = 1 if server == 2 else 2

        if g1 >= 6 and g1 - g2 >= 2: return g1, g2, tb
        if g2 >= 6 and g2 - g1 >= 2: return g1, g2, tb

        if g1 == 6 and g2 == 6:
            tb = True
            p_tb = hold1 / (hold1 + hold2)

            # v16: boost inteligente según saque/hold/superficie
            p_tb += tb_boost

            p_tb = np.clip(p_tb, 0.30, 0.82)
            return (7, 6, tb) if random.random() < p_tb else (6, 7, tb)


def smart_set_dynamics(p1_win, p2_win, hold1, hold2, tb_rate, vol, surface, circuito="ATP", rank1=999, rank2=999):
    """
    v19 Smart Set Dynamics.
    ATP: mantiene estructura.
    WTA: evita coinflips artificiales y dog-set exagerado.
    """
    fav = max(p1_win, p2_win)
    hold_gap = abs(hold1 - hold2)

    dog_set_suppress = 0.0
    fav20_boost = 0.0
    long_match_adj = 0.0

    if fav >= 0.72:
        fav20_boost += 0.08
        dog_set_suppress += 0.10
        long_match_adj -= 0.05
    elif fav >= 0.66:
        fav20_boost += 0.05
        dog_set_suppress += 0.06
        long_match_adj -= 0.03

    if hold_gap >= 0.09:
        fav20_boost += 0.05
        dog_set_suppress += 0.06
    elif hold_gap >= 0.06:
        fav20_boost += 0.03
        dog_set_suppress += 0.03

    if tb_rate <= 0.24:
        fav20_boost += 0.04
        dog_set_suppress += 0.04
        long_match_adj -= 0.04

    if surface == "Clay":
        dog_set_suppress *= 0.85
        fav20_boost *= 0.90

    if vol >= 0.035:
        dog_set_suppress *= 0.70
        fav20_boost *= 0.75
        long_match_adj += 0.05

    if circuito == "WTA":
        top_player = rank1 <= 15 or rank2 <= 15
        rank_gap = abs(rank1 - rank2)

        if fav >= 0.68 and top_player:
            dog_set_suppress += 0.055
            fav20_boost += 0.040
            long_match_adj -= 0.035
        elif fav >= 0.64 and top_player:
            dog_set_suppress += 0.035
            fav20_boost += 0.025
            long_match_adj -= 0.020

        if rank_gap >= 40 and fav >= 0.62:
            dog_set_suppress += 0.035
            fav20_boost += 0.020
            long_match_adj -= 0.020

        if surface == "Clay":
            dog_set_suppress *= 0.92
            fav20_boost *= 0.92
            long_match_adj *= 0.85

        if fav < 0.58:
            dog_set_suppress *= 0.50
            fav20_boost *= 0.45
            long_match_adj += 0.025

    return {
        "dog_set_suppress": float(np.clip(dog_set_suppress, 0, 0.22)),
        "fav20_boost": float(np.clip(fav20_boost, 0, 0.20)),
        "long_match_adj": float(np.clip(long_match_adj, -0.12, 0.12))
    }







def straight_sets_guard_engine(circuito, surface, fav_prob, hold1, hold2, tb_rate, vol, rating_sanity, fav20, dogset, longm):
    """
    v22.20 Straight Sets Guard.
    Separa dos ideas que antes se mezclaban demasiado:
    - Over bajo útil: 6-4 6-4 / 7-5 6-3 puede ser buen over.
    - Partido largo / dog gana set: no debe inflarse si el favorito tiene control real.

    No toca ganador ni mercados over calculados por games. Solo ajusta:
    - favorito 2-0
    - underdog gana set
    - partido largo
    """
    out = {
        "active": False,
        "profile": "neutral",
        "fav20_boost": 0.0,
        "dogset_cut": 0.0,
        "long_cut": 0.0,
        "hold_gap": float(abs(hold1 - hold2)),
        "notes": []
    }

    if circuito != "ATP":
        return fav20, dogset, longm, out

    rs = rating_sanity or {}
    p1_rs = rs.get("p1", {}) or {}
    p2_rs = rs.get("p2", {}) or {}
    min_conf = min(p1_rs.get("confidence", 1.0), p2_rs.get("confidence", 1.0))
    min_surface = min(p1_rs.get("matches_surface", 99), p2_rs.get("matches_surface", 99))
    min_stability = min(p1_rs.get("stability", 1.0), p2_rs.get("stability", 1.0))

    hold_gap = abs(hold1 - hold2)

    # Condiciones de activación: favorito real, datos fiables y diferencia clara de hold.
    if fav_prob < 0.615:
        return fav20, dogset, longm, out
    # v22.21: favoritos muy claros también necesitan guardia si el hold gap es enorme.
    # Antes se excluían >73.5% y por eso spots tipo 6-4 6-3 podían inflar "dog gana set".
    clear_fav_mode = fav_prob >= 0.735

    if clear_fav_mode:
        if min_conf < 0.60 or min_surface < 18 or min_stability < 0.55:
            return fav20, dogset, longm, out
        if hold_gap < 0.115:
            return fav20, dogset, longm, out
        if tb_rate >= 0.32:
            return fav20, dogset, longm, out
    else:
        if min_conf < 0.74 or min_surface < 35 or min_stability < 0.72:
            return fav20, dogset, longm, out
        if hold_gap < 0.070:
            return fav20, dogset, longm, out
        if tb_rate >= 0.34:
            # Si hay entorno fuerte de tie-break, no forzamos 2-0.
            return fav20, dogset, longm, out

    boost = 0.030
    if fav_prob >= 0.64: boost += 0.012
    if fav_prob >= 0.67: boost += 0.010
    if fav_prob >= 0.735: boost += 0.018
    if fav_prob >= 0.78: boost += 0.010
    if hold_gap >= 0.085: boost += 0.018
    if hold_gap >= 0.105: boost += 0.012
    if hold_gap >= 0.140: boost += 0.014
    if min_conf >= 0.82: boost += 0.010
    if min_surface >= 100: boost += 0.008
    if tb_rate <= 0.27: boost += 0.006
    if surface == "Clay": boost *= 0.95

    boost = float(np.clip(boost, 0.0, 0.125))
    dog_cut = float(np.clip(boost * 1.10, 0.0, 0.135))
    long_cut = float(np.clip(boost * 0.82, 0.0, 0.105))

    out["active"] = True
    out["profile"] = "clear_favorite_control" if fav_prob >= 0.735 else "controlled_favorite"
    out["fav20_boost"] = boost
    out["dogset_cut"] = dog_cut
    out["long_cut"] = long_cut
    out["notes"].append("favorito fiable con ventaja clara de hold")
    out["notes"].append("over bajo no implica partido largo")

    fav20 = float(np.clip(fav20 + boost, 0.0, 0.95))
    dogset = float(np.clip(dogset - dog_cut, 0.05, 0.95))
    longm = float(np.clip(longm - long_cut, 0.05, 0.95))

    return fav20, dogset, longm, out


def favorite_2_0_sanity_guard(circuito, surface, fav_prob, fav20, dogset, longm, rating_sanity):
    """
    v22.25 Favorite 2-0 Sanity Guard.
    Evita probabilidades absurdamente altas de 2-0 cuando la confianza
    real o la calidad del histórico no justifican tanta seguridad.
    No toca ML ni overs; solo mercados derivados de sets.
    """
    out = {
        "active": False,
        "fav20_cap": None,
        "dogset_floor": None,
        "long_floor": None,
        "notes": []
    }

    if circuito != "ATP":
        return fav20, dogset, longm, out

    rs = rating_sanity or {}
    p1 = rs.get("p1", {}) or {}
    p2 = rs.get("p2", {}) or {}
    min_conf = min(p1.get("confidence", 1.0), p2.get("confidence", 1.0))
    min_quality = min(p1.get("tour_quality", 1.0), p2.get("tour_quality", 1.0))
    min_surface = min(p1.get("matches_surface", 99), p2.get("matches_surface", 99))
    min_stability = min(p1.get("stability", 1.0), p2.get("stability", 1.0))

    flags = []
    flags += list(p1.get("flags", []) or [])
    flags += list(p2.get("flags", []) or [])
    challenger_flag = any("challenger" in str(x).lower() or "qualy" in str(x).lower() for x in flags)

    cap = None
    floor = None
    long_floor = None

    # Con confianza real limitada no puede existir un 2-0 del 90-95%.
    if min_conf < 0.60:
        cap = 0.66
        floor = 0.26
        long_floor = 0.24
        out["notes"].append("cap 2-0 por confianza limitada")
    elif min_conf < 0.68:
        cap = 0.74
        floor = 0.20
        long_floor = 0.22
        out["notes"].append("cap 2-0 por confianza media")

    # Mucho Challenger/Qualy o calidad <70%: menos seguridad en sets corridos.
    if min_quality < 0.70 or challenger_flag:
        cap = min(cap if cap is not None else 0.82, 0.76)
        floor = max(floor if floor is not None else 0.16, 0.18)
        long_floor = max(long_floor if long_floor is not None else 0.18, 0.20)
        out["notes"].append("cap 2-0 por calidad Challenger/Qualy")

    # Muestra en superficie suficiente pero no elite: no bloquear, solo evitar 95%.
    if min_surface < 60 or min_stability < 0.78:
        cap = min(cap if cap is not None else 0.84, 0.80)
        floor = max(floor if floor is not None else 0.14, 0.16)
        out["notes"].append("cap 2-0 por muestra/stability no elite")

    # Incluso un favorito ATP clay al 80% rara vez merece 95% de 2-0 salvo confianza extrema.
    if fav_prob < 0.82:
        cap = min(cap if cap is not None else 0.86, 0.82)

    if cap is not None and fav20 > cap:
        out["active"] = True
        out["fav20_cap"] = float(cap)
        fav20 = float(cap)

    if floor is not None and dogset < floor:
        out["active"] = True
        out["dogset_floor"] = float(floor)
        dogset = float(floor)

    if long_floor is not None and longm < long_floor:
        out["active"] = True
        out["long_floor"] = float(long_floor)
        longm = float(long_floor)

    return float(np.clip(fav20, 0.0, 0.95)), float(np.clip(dogset, 0.05, 0.95)), float(np.clip(longm, 0.05, 0.95)), out


def upset_risk_guard_engine(circuito, surface, p1_cal, p1_raw, fav20, dogset, longm, rating_sanity):
    """
    v22.25 Upset Risk Guard.
    Evita que favoritos ATP Clay aparentemente claros se vendan como spots limpios
    cuando el propio sanity marca inflación/confianza limitada/calidad Challenger.
    No cambia el ganador bruto ni los overs; solo calibra ML final y mercados de sets.
    """
    out = {
        "active": False,
        "profile": "neutral",
        "ml_cap": None,
        "fav20_cap": None,
        "dogset_floor": None,
        "long_floor": None,
        "notes": []
    }

    if circuito != "ATP" or surface != "Clay":
        return p1_cal, fav20, dogset, longm, out

    rs = rating_sanity or {}
    p1 = rs.get("p1", {}) or {}
    p2 = rs.get("p2", {}) or {}

    min_conf = min(p1.get("confidence", 1.0), p2.get("confidence", 1.0))
    min_quality = min(p1.get("tour_quality", 1.0), p2.get("tour_quality", 1.0))
    min_surface = min(p1.get("matches_surface", 99), p2.get("matches_surface", 99))
    min_stability = min(p1.get("stability", 1.0), p2.get("stability", 1.0))

    flags = []
    flags += list(p1.get("flags", []) or [])
    flags += list(p2.get("flags", []) or [])
    flag_text = " | ".join(str(x).lower() for x in flags)

    fav_prob = max(p1_cal, 1 - p1_cal)
    raw_cal_gap = abs(float(p1_raw) - float(p1_cal))

    challenger_pressure = (min_quality < 0.72) or ("challenger" in flag_text) or ("qualy" in flag_text)
    rating_inflation = (raw_cal_gap >= 0.050) or ("inflado" in flag_text) or ("incoherente" in flag_text)
    data_limited = (min_conf < 0.65) or (min_stability < 0.82)
    # v22.27: en clay, un dog con 10-24 partidos de superficie NO es "sin datos";
    # es una muestra corta pero suficiente para exigir prudencia si el favorito sale demasiado limpio.
    dog_has_clay_data = min_surface >= 10
    short_surface_sample = 10 <= min_surface < 25

    # Caso típico: favorito 75-82%, pero con señales internas de fragilidad.
    # v22.27 relaja el trigger para capturar spots tipo big-server / favorito frágil en clay,
    # donde el Elo favorece mucho pero el underdog tiene muestra clay corta y potencial de swing.
    trigger = (
        fav_prob >= 0.75
        and dog_has_clay_data
        and (data_limited or challenger_pressure or short_surface_sample)
        and (rating_inflation or min_quality < 0.70 or min_conf < 0.70 or short_surface_sample)
    )

    if not trigger:
        return p1_cal, fav20, dogset, longm, out

    ml_cap = 0.765
    if min_conf < 0.60:
        ml_cap = 0.745
    if min_quality < 0.66:
        ml_cap -= 0.010
    if raw_cal_gap >= 0.070:
        ml_cap -= 0.010
    # v22.27: muestra clay corta del underdog = no permitir lectura de favorito 80%+ limpia.
    if short_surface_sample and fav_prob >= 0.78:
        ml_cap = min(ml_cap, 0.745)
    ml_cap = float(np.clip(ml_cap, 0.70, 0.77))

    fav20_cap = 0.62
    if min_conf < 0.60 or min_quality < 0.66:
        fav20_cap = 0.58
    if raw_cal_gap >= 0.070:
        fav20_cap -= 0.02
    if short_surface_sample and fav_prob >= 0.78:
        fav20_cap = min(fav20_cap, 0.58)
    fav20_cap = float(np.clip(fav20_cap, 0.52, 0.64))

    dog_floor = 0.34
    if min_surface >= 80:
        dog_floor = 0.36
    if min_conf < 0.58:
        dog_floor += 0.02
    if short_surface_sample and fav_prob >= 0.78:
        dog_floor = max(dog_floor, 0.38)
    dog_floor = float(np.clip(dog_floor, 0.30, 0.42))

    long_floor = 0.30
    if min_surface >= 80:
        long_floor = 0.32
    if short_surface_sample and fav_prob >= 0.78:
        long_floor = 0.34

    fav_is_p1 = p1_cal >= 0.50
    if fav_prob > ml_cap:
        p1_cal = ml_cap if fav_is_p1 else 1 - ml_cap

    fav20 = min(float(fav20), fav20_cap)
    dogset = max(float(dogset), dog_floor)
    longm = max(float(longm), long_floor)

    out["active"] = True
    out["profile"] = "inflated_favorite_upset_risk"
    out["ml_cap"] = ml_cap
    out["fav20_cap"] = fav20_cap
    out["dogset_floor"] = dog_floor
    out["long_floor"] = long_floor
    out["notes"].append("riesgo upset por favorito inflado/confianza limitada")
    if challenger_pressure:
        out["notes"].append("calidad Challenger/Qualy reduce seguridad")
    if rating_inflation:
        out["notes"].append("rating comprimido/inflado")
    if short_surface_sample:
        out["notes"].append("muestra clay corta: favorito vulnerable")

    return float(np.clip(p1_cal, 0.05, 0.95)), float(np.clip(fav20, 0.0, 0.95)), float(np.clip(dogset, 0.05, 0.95)), float(np.clip(longm, 0.05, 0.95)), out



def low_over_long_match_split_guard(circuito, surface, p1_cal, e1, e2, hold1, hold2, raw_tb, rating_sanity, fav20, dogset, longm, upset_risk_guard=None):
    """
    v22.30 Low Over vs Long Match Split.
    Separa una señal fuerte de Over 18.5 de la lectura de partido largo/dog set.
    Caso objetivo: favorito clay fiable 56-64%, mucha muestra, Elo clay elite,
    TB bajo/medio. No toca ML ni overs; solo mercados derivados de sets.
    """
    out = {
        "active": False,
        "profile": "neutral",
        "fav20_boost": 0.0,
        "dogset_cut": 0.0,
        "long_cut": 0.0,
        "notes": []
    }

    if circuito != "ATP" or surface != "Clay":
        return fav20, dogset, longm, out

    if upset_risk_guard and upset_risk_guard.get("active", False):
        return fav20, dogset, longm, out

    rs = rating_sanity or {}
    p1 = rs.get("p1", {}) or {}
    p2 = rs.get("p2", {}) or {}
    min_conf = min(p1.get("confidence", 1.0), p2.get("confidence", 1.0))
    min_surface = min(p1.get("matches_surface", 99), p2.get("matches_surface", 99))
    min_quality = min(p1.get("tour_quality", 1.0), p2.get("tour_quality", 1.0))

    fav_prob = max(float(p1_cal), 1.0 - float(p1_cal))
    fav_is_p1 = p1_cal >= 0.50
    fav_elo = e1 if fav_is_p1 else e2
    dog_elo = e2 if fav_is_p1 else e1
    fav_hold = hold1 if fav_is_p1 else hold2
    dog_hold = hold2 if fav_is_p1 else hold1

    elo_gap = fav_elo - dog_elo
    hold_gap = fav_hold - dog_hold

    # Solo favoritos clay realmente fiables, no favoritos de hard ni favoritos vulnerables.
    trigger = (
        0.555 <= fav_prob <= 0.645
        and fav_elo >= 1900
        and elo_gap >= 65
        and min_conf >= 0.80
        and min_surface >= 80
        and min_quality >= 0.74
        and raw_tb <= 0.32
        and hold_gap >= 0.025
    )

    if not trigger:
        return fav20, dogset, longm, out

    boost = 0.055
    dog_cut = 0.055
    long_cut = 0.035

    if fav_prob >= 0.59:
        boost += 0.015
        dog_cut += 0.010

    if fav_elo >= 1930 and elo_gap >= 80:
        boost += 0.015
        dog_cut += 0.010
        long_cut += 0.010

    # No convertirlo en "spot 2-0 fuerte"; solo corregir el sesgo de dog set/long.
    fav20 = float(np.clip(fav20 + boost, 0.0, 0.58))
    dogset = float(np.clip(dogset - dog_cut, 0.34, 0.95))
    longm = float(np.clip(longm - long_cut, 0.34, 0.95))

    out["active"] = True
    out["profile"] = "low_over_not_long_match"
    out["fav20_boost"] = float(boost)
    out["dogset_cut"] = float(dog_cut)
    out["long_cut"] = float(long_cut)
    out["notes"].append("over bajo fuerte no implica dog set/partido largo")
    out["notes"].append("favorito clay fiable con muestra alta")

    return fav20, dogset, longm, out


def elite_clay_opponent_guard(circuito, surface, p1_cal, e1, e2, rating_sanity, fav20, dogset, longm):
    """
    v22.31 Elite Clay Opponent Guard.
    Evita 2-0 demasiado extremos cuando el underdog también es jugador clay fuerte.
    No toca ML ni overs; solo mercados derivados de sets.
    """
    out = {
        "active": False,
        "profile": "neutral",
        "fav20_cap": None,
        "dogset_floor": None,
        "notes": []
    }

    if circuito != "ATP" or surface != "Clay":
        return fav20, dogset, longm, out

    rs = rating_sanity or {}
    p1 = rs.get("p1", {}) or {}
    p2 = rs.get("p2", {}) or {}

    min_conf = min(p1.get("confidence", 1.0), p2.get("confidence", 1.0))
    min_surface = min(p1.get("matches_surface", 99), p2.get("matches_surface", 99))
    min_quality = min(p1.get("tour_quality", 1.0), p2.get("tour_quality", 1.0))

    fav_prob = max(float(p1_cal), 1.0 - float(p1_cal))
    fav_is_p1 = p1_cal >= 0.50

    fav_elo = e1 if fav_is_p1 else e2
    dog_elo = e2 if fav_is_p1 else e1
    elo_gap = fav_elo - dog_elo

    # Rival clay fuerte: no tratar su set como casi imposible.
    trigger = (
        circuito == "ATP"
        and surface == "Clay"
        and 0.68 <= fav_prob <= 0.80
        and fav20 >= 0.74
        and dog_elo >= 1840
        and elo_gap <= 115
        and min_conf >= 0.78
        and min_surface >= 90
        and min_quality >= 0.72
    )

    if not trigger:
        return fav20, dogset, longm, out

    cap = 0.72
    floor = 0.26

    # Si el favorito no llega al 75%, todavía menos agresivo con el 2-0.
    if fav_prob < 0.75:
        cap = 0.70
        floor = 0.28

    # Si el underdog clay está por encima de 1870, subir un poco su capacidad de set.
    if dog_elo >= 1870:
        cap = min(cap, 0.70)
        floor = max(floor, 0.28)

    fav20 = float(min(fav20, cap))
    dogset = float(max(dogset, floor))

    # Partido largo no lo subimos demasiado: puede ser 7-6 6-4 como Musetti/Cerundolo.
    longm = float(max(longm, 0.30))

    out["active"] = True
    out["profile"] = "elite_clay_opponent"
    out["fav20_cap"] = float(cap)
    out["dogset_floor"] = float(floor)
    out["notes"].append("underdog clay fuerte: 2-0 no debe ser extremo")

    return float(np.clip(fav20, 0.0, 0.95)), float(np.clip(dogset, 0.05, 0.95)), float(np.clip(longm, 0.05, 0.95)), out

def wta_match_script_engine(d1, d2, s1, s2, hold1, hold2, ret1, ret2, surface, circuito):
    """
    v19.2 Rating Sanity Engine
    Detecta perfiles reales WTA:
    - dominadora top
    - clay grinder
    - streaky hitter
    - dangerous dog
    """
    out = {
        "active": False,
        "script": "neutral",
        "fav20_mult": 1.0,
        "dogset_mult": 1.0,
        "long_mult": 1.0,
        "vol_mult": 1.0
    }

    if circuito != "WTA":
        return out

    r1 = d1.get("Rank", 999)
    r2 = d2.get("Rank", 999)

    e1 = d1.get(surface, 1500)
    e2 = d2.get(surface, 1500)

    fav_is_p1 = e1 >= e2
    fav_rank = r1 if fav_is_p1 else r2
    dog_rank = r2 if fav_is_p1 else r1

    fav_hold = hold1 if fav_is_p1 else hold2
    fav_ret = ret1 if fav_is_p1 else ret2

    dog_hold = hold2 if fav_is_p1 else hold1

    hold_gap = abs(hold1 - hold2)

    ace1 = s1.get("Ace%", 0)
    ace2 = s2.get("Ace%", 0)

    # 1) Dominadora top WTA
    if fav_rank <= 12 and fav_ret >= 0.36 and hold_gap >= 0.06:
        out["active"] = True
        out["script"] = "top_dominator"
        out["fav20_mult"] = 1.22
        out["dogset_mult"] = 0.78
        out["long_mult"] = 0.78
        out["vol_mult"] = 0.82

    # 2) Clay grinder
    elif surface == "Clay" and fav_ret >= 0.34 and dog_hold <= 0.58:
        out["active"] = True
        out["script"] = "clay_grinder"
        out["fav20_mult"] = 1.14
        out["dogset_mult"] = 0.84
        out["long_mult"] = 0.88
        out["vol_mult"] = 0.90

    # 3) Streaky hitter
    elif ace1 >= 8 or ace2 >= 8:
        out["active"] = True
        out["script"] = "streaky_hitter"
        out["fav20_mult"] = 0.92
        out["dogset_mult"] = 1.10
        out["long_mult"] = 1.08
        out["vol_mult"] = 1.10

    # 4) Dangerous underdog
    elif dog_rank <= 35 and hold_gap <= 0.04:
        out["active"] = True
        out["script"] = "dangerous_dog"
        out["fav20_mult"] = 0.88
        out["dogset_mult"] = 1.14
        out["long_mult"] = 1.12
        out["vol_mult"] = 1.12

    return out


def elite_wta_separation(d1, d2, hold1, hold2, ret1, ret2, surface, circuito):
    """
    v19.1 Rating Sanity Engine.
    La WTA no solo es caos: las top consolidan ventajas y castigan más.
    Ajusta solo WTA y protege ATP.
    """
    info = {
        "active": False,
        "hold_adj1": 0.0,
        "hold_adj2": 0.0,
        "ret_adj1": 0.0,
        "ret_adj2": 0.0,
        "vol_mult": 1.0,
        "top_player": "",
        "edge_label": "normal"
    }

    if circuito != "WTA":
        return hold1, hold2, ret1, ret2, info

    r1 = d1.get("Rank", 999)
    r2 = d2.get("Rank", 999)
    e1 = d1.get(surface, 1500)
    e2 = d2.get(surface, 1500)

    elo_gap = abs(e1 - e2)
    rank_gap = abs(r1 - r2)

    p1_top = r1 <= 15
    p2_top = r2 <= 15

    # Jugadora top con ventaja Elo/ranking: más separación real.
    if p1_top and (e1 >= e2 + 80 or r1 + 20 <= r2):
        info["active"] = True
        info["top_player"] = d1.get("Player", "Jugador 1")
        info["hold_adj1"] += 0.020
        info["ret_adj1"] += 0.020
        info["hold_adj2"] -= 0.010
        info["vol_mult"] *= 0.82
        info["edge_label"] = "top_wta_p1"

    if p2_top and (e2 >= e1 + 80 or r2 + 20 <= r1):
        info["active"] = True
        info["top_player"] = d2.get("Player", "Jugador 2")
        info["hold_adj2"] += 0.020
        info["ret_adj2"] += 0.020
        info["hold_adj1"] -= 0.010
        info["vol_mult"] *= 0.82
        info["edge_label"] = "top_wta_p2"

    # Diferencia grande aunque no sea top15.
    if rank_gap >= 45 and elo_gap >= 120:
        info["active"] = True
        if e1 > e2:
            info["hold_adj1"] += 0.012
            info["ret_adj1"] += 0.012
            info["hold_adj2"] -= 0.006
        else:
            info["hold_adj2"] += 0.012
            info["ret_adj2"] += 0.012
            info["hold_adj1"] -= 0.006
        info["vol_mult"] *= 0.90
        if info["edge_label"] == "normal":
            info["edge_label"] = "rank_elo_gap"

    # En clay dejamos algo más de swing, pero no anulamos élite.
    if surface == "Clay":
        info["vol_mult"] = min(1.0, info["vol_mult"] * 1.04)

    hold1 = np.clip(hold1 + info["hold_adj1"], 0.42, 0.86)
    hold2 = np.clip(hold2 + info["hold_adj2"], 0.42, 0.86)
    ret1 = np.clip(ret1 + info["ret_adj1"], 0.10, 0.42)
    ret2 = np.clip(ret2 + info["ret_adj2"], 0.10, 0.42)

    return hold1, hold2, ret1, ret2, info



def clay_return_weight_engine(d1, d2, s1, s2, hold1, hold2, ret1, ret2, surface, circuito):
    out = {
        "active": False, "profile": "neutral", "vol_mult": 1.0,
        "hold_adj1": 0.0, "hold_adj2": 0.0,
        "ret_adj1": 0.0, "ret_adj2": 0.0,
        "notes": []
    }
    if circuito != "ATP" or surface != "Clay":
        return hold1, hold2, ret1, ret2, out

    e1, e2 = d1.get("Clay",1500), d2.get("Clay",1500)
    r1, r2 = d1.get("Rank",999), d2.get("Rank",999)
    gap = abs(e1 - e2)
    ace1, ace2 = s1.get("ace",0.05), s2.get("ace",0.05)

    if ret1 >= 0.335 and hold1 <= 0.705:
        out["active"] = True
        out["profile"] = "clay_grinder_p1"
        out["ret_adj1"] += 0.018
        out["hold_adj1"] -= 0.006
        out["vol_mult"] *= 0.94
        out["notes"].append("P1 grinder")

    if ret2 >= 0.335 and hold2 <= 0.705:
        out["active"] = True
        out["profile"] = "clay_grinder_p2"
        out["ret_adj2"] += 0.018
        out["hold_adj2"] -= 0.006
        out["vol_mult"] *= 0.94
        out["notes"].append("P2 grinder")

    if gap >= 115:
        out["active"] = True
        if e1 > e2:
            out["profile"] = "clay_specialist_p1"
            out["ret_adj1"] += 0.018
            out["hold_adj1"] += 0.008
            out["hold_adj2"] -= 0.006
            out["vol_mult"] *= 0.89
            out["notes"].append("P1 clay edge")
        else:
            out["profile"] = "clay_specialist_p2"
            out["ret_adj2"] += 0.018
            out["hold_adj2"] += 0.008
            out["hold_adj1"] -= 0.006
            out["vol_mult"] *= 0.89
            out["notes"].append("P2 clay edge")

    if e1 > e2 + 80 and r1 + 35 <= r2:
        out["active"] = True
        out["profile"] = "clay_favorite_p1"
        out["ret_adj1"] += 0.012
        out["hold_adj1"] += 0.006
        out["vol_mult"] *= 0.92
        out["notes"].append("P1 rank+clay edge")

    if e2 > e1 + 80 and r2 + 35 <= r1:
        out["active"] = True
        out["profile"] = "clay_favorite_p2"
        out["ret_adj2"] += 0.012
        out["hold_adj2"] += 0.006
        out["vol_mult"] *= 0.92
        out["notes"].append("P2 rank+clay edge")

    if ace1 >= 0.105 and ret1 <= 0.275:
        out["active"] = True
        out["hold_adj1"] -= 0.014
        out["vol_mult"] *= 1.04
        out["notes"].append("P1 big server clay nerf")
        if out["profile"] == "neutral":
            out["profile"] = "big_server_nerf_p1"

    if ace2 >= 0.105 and ret2 <= 0.275:
        out["active"] = True
        out["hold_adj2"] -= 0.014
        out["vol_mult"] *= 1.04
        out["notes"].append("P2 big server clay nerf")
        if out["profile"] == "neutral":
            out["profile"] = "big_server_nerf_p2"

    out["hold_adj1"] = float(np.clip(out["hold_adj1"], -0.025, 0.030))
    out["hold_adj2"] = float(np.clip(out["hold_adj2"], -0.025, 0.030))
    out["ret_adj1"] = float(np.clip(out["ret_adj1"], -0.005, 0.035))
    out["ret_adj2"] = float(np.clip(out["ret_adj2"], -0.005, 0.035))
    out["vol_mult"] = float(np.clip(out["vol_mult"], 0.82, 1.10))

    hold1 = np.clip(hold1 + out["hold_adj1"], 0.42, 0.84)
    hold2 = np.clip(hold2 + out["hold_adj2"], 0.42, 0.84)
    ret1 = np.clip(ret1 + out["ret_adj1"], 0.10, 0.42)
    ret2 = np.clip(ret2 + out["ret_adj2"], 0.10, 0.42)

    return hold1, hold2, ret1, ret2, out



def elite_atp_clay_protection(d1, d2, hold1, hold2, ret1, ret2, surface, circuito):
    """
    v21.1 Protección para élites ATP en clay.
    Evita compresión artificial tipo coinflip para top5 con clara ventaja clay.
    """
    out = {
        "active": False,
        "fav": "",
        "hold_boost": 0.0,
        "ret_boost": 0.0,
        "dog_hold_nerf": 0.0,
        "vol_mult": 1.0
    }

    if circuito != "ATP" or surface != "Clay":
        return hold1, hold2, ret1, ret2, out

    r1, r2 = d1.get("Rank", 999), d2.get("Rank", 999)
    e1, e2 = d1.get("Clay", 1500), d2.get("Clay", 1500)

    gap = abs(e1 - e2)
    rank_gap = abs(r1 - r2)

    p1_elite = r1 <= 5 and e1 >= 1950
    p2_elite = r2 <= 5 and e2 >= 1950

    if p1_elite and e1 > e2 and gap >= 130 and rank_gap >= 20:
        out["active"] = True
        out["fav"] = "p1"
        out["hold_boost"] = 0.018
        out["ret_boost"] = 0.010
        out["dog_hold_nerf"] = -0.010
        out["vol_mult"] = 0.84

    elif p2_elite and e2 > e1 and gap >= 130 and rank_gap >= 20:
        out["active"] = True
        out["fav"] = "p2"
        out["hold_boost"] = 0.018
        out["ret_boost"] = 0.010
        out["dog_hold_nerf"] = -0.010
        out["vol_mult"] = 0.84

    if out["active"]:
        if out["fav"] == "p1":
            hold1 += out["hold_boost"]
            ret1 += out["ret_boost"]
            hold2 += out["dog_hold_nerf"]
        else:
            hold2 += out["hold_boost"]
            ret2 += out["ret_boost"]
            hold1 += out["dog_hold_nerf"]

    hold1 = np.clip(hold1, 0.42, 0.86)
    hold2 = np.clip(hold2, 0.42, 0.86)
    ret1 = np.clip(ret1, 0.10, 0.42)
    ret2 = np.clip(ret2, 0.10, 0.42)

    return hold1, hold2, ret1, ret2, out


def clay_market_adjustments(circuito, surface, clay_engine, fav20, dogset, longm, p1_cal):
    if circuito != "ATP" or surface != "Clay" or not clay_engine.get("active", False):
        return fav20, dogset, longm

    profile = clay_engine.get("profile", "")
    fav_prob = max(p1_cal, 1 - p1_cal)

    if "grinder" in profile:
        longm *= 1.06
        dogset *= 1.03

    if "specialist" in profile or "favorite" in profile:
        if fav_prob >= 0.62:
            fav20 *= 1.08
            dogset *= 0.94
            longm *= 0.96
        if fav_prob >= 0.70:
            fav20 *= 1.10
            dogset *= 0.88
            longm *= 0.92

    if "big_server_nerf" in profile:
        dogset *= 1.06
        longm *= 1.04

    return (
        float(np.clip(fav20, 0.0, 0.95)),
        float(np.clip(dogset, 0.05, 0.95)),
        float(np.clip(longm, 0.05, 0.95))
    )



def clay_dominance_preservation(d1, d2, hold1, hold2, ret1, ret2, surface, circuito):
    out = {
        "active": False, "fav": "", "profile": "neutral", "vol_mult": 1.0,
        "market_fav20_mult": 1.0, "market_dogset_mult": 1.0, "market_long_mult": 1.0
    }
    if circuito != "ATP" or surface != "Clay":
        return hold1, hold2, ret1, ret2, out

    r1, r2 = d1.get("Rank",999), d2.get("Rank",999)
    e1, e2 = d1.get("Clay",1500), d2.get("Clay",1500)
    h1elo, h2elo = d1.get("Hard",1500), d2.get("Hard",1500)

    if e1 >= e2:
        fav = "p1"; fav_rank = r1; fav_elo = e1; dog_elo = e2; fav_delta = e1-h1elo; dog_delta = e2-h2elo
    else:
        fav = "p2"; fav_rank = r2; fav_elo = e2; dog_elo = e1; fav_delta = e2-h2elo; dog_delta = e1-h1elo

    elo_gap = fav_elo - dog_elo
    elite_specialist = fav_elo >= 1920 and elo_gap >= 70
    clay_identity = fav_delta >= dog_delta + 45 and elo_gap >= 60
    proven = fav_rank <= 30 and elo_gap >= 75

    if elite_specialist or clay_identity or proven:
        out["active"] = True
        out["fav"] = fav
        out["profile"] = "clay_dominator"
        hold_boost = 0.022 if elite_specialist else 0.016
        ret_boost = 0.016 if elite_specialist else 0.012
        dog_hold_nerf = -0.012
        dog_ret_nerf = -0.006
        out["vol_mult"] = 0.82 if elite_specialist else 0.88
        out["market_fav20_mult"] = 1.24 if elite_specialist else 1.16
        out["market_dogset_mult"] = 0.72 if elite_specialist else 0.82
        out["market_long_mult"] = 0.76 if elite_specialist else 0.84

        if fav == "p1":
            hold1 += hold_boost; ret1 += ret_boost; hold2 += dog_hold_nerf; ret2 += dog_ret_nerf
        else:
            hold2 += hold_boost; ret2 += ret_boost; hold1 += dog_hold_nerf; ret1 += dog_ret_nerf

    return (
        np.clip(hold1,0.42,0.86),
        np.clip(hold2,0.42,0.86),
        np.clip(ret1,0.10,0.42),
        np.clip(ret2,0.10,0.42),
        out
    )


# =========================================================
# v22 RATING SANITY ENGINE
# =========================================================

def rating_sanity_engine(d1, d2, surface, circuito):
    """
    v22.1 Match Count + Tour Quality Engine.
    Combina el sanity v22 con evidencia real de históricos:
    - nº partidos por superficie
    - calidad Tour / Challenger / ITF
    - stability score
    - confidence real del Elo
    - penalización de ratings pequeños o inflados
    """
    def player_sanity(d):
        rank = d.get("Rank", 999)
        elo_surface = d.get(surface, 1500)
        elo_hard = d.get("Hard", 1500)
        elo_clay = d.get("Clay", 1500)
        elo_grass = d.get("Grass", 1500)
        q = d.get("Quality", {}) or {}
        # v22.18: fuerza un conteo directo cacheado y lo usa si aporta más partidos.
        # Esto corrige el caso en el que el QualityMap global queda indexado solo con ATP Tour
        # aunque los CSV Challenger estén cargados. Al estar cacheado, no bloquea la simulación.
        q_direct = buscar_quality_directo_historicos(d.get("Player", ""), circuito)
        try:
            q_total = int(q.get("matches_total", 0) or 0)
            d_total = int(q_direct.get("matches_total", 0) or 0)
            q_ch = int((q.get("level_counts", {}) or {}).get("challenger", 0) or 0) + int((q.get("level_counts", {}) or {}).get("qualy", 0) or 0)
            d_ch = int((q_direct.get("level_counts", {}) or {}).get("challenger", 0) or 0) + int((q_direct.get("level_counts", {}) or {}).get("qualy", 0) or 0)
        except Exception:
            q_total, d_total, q_ch, d_ch = 0, 0, 0, 0
        if (d_total > q_total) or (d_ch > q_ch) or (not q.get("quality_meta")) or q.get("matched_name") in [None, "", "NO ENCONTRADO"]:
            q = q_direct

        matches_total = q.get("matches_total", 0)
        surface_counts = q.get("matches_surface", {}) if isinstance(q.get("matches_surface", {}), dict) else {}
        matches_surface = surface_counts.get(surface, 0)
        level_counts = q.get("level_counts", {})
        tour_quality = q.get("tour_quality", 0.45)
        stability = q.get("stability", {}).get(surface, 0.05)
        data_conf = q.get("confidence", {}).get(surface, 0.32)

        confidence = float(data_conf)
        flags = []

        if matches_total == 0 and q.get("match_score", 0.0) == 0.0:
            flags.append("jugador no encontrado en históricos cargados")

        if matches_surface < 5:
            confidence -= 0.16
            flags.append("muy pocos partidos superficie")
        elif matches_surface < 10:
            confidence -= 0.08
            flags.append("muestra superficie baja")
        elif matches_surface < 18:
            confidence -= 0.04
            flags.append("muestra superficie media")

        if matches_total < 20:
            confidence -= 0.06
            flags.append("pocos partidos totales")

        if matches_surface < 10 and matches_total < 30 and level_counts.get("tour", 0) >= max(1, int(matches_total * 0.80)):
            flags.append("muestra ATP pequeña")

        if tour_quality < 0.52:
            confidence -= 0.10
            flags.append("calidad ITF/unknown alta")
        elif tour_quality < 0.68:
            confidence -= 0.05
            flags.append("calidad challenger/qualy")

        # Elo muy alto con ranking no equivalente
        if elo_surface >= 1900 and rank > 50:
            confidence -= 0.18
            flags.append("elo_surface_inflado")

        if elo_surface >= 1850 and rank > 80:
            confidence -= 0.22
            flags.append("elo/rank incoherente")

        # Delta clay-hard sospechoso
        if surface == "Clay" and (elo_clay - elo_hard) >= 180:
            confidence -= 0.13
            flags.append("delta clay-hard alto")

        # Jugador fuera top100 con Elo top
        if rank > 100 and elo_surface >= 1800:
            confidence -= 0.18
            flags.append("posible challenger/junior inflation")

        confidence = float(np.clip(confidence, 0.25, 1.00))

        # Elo efectivo: si la confianza es baja, reduce ventaja extrema hacia su Elo general medio.
        baseline = np.mean([elo_hard, elo_clay, elo_grass])
        shrink = (1.0 - confidence) * 0.52
        elo_effective = elo_surface - (elo_surface - baseline) * shrink

        # Penalización adicional para Elo alto con muestra débil.
        elo_penalty = 0.0
        if elo_surface >= 1800 and confidence < 0.62:
            elo_penalty = min(70, (0.62 - confidence) * 150)
            elo_effective -= elo_penalty

        return {
            "confidence": confidence,
            "flags": flags,
            "rank": rank,
            "elo_surface": elo_surface,
            "elo_effective": float(elo_effective),
            "elo_penalty": float(elo_surface - elo_effective),
            "matches_total": int(matches_total),
            "matches_surface": int(matches_surface),
            "matches_by_surface": {"Hard": int(surface_counts.get("Hard", 0)), "Clay": int(surface_counts.get("Clay", 0)), "Grass": int(surface_counts.get("Grass", 0))},
            "tour_quality": float(tour_quality),
            "stability": float(stability),
            "level_counts": level_counts,
            "matched_name": q.get("matched_name", "N/A"),
            "match_score": float(q.get("match_score", 0.0)),
            "raw_names": q.get("raw_names", []),
            "source_files": q.get("source_files", []),
            "quality_map_size": int(q.get("quality_map_size", 0)),
            "quality_meta": q.get("quality_meta", {}),
            "candidate_matches": q.get("candidate_matches", [])
        }

    s1 = player_sanity(d1)
    s2 = player_sanity(d2)

    min_conf = min(s1["confidence"], s2["confidence"])
    vol_mult = 1.0
    if min_conf < 0.45:
        vol_mult = 1.16
    elif min_conf < 0.60:
        vol_mult = 1.11
    elif min_conf < 0.75:
        vol_mult = 1.06

    return {
        "p1": s1,
        "p2": s2,
        "vol_mult": vol_mult,
        "active": bool(s1["flags"] or s2["flags"]),
        "version": APP_VERSION
    }


def sim_match(d1, d2, surface, circuito, best_of=3, n=5000, context_row=None, progress_callback=None):
    e1, e2 = d1[surface], d2[surface]

    # v22.1 Match Count + Tour Quality Engine
    rating_sanity = rating_sanity_engine(d1, d2, surface, circuito)
    e1_eff = rating_sanity.get("p1", {}).get("elo_effective", e1)
    e2_eff = rating_sanity.get("p2", {}).get("elo_effective", e2)
    elo_diff = e1_eff - e2_eff

    # v15: usa stats específicas de superficie si existen
    s1 = get_stats_surface(d1, surface)
    s2 = get_stats_surface(d2, surface)

    raw1 = calc_hold(s1, elo_diff, surface, circuito)
    raw2 = calc_hold(s2, -elo_diff, surface, circuito)
    ret1 = calc_return_strength(s1, e1, surface)
    ret2 = calc_return_strength(s2, e2, surface)

    # v19 WTA Return Compression:
    # evita que todas las jugadoras parezcan returners elite y vuelvan el partido coinflip.
    if circuito == "WTA":
        ret1 *= 0.90
        ret2 *= 0.90

        if d1.get("Rank", 999) <= 10:
            ret1 *= 1.06
        elif d1.get("Rank", 999) <= 20:
            ret1 *= 1.03

        if d2.get("Rank", 999) <= 10:
            ret2 *= 1.06
        elif d2.get("Rank", 999) <= 20:
            ret2 *= 1.03

        ret1 = np.clip(ret1, 0.10, 0.38)
        ret2 = np.clip(ret2, 0.10, 0.38)

    hold1, hold2 = aplicar_return_pressure(raw1, raw2, ret1, ret2, surface)

    # v21 Rating Sanity Engine
    hold1, hold2, ret1, ret2, clay_engine = clay_return_weight_engine(
        d1, d2, s1, s2, hold1, hold2, ret1, ret2, surface, circuito
    )

    # v21.1 Elite ATP Clay Protection
    hold1, hold2, ret1, ret2, elite_clay = elite_atp_clay_protection(
        d1, d2, hold1, hold2, ret1, ret2, surface, circuito
    )

    # v21.2 Clay Dominance Preservation
    hold1, hold2, ret1, ret2, dominance_clay = clay_dominance_preservation(
        d1, d2, hold1, hold2, ret1, ret2, surface, circuito
    )

    # v17 Fatigue Engine
    fatigue1 = d1.get("Fatigue", {})
    fatigue2 = d2.get("Fatigue", {})
    fat_adj1, fat_adj2, fatigue_vol_extra = fatigue_adjustments(fatigue1, fatigue2, surface, circuito, d1.get('Rank',999), d2.get('Rank',999))

    hold1 = np.clip(hold1 + fat_adj1, 0.42, 0.84)
    hold2 = np.clip(hold2 + fat_adj2, 0.42, 0.84)

    ret1 = np.clip(ret1 + fat_adj1 * 0.35, 0.10, 0.40)
    ret2 = np.clip(ret2 + fat_adj2 * 0.35, 0.10, 0.40)

    # v18 Tournament Engine
    ctx = tournament_context_adjustments(
        context_row,
        p1_name=s1.get("raw_name_stats", ""),
        p2_name=s2.get("raw_name_stats", "")
    )

    hold1 = np.clip(hold1 + ctx["p1_adj"], 0.42, 0.86)
    hold2 = np.clip(hold2 + ctx["p2_adj"], 0.42, 0.86)

    # v19.1 Elite WTA Separation
    hold1, hold2, ret1, ret2, wta_sep = elite_wta_separation(
        d1, d2, hold1, hold2, ret1, ret2, surface, circuito
    )

    # v19.2 Match Script Engine
    wta_script = wta_match_script_engine(
        d1, d2, s1, s2,
        hold1, hold2,
        ret1, ret2,
        surface, circuito
    )

    p1_profile = s1.get("serve_profile", "normal")
    p2_profile = s2.get("serve_profile", "normal")
    p1_big = p1_profile in ["big_server", "elite_server"]
    p2_big = p2_profile in ["big_server", "elite_server"]

    tb_intel_boost = calcular_tiebreak_boost(s1, s2, hold1, hold2, surface, p1_big, p2_big)
    tb_intel_boost += ctx.get("tb_adj", 0)

    pressure_skill1, pressure_skill2 = pressure_collapse_params(s1, s2, surface)

    sets_to_win = 3 if best_of == 5 else 2
    fav_est = max(elo_prob(e1_eff, e2_eff), 1 - elo_prob(e1_eff, e2_eff))

    vol = calcular_match_volatility(e1, e2, surface, fav_est)
    vol *= rating_sanity.get("vol_mult", 1.0)
    if p1_big or p2_big:
        vol += 0.006
    vol += fatigue_vol_extra
    vol += ctx.get("vol_adj", 0)

    if circuito == "ATP" and surface == "Clay":
        vol *= clay_engine.get("vol_mult", 1.0)
        vol *= elite_clay.get("vol_mult", 1.0)
        vol *= dominance_clay.get("vol_mult", 1.0)
        vol = float(np.clip(vol, 0.016, 0.070))

    # v19 WTA Controlled Chaos:
    # WTA sigue siendo variable, pero protegemos top players y gaps grandes.
    if circuito == "WTA":
        rank_gap = abs(d1.get("Rank",999) - d2.get("Rank",999))

        if d1.get("Rank",999) <= 15 or d2.get("Rank",999) <= 15:
            vol *= 0.78

        if rank_gap >= 40 and fav_est >= 0.62:
            vol *= 0.86

        if surface == "Clay":
            vol *= 1.04
        elif surface == "Grass":
            vol *= 0.94

        vol *= wta_sep.get("vol_mult", 1.0)
        vol *= wta_script.get("vol_mult", 1.0)
        vol = float(np.clip(vol, 0.014, 0.060))

    res = {
        "p1": 0, "p2": 0, "set3": 0, "tb": 0, "games": [],
        # v22.35: juegos esperados por jugador
        "games_p1": [], "games_p2": [],
        "p1_fs": 0, "p2_fs": 0,
        "fav_under22": 0, "dog_over20": 0,
        "fav_2_0": 0, "dog_wins_set": 0, "long_match": 0,
        # v22.25 Favorite Identity Engine:
        # count straight sets / set wins by player, then decide favorite by final model probability,
        # not by raw surface Elo. This avoids labels such as "underdog wins set" being tied to Elo
        # when the final model has flipped the favorite.
        "p1_2_0": 0, "p2_2_0": 0,
        "p1_wins_set_any": 0, "p2_wins_set_any": 0,
        # v23.38 Grand Slam BO5 Engine: contadores separados para 5 sets.
        # No modifica el motor Over BO3; solo expone resultados simulados cuando best_of=5.
        "set4_bo5": 0, "set5_bo5": 0,
        "p1_3_0": 0, "p2_3_0": 0,
        "fav_3_0_bo5": 0, "dog_wins_set_bo5": 0
    }

    # v22.36: progreso más suave sin saturar Streamlit
    progress_step = max(1, min(250, n // 40))

    if progress_callback is not None:
        try:
            progress_callback(0, n)
        except Exception:
            pass

    for sim_i in range(n):
        sets1 = sets2 = games = 0
        games_p1 = games_p2 = 0
        tb_seen = False
        first_done = False
        shift = np.random.normal(0, vol)

        while sets1 < sets_to_win and sets2 < sets_to_win:
            g1, g2, tb = sim_set(
                hold1, hold2, surface, shift, p1_big, p2_big, fav_est,
                stats1=s1, stats2=s2
            )

            games += g1 + g2
            games_p1 += g1
            games_p2 += g2

            if tb:
                tb_seen = True

            if not first_done:
                if g1 > g2:
                    res["p1_fs"] += 1
                else:
                    res["p2_fs"] += 1
                first_done = True

            if g1 > g2:
                sets1 += 1
            else:
                sets2 += 1

        p1_wins = sets1 > sets2

        if p1_wins:
            res["p1"] += 1
        else:
            res["p2"] += 1

        if (sets1, sets2) in [(2, 1), (1, 2)]:
            res["set3"] += 1

        if tb_seen:
            res["tb"] += 1

        # v22.25: player-specific set outcomes for final-model favorite markets.
        if sets1 == sets_to_win and sets2 == 0:
            res["p1_2_0"] += 1
        if sets2 == sets_to_win and sets1 == 0:
            res["p2_2_0"] += 1
        if sets1 >= 1:
            res["p1_wins_set_any"] += 1
        if sets2 >= 1:
            res["p2_wins_set_any"] += 1

        # v23.38 Grand Slam BO5 Engine: resultados específicos para partidos al mejor de 5.
        if sets_to_win == 3:
            total_sets_played = sets1 + sets2
            if total_sets_played == 4:
                res["set4_bo5"] += 1
            elif total_sets_played == 5:
                res["set5_bo5"] += 1
            if sets1 == 3 and sets2 == 0:
                res["p1_3_0"] += 1
            if sets2 == 3 and sets1 == 0:
                res["p2_3_0"] += 1

        p1_is_fav = e1 >= e2
        fav_wins = p1_wins if p1_is_fav else (not p1_wins)
        dog_wins = not fav_wins

        if fav_wins and games < 22.5:
            res["fav_under22"] += 1

        if dog_wins and games > 20.5:
            res["dog_over20"] += 1

        if p1_is_fav:
            dog_sets = sets2
        else:
            dog_sets = sets1

        if fav_wins and dog_sets == 0:
            res["fav_2_0"] += 1

        if dog_sets >= 1:
            res["dog_wins_set"] += 1

        if sets_to_win == 3:
            if fav_wins and dog_sets == 0:
                res["fav_3_0_bo5"] += 1
            if dog_sets >= 1:
                res["dog_wins_set_bo5"] += 1

        if games > 22.5 or tb_seen or ((sets1, sets2) in [(2, 1), (1, 2)]):
            res["long_match"] += 1

        res["games"].append(games)
        res["games_p1"].append(games_p1)
        res["games_p2"].append(games_p2)

        if progress_callback is not None and ((sim_i + 1) % progress_step == 0 or (sim_i + 1) == n):
            try:
                progress_callback(sim_i + 1, n)
            except Exception:
                pass

    p1_raw = res["p1"] / n
    # v22.29 compatibility aliases: older blocks used p1_win/p2_win names
    p1_win = p1_raw
    p2_win = 1 - p1_raw
    p1_cal = calibrar_probabilidad(p1_raw, surface)

    raw_tb = res["tb"] / n

    # v22.25 Favorite Identity Engine:
    # Favorite/underdog derived markets follow the calibrated model favorite, not raw Elo.
    model_fav_is_p1 = p1_cal >= 0.50
    model_fav_name = d1.get("Player", "Jugador 1") if model_fav_is_p1 else d2.get("Player", "Jugador 2")
    model_dog_name = d2.get("Player", "Jugador 2") if model_fav_is_p1 else d1.get("Player", "Jugador 1")
    raw_fav20 = (res["p1_2_0"] if model_fav_is_p1 else res["p2_2_0"]) / n
    raw_dogset = (res["p2_wins_set_any"] if model_fav_is_p1 else res["p1_wins_set_any"]) / n
    raw_long = res["long_match"] / n

    # v18.1 Rating Sanity Engine: se aplica después de simular, no antes.
    set_dyn = smart_set_dynamics(
        p1_raw, 1 - p1_raw,
        hold1, hold2,
        raw_tb,
        vol,
        surface,
        circuito,
        d1.get("Rank",999),
        d2.get("Rank",999)
    )

    fav20 = float(np.clip(raw_fav20 + set_dyn["fav20_boost"], 0.0, 0.95))
    dogset = float(np.clip(raw_dogset - set_dyn["dog_set_suppress"], 0.05, 0.95))
    longm = float(np.clip(raw_long + set_dyn["long_match_adj"], 0.05, 0.95))

    # v19 WTA caps: corrige exceso de "dog gana set" en favoritas claras.
    if circuito == "WTA":
        fav_prob = max(p1_cal, 1 - p1_cal)
        top_player = d1.get("Rank",999) <= 15 or d2.get("Rank",999) <= 15
        rank_gap = abs(d1.get("Rank",999) - d2.get("Rank",999))

        if fav_prob >= 0.70 and top_player:
            dogset *= 0.78
            fav20 = min(0.90, fav20 * 1.12)
            longm *= 0.88
        elif fav_prob >= 0.65 and top_player:
            dogset *= 0.85
            fav20 = min(0.88, fav20 * 1.08)
            longm *= 0.92

        if rank_gap >= 40 and fav_prob >= 0.63:
            dogset *= 0.88
            longm *= 0.94

        # v19.1 Elite WTA set separation:
        # Si hay top WTA con ventaja, no convertir automáticamente en partido largo.
        if wta_sep.get("active", False):
            fav_prob2 = max(p1_cal, 1 - p1_cal)
            if fav_prob2 >= 0.60:
                dogset *= 0.82
                fav20 = min(0.92, fav20 * 1.16)
                longm *= 0.82
            elif fav_prob2 >= 0.56:
                dogset *= 0.90
                fav20 = min(0.90, fav20 * 1.08)
                longm *= 0.90

        # v19.2 Match Script multipliers
        fav20 *= wta_script.get("fav20_mult", 1.0)
        dogset *= wta_script.get("dogset_mult", 1.0)
        longm *= wta_script.get("long_mult", 1.0)

        dogset = float(np.clip(dogset, 0.05, 0.90))
        fav20 = float(np.clip(fav20, 0.0, 0.92))
        longm = float(np.clip(longm, 0.05, 0.92))

    # v21 ATP Clay market logic
    fav20, dogset, longm = clay_market_adjustments(
        circuito, surface, clay_engine, fav20, dogset, longm, p1_cal
    )

    # v21.1 Elite ATP protection market logic
    if circuito == "ATP" and surface == "Clay" and elite_clay.get("active", False):
        fav20 *= 1.18
        dogset *= 0.78
        longm *= 0.84

        fav20 = float(np.clip(fav20, 0.0, 0.95))
        dogset = float(np.clip(dogset, 0.05, 0.95))
        longm = float(np.clip(longm, 0.05, 0.95))

    # v21.2 Clay Dominance market logic
    if circuito == "ATP" and surface == "Clay" and dominance_clay.get("active", False):
        fav20 *= dominance_clay.get("market_fav20_mult", 1.0)
        dogset *= dominance_clay.get("market_dogset_mult", 1.0)
        longm *= dominance_clay.get("market_long_mult", 1.0)
        fav20 = float(np.clip(fav20, 0.0, 0.95))
        dogset = float(np.clip(dogset, 0.05, 0.95))
        longm = float(np.clip(longm, 0.05, 0.95))

    # v22.20 Straight Sets Guard: buen over bajo no debe inflar dog set/partido largo.
    straight_sets_guard = {"active": False, "profile": "neutral"}
    fav_prob_guard = max(p1_cal, 1 - p1_cal)
    fav20, dogset, longm, straight_sets_guard = straight_sets_guard_engine(
        circuito, surface, fav_prob_guard, hold1, hold2, raw_tb, vol,
        rating_sanity, fav20, dogset, longm
    )

    fav20_sanity_guard = {"active": False}
    fav20, dogset, longm, fav20_sanity_guard = favorite_2_0_sanity_guard(
        circuito, surface, fav_prob_guard, fav20, dogset, longm, rating_sanity
    )

    upset_risk_guard = {"active": False, "profile": "neutral"}
    p1_cal, fav20, dogset, longm, upset_risk_guard = upset_risk_guard_engine(
        circuito, surface, p1_cal, p1_raw, fav20, dogset, longm, rating_sanity
    )

    low_over_split_guard = {"active": False, "profile": "neutral"}
    fav20, dogset, longm, low_over_split_guard = low_over_long_match_split_guard(
        circuito, surface, p1_cal, e1, e2, hold1, hold2, raw_tb,
        rating_sanity, fav20, dogset, longm, upset_risk_guard
    )

    elite_clay_opponent_guard_info = {"active": False, "profile": "neutral"}
    fav20, dogset, longm, elite_clay_opponent_guard_info = elite_clay_opponent_guard(
        circuito, surface, p1_cal, e1, e2, rating_sanity, fav20, dogset, longm
    )

    return {
        "p1": p1_raw,
        "p2": 1 - p1_raw,
        "p1_cal": p1_cal,
        "p2_cal": 1 - p1_cal,
        "p1_fs": res["p1_fs"] / n,
        "p2_fs": res["p2_fs"] / n,
        "set3": res["set3"] / n,
        "tb": raw_tb,
        "fav_under22": res["fav_under22"] / n,
        "dog_over20": res["dog_over20"] / n,
        "fav_2_0": fav20,
        "dog_wins_set": dogset,
        # v23.37: probabilidades exactas de "gana al menos 1 set" por jugador.
        # No toca el motor Over; solo expone resultados ya simulados para el selector de máximo acierto.
        "p1_wins_set_any": res["p1_wins_set_any"] / n,
        "p2_wins_set_any": res["p2_wins_set_any"] / n,
        "p1_2_0": res["p1_2_0"] / n,
        "p2_2_0": res["p2_2_0"] / n,
        # v23.38 Grand Slam BO5 Engine outputs.
        "set4_bo5": res["set4_bo5"] / n,
        "set5_bo5": res["set5_bo5"] / n,
        "p1_3_0": res["p1_3_0"] / n,
        "p2_3_0": res["p2_3_0"] / n,
        "fav_3_0_bo5": res["fav_3_0_bo5"] / n,
        "dog_wins_set_bo5": res["dog_wins_set_bo5"] / n,
        "long_match": longm,
        "games": res["games"],
        "games_p1": res["games_p1"],
        "games_p2": res["games_p2"],
        "hold1": hold1,
        "hold2": hold2,
        "raw_hold1": raw1,
        "raw_hold2": raw2,
        "ret1": ret1,
        "ret2": ret2,
        "p1_profile": p1_profile,
        "p2_profile": p2_profile,
        # v23.39.4: exponer ranking al selector WTA. No cambia simulación ni Over.
        "rank1": d1.get("Rank", 999),
        "rank2": d2.get("Rank", 999),
        "rank_gap": abs(d1.get("Rank", 999) - d2.get("Rank", 999)),
        "vol": vol,
        "fav_raw_est": fav_est,
        "model_fav_is_p1": model_fav_is_p1,
        "model_fav_name": model_fav_name,
        "model_dog_name": model_dog_name,
        "elo_effective1": e1_eff,
        "elo_effective2": e2_eff,
        "tb_intel_boost": tb_intel_boost,
        "pressure_skill1": pressure_skill1,
        "pressure_skill2": pressure_skill2,
        "fatigue1": fatigue1,
        "fatigue2": fatigue2,
        "fatigue_adj1": fat_adj1,
        "fatigue_adj2": fat_adj2,
        "fatigue_vol_extra": fatigue_vol_extra,
        "tournament_ctx": ctx,
        "set_dynamics": set_dyn,
        "wta_engine_active": circuito == "WTA",
        "clay_engine": clay_engine,
        "rating_sanity": rating_sanity,
        "elite_clay": elite_clay,
        "dominance_clay": dominance_clay,
        "straight_sets_guard": straight_sets_guard,
        "fav20_sanity_guard": fav20_sanity_guard,
        "upset_risk_guard": upset_risk_guard,
        "low_over_split_guard": low_over_split_guard,
        "elite_clay_opponent_guard": elite_clay_opponent_guard_info,
        "wta_separation": wta_sep,
        "wta_script": wta_script
    }


def _first_nonempty_from_columns(df, candidates):
    """v22.17: crea una serie unificada usando la primera columna no vacía por fila."""
    existing = [c for c in candidates if c in df.columns]
    if not existing:
        return pd.Series([""] * len(df), index=df.index)
    tmp = df[existing].copy()
    for c in existing:
        tmp[c] = tmp[c].apply(normalizar_texto)
        tmp[c] = tmp[c].replace({"nan": "", "NaN": "", "None": ""})
    return tmp.replace("", np.nan).bfill(axis=1).iloc[:, 0].fillna("")


def normalizar_schema_historicos(df):
    """
    v22.17 Unified Historical Schema.
    Al concatenar Excel ATP Tour con CSV qual_chall, pandas deja columnas separadas:
    Winner/Loser para Excel y winner_name/loser_name para CSV. Antes buscar_columna elegía Winner
    y las filas CSV quedaban vacías. Esta función crea columnas MC_* para que todos los loaders
    cuenten Tour + Challenger/Qualy en el mismo mapa.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    out["MC_Winner"] = _first_nonempty_from_columns(out, [
        "MC_Winner", "Winner", "Ganador", "Player1", "Player 1", "WName",
        "winner_name", "winner_name_clean", "Jugador1", "Home", "P1"
    ])
    out["MC_Loser"] = _first_nonempty_from_columns(out, [
        "MC_Loser", "Loser", "Perdedor", "Player2", "Player 2", "LName",
        "loser_name", "loser_name_clean", "Jugador2", "Away", "P2"
    ])
    out["MC_Surface"] = _first_nonempty_from_columns(out, [
        "MC_Surface", "Surface", "surface", "Superficie", "Court Surface", "Court", "surface_name"
    ])
    # Mantiene columnas originales para detectar nivel, pero añade alias unificados útiles para debug.
    out["MC_Level"] = _first_nonempty_from_columns(out, [
        "MC_Level", "tourney_level", "Tourney Level", "Series", "Level", "Tour", "Category", "Circuit", "Event Type"
    ])
    out["MC_Tournament"] = _first_nonempty_from_columns(out, [
        "MC_Tournament", "tourney_name", "Tournament", "Event", "Location"
    ])
    out["MC_Round"] = _first_nonempty_from_columns(out, [
        "MC_Round", "round", "Round", "ronda"
    ])

    # v23.39.3 WTA Analyzer Rank/Date Fix.
    # Normaliza fecha y rankings para CSV WTA tipo Jeff Sackmann:
    # tourney_date / winner_rank / loser_rank.
    # También conserva compatibilidad con Tennis-Data: Date / WRank / LRank.
    out["MC_Date"] = _first_nonempty_from_columns(out, [
        "MC_Date", "Date", "Fecha", "date", "match_date", "tourney_date", "Tourney Date"
    ])
    out["MC_WinnerRank"] = _first_nonempty_from_columns(out, [
        "MC_WinnerRank", "WRank", "WinnerRank", "winner_rank", "winner_rank_num", "winner_rank_points_rank"
    ])
    out["MC_LoserRank"] = _first_nonempty_from_columns(out, [
        "MC_LoserRank", "LRank", "LoserRank", "loser_rank", "loser_rank_num", "loser_rank_points_rank"
    ])
    return out

def _historical_folders(circuito):
    base = rutas(circuito)["base"]
    return [
        rutas(circuito)["historicos"],
        os.path.join(base, "challenger"),
        os.path.join(base, "qual_chall"),
        os.path.join(base, "historicos_challenger"),
        os.path.join(base, "historicos_qual_chall"),
    ]


def _is_supported_history_file(path):
    ext = os.path.splitext(str(path))[1].lower()
    return ext in [".xlsx", ".xls", ".csv"]


@st.cache_data(show_spinner=False)
def descubrir_archivos_historicos(circuito, cache_version=QUALITY_ENGINE_VERSION):
    """
    v22.17 Deep File Scanner.
    - Escanea carpetas esperadas.
    - Además escanea TODO datos/atp de forma recursiva por si los CSV están en otra subcarpeta.
    - Detecta extensiones en mayúsculas: .CSV, .XLSX, .XLS.
    """
    base = rutas(circuito)["base"]
    folders = _historical_folders(circuito)
    files = []

    # 1) Carpetas esperadas, recursivo y case-insensitive por extensión.
    for folder in folders:
        if not os.path.exists(folder):
            continue
        for root, _, names in os.walk(folder):
            for name in names:
                fp = os.path.join(root, name)
                if _is_supported_history_file(fp):
                    files.append(fp)

    # 2) v22.17 Fast mode: NO escanea todo datos/atp en cada ejecución.
    # Si los CSV están en datos/atp/challenger o qual_chall, ya entran arriba de forma recursiva.
    # Esto evita cuelgues al pulsar Simular con repos grandes.
    files = sorted(set(files))

    folder_counts = {}
    folder_samples = {}
    for fd in folders:
        fd_name = os.path.relpath(fd, base) if os.path.exists(base) else fd
        found = [f for f in files if os.path.commonpath([os.path.abspath(fd), os.path.abspath(f)]) == os.path.abspath(fd)] if os.path.exists(fd) else []
        folder_counts[fd_name] = len(found)
        folder_samples[fd_name] = [os.path.relpath(x, fd) for x in found[:8]] if os.path.exists(fd) else []

    deep_extra = []
    expected_abs = [os.path.abspath(x) for x in folders if os.path.exists(x)]
    for f in files:
        af = os.path.abspath(f)
        in_expected = False
        for fd in expected_abs:
            try:
                if os.path.commonpath([fd, af]) == fd:
                    in_expected = True
                    break
            except Exception:
                pass
        if not in_expected:
            deep_extra.append(f)
    folder_counts["deep_extra"] = len(deep_extra)
    folder_samples["deep_extra"] = [os.path.relpath(x, base) if os.path.exists(base) else os.path.basename(x) for x in deep_extra[:8]]

    return files, folder_counts, folder_samples


@st.cache_data(show_spinner=False)
def cargar_historicos(circuito, cache_version=QUALITY_ENGINE_VERSION):
    """
    v22.17 Historical Loader.

    Lee históricos Tour y CSV Challenger/Qualy desde carpetas conocidas y también mediante
    escaneo profundo de datos/atp. Soporta .xlsx, .xls, .csv, también con extensión mayúscula.
    """
    base = rutas(circuito)["base"]
    files, _, _ = descubrir_archivos_historicos(circuito, cache_version=cache_version)

    dfs = []
    for f in files:
        try:
            ext = os.path.splitext(f)[1].lower()
            if ext in [".xlsx", ".xls"]:
                df = pd.read_excel(f)
            elif ext == ".csv":
                try:
                    df = pd.read_csv(f)
                except Exception:
                    df = pd.read_csv(f, sep=";")
            else:
                continue
            try:
                rel = os.path.relpath(f, base)
            except Exception:
                rel = os.path.basename(f)
            df["SourceFile"] = rel.replace("\\", "/")
            dfs.append(df)
        except Exception:
            pass
    if not dfs:
        return pd.DataFrame()
    return normalizar_schema_historicos(pd.concat(dfs, ignore_index=True))


@st.cache_data(show_spinner=False)
def historicos_diagnostics(circuito, cache_version=QUALITY_ENGINE_VERSION):
    """Diagnóstico visible para saber si el Match Count está leyendo archivos/columnas."""
    base = rutas(circuito)["base"]
    folder = rutas(circuito)["historicos"]
    files, folder_counts, folder_samples = descubrir_archivos_historicos(circuito, cache_version=cache_version)
    hist = cargar_historicos(circuito, cache_version=cache_version)

    if hist.empty:
        return {
            "folder": folder, "folder_exists": os.path.exists(folder),
            "files_count": len(files), "rows": 0, "columns": [],
            "winner_col": None, "loser_col": None, "surface_col": None,
            "sample_players": [], "sample_files": [os.path.relpath(x, base) if os.path.exists(base) else os.path.basename(x) for x in files[:12]],
            "folder_counts": folder_counts, "folder_samples": folder_samples,
        }

    col_winner = buscar_columna(hist, ["MC_Winner", "Winner", "Ganador", "Player1", "Player 1", "WName", "winner_name", "winner_name_clean", "Jugador1", "Home", "P1"] )
    col_loser = buscar_columna(hist, ["MC_Loser", "Loser", "Perdedor", "Player2", "Player 2", "LName", "loser_name", "loser_name_clean", "Jugador2", "Away", "P2"] )
    col_surface = buscar_columna(hist, ["MC_Surface", "Surface", "surface", "Superficie", "Court Surface", "Court", "surface_name"] )
    level_cols_found = [c for c in ["MC_Level", "MC_Tournament", "MC_Round", "Series", "Level", "Tour", "Category", "Circuit", "Event Type", "Tournament", "Event", "Round", "Comment", "SourceFile", "tourney_level", "tourney_name", "round", "winner_entry", "loser_entry"] if c in hist.columns]
    level_samples = {}
    for c in level_cols_found[:8]:
        try:
            vals = [normalizar_texto(x) for x in hist[c].dropna().astype(str).unique().tolist()[:6]]
            level_samples[c] = vals
        except Exception:
            level_samples[c] = []

    samples = []
    if col_winner is not None:
        samples += [normalizar_texto(x) for x in hist[col_winner].dropna().astype(str).head(5).tolist()]
    if col_loser is not None:
        samples += [normalizar_texto(x) for x in hist[col_loser].dropna().astype(str).head(5).tolist()]

    col_date = buscar_columna(hist, ["Date", "Fecha", "date", "match_date", "tourney_date", "Tourney Date"])
    date_min = date_max = "N/A"
    if col_date is not None:
        raw = hist[col_date]
        # Jeff Sackmann usa YYYYMMDD; pandas con dayfirst puede interpretarlo mal.
        if raw.astype(str).str.fullmatch(r"\d{8}").any():
            dt = pd.to_datetime(raw.astype(str), format="%Y%m%d", errors="coerce")
        else:
            dt = pd.to_datetime(raw, dayfirst=True, errors="coerce")
        if dt.notna().any():
            date_min = str(dt.min().date())
            date_max = str(dt.max().date())

    return {
        "folder": folder, "folder_exists": os.path.exists(folder),
        "files_count": len(files), "rows": int(len(hist)),
        "columns": list(hist.columns)[:20],
        "winner_col": col_winner, "loser_col": col_loser, "surface_col": col_surface,
        "date_min": date_min, "date_max": date_max,
        "sample_players": samples[:8],
        "sample_files": [os.path.relpath(x, base) if os.path.exists(base) else os.path.basename(x) for x in files[:12]],
        "folder_counts": folder_counts, "folder_samples": folder_samples,
        "level_cols": level_cols_found,
        "level_samples": level_samples,
    }


@st.cache_data
def crear_fatigue_map(circuito):
    """
    v17 Fatigue Engine.
    Calcula desgaste reciente desde históricos:
    - partidos últimos 7 días
    - games últimos 7 días
    - sets últimos 7 días
    - tie-breaks últimos 7 días
    - back-to-back
    - descanso estimado
    Nota: para predictor usa la fecha máxima del histórico cargado como referencia.
    """
    hist = cargar_historicos(circuito)
    fatigue = {}

    if hist.empty or "Date" not in hist.columns:
        return fatigue

    df = hist.copy()
    if "Comment" in df.columns:
        completed_mask = df["Comment"].astype(str).str.contains("Completed", case=False, na=False)
        if completed_mask.sum() > 0:
            df = df[completed_mask]
    df["DateParsed"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["DateParsed", "Winner", "Loser"])
    if df.empty:
        return fatigue

    ref_date = df["DateParsed"].max()

    def row_games(row):
        total = 0
        sets = 0
        tb = 0
        for i in range(1, 6):
            w = row.get(f"W{i}", np.nan)
            l = row.get(f"L{i}", np.nan)
            if not pd.isna(w) and not pd.isna(l):
                try:
                    wi, li = int(w), int(l)
                    total += wi + li
                    sets += 1
                    if (wi, li) in [(7, 6), (6, 7)]:
                        tb += 1
                except:
                    pass
        return total, sets, tb

    player_rows = {}

    for _, row in df.iterrows():
        date = row["DateParsed"]
        surface = str(row.get("Surface", "Hard"))
        games, sets, tbs = row_games(row)
        for player in [normalizar_texto(row.get("Winner", "")), normalizar_texto(row.get("Loser", ""))]:
            pid = limpiar(player)
            if not pid:
                continue
            player_rows.setdefault(pid, []).append({
                "date": date,
                "surface": surface,
                "games": games,
                "sets": sets,
                "tbs": tbs
            })

    for pid, rows in player_rows.items():
        rows = sorted(rows, key=lambda x: x["date"])
        recent7 = [r for r in rows if 0 <= (ref_date - r["date"]).days <= 7]
        recent14 = [r for r in rows if 0 <= (ref_date - r["date"]).days <= 14]

        last_date = rows[-1]["date"]
        rest_days = int((ref_date - last_date).days)

        matches7 = len(recent7)
        games7 = sum(r["games"] for r in recent7)
        sets7 = sum(r["sets"] for r in recent7)
        tbs7 = sum(r["tbs"] for r in recent7)

        matches14 = len(recent14)
        clay_games7 = sum(r["games"] for r in recent7 if r["surface"] == "Clay")

        fatigue_score = 0.0
        fatigue_score += matches7 * 0.008
        fatigue_score += max(0, games7 - 45) * 0.0007
        fatigue_score += max(0, sets7 - 6) * 0.004
        fatigue_score += tbs7 * 0.004
        fatigue_score += max(0, clay_games7 - 35) * 0.0005

        if rest_days <= 1:
            fatigue_score += 0.018
        elif rest_days <= 2:
            fatigue_score += 0.010
        elif rest_days >= 7:
            fatigue_score -= 0.006
        elif rest_days >= 4:
            fatigue_score -= 0.003

        fatigue_score = float(np.clip(fatigue_score, -0.010, 0.055))

        fatigue[pid] = {
            "fatigue_score": fatigue_score,
            "matches7": matches7,
            "matches14": matches14,
            "games7": games7,
            "sets7": sets7,
            "tbs7": tbs7,
            "rest_days": rest_days,
            "latest_date": str(last_date.date())
        }

    return fatigue


def buscar_fatigue(nombre, fatigue_map):
    nombre_alias = _ta_alias_resolve(nombre)
    nid = limpiar(nombre_alias)
    if nid in fatigue_map:
        return fatigue_map[nid]

    mejor, best = None, 0
    for fid, data in fatigue_map.items():
        score = max(similitud_nombre(nombre_alias, fid), similitud_nombre(nombre, fid))
        if score > best:
            best, mejor = score, data

    if mejor is not None and best >= 0.72:
        return mejor

    return {
        "fatigue_score": 0.0,
        "matches7": 0,
        "matches14": 0,
        "games7": 0,
        "sets7": 0,
        "tbs7": 0,
        "rest_days": 7,
        "latest_date": "N/A"
    }



def tournament_context_adjustments(context_row, p1_name="", p2_name=""):
    """
    v18 Rating Sanity Engine
    """
    import numpy as np

    if context_row is None:
        context_row = {}

    series = str(context_row.get("Series", ""))
    rnd = str(context_row.get("Round", ""))
    court = str(context_row.get("Court", "Outdoor"))
    location = str(context_row.get("Location", ""))

    p1_adj = 0.0
    p2_adj = 0.0
    vol_adj = 0.0
    tb_adj = 0.0

    # Tournament importance
    if "Grand Slam" in series:
        vol_adj -= 0.004
    elif "Masters" in series or "1000" in series:
        vol_adj -= 0.002
    elif "250" in series:
        vol_adj += 0.003

    # Round pressure
    if "Final" in rnd:
        vol_adj -= 0.004
    elif "Semi" in rnd:
        vol_adj -= 0.002
    elif "1st" in rnd:
        vol_adj += 0.003

    # Indoor boosts
    if "Indoor" in court:
        tb_adj += 0.025

    # Tiny home boosts
    home_map = {
        "Paris": "FRA",
        "Lyon": "FRA",
        "Marseille": "FRA",
        "Madrid": "ESP",
        "Barcelona": "ESP",
        "Mallorca": "ESP",
        "Rome": "ITA",
        "Turin": "ITA",
        "Milan": "ITA",
        "Munich": "GER",
        "Hamburg": "GER"
    }

    hc = home_map.get(location)

    if hc:
        if f"[{hc}]" in str(p1_name):
            p1_adj += 0.006
        if f"[{hc}]" in str(p2_name):
            p2_adj += 0.006

    return {
        "p1_adj": float(np.clip(p1_adj, -0.01, 0.01)),
        "p2_adj": float(np.clip(p2_adj, -0.01, 0.01)),
        "vol_adj": float(np.clip(vol_adj, -0.01, 0.01)),
        "tb_adj": float(np.clip(tb_adj, 0, 0.04)),
        "series": series,
        "round": rnd,
        "court": court
    }


def fatigue_adjustments(f1, f2, surface, circuito="ATP", rank1=999, rank2=999):
    """
    v19 Fatigue Engine.
    ATP: lógica anterior.
    WTA: menos castigo de fatiga y protección de top players.
    """
    fat1 = f1.get("fatigue_score", 0.0)
    fat2 = f2.get("fatigue_score", 0.0)

    if circuito == "WTA":
        if rank1 <= 15:
            fat1 *= 0.62
        elif rank1 <= 30:
            fat1 *= 0.75

        if rank2 <= 15:
            fat2 *= 0.62
        elif rank2 <= 30:
            fat2 *= 0.75

        if surface == "Clay":
            mult = 0.72
        elif surface == "Grass":
            mult = 0.48
        else:
            mult = 0.58
    else:
        if surface == "Clay":
            mult = 1.18
        elif surface == "Grass":
            mult = 0.82
        else:
            mult = 1.0

    adj1 = -fat1 * mult
    adj2 = -fat2 * mult

    vol_extra = abs(fat1 - fat2) * (0.20 if circuito == "WTA" else 0.35)

    if circuito == "ATP":
        if fat1 > 0.030 or fat2 > 0.030:
            vol_extra += 0.003
    else:
        if (fat1 > 0.040 or fat2 > 0.040) and rank1 > 30 and rank2 > 30:
            vol_extra += 0.002

    return (
        float(np.clip(adj1, -0.032 if circuito == "WTA" else -0.055, 0.010)),
        float(np.clip(adj2, -0.032 if circuito == "WTA" else -0.055, 0.010)),
        float(np.clip(vol_extra, 0, 0.010 if circuito == "WTA" else 0.018))
    )


def encontrar_jugador(nombre_hist, db):
    # v23.47.2: resuelve alias antes de buscar en DB.
    nombre_busqueda = _ta_alias_resolve(nombre_hist)
    if nombre_busqueda in db: return nombre_busqueda
    if nombre_hist in db: return nombre_hist
    mejor, best = None, 0
    for nombre in db:
        score = max(similitud_nombre(nombre_busqueda, nombre), similitud_nombre(nombre_hist, nombre))
        if score > best:
            best, mejor = score, nombre
    return mejor if mejor is not None and best >= 0.70 else None

def _hist_score_text(row):
    """Score histórico compatible con Tennis-Data/Jeff Sackmann: score, Score, Marcador."""
    for c in ["score", "Score", "Marcador", "Resultado", "Result"]:
        try:
            v = row.get(c, "")
            if normalizar_texto(v):
                return normalizar_texto(v)
        except Exception:
            pass
    return ""

def _score_pairs_from_row(row):
    """Devuelve pares de games desde columnas W1/L1... o desde score tipo '3-6 6-0 6-3'.
    Los pares siempre están en perspectiva Winner/Loser cuando vienen de Tennis Abstract/Jeff.
    """
    pairs = []
    for i in range(1, 6):
        w, l = row.get(f"W{i}", np.nan), row.get(f"L{i}", np.nan)
        if not pd.isna(w) and not pd.isna(l):
            try:
                pairs.append((int(w), int(l)))
            except Exception:
                pass
    if pairs:
        return pairs

    txt = _hist_score_text(row).upper()
    if not txt or any(x in txt for x in ["W/O", "WALKOVER", "RET", "DEF", "ABD", "DEFAULT"]):
        return []
    out = []
    # Acepta 7-6(4), 6-7 [10-8], etc. Para Analyzer solo necesitamos games base.
    for a, b in re.findall(r"(\d+)\s*-\s*(\d+)", txt):
        try:
            ga, gb = int(a), int(b)
            # Evita capturar super tie-breaks como sets normales si aparecen como [10-8].
            if ga <= 9 and gb <= 9:
                out.append((ga, gb))
        except Exception:
            pass
    return out

def _sets_winner_loser_row(row):
    try:
        ws = row.get("Wsets", np.nan)
        ls = row.get("Lsets", np.nan)
        if not pd.isna(ws) and not pd.isna(ls):
            return int(ws), int(ls)
    except Exception:
        pass
    wsets = lsets = 0
    for w, l in _score_pairs_from_row(row):
        if w > l:
            wsets += 1
        elif l > w:
            lsets += 1
    return wsets, lsets

def total_games_row(row):
    return int(sum(w + l for w, l in _score_pairs_from_row(row)))

def hay_tiebreak_row(row):
    for w, l in _score_pairs_from_row(row):
        if (int(w), int(l)) in [(7, 6), (6, 7)]:
            return True
    return False


def _row_first_value(row, candidates, default=""):
    """Devuelve el primer valor no vacío de una fila, tolerando nombres exactos o case-insensitive."""
    try:
        index_map = {str(c).lower(): c for c in row.index}
    except Exception:
        index_map = {}
    for c in candidates:
        val = row.get(c, None)
        if val is None and str(c).lower() in index_map:
            val = row.get(index_map[str(c).lower()], None)
        if val is None:
            continue
        try:
            if pd.isna(val):
                continue
        except Exception:
            pass
        txt = normalizar_texto(val)
        if txt and txt.lower() not in ["nan", "none", "nat"]:
            return val
    return default


def _parse_hist_date_value(v):
    """Normaliza fechas históricas para Analyzer; soporta YYYYMMDD, Date Excel y texto."""
    try:
        if v is None or pd.isna(v):
            return ""
    except Exception:
        pass
    raw = str(v).strip()
    if raw == "" or raw.lower() in ["nan", "none", "nat"]:
        return ""
    # Evita que 20260105.0 rompa el parser.
    raw_clean = raw.replace(".0", "") if re.fullmatch(r"\d+\.0", raw) else raw
    try:
        if re.fullmatch(r"\d{8}", raw_clean):
            dt = pd.to_datetime(raw_clean, format="%Y%m%d", errors="coerce")
        else:
            dt = pd.to_datetime(raw_clean, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return str(dt.date())
    except Exception:
        pass
    return raw


def _row_float_value(row, candidates, default=np.nan):
    v = _row_first_value(row, candidates, default=None)
    if v is None:
        return default
    return leer_float(v, default)

def validar_historico(db, hist_df, circuito, surface_filter, max_matches, sims_bt):
    rows = []
    df = hist_df.copy()

    # v23.39.2 Analyzer WTA Fix:
    # El Analyzer ahora usa columnas normalizadas MC_* y soporta CSV WTA tipo Jeff Sackmann
    # (winner_name/loser_name/surface/score/best_of) sin exigir Comment='Completed'.
    if "MC_Surface" not in df.columns:
        df = normalizar_schema_historicos(df)

    if surface_filter != "Todas":
        surf_col = "MC_Surface" if "MC_Surface" in df.columns else "Surface"
        df = df[df[surf_col].apply(lambda x: normalizar_superficie_hist(x, default="Hard")).astype(str) == surface_filter]

    if "Comment" in df.columns:
        comment_txt = df["Comment"].apply(normalizar_texto).replace({"nan": "", "NaN": "", "None": ""})
        has_comment = comment_txt.str.strip().ne("")
        completed_mask = comment_txt.str.contains("Completed", case=False, na=False)
        # Solo filtramos Completed si existe información real de Comment.
        # Los CSV WTA 2026 no traen Comment y antes el Analyzer los eliminaba o fallaba.
        if has_comment.any() and completed_mask.sum() > 0:
            df = df[(~has_comment) | completed_mask]

    # Quita walkovers/retiradas si vienen en score; si no hay score, deja pasar Tennis-Data con W1/L1.
    if "score" in df.columns or "Score" in df.columns:
        df = df[df.apply(lambda r: bool(_score_pairs_from_row(r)), axis=1)]

    if max_matches:
        df = df.tail(max_matches)

    progress = st.progress(0)
    total = len(df)

    for idx, (_, row) in enumerate(df.iterrows()):
        winner_hist = normalizar_texto(row.get("MC_Winner", row.get("Winner", row.get("winner_name", ""))))
        loser_hist = normalizar_texto(row.get("MC_Loser", row.get("Loser", row.get("loser_name", ""))))
        surface = normalizar_superficie_hist(row.get("MC_Surface", row.get("Surface", row.get("surface", "Hard"))), default="Hard")
        best_of = int(leer_float(row.get("Best of", row.get("best_of", 3)), 3))

        if not winner_hist or not loser_hist:
            continue

        p_win = encontrar_jugador(winner_hist, db)
        p_los = encontrar_jugador(loser_hist, db)
        if p_win is None or p_los is None:
            continue

        d_win, d_los = db[p_win], db[p_los]
        sim = sim_match(d_win, d_los, surface, circuito, best_of, sims_bt, context_row=row)

        p_raw, p_cal = sim["p1"], sim["p1_cal"]
        games_real = total_games_row(row)
        if games_real <= 0:
            continue
        games_model = sim["games"]
        w_sets, l_sets = _sets_winner_loser_row(row)
        set3_real = (w_sets + l_sets) >= 3 if best_of == 3 else (w_sets + l_sets) >= 4

        model_fav_is_winner = p_cal >= 0.50
        real_fav_wins_set = True if model_fav_is_winner else (l_sets >= 1)
        # Si el favorito del modelo era el ganador histórico, el dog histórico fue el perdedor
        # y solo gana set si el perdedor hizo algún set. Si el favorito del modelo era el
        # perdedor histórico, el dog fue el ganador real y por tanto ganó al menos un set.
        real_dog_wins_set = (l_sets >= 1) if model_fav_is_winner else True
        model_fav_wins_set = sim.get("p1_wins_set_any", 0.0) if model_fav_is_winner else sim.get("p2_wins_set_any", 0.0)

        fav_under22_real = (p_raw >= 0.50 and games_real < 22.5)
        dog_over20_real = (p_raw < 0.50 and games_real > 20.5)

        hist_date = _parse_hist_date_value(_row_first_value(row, ["MC_Date", "Date", "Fecha", "date", "match_date", "tourney_date", "Tourney Date"], ""))
        winner_rank = _row_float_value(row, ["MC_WinnerRank", "WRank", "WinnerRank", "winner_rank"], np.nan)
        loser_rank = _row_float_value(row, ["MC_LoserRank", "LRank", "LoserRank", "loser_rank"], np.nan)
        rank_gap = abs(winner_rank - loser_rank) if not pd.isna(winner_rank) and not pd.isna(loser_rank) else np.nan

        rows.append({
            "Date": hist_date, "Tournament": row.get("MC_Tournament", row.get("Tournament", row.get("tourney_name", ""))),
            "Series": row.get("Series", row.get("MC_Level", row.get("tourney_level", ""))), "Round": row.get("MC_Round", row.get("Round", row.get("round", ""))),
            "Surface": surface, "WinnerHist": winner_hist, "LoserHist": loser_hist,
            "WinnerMatched": p_win, "LoserMatched": p_los,
            "WinnerRank": winner_rank,
            "LoserRank": loser_rank,
            "RankGap": rank_gap,
            "ResultadoScore": _hist_score_text(row),
            "WinnerSets": w_sets, "LoserSets": l_sets,
            "ModelWinnerProbRaw": p_raw, "ModelWinnerProbCal": p_cal,
            "RawFavWasWinner": p_raw >= 0.50, "CalFavWasWinner": p_cal >= 0.50,
            "RealGames": games_real, "ModelAvgGames": np.mean(games_model),
            "Real3Sets": set3_real, "Model3Sets": sim["set3"],
            "RealTB": hay_tiebreak_row(row), "ModelTB": sim["tb"],
            "RealOver17": games_real > 17.5, "ModelOver17": sum(x > 17.5 for x in games_model)/sims_bt,
            "RealOver18": games_real > 18.5, "ModelOver18": sum(x > 18.5 for x in games_model)/sims_bt,
            "RealOver19": games_real > 19.5, "ModelOver19": sum(x > 19.5 for x in games_model)/sims_bt,
            "RealOver20": games_real > 20.5, "ModelOver20": sum(x > 20.5 for x in games_model)/sims_bt,
            "RealOver22": games_real > 22.5, "ModelOver22": sum(x > 22.5 for x in games_model)/sims_bt,
            "RealFavWinsSet": real_fav_wins_set, "ModelFavWinsSet": model_fav_wins_set,
            "RealDogWinsSet": real_dog_wins_set, "ModelDogWinsSet": sim.get("dog_wins_set", 0),
            "RealFavUnder22": fav_under22_real, "ModelFavUnder22": sim["fav_under22"],
            "RealDogOver20": dog_over20_real, "ModelDogOver20": sim["dog_over20"],
            "ModelFav2_0": sim.get("fav_2_0", 0),
            "ModelLongMatch": sim.get("long_match", 0),
            "WinnerFatigueScore": sim.get("fatigue1", {}).get("fatigue_score", 0),
            "LoserFatigueScore": sim.get("fatigue2", {}).get("fatigue_score", 0),
            "WinnerStats": get_stats_surface(d_win, surface).get("match_type", "N/A"), "LoserStats": get_stats_surface(d_los, surface).get("match_type", "N/A"),
            "WinnerHold": sim["hold1"], "LoserHold": sim["hold2"],
            "WinnerReturn": sim["ret1"], "LoserReturn": sim["ret2"],
            "WinnerServeProfile": sim["p1_profile"], "LoserServeProfile": sim["p2_profile"],
            "FavRawEst": sim["fav_raw_est"],
            "TBIntelBoost": sim.get("tb_intel_boost", 0),
            "WinnerPressureSkill": sim.get("pressure_skill1", 0),
            "LoserPressureSkill": sim.get("pressure_skill2", 0)
        })

        if total:
            progress.progress(min((idx+1)/total, 1.0))
    progress.empty()
    return pd.DataFrame(rows)

def market_hit_rate(df, model_col, real_col, threshold):
    sub = df[df[model_col] >= threshold].copy()
    if len(sub) == 0: return None
    return {"Casos": len(sub), "Prob media modelo": sub[model_col].mean(), "Acierto real": sub[real_col].mean()}

def crear_analyzer_tables(val, min_casos=20):
    val = val.copy()
    tables = {}
    val["FavProbCal"] = val["ModelWinnerProbCal"].apply(lambda x: max(x, 1-x))
    val["MLBin"] = pd.cut(val["FavProbCal"], bins=[0.50,0.55,0.60,0.65,0.70,0.75,0.80,0.90], labels=["50-55","55-60","60-65","65-70","70-75","75-80","80-90"])
    ml = val.groupby("MLBin", observed=True).agg(Casos=("FavProbCal","count"), ProbMedia=("FavProbCal","mean"), AciertoReal=("CalFavWasWinner","mean")).reset_index()
    tables["ML por tramos"] = ml[ml["Casos"] >= min_casos]

    rows = []
    for name, model, real in [
        ("Over 17.5","ModelOver17","RealOver17"),
        ("Over 18.5","ModelOver18","RealOver18"),
        ("Over 19.5","ModelOver19","RealOver19"),
        ("Over 20.5","ModelOver20","RealOver20"),
        ("Over 22.5","ModelOver22","RealOver22"),
        ("3 Sets","Model3Sets","Real3Sets"),
        ("Tie-break","ModelTB","RealTB"),
        ("Favorito + Under 22.5","ModelFavUnder22","RealFavUnder22"),
        ("Dog + Over 20.5","ModelDogOver20","RealDogOver20"),
        ("Favorito gana al menos 1 set","ModelFavWinsSet","RealFavWinsSet"),
        ("Favorito 2-0","ModelFav2_0","CalFavWasWinner"),
        ("Underdog gana set","ModelDogWinsSet","RealDogWinsSet"),
        ("Partido largo","ModelLongMatch","RealOver22"),
    ]:
        for th in [0.40,0.45,0.50,0.55,0.60,0.65,0.70]:
            r = market_hit_rate(val, model, real, th)
            if r and r["Casos"] >= min_casos:
                rows.append({"Mercado": name, "Umbral modelo": th, **r})
    tables["Mercados por umbral"] = pd.DataFrame(rows)

    surface = val.groupby("Surface").agg(
        Casos=("Surface","count"), MLAccuracy=("CalFavWasWinner","mean"),
        Over17Hit=("RealOver17","mean"), Over18Hit=("RealOver18","mean"), Over20Hit=("RealOver20","mean"),
        Over22Hit=("RealOver22","mean"), ThreeSetsReal=("Real3Sets","mean"),
        TBReal=("RealTB","mean"), TBModel=("ModelTB","mean"),
        GamesReal=("RealGames","mean"), GamesModel=("ModelAvgGames","mean")
    ).reset_index()
    tables["Por superficie"] = surface[surface["Casos"] >= min_casos]

    val["AnyBigServer"] = val["WinnerServeProfile"].isin(["big_server","elite_server"]) | val["LoserServeProfile"].isin(["big_server","elite_server"])
    server = val.groupby("AnyBigServer").agg(
        Casos=("AnyBigServer","count"), MLAccuracy=("CalFavWasWinner","mean"),
        TBReal=("RealTB","mean"), TBModel=("ModelTB","mean"),
        Over22Real=("RealOver22","mean"), Over22Model=("ModelOver22","mean"),
        GamesReal=("RealGames","mean"), GamesModel=("ModelAvgGames","mean")
    ).reset_index()
    tables["Big server"] = server[server["Casos"] >= min_casos]

    # v23.39.3: ranking-gap robusto. Si el CSV ya trae RankGap lo respeta; si no, lo calcula.
    if "RankGap" not in val.columns:
        val["RankGap"] = abs(val["WinnerRank"] - val["LoserRank"])
    val["RankGap"] = pd.to_numeric(val["RankGap"], errors="coerce")
    val_rank = val.dropna(subset=["RankGap"]).copy()
    if not val_rank.empty:
        val_rank["RankGapBin"] = pd.cut(
            val_rank["RankGap"],
            bins=[-0.01,20,50,100,200,10000],
            labels=["0-20","21-50","51-100","101-200","200+"]
        )
        rg = val_rank.groupby("RankGapBin", observed=True).agg(
            Casos=("RankGap","count"), MLAccuracy=("CalFavWasWinner","mean"),
            FavWinsSetReal=("RealFavWinsSet","mean"),
            Over17Hit=("RealOver17","mean"), Over18Hit=("RealOver18","mean"),
            ThreeSetsReal=("Real3Sets","mean"), ThreeSetsModel=("Model3Sets","mean"),
            GamesReal=("RealGames","mean"), GamesModel=("ModelAvgGames","mean")
        ).reset_index()
        tables["Ranking gap"] = rg[rg["Casos"] >= min_casos]

        rows_rg = []
        for gap_label, sub_gap in val_rank.groupby("RankGapBin", observed=True):
            if len(sub_gap) < min_casos:
                continue
            for th in [0.70,0.75,0.80,0.85,0.90]:
                sub = sub_gap[sub_gap["ModelOver17"] >= th]
                if len(sub) >= min_casos:
                    rows_rg.append({
                        "RankGapBin": str(gap_label),
                        "Mercado": "Over 17.5",
                        "Umbral modelo": th,
                        "Casos": len(sub),
                        "Prob media modelo": sub["ModelOver17"].mean(),
                        "Acierto real": sub["RealOver17"].mean(),
                        "3Sets real": sub["Real3Sets"].mean(),
                        "Games real": sub["RealGames"].mean(),
                    })
        tables["Over17 por ranking gap"] = pd.DataFrame(rows_rg)
    else:
        tables["Ranking gap"] = pd.DataFrame()
        tables["Over17 por ranking gap"] = pd.DataFrame()
    return tables


# =========================================================
# v23.46.4 Backtest Oficial / Observar + Dog gana set sobre históricos
# No toca motor Over ni simulaciones. Solo clasifica las filas ya validadas
# para medir qué habría pasado con señales tipo OFICIAL y OBSERVAR.
# =========================================================
def _bt_safe_float(x, default=0.0):
    try:
        if x is None or pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def _bt_signal_rows_from_val_row(row, circuito):
    signals = []

    fav_prob = max(
        _bt_safe_float(row.get("ModelWinnerProbCal", 0.5), 0.5),
        1.0 - _bt_safe_float(row.get("ModelWinnerProbCal", 0.5), 0.5),
    )
    over17 = _bt_safe_float(row.get("ModelOver17", 0.0), 0.0)
    over18 = _bt_safe_float(row.get("ModelOver18", 0.0), 0.0)
    over19 = _bt_safe_float(row.get("ModelOver19", 0.0), 0.0)
    over20 = _bt_safe_float(row.get("ModelOver20", 0.0), 0.0)
    fav_set = _bt_safe_float(row.get("ModelFavWinsSet", 0.0), 0.0)
    dog_set = _bt_safe_float(row.get("ModelDogWinsSet", 0.0), 0.0)
    three_sets = _bt_safe_float(row.get("Model3Sets", 0.0), 0.0)

    def add(tipo, mercado, prob, real_hit, motivo):
        signals.append({
            "Tipo seguimiento": tipo,
            "Mercado": mercado,
            "Probabilidad modelo": float(prob),
            "Acierto": bool(real_hit),
            "Motivo regla": motivo,
            "Date": row.get("Date", ""),
            "Surface": row.get("Surface", ""),
            "Partido histórico": f"{row.get('WinnerHist','')} vs {row.get('LoserHist','')}",
            "ResultadoScore": row.get("ResultadoScore", ""),
            "RealGames": row.get("RealGames", np.nan),
            "WinnerHist": row.get("WinnerHist", ""),
            "LoserHist": row.get("LoserHist", ""),
            "RankGap": row.get("RankGap", np.nan),
        })

    # OFICIAL: zona jugable estricta usada por el selector/analyzer actual.
    if fav_prob >= 0.70:
        add("OFICIAL", "ML favorito", fav_prob, row.get("CalFavWasWinner", False), "ML favorito >=70%")

    if over18 >= 0.80:
        add("OFICIAL", "Over 18.5", over18, row.get("RealOver18", False), "Over 18.5 >=80%")

    # WTA Over17 en la app queda como watch/premium watch, no oficial duro.
    if str(circuito).upper() == "WTA" and over17 >= 0.80:
        add("OBSERVAR", "WTA Over 17.5", over17, row.get("RealOver17", False), "WTA Over17 >=80% watch")

    # OBSERVAR: señales interesantes que no pasan a oficial sin confirmación extra.
    if 0.76 <= over18 < 0.80:
        add("OBSERVAR", "Over 18.5", over18, row.get("RealOver18", False), "Over18 76%-79.9%")

    if over19 >= 0.74:
        add("OBSERVAR", "Over 19.5", over19, row.get("RealOver19", False), "Over19 >=74% secundario")

    if over20 >= 0.74:
        add("OBSERVAR", "Over 20.5", over20, row.get("RealOver20", False), "Over20 >=74% secundario")

    if 0.84 <= fav_set < 0.94:
        add("OBSERVAR", "Favorito gana al menos 1 set", fav_set, row.get("RealFavWinsSet", False), "Gana-set favorito 84%-93.9%")
    elif fav_set >= 0.94 and fav_prob >= 0.68:
        add("OFICIAL", "Favorito gana al menos 1 set", fav_set, row.get("RealFavWinsSet", False), "Gana-set favorito elite >=94% + ML >=68%")

    # Mercado que queríamos medir: underdog gana al menos 1 set.
    # En zona alta se marca como OFICIAL; la zona media queda como OBSERVAR.
    if 0.55 <= dog_set < 0.70:
        add("OBSERVAR", "Dog gana al menos 1 set", dog_set, row.get("RealDogWinsSet", False), "Dog gana set 55%-69.9%")
    elif dog_set >= 0.70:
        add("OFICIAL", "Dog gana al menos 1 set", dog_set, row.get("RealDogWinsSet", False), "Dog gana set fuerte >=70%")

    if three_sets >= 0.55:
        add("OBSERVAR", "3 sets", three_sets, row.get("Real3Sets", False), "3 sets >=55% solo observar")

    return signals


def crear_backtest_oficial_observar(val, circuito, min_casos=10):
    if val is None or val.empty:
        return pd.DataFrame(), pd.DataFrame()

    rows = []
    for _, r in val.iterrows():
        rows.extend(_bt_signal_rows_from_val_row(r, circuito))

    detail = pd.DataFrame(rows)
    if detail.empty:
        return pd.DataFrame(), detail

    summary = (
        detail.groupby(["Tipo seguimiento", "Mercado"], dropna=False)
        .agg(
            Casos=("Acierto", "count"),
            Aciertos=("Acierto", "sum"),
            Porcentaje_acierto=("Acierto", "mean"),
            Prob_media_modelo=("Probabilidad modelo", "mean"),
            Games_media_real=("RealGames", "mean"),
        )
        .reset_index()
    )
    summary = summary[summary["Casos"] >= int(min_casos)].copy()
    if not summary.empty:
        summary = summary.sort_values(["Tipo seguimiento", "Porcentaje_acierto", "Casos"], ascending=[True, False, False])
    return summary, detail


# =========================================================
# v22.2 MARKET SANITY CAPS
# =========================================================

def aplicar_market_sanity_caps(sim, circuito, surface, over18, over19, over20, over22):
    """
    v22.2: suaviza mercados de games cuando el propio Rating Sanity
    dice que la confianza del Elo/superficie es baja.
    No cambia el resultado simulado; cambia la probabilidad operativa mostrada.
    """
    rs = sim.get("rating_sanity", {}) or {}
    p1_rs = rs.get("p1", {}) or {}
    p2_rs = rs.get("p2", {}) or {}
    min_conf = min(p1_rs.get("confidence", 1.0), p2_rs.get("confidence", 1.0))
    min_surface_matches = min(p1_rs.get("matches_surface", 99), p2_rs.get("matches_surface", 99))
    fav_prob = max(sim.get("p1_cal", 0.5), sim.get("p2_cal", 0.5))
    elo_pure_gap = abs(sim.get("fav_raw_est", 0.5) - fav_prob)

    notes = []
    o18, o19, o20, o22 = over18, over19, over20, over22

    if circuito == "ATP" and surface == "Clay":
        # Caso detectado: Over 18.5 salía demasiado alto como señal principal
        # aunque ambos jugadores tuvieran confidence ~25% y pocos partidos clay.
        if min_conf < 0.35 or min_surface_matches < 5:
            cap18 = 0.72 if fav_prob < 0.58 else 0.74
            if o18 > cap18:
                o18 = cap18
                notes.append(f"Over 18.5 capado a {cap18:.0%} por baja confianza clay")

            cap19 = 0.68 if fav_prob < 0.58 else 0.70
            if o19 > cap19:
                o19 = cap19
                notes.append(f"Over 19.5 capado a {cap19:.0%} por baja confianza clay")

        # v22.11: muestra aceptable pero todavía corta + Elo puro muy separado del modelo.
        # Ejemplo: dog con 8 partidos clay y gap ~15-16%; no debe escalar a SPOT APTO.
        if min_surface_matches < 10 and elo_pure_gap >= 0.14:
            cap18 = 0.74 if fav_prob < 0.58 else 0.75
            if o18 > cap18:
                o18 = cap18
                notes.append(f"Over 18.5 capado a {cap18:.0%} por muestra clay corta + gap Elo/modelo")

            cap19 = 0.68 if fav_prob < 0.58 else 0.69
            if o19 > cap19:
                o19 = cap19
                notes.append(f"Over 19.5 capado a {cap19:.0%} por muestra clay corta + gap Elo/modelo")

        # Si el favorito es muy débil, evitamos transformar un partido abierto
        # automáticamente en spot fuerte de over bajo.
        if fav_prob < 0.56 and o18 > 0.72:
            o18 = 0.72
            notes.append("Over 18.5 limitado por favorito débil")

        # v22.33 Elite Blowout Guard:
        # cuando el partido ya está en modo dominio claro del favorito, el over bajo
        # no debe seguir apareciendo automáticamente como señal principal.
        fav20 = sim.get("fav_2_0", 0.0)
        dogset = sim.get("dog_wins_set", 1.0)
        e1_eff = sim.get("elo_effective1", 1500)
        e2_eff = sim.get("elo_effective2", 1500)
        fav_is_p1 = sim.get("p1_cal", 0.5) >= sim.get("p2_cal", 0.5)
        fav_elo = e1_eff if fav_is_p1 else e2_eff
        dog_elo = e2_eff if fav_is_p1 else e1_eff
        elo_gap_eff = fav_elo - dog_elo

        elite_blowout = (
            fav_prob >= 0.80
            and fav20 >= 0.78
            and dogset <= 0.22
            and elo_gap_eff >= 250
            and min_conf >= 0.75
            and min_surface_matches >= 60
        )

        if elite_blowout:
            cap18 = 0.56
            cap19 = 0.48
            cap20 = 0.42
            cap22 = 0.30

            # Si el 2-0 es casi total y el dog-set está hundido, cap más fuerte.
            if fav20 >= 0.90 and dogset <= 0.10 and elo_gap_eff >= 350:
                cap18 = 0.54
                cap19 = 0.46
                cap20 = 0.40
                cap22 = 0.28

            if o18 > cap18:
                o18 = cap18
                notes.append(f"Over 18.5 capado a {cap18:.0%} por dominio élite/blowout")
            if o19 > cap19:
                o19 = cap19
                notes.append(f"Over 19.5 capado a {cap19:.0%} por dominio élite/blowout")
            if o20 > cap20:
                o20 = cap20
                notes.append(f"Over 20.5 capado a {cap20:.0%} por dominio élite/blowout")
            if o22 > cap22:
                o22 = cap22
                notes.append(f"Over 22.5 capado a {cap22:.0%} por dominio élite/blowout")

    # v23.20 WTA Over Watchlist:
    # Recupera solo perfiles concretos detectados como útiles:
    # - Over >=72% con favorita <=64%.
    # - Over >=68% con favorita 64-67%.
    # - Over >=66% con favorita 70-72%.
    # Evita abrir de forma general todos los partidos 55-67%.
    if circuito == "WTA" and surface == "Clay":
        dogset = sim.get("dog_wins_set", 0.0)
        set3 = sim.get("set3", 0.0)
        longm = sim.get("long_match", 0.0)

        wta_close_match = (
            fav_prob <= 0.64
            or (0.64 <= fav_prob < 0.67 and dogset >= 0.45)
            or (dogset >= 0.58 and (set3 >= 0.42 or longm >= 0.50))
        )

        if fav_prob >= 0.75:
            cap18, cap19, cap20, cap22 = 0.62, 0.54, 0.48, 0.38
            reason = "WTA v23.20 favorita dominante, over muy protegido"
        elif fav_prob >= 0.72:
            cap18, cap19, cap20, cap22 = 0.64, 0.56, 0.50, 0.40
            reason = "WTA v23.20 favorita muy fuerte"
        elif fav_prob >= 0.70:
            cap18, cap19, cap20, cap22 = 0.66, 0.58, 0.52, 0.42
            reason = "WTA v23.20 zona 70-72, over 66 recuperable"
        elif fav_prob >= 0.68:
            cap18, cap19, cap20, cap22 = 0.66, 0.58, 0.52, 0.42
            reason = "WTA v23.20 favorita clara, cautela"
        elif fav_prob >= 0.64:
            cap18, cap19, cap20, cap22 = 0.68, 0.60, 0.54, 0.44
            reason = "WTA v23.20 zona 64-67, over 68 recuperable"
        elif wta_close_match:
            cap18, cap19, cap20, cap22 = 0.72, 0.64, 0.58, 0.48
            reason = "WTA v23.20 partido igualado, over 72 recuperable"
        else:
            cap18, cap19, cap20, cap22 = 0.68, 0.60, 0.54, 0.44
            reason = "WTA v23.20 neutro, sin recuperación extra"

        if o18 > cap18:
            o18 = cap18
            notes.append(f"Over 18.5 capado a {cap18:.0%} por {reason}")
        if o19 > cap19:
            o19 = cap19
            notes.append(f"Over 19.5 capado a {cap19:.0%} por {reason}")
        if o20 > cap20:
            o20 = cap20
            notes.append(f"Over 20.5 capado a {cap20:.0%} por {reason}")
        if o22 > cap22:
            o22 = cap22
            notes.append(f"Over 22.5 capado a {cap22:.0%} por {reason}")

    return {
        "over18": float(np.clip(o18, 0.0, 0.95)),
        "over19": float(np.clip(o19, 0.0, 0.95)),
        "over20": float(np.clip(o20, 0.0, 0.95)),
        "over22": float(np.clip(o22, 0.0, 0.95)),
        "notes": notes
    }


# =========================================================
# v20 BETTING FILTERS ENGINE
# =========================================================

def betting_filter_engine(circuito, surface, sim, p1_name, p2_name):
    p1c = sim.get("p1_cal", 0.5)
    p2c = sim.get("p2_cal", 0.5)
    fav_prob = max(p1c, p2c)
    fav_name = p1_name if p1c >= p2c else p2_name

    # v23.39.4: reglas WTA derivadas del Analyzer histórico.
    # Solo afectan al selector/etiqueta; no modifican el motor Over ni las simulaciones.
    rank_gap = sim.get("rank_gap", None)
    try:
        rank_gap = abs(float(rank_gap))
    except Exception:
        r1 = sim.get("rank1", None)
        r2 = sim.get("rank2", None)
        try:
            rank_gap = abs(float(r1) - float(r2))
        except Exception:
            rank_gap = None
    wta_rankgap_ideal_over17 = (str(circuito).upper().strip() == "WTA" and rank_gap is not None and 21 <= rank_gap <= 50)

    games = sim.get("games", [])
    if sim.get("market_over18") is not None:
        over17 = sim.get("market_over17", 0.0)
        over18 = sim.get("market_over18", 0.0)
        over19 = sim.get("market_over19", 0.0)
        over20 = sim.get("market_over20", 0.0)
        over22 = sim.get("market_over22", 0.0)
        under22 = 1 - over22
    elif games:
        over17 = sum(x > 17.5 for x in games) / len(games)
        over18 = sum(x > 18.5 for x in games) / len(games)
        over19 = sum(x > 19.5 for x in games) / len(games)
        over20 = sum(x > 20.5 for x in games) / len(games)
        over22 = sum(x > 22.5 for x in games) / len(games)
        under22 = 1 - over22
    else:
        over17 = over18 = over19 = over20 = over22 = under22 = 0.0

    fav20 = sim.get("fav_2_0", 0)
    dogset = sim.get("dog_wins_set", 0)
    longm = sim.get("long_match", 0)
    tb = sim.get("tb", 0)
    set3 = sim.get("set3", sim.get("market_3sets", sim.get("prob_3sets", 0.0)))
    vol = sim.get("vol", 0)

    # v22.2 Signal Trust Gate: una probabilidad puede ser buena,
    # pero no debe etiquetarse como fuerte si el Elo viene de poca muestra.
    rs = sim.get("rating_sanity", {}) or {}
    p1_rs = rs.get("p1", {}) or {}
    p2_rs = rs.get("p2", {}) or {}
    min_conf = min(p1_rs.get("confidence", 1.0), p2_rs.get("confidence", 1.0))
    avg_conf = (p1_rs.get("confidence", 1.0) + p2_rs.get("confidence", 1.0)) / 2
    min_surface_matches = min(p1_rs.get("matches_surface", 99), p2_rs.get("matches_surface", 99))
    elo_pure_gap = abs(sim.get("fav_raw_est", 0.5) - fav_prob)
    upset_guard = sim.get("upset_risk_guard", {}) or {}

    # =========================================================
    # v23.29 OVER QUALITY GUARD + UNDER 2.5 RESCUE
    # Objetivo: bloquear falsos Over 18.5/19.5 que salen altos por
    # supuesta resistencia, pero con datos débiles o perfil de 2-0 corto.
    # Patrones detectados en tus fallos: Over 72-74% clay + rating sanity +
    # baja muestra/confianza, o favorito 64%+ con dog-set/3-sets bajos.
    # =========================================================
    rating_active = bool(rs.get("active", False))
    over_quality_reasons = []
    over_quality_block = False
    over_quality_watch = False

    # Bloque 1: falso over por calidad de datos.
    # v23.29.1: menos agresivo. Ya NO bloquea solo por una alerta aislada.
    # Bloquea únicamente cuando coinciden varias señales malas.
    quality_block_hard = False
    quality_watch = False
    if circuito == "ATP" and surface == "Clay":
        very_low_conf = min_conf <= 0.35
        low_conf = min_conf <= 0.50
        very_short_surface = min_surface_matches < 5
        short_surface = min_surface_matches < 10
        mid_gap = elo_pure_gap >= 0.12
        high_gap = elo_pure_gap >= 0.18

        if very_low_conf and very_short_surface:
            quality_block_hard = True
            over_quality_reasons.append("confianza <=35% + muestra <5")
        if rating_active and very_short_surface and low_conf:
            quality_block_hard = True
            over_quality_reasons.append("rating sanity + muestra <5 + confianza baja")
        if high_gap and (low_conf or short_surface):
            quality_block_hard = True
            over_quality_reasons.append("gap Elo/modelo >=18% con fiabilidad baja")

        # Watch: alerta visual, pero no bloquea automáticamente el Over.
        if rating_active and (short_surface or min_conf < 0.60):
            quality_watch = True
            over_quality_reasons.append("rating sanity + fiabilidad mejorable")
        if mid_gap:
            quality_watch = True
            over_quality_reasons.append("gap Elo/modelo elevado")
        if 0.70 <= over18 <= 0.75 and rating_active and (short_surface or min_conf < 0.55):
            quality_watch = True
            over_quality_reasons.append("zona vigilancia Over 70-75% en clay")

    if quality_block_hard:
        over_quality_block = True
    elif quality_watch:
        over_quality_watch = True

    # Bloque 2: falso over por favorito 2-0 / marcador corto.
    straight_sets_risk = (
        fav_prob >= 0.64
        and dogset <= 0.42
        and set3 <= 0.45
        and tb <= 0.28
    )
    straight_sets_clear = (
        fav_prob >= 0.70
        and dogset <= 0.38
        and set3 <= 0.42
        and tb <= 0.25
    )
    if straight_sets_risk:
        over_quality_watch = True
        over_quality_reasons.append("favorito 2-0: dog set bajo + 3 sets bajo + tie-break bajo")
        # v23.29.1: solo bloquea si el patrón 2-0 es muy claro.
        if straight_sets_clear and over18 < 0.74:
            over_quality_block = True
            over_quality_reasons.append("perfil 2-0 claro con Over insuficiente")

    # Bloque 3: zona roja observada en tus fallos.
    # v23.29.1: ya no bloquea todos los Over 70-75%; solo si hay calidad dura mala.
    if circuito == "ATP" and surface == "Clay" and 0.70 <= over18 <= 0.75:
        if quality_block_hard:
            over_quality_block = True
            over_quality_reasons.append("zona roja Over 70-75% + calidad muy baja")
        elif over_quality_watch:
            over_quality_reasons.append("zona watch: Over 70-75%, revisar antes de combinar")

    # Ajuste simple para convertir 3 sets bruto en lectura Under 2.5.
    # v23.29.2: Under Rescue prudente; casi todo Under queda como WATCH hasta validar.
    under25_raw = 1.0 - float(set3 or 0.0)
    under25_boost = 0.0
    if over_quality_block:
        under25_boost += 0.08
    elif over_quality_watch:
        under25_boost += 0.04
    if straight_sets_risk:
        under25_boost += 0.05
    if fav_prob >= 0.70 and dogset <= 0.38:
        under25_boost += 0.03
    under25_adjusted = float(np.clip(under25_raw + under25_boost, 0.0, 0.84))

    if over_quality_block:
        over_guard_label = "🚫 OVER BLOQUEADO"
    elif over_quality_watch:
        over_guard_label = "⚠️ OVER WATCH / NO COMBI"
    else:
        over_guard_label = ""

    # v23.29.2: no convertimos el falso Over automáticamente en apuesta Under.
    # Solo damos ✅ Under con perfil MUY claro; si no, queda como WATCH.
    under25_strong_context = (
        over_quality_block
        and straight_sets_clear
        and fav20 >= 0.70
        and over18 <= 0.68
        and under25_adjusted >= 0.76
    )
    if under25_strong_context:
        under25_label = "✅ MIRAR UNDER 2.5 SETS"
    elif under25_adjusted >= 0.63:
        under25_label = "👀 WATCH UNDER 2.5 SETS"
    else:
        under25_label = ""

    risk_notes = []
    if upset_guard.get("active", False):
        risk_notes.append("riesgo upset por favorito inflado")
    if min_conf < 0.35:
        risk_notes.append("confianza de datos muy baja")
    elif min_conf < 0.50:
        risk_notes.append("confianza de datos baja")
    elif min_conf < 0.65:
        risk_notes.append("confianza insuficiente para ML/2-0 fuerte")
    if min_surface_matches < 5:
        risk_notes.append("muestra de superficie insuficiente")
    elif min_surface_matches < 10:
        risk_notes.append("muestra de superficie corta")
    if elo_pure_gap >= 0.14:
        risk_notes.append("Elo puro y modelo final muy separados")

    zone_score = 0
    if circuito == "ATP" and surface == "Hard":
        zone_score = 3
    elif circuito == "ATP" and surface == "Clay":
        zone_score = 2
    elif circuito == "WTA":
        zone_score = 1
        risk_notes.append("WTA: filtros más exigentes")
    else:
        zone_score = 1

    if vol >= 0.060:
        risk_notes.append("volatilidad alta")
    if fav_prob < 0.56:
        risk_notes.append("favorito débil")
    if tb >= 0.34:
        risk_notes.append("tie-break alto")

    f1 = sim.get("fatigue1", {}).get("fatigue_score", 0)
    f2 = sim.get("fatigue2", {}).get("fatigue_score", 0)
    if f1 >= 0.04 or f2 >= 0.04:
        risk_notes.append("fatiga relevante")

    signals = []

    def add_signal(name, prob, min_prob, market_type, reason, bonus=0):
        score = zone_score + bonus

        # v22.2: no permitimos A+ fácil con muestra pobre.
        if min_conf < 0.35:
            score -= 2
        elif min_conf < 0.50:
            score -= 1

        if min_surface_matches < 5:
            score -= 1

        if elo_pure_gap >= 0.14:
            score -= 1

        # Over bajo WTA/ATP: controlar que no salga como spot fuerte con baja muestra.
        if name in ["Over 17.5", "Over 18.5"] and min_conf < 0.45:
            score -= 1
        if name == "Over 18.5" and fav_prob < 0.58:
            score -= 1

        if prob >= min_prob:
            score += 2
        if prob >= min_prob + 0.06:
            score += 1
        if prob >= min_prob + 0.12:
            score += 1

        if circuito == "WTA" and market_type in ["over", "long", "wta_over17"]:
            # v23.20: Over 17.5 solo para WTA + Watchlist Tight.
            # ATP/Challenger no se tocan. El 18.5 queda más selectivo y el 17.5
            # sirve como mercado puente cuando hay favorita clara pero no paliza total.
            if surface == "Clay" and name == "Over 18.5":
                if fav_prob >= 0.75:
                    score -= 3
                elif fav_prob >= 0.72:
                    score -= 2
                elif fav_prob >= 0.68:
                    score -= 1

                selective_recovery = False
                # Quitamos la antigua regla floja: Over 72% con favorita 50-60.
                # Ese patrón pasa a WATCHLIST/Over 17.5, no a señal principal 18.5.
                if 0.64 < fav_prob < 0.67 and prob >= 0.68:
                    selective_recovery = True
                    reason = reason + " · v23.20: Over 68% con favorita 64-67%"
                elif 0.70 <= fav_prob < 0.72 and prob >= 0.66:
                    selective_recovery = True
                    reason = reason + " · v23.20: Over 66% recuperado en zona 70-72%"

                if selective_recovery:
                    score += 2

            elif surface == "Clay" and name == "Over 17.5":
                if fav_prob >= 0.78:
                    score -= 3
                elif fav_prob >= 0.72:
                    score -= 1
                elif 0.64 <= fav_prob < 0.72:
                    score += 1

                over17_recovery = False
                if 0.64 <= fav_prob < 0.72 and prob >= 0.74:
                    over17_recovery = True
                    reason = reason + " · v23.20: Over 17.5 ideal con favorita 64-72%"
                elif 0.50 <= fav_prob < 0.64 and prob >= 0.78:
                    over17_recovery = True
                    reason = reason + " · v23.20: Over 17.5 alto con partido igualado"

                if over17_recovery:
                    score += 2

            else:
                score -= 2
                if fav_prob >= 0.65:
                    score -= 1
        if vol >= 0.060 and market_type in ["ml", "fav20"]:
            score -= 1
        if fav_prob < 0.56 and market_type in ["ml", "fav20"]:
            score -= 2

        if score >= 6:
            grade = "🔥 A+"
            action = "APTO fuerte"
        elif score >= 5:
            grade = "✅ A"
            action = "APTO"
        elif score >= 4:
            grade = "⚖️ B"
            action = "Solo si acompaña lectura"
        else:
            grade = "⚠️ C"
            action = "Evitar / observar"

        # v23.39.4 WTA Analyzer Rules Upgrade:
        # El Analyzer mostró que Over 17.5 funciona especialmente bien con RankGap 21-50
        # y modelo >=80%. Permitimos que ese perfil suba a APTO; el resto sigue prudente.
        if circuito == "WTA" and name == "Over 17.5" and prob >= 0.80 and wta_rankgap_ideal_over17:
            grade = "✅ A"
            action = "APTO"
            reason = reason + " · v23.39.4 Analyzer: RankGap 21-50 + Over17 >=80%"

        # WTA Clay Over Recovery debe producir DUDOSAS/Watchlist útiles, no APTA agresiva,
        # salvo la regla histórica nueva del Analyzer para Over 17.5.
        if (
            circuito == "WTA" and surface == "Clay" and name in ["Over 17.5", "Over 18.5"]
            and action in ["APTO fuerte", "APTO"]
            and not (name == "Over 17.5" and prob >= 0.80 and wta_rankgap_ideal_over17)
        ):
            grade = "⚖️ B"
            action = "Solo si acompaña lectura"
            reason = reason + " · v23.20 WTA Over: señal válida pero no APTA automática"

        # v22.25: con confianza <65% no dejamos señales agresivas ML/fav20.
        # Evita casos tipo favorito 80% + 2-0 95% con calidad Challenger/Qualy.
        if market_type in ["ml", "fav20"] and min_conf < 0.65 and action in ["APTO fuerte", "APTO"]:
            grade = "⚖️ B"
            action = "Solo si acompaña lectura"
            reason = reason + " · degradado por confianza insuficiente para ML/2-0"

        # v22.25: upset risk activo degrada cualquier ML/2-0 agresivo.
        if upset_guard.get("active", False) and market_type in ["ml", "fav20"] and action in ["APTO fuerte", "APTO"]:
            grade = "⚖️ B"
            action = "Solo si acompaña lectura"
            reason = reason + " · degradado por riesgo upset"

        # v23.29: Over Quality Guard. Si se activa, ningún Over/3 sets puede salir APTO.
        if (over_quality_block or over_quality_watch) and market_type in ["over", "long", "set3"] and action in ["APTO fuerte", "APTO"]:
            if over_quality_block:
                grade = "⚠️ C"
                action = "Evitar / observar"
                reason = reason + " · v23.29 Over bloqueado: " + "; ".join(over_quality_reasons[:3])
            else:
                grade = "⚖️ B"
                action = "Solo si acompaña lectura"
                reason = reason + " · v23.29 Over watch: " + "; ".join(over_quality_reasons[:3])

        # Trust Gate final: con confianza muy baja, como máximo B.
        if min_conf < 0.35 and action in ["APTO fuerte", "APTO"]:
            grade = "⚖️ B"
            action = "Solo si acompaña lectura"
            reason = reason + " · degradado por baja confianza de datos"
        elif min_conf < 0.50 and action == "APTO fuerte":
            grade = "✅ A"
            action = "APTO"
            reason = reason + " · sin A+ por confianza limitada"

        # v22.11: guardia de calidad combinada. Aunque la probabilidad sea alta,
        # una muestra clay corta y un gap Elo/modelo alto no pueden salir como APTO.
        if min_surface_matches < 10 and elo_pure_gap >= 0.14 and action in ["APTO fuerte", "APTO"]:
            grade = "⚖️ B"
            action = "Solo si acompaña lectura"
            reason = reason + " · degradado por muestra clay corta + gap Elo/modelo"

        signals.append({
            "Mercado": name,
            "Probabilidad": prob,
            "Umbral mínimo": min_prob,
            "Grade": grade,
            "Acción": action,
            "Motivo": reason
        })

    # =========================================================
    # v23.25 OVER FOCUS ENGINE
    # Cambio de filosofía: el ML queda como contexto visible, pero NO entra
    # como recomendación principal. El motor prioriza juegos y 3 sets.
    # =========================================================
    if circuito == "ATP" and surface == "Hard":
        add_signal("Over 18.5", over18, 0.73, "over", "v23.25 ATP Hard: foco principal en Over 18.5", 3)
        add_signal("Over 19.5", over19, 0.67, "over", "v23.25 ATP Hard: alternativa con mejor cuota si acompaña", 2)
        add_signal("Partido a 3 sets", set3, 0.43, "set3", "v23.25 ATP Hard: señal de partido largo / resistencia", 1)
        add_signal("Over 20.5", over20, 0.64, "over", "v23.25 ATP Hard: solo si la línea larga acompaña", 0)

    elif circuito == "ATP" and surface == "Clay":
        add_signal("Over 18.5", over18, 0.73, "over", "v23.25 ATP Clay: mercado principal por estabilidad", 3)
        add_signal("Over 19.5", over19, 0.66, "over", "v23.25 ATP Clay: segundo mercado si hay resistencia", 2)
        add_signal("Partido a 3 sets", set3, 0.44, "set3", "v23.25 ATP Clay: apoyo a lectura de over", 1)
        add_signal("Over 20.5", over20, 0.63, "over", "v23.25 ATP Clay: línea larga con cautela", 0)

    elif circuito == "WTA":
        # v23.39.4: el Analyzer validó favorito gana al menos 1 set >=90% como mercado WTA estable.
        fav_set_prob = sim.get("p1_wins_set_any", 0.0) if p1c >= p2c else sim.get("p2_wins_set_any", 0.0)
        add_signal(f"{fav_name} gana al menos 1 set", fav_set_prob, 0.90, "set_fav", "v23.39.4 WTA Analyzer: favorito gana set >=90%", 4)

        if surface == "Clay":
            # WTA mantiene la ventaja observada del Over 17.5.
            if fav_prob >= 0.78:
                over17_min, over17_bonus, over17_reason = 0.82, -1, "v23.25 WTA Clay: favorita dominante, Over 17.5 solo excepcional"
            elif fav_prob >= 0.72:
                over17_min, over17_bonus, over17_reason = 0.78, 1, "v23.25 WTA Clay: Over 17.5 conservador con favorita fuerte"
            elif fav_prob >= 0.64:
                over17_min, over17_bonus, over17_reason = 0.74, 3, "v23.25 WTA Clay: zona ideal Over 17.5"
            else:
                over17_min, over17_bonus, over17_reason = 0.78, 2, "v23.25 WTA Clay: partido igualado, mirar Over 17.5"
            add_signal("Over 17.5", over17, over17_min, "wta_over17", over17_reason, over17_bonus)

            if fav_prob >= 0.75:
                over18_min, over18_bonus, over18_reason = 0.74, -2, "v23.25 WTA Clay: Over 18.5 solo si sale muy claro"
            elif fav_prob >= 0.70:
                over18_min, over18_bonus, over18_reason = 0.66, 1, "v23.25 WTA Clay: Over 18.5 recuperable en zona 70-75%"
            elif fav_prob >= 0.64:
                over18_min, over18_bonus, over18_reason = 0.68, 1, "v23.25 WTA Clay: Over 18.5 secundario"
            else:
                over18_min, over18_bonus, over18_reason = 0.74, -1, "v23.25 WTA Clay: Over 18.5 exige mucha claridad en igualados"
            add_signal("Over 18.5", over18, over18_min, "over", over18_reason, over18_bonus)
        else:
            add_signal("Over 17.5", over17, 0.78, "wta_over17", "v23.25 WTA: Over 17.5 solo si señal muy clara", 1)
            add_signal("Over 18.5", over18, 0.74, "over", "v23.25 WTA: Over 18.5 secundario", 0)

        add_signal("Partido a 3 sets", set3, 0.44, "set3", "v23.25 WTA: apoyo a lectura de partido largo", 1)

    else:
        # Challenger / otros: no usar ML como pick principal. Over primero, 3 sets como apoyo.
        add_signal("Over 18.5", over18, 0.72, "over", "v23.25 Challenger: mercado principal, más fiable que ML", 2)
        add_signal("Over 19.5", over19, 0.67, "over", "v23.25 Challenger: alternativa si la cuota compensa", 1)
        add_signal("Partido a 3 sets", set3, 0.45, "set3", "v23.25 Challenger: watch de partido largo", 1)
        add_signal("Over 20.5", over20, 0.64, "over", "v23.25 Challenger: línea larga solo con cautela", 0)

    grade_order = {"🔥 A+": 4, "✅ A": 3, "⚖️ B": 2, "⚠️ C": 1}
    signals = sorted(signals, key=lambda x: (grade_order.get(x["Grade"], 0), x["Probabilidad"]), reverse=True)

    # v23.25.2 Over Focus Priority Fix:
    # La señal principal no debe ser ML/gana set si los Overs muestran estructura larga.
    # Se promueve el mercado de juegos de forma explícita antes de elegir el main.
    def _get_signal(market_name):
        return next((x for x in signals if x.get("Mercado") == market_name), None)

    def _promote_signal(signal, grade="✅ A", action="APTO", note=""):
        if signal is None:
            return None
        signal["Grade"] = grade
        signal["Acción"] = action
        if note:
            signal["Motivo"] = str(signal.get("Motivo", "")) + " · " + note
        return signal

    priority_main = None
    circuito_norm = str(circuito).upper().strip()
    if circuito_norm != "WTA" and not over_quality_block and not over_quality_watch:
        # Over 18.5 alto sigue siendo lectura principal solo si v23.29 no detecta falso Over.
        if over18 >= 0.73:
            priority_main = _promote_signal(
                _get_signal("Over 18.5"),
                grade="🔥 A+",
                action="APTO fuerte",
                note="v23.25.2: Over 18.5 no puede ser tapado por ML/contexto"
            )
        # Si el 18.5 está fuerte y el 19.5 acompaña, destacamos el 19.5 como mercado de valor.
        elif over18 >= 0.70 and over19 >= 0.66:
            priority_main = _promote_signal(
                _get_signal("Over 19.5"),
                grade="✅ A",
                action="APTO",
                note="v23.25.2: Over 19.5 priorizado por estructura larga"
            )
        # Si el 3 sets acompaña, mantenemos Over 18.5 como principal y 3 sets como apoyo.
        elif over18 >= 0.70 and set3 >= 0.45:
            priority_main = _promote_signal(
                _get_signal("Over 18.5"),
                grade="✅ A",
                action="APTO",
                note="v23.25.2: Over apoyado por probabilidad de 3 sets"
            )

    main = priority_main
    if main is None:
        for s in signals:
            if s["Acción"] in ["APTO fuerte", "APTO"]:
                main = s
                break

    # v23.25 WTA Over17 Priority:
    # Si no hay APTA clara y el Over 17.5 WTA es alto, lo priorizamos como lectura
    # conservadora antes que forzar ML u Over 18.5. No convierte la señal en APTA;
    # sigue saliendo como SPOT DUDOSO si el Signal Trust lo marca como B.
    if main is None and circuito == "WTA" and surface == "Clay" and over17 >= 0.77:
        over17_signal = next((x for x in signals if x.get("Mercado") == "Over 17.5"), None)
        if over17_signal is not None and over17_signal.get("Acción", "").startswith("Solo"):
            over17_signal["Motivo"] = str(over17_signal.get("Motivo", "")) + " · v23.25: prioridad conservadora Over 17.5"
            main = over17_signal

    if main is None and signals:
        main = signals[0]

    if main and main["Acción"] == "APTO fuerte":
        status = "🔥 SPOT FUERTE"
    elif main and main["Acción"] == "APTO":
        status = "✅ SPOT APTO"
    elif main and main["Acción"].startswith("Solo"):
        status = "⚖️ SPOT DUDOSO"
    else:
        status = "⚠️ NO BET / SOLO OBSERVAR"

    # v22.27 Upset Label Cleanup:
    # Si el guardia anti-upset está activo, no mostramos una lectura limpia de ML/2-0.
    # No cambia probabilidades; solo evita una etiqueta contradictoria en la señal final.
    if upset_guard.get("active", False) and main:
        mt = str(main.get("Mercado", "")).lower()
        if "ml favorito" in mt or "favorito 2-0" in mt:
            status = "⚠️ NO BET / RIESGO UPSET"

    # v23.29: si bloqueamos Over y hay señal Under 2.5, mostrarlo como rescate.
    if over_quality_block:
        status = "🚫 OVER BLOQUEADO / MIRAR UNDER 2.5" if under25_label else "🚫 OVER BLOQUEADO / NO BET"
    elif over_quality_watch and "OVER" in str(status).upper():
        status = "⚠️ OVER WATCH / NO COMBI"

    return {
        "status": status,
        "main": main,
        "signals": signals,
        "risk_notes": risk_notes,
        "zone_score": zone_score,
        "min_confidence": min_conf,
        "avg_confidence": avg_conf,
        "min_surface_matches": min_surface_matches,
        "elo_pure_gap": elo_pure_gap,
        "over_quality_block": over_quality_block,
        "over_quality_watch": over_quality_watch,
        "over_quality_reasons": over_quality_reasons,
        "over_guard_label": over_guard_label,
        "under25_raw": under25_raw,
        "under25_adjusted": under25_adjusted,
        "under25_label": under25_label,
        "straight_sets_risk": straight_sets_risk
    }


def render_betting_filters(filters):
    st.divider()
    st.subheader("💎 Signal Trust Engine")

    main = filters.get("main")
    status = filters.get("status", "⚠️ NO BET")

    if main:
        if "FUERTE" in status or "APTO" in status:
            st.success(f"{status} · {main['Mercado']} → {main['Probabilidad']:.1%}")
        elif "DUDOSO" in status:
            st.warning(f"{status} · {main['Mercado']} → {main['Probabilidad']:.1%}")
        else:
            st.error(f"{status} · mejor señal: {main['Mercado']} → {main['Probabilidad']:.1%}")
        st.caption(main.get("Motivo", ""))

    if filters.get("risk_notes"):
        st.warning("Riesgos: " + " · ".join(filters["risk_notes"]))

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Confianza mínima", f"{filters.get('min_confidence', 1.0):.0%}")
    with c2:
        st.metric("Mín. partidos superficie", f"{filters.get('min_surface_matches', 0)}")
    with c3:
        st.metric("Gap Elo puro/modelo", f"{filters.get('elo_pure_gap', 0):.1%}")

    rows = []
    for s in filters.get("signals", []):
        rows.append({
            "Mercado": s["Mercado"],
            "Prob": f"{s['Probabilidad']:.1%}",
            "Mín": f"{s['Umbral mínimo']:.1%}",
            "Grade": s["Grade"],
            "Acción": s["Acción"],
            "Motivo": s["Motivo"]
        })

    if rows:
        with st.expander("📋 Ver todas las señales del filtro", expanded=False):
            st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True)



# =========================================================
# v23.0 BATCH PASTE ANALYZER
# =========================================================

def _parse_decimal_odd(x):
    s = str(x).strip().replace(",", ".")
    try:
        v = float(s)
        if 1.01 <= v <= 100:
            return v
    except Exception:
        pass
    return None


def _abbr_match_score(abbr, full_name):
    """
    Compara abreviaturas tipo "S. Baez" o "I. Montes de la Torr"
    con nombres completos de la línea de partido.
    """
    a = normalizar_texto(abbr).lower().replace(".", " ")
    f = normalizar_texto(full_name).lower()
    a_parts = [p for p in re.split(r"\s+", a) if p]
    f_parts = [p for p in re.split(r"\s+", f) if p]
    if not a_parts or not f_parts:
        return 0.0

    # Si trae inicial + apellido aproximado.
    if len(a_parts) >= 2 and len(f_parts) >= 2 and len(a_parts[0]) == 1:
        initial_ok = f_parts[0].startswith(a_parts[0])
        last_abbr = " ".join(a_parts[1:])
        last_full = " ".join(f_parts[1:])
        last_score = SequenceMatcher(None, last_abbr, last_full).ratio()
        return (0.35 if initial_ok else 0.0) + 0.65 * last_score

    return SequenceMatcher(None, a, f).ratio()


def _match_quoted_player(quoted, p1_raw, p2_raw):
    quoted = re.sub(r"(?i).*ganador", "", str(quoted)).strip()
    quoted = quoted.replace("Guarantee Logo", "").strip()
    if not quoted:
        return None

    s1 = max(
        SequenceMatcher(None, normalizar_texto(quoted).lower(), normalizar_texto(p1_raw).lower()).ratio(),
        _abbr_match_score(quoted, p1_raw)
    )
    s2 = max(
        SequenceMatcher(None, normalizar_texto(quoted).lower(), normalizar_texto(p2_raw).lower()).ratio(),
        _abbr_match_score(quoted, p2_raw)
    )
    if max(s1, s2) < 0.45:
        return None
    return 1 if s1 >= s2 else 2


def parse_datos_extra_paste(raw_text):
    """
    Acepta formatos como:
    Jugador A - Jugador B
    Guarantee LogoGanadorS. Baez
    1,32

    También acepta:
    Jugador A - Jugador B | 1,50 | 2,70
    """
    lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []
    i = 0
    while i < len(lines):
        line = lines[i]
        clean = line.replace("–", "-").replace("—", "-")

        if " - " in clean:
            left, right = clean.split(" - ", 1)
            p1_raw = left.strip()
            rest = right.strip()

            odd1 = odd2 = None
            quoted_side = None
            quoted_odd = None
            quoted_text = None

            # Formato: P1 - P2 | cuota1 | cuota2
            parts = [p.strip() for p in rest.split("|")]
            p2_raw = parts[0].strip()
            if len(parts) >= 3:
                odd1 = _parse_decimal_odd(parts[1])
                odd2 = _parse_decimal_odd(parts[2])

            # Mirar 1-3 líneas siguientes por "Ganador..." y cuota.
            look = lines[i+1:i+4]
            for j, ln in enumerate(look):
                if "ganador" in normalizar_texto(ln).lower():
                    quoted_text = ln
                    quoted_side = _match_quoted_player(ln, p1_raw, p2_raw)
                    # cuota suele venir justo después
                    if i + 1 + j + 1 < len(lines):
                        quoted_odd = _parse_decimal_odd(lines[i + 1 + j + 1])
                    break

            # Si no hay "Ganador", intentar cuota directa siguiente.
            if quoted_odd is None and odd1 is None and i + 1 < len(lines):
                q = _parse_decimal_odd(lines[i+1])
                if q is not None:
                    quoted_odd = q

            matches.append({
                "raw": line,
                "p1_raw": p1_raw,
                "p2_raw": p2_raw,
                "odd1": odd1,
                "odd2": odd2,
                "quoted_side": quoted_side,
                "quoted_odd": quoted_odd,
                "quoted_text": quoted_text
            })
        i += 1

    return matches



def parse_simple_match_list(raw_text):
    """
    v23.2: parser limpio sin cuotas.
    Usa solo líneas con separador de partido:
    Jugador A - Jugador B
    Ignora cuotas, "Ganador", logos y líneas sueltas.
    """
    matches = []
    for ln in str(raw_text).splitlines():
        line = ln.strip()
        if not line:
            continue
        clean = line.replace("–", "-").replace("—", "-")
        low = normalizar_texto(clean).lower()
        if "ganador" in low or "guarantee" in low or "logo" in low:
            continue
        # ignorar cuotas sueltas
        if _parse_decimal_odd(clean) is not None:
            continue
        if " - " in clean:
            left, right = clean.split(" - ", 1)
            p1 = left.strip()
            p2 = right.strip()
            # limpiar posibles restos tras separadores
            p2 = p2.split("|")[0].strip()
            if p1 and p2:
                matches.append({
                    "raw": line,
                    "p1_raw": p1,
                    "p2_raw": p2,
                    "odd1": None,
                    "odd2": None,
                    "quoted_side": None,
                    "quoted_odd": None,
                    "quoted_text": None
                })
    return matches


def is_time_line_sofa(x):
    return bool(re.match(r"^\d{1,2}:\d{2}$", str(x).strip()))


def is_country_line_sofa(x):
    countries = {
        "andorra","argentina","australia","austria","belgium","bolivia","brazil","canada","chile","china","colombia",
        "croatia","czechia","denmark","finland","france","georgia","germany","hungary","italy","kazakhstan",
        "latvia","lebanon","lithuania","luxembourg","monaco","netherlands","paraguay","peru","portugal","russia","serbia",
        "spain","switzerland","tunisia","united kingdom","uruguay","usa","ukraine","poland","slovakia",
        "slovenia","romania","bulgaria","turkey","turkiye","türkiye","greece","norway","sweden","japan","india","south korea",
        "latvia","belarus","cyprus","malta","albania","armenia","azerbaijan","montenegro","north macedonia",
        "macedonia","kosovo","iran","iraq","qatar","uae","united arab emirates","saudi arabia","vietnam",
        "philippines","malaysia","singapore","uzbekistan","costa rica","puerto rico","morocco","czech republic","great britain"
    }
    return normalizar_texto(x).lower().strip() in countries


def is_sofa_meta_line(x):
    t = normalizar_texto(x).lower().strip()
    if not t:
        return True
    meta_exact = {"-", "atp", "wta", "challenger", "itf", "tierra batida", "hard", "grass"}
    if t in meta_exact:
        return True
    if re.match(r"^\d+$", t):
        return True
    if "doubles" in t:
        return False  # lo usamos para activar skip doubles
    if any(k in t for k in ["masters", "rome", "valencia", "bordeaux", "cordoba", "oeiras", "tunis", "zagreb"]):
        return True
    if "challenger" in t or "wta 1000" in t or "atp 1000" in t or "atp 250" in t or "atp 500" in t:
        return True
    if "," in t and len(t) > 6:
        return True
    return False


def looks_like_player_line_sofa(x):
    s = str(x).strip()
    if not s:
        return False
    if is_time_line_sofa(s) or s == "-":
        return False
    low = normalizar_texto(s).lower()
    if is_country_line_sofa(s):
        return False
    if is_sofa_meta_line(s):
        return False
    if "/" in s:
        return False  # dobles
    # Normalmente jugadores: inicial + apellido, o nombre abreviado.
    return bool(re.search(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]", s))



def is_pending_opponent_name(name):
    t = normalizar_texto(name).lower().strip()
    if not t:
        return True
    t = re.sub(r"\s+", "", t)
    if re.match(r"^(qf|sf|f|r\d+|r\d+p\d+|q\d+|w\d+|ll\d+|bye)(p\d+)?$", t):
        return True
    if re.match(r"^r\d+p\d+$", t):
        return True
    if t in {"qf1","qf2","qf3","qf4","sf1","sf2","winner","qualifier","luckyloser"}:
        return True
    return False


def is_country_like_name(name):
    t = normalizar_texto(name).lower().strip()
    if is_country_line_sofa(t):
        return True
    country_like = {
        "andorra", "bosnia & herzegovina", "bosnia and herzegovina", "great britain",
        "united states", "dominican republic", "south africa", "new zealand",
        "estonia", "bosnia", "herzegovina", "moldova", "moldavia", "israel",
        "ireland", "mexico", "ecuador", "venezuela", "morocco", "egypt",
        "taiwan", "hong kong", "thailand", "indonesia", "latvia", "belarus",
        "cyprus", "malta", "qatar", "uae", "uzbekistan", "vietnam", "philippines",
        "malaysia", "singapore", "costa rica", "puerto rico"
    }
    return t in country_like


def name_words_keep_initials(name):
    # Importante: NO usar limpiar(), porque elimina puntos/espacios y rompe "C. Ruud".
    t = normalizar_texto(name).lower().replace(".", " ")
    return [p for p in re.findall(r"[a-z]+", t) if p]


def surname_tokens_for_match(name):
    parts = name_words_keep_initials(name)
    if not parts:
        return []
    if len(parts[0]) == 1 and len(parts) >= 2:
        return parts[1:]
    if len(parts) >= 2:
        return parts[1:]
    return parts


def strict_abbrev_surname_compatible(query, full_name):
    """
    v23.20 Strict Surname Fix.
    Si la entrada viene como inicial + apellido, por ejemplo "D. Chiesa",
    solo se acepta un candidato cuyo apellido contenga exactamente Chiesa.
    Esto bloquea falsos positivos como Danielle Collins aunque el fuzzy sea medio.
    """
    qparts = name_words_keep_initials(query)
    fparts = name_words_keep_initials(full_name)
    if len(qparts) < 2 or len(fparts) < 2:
        return True
    if len(qparts[0]) != 1:
        return True

    q_surname_tokens = qparts[1:]
    f_surname_tokens = fparts[1:]
    if not q_surname_tokens or not f_surname_tokens:
        return True

    q_surname = " ".join(q_surname_tokens).strip()
    f_surname = " ".join(f_surname_tokens).strip()

    # Apellido simple: Chiesa solo puede casar con Chiesa, no Collins.
    if len(q_surname_tokens) == 1:
        return q_surname_tokens[0] in f_surname_tokens

    # Apellido compuesto: permitimos igualdad o contención de frase completa.
    return q_surname == f_surname or q_surname in f_surname or f_surname in q_surname


def abbreviation_player_score(query, full_name):
    """
    v23.8:
    C. Ruud -> Casper Ruud
    B. Van de Zandschulp -> Botic Van De Zandschulp
    T. Seyboth Wild -> Thiago Seyboth Wild
    J. Pinnington Jones -> Jack Pinnington Jones
    """
    q_parts = name_words_keep_initials(query)
    f_parts = name_words_keep_initials(full_name)
    if not q_parts or not f_parts:
        return 0.0

    # Inicial + apellido(s)
    if len(q_parts[0]) == 1 and len(q_parts) >= 2 and len(f_parts) >= 2:
        initial_ok = f_parts[0].startswith(q_parts[0])
        q_surname = " ".join(q_parts[1:])
        f_surname = " ".join(f_parts[1:])
        surname_score = SequenceMatcher(None, q_surname, f_surname).ratio()

        q_last = q_parts[-1]
        f_last = f_parts[-1]
        last_score = SequenceMatcher(None, q_last, f_last).ratio()

        # apellido compuesto exacto/contención
        if initial_ok and q_surname == f_surname:
            return 0.99
        if initial_ok and q_surname and (q_surname in f_surname or f_surname in q_surname):
            return 0.97
        if initial_ok and q_last == f_last:
            return max(0.94, 0.40 + 0.55 * surname_score)
        if initial_ok:
            return min(0.93, 0.38 + 0.55 * max(surname_score, last_score))

        return 0.55 * max(surname_score, last_score)

    # Apellido solo o nombre parcial.
    q_surname_tokens = surname_tokens_for_match(query)
    f_surname_tokens = surname_tokens_for_match(full_name)
    if q_surname_tokens and f_surname_tokens:
        qs = " ".join(q_surname_tokens)
        fs = " ".join(f_surname_tokens)
        if qs == fs:
            return 0.94
        if qs in fs or fs in qs:
            return 0.88
        return 0.70 * SequenceMatcher(None, qs, fs).ratio()

    return SequenceMatcher(None, " ".join(q_parts), " ".join(f_parts)).ratio() * 0.75


def parse_sofascore_paste(raw_text):
    """
    v23.4 parser Sofascore robusto.
    Detecta cualquier bloque:
    HORA / - / país / jugador / país / jugador

    No depende del nombre del torneo y no bloquea todo el bloque de dobles.
    Si una pareja contiene "/", se ignora solo ese partido.
    """
    raw_lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []
    i = 0

    while i < len(raw_lines):
        line = raw_lines[i]

        if not is_time_line_sofa(line):
            i += 1
            continue

        hora = line
        j = i + 1
        candidates = []
        doubles_match = False

        # Analizar hasta la próxima hora o hasta recoger 2 jugadores.
        while j < len(raw_lines) and not is_time_line_sofa(raw_lines[j]) and len(candidates) < 2:
            ln = raw_lines[j].strip()
            low = normalizar_texto(ln).lower()

            if ln == "-" or is_country_line_sofa(ln) or is_country_like_name(ln):
                j += 1
                continue

            # Si aparece rival pendiente dentro del partido, ignorar esta hora entera.
            if is_pending_opponent_name(ln):
                doubles_match = True
                break

            # Si justo después de la hora aparecen metadatos de torneo, cortar.
            if len(candidates) == 0 and is_sofa_meta_line(ln):
                j += 1
                continue

            if "/" in ln:
                doubles_match = True
                break

            if looks_like_player_line_sofa(ln) and not is_country_like_name(ln):
                candidates.append(ln)

            j += 1

        if len(candidates) >= 2 and not doubles_match:
            p1_raw, p2_raw = candidates[0], candidates[1]
            matches.append({
                "raw": f"{hora} · {p1_raw} - {p2_raw}",
                "time": hora,
                "p1_raw": p1_raw,
                "p2_raw": p2_raw,
                "odd1": None,
                "odd2": None,
                "quoted_side": None,
                "quoted_odd": None,
                "quoted_text": None
            })

        i = max(j, i + 1)

    return matches



def normalizar_superficie_pegada(x, default="Clay"):
    """Normaliza superficies copiadas de Sofascore/Flashscore sin tocar cálculos."""
    t = normalizar_texto(x).lower()
    if any(k in t for k in ["tierra", "clay", "arcilla"]):
        return "Clay"
    if any(k in t for k in ["dura", "hard", "cemento"]):
        return "Hard"
    if any(k in t for k in ["hierba", "grass"]):
        return "Grass"
    return default


def clasificar_bloque_torneo_pegado(tournament, meta_lines):
    """Clasifica encabezados pegados: ATP/WTA/Challenger y dobles."""
    txt = " ".join([str(tournament)] + [str(x) for x in meta_lines])
    clean = normalizar_texto(txt).upper()

    if "DOUBLES" in clean or "DOBLES" in clean:
        return "IGNORAR_DOBLES"
    if "WTA" in clean:
        if "125" in clean:
            return "WTA_125"
        return "WTA"
    if "CHALLENGER" in clean:
        # En Sofascore, Challenger sin WTA suele ser masculino.
        if "WOMEN" in clean or "MUJER" in clean or "WTA" in clean:
            return "CHALLENGER_WTA"
        return "CHALLENGER_ATP"
    # v23.26.7: Grand Slam/Qualifying es categoría, no circuito.
    # Debe heredar ATP/WTA del bloque. Si solo aparece Grand Slam sin ATP/WTA,
    # lo dejamos desconocido para no mezclar cuadros masculino/femenino.
    if "ATP" in clean:
        return "ATP"
    if "ITF" in clean:
        if "WOMEN" in clean or "MUJER" in clean:
            return "ITF_WTA"
        return "ITF_ATP"
    return "DESCONOCIDO"



def es_rival_pendiente_pegado(line):
    """Detecta placeholders tipo WQF1, WSF2, Qualifier, TBD."""
    t = limpiar(line)
    if not t:
        return True
    if is_pending_opponent_name(str(line)):
        return True
    if re.match(r"^(W|L)?(QF|SF|F|R\d|Q)\d*$", t):
        return True
    if re.match(r"^W(QF|SF|F)\d*$", t):
        return True
    if t in {"TBD", "BYE", "QUALIFIER", "LUCKYLOSER", "LL"}:
        return True
    return False


def es_linea_torneo_pegado(line):
    """Detecta títulos de torneo en pegados largos de Sofascore/Flashscore."""
    t = normalizar_texto(line).strip()
    u = t.upper()
    if not t or is_date_line_sofa_result(t) or is_time_line_sofa(t):
        return False
    if is_country_line_sofa(t) or is_country_like_name(t):
        return False
    if any(k in u for k in ["ATP", "WTA", "CHALLENGER", "ITF"]):
        return True
    if "," in t and not looks_like_player_line_sofa(t):
        return True
    return False


def filtrar_matches_por_circuito_pegado(matches, circuito, incluir_itf=False, incluir_challenger=True):
    """
    Filtro estricto por el circuito elegido en el menú lateral.

    v23.36.3:
    - Si eliges ATP: SOLO deja ATP + opcionalmente Challenger ATP + opcionalmente ITF ATP.
      Nunca mete WTA/WTA125.
    - Si eliges WTA: SOLO deja WTA + opcionalmente WTA125/Challenger WTA + opcionalmente ITF WTA.
      Nunca mete ATP/Challenger ATP.
    - WTA125 ya NO entra si el interruptor de secundarios está apagado.
    """
    circuito = str(circuito).upper().strip()
    if circuito == "ATP":
        allowed = {"ATP"}
        if incluir_challenger:
            allowed.add("CHALLENGER_ATP")
        if incluir_itf:
            allowed.add("ITF_ATP")
    elif circuito == "CHALLENGER":
        # v23.48.0: Challenger debe poder analizarse separado.
        # Usa el mismo parser que ATP/SofaScore, pero solo deja bloques Challenger.
        allowed = {"CHALLENGER_ATP"}
        if incluir_itf:
            allowed.add("ITF_ATP")
    elif circuito == "WTA":
        allowed = {"WTA"}
        if incluir_challenger:
            allowed.update({"WTA_125", "CHALLENGER_WTA"})
        if incluir_itf:
            allowed.add("ITF_WTA")
    else:
        allowed = {"ATP", "WTA", "CHALLENGER_ATP"}
    return [m for m in matches if str(m.get("circuito_detectado", "")).upper().strip() in allowed]


def _is_schedule_header_candidate(line, following=None):
    """v23.25.6: detecta cabeceras de torneos en pegados de horarios sin fecha."""
    t = normalizar_texto(line).strip()
    if not t or is_time_line_sofa(t) or t == "-":
        return False
    if is_country_line_sofa(t) or is_country_like_name(t):
        return False
    if is_number_line_sofa_schedule(t):
        return False
    u = t.upper()
    if any(k in u for k in ["ATP", "WTA", "CHALLENGER", "ITF", "DOUBLES"]):
        return True
    if "," in t:
        return True
    following = following or []
    ftxt = " ".join(normalizar_texto(x).upper() for x in following[:6])
    if any(k in ftxt for k in ["ATP", "WTA", "CHALLENGER", "ITF", "TIERRA", "DURA", "HARD", "CLAY", "GRASS"]):
        return True
    return False


def is_number_line_sofa_schedule(x):
    return bool(re.match(r"^\d+$", str(x).strip()))


def parse_sofascore_schedule_no_date_paste(raw_text):
    """
    v23.25.6 Schedule Parser Fix.
    Lee pegados de Sofascore/Flashscore SIN fecha:
      HORA / - / País / Jugador / País / Jugador
    Mantiene contexto ATP/WTA/Challenger, superficie y torneo.
    Ignora cabeceras vacías, dobles, números sueltos y partidos incompletos.
    """
    lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []

    current_tournament = ""
    current_meta = []
    current_circuit = "DESCONOCIDO"
    current_surface = "Clay"
    current_ignore_doubles = False

    def refresh_context():
        nonlocal current_circuit, current_surface, current_ignore_doubles
        current_circuit = clasificar_bloque_torneo_pegado(current_tournament, current_meta)
        current_ignore_doubles = current_circuit == "IGNORAR_DOBLES"
        for ml in current_meta + [current_tournament]:
            surf = normalizar_superficie_pegada(ml, default="")
            if surf in ["Hard", "Clay", "Grass"]:
                current_surface = surf

    def add_meta(ln):
        nonlocal current_meta
        if ln not in current_meta:
            current_meta.append(ln)
        if len(current_meta) > 8:
            current_meta = current_meta[-8:]
        refresh_context()

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Partido estricto: hora, -, país, jugador, país, jugador.
        if is_time_line_sofa(line):
            j = i + 1
            if j < len(lines) and lines[j].strip() == "-":
                j += 1

            if j + 3 < len(lines):
                pais1 = lines[j].strip()
                jugador1 = lines[j + 1].strip()
                pais2 = lines[j + 2].strip()
                jugador2 = lines[j + 3].strip()

                valid_country_pair = (
                    (is_country_line_sofa(pais1) or is_country_like_name(pais1)) and
                    (is_country_line_sofa(pais2) or is_country_like_name(pais2))
                )
                valid_players = (
                    looks_like_player_line_sofa(jugador1) and looks_like_player_line_sofa(jugador2) and
                    not is_country_like_name(jugador1) and not is_country_like_name(jugador2) and
                    not es_rival_pendiente_pegado(jugador1) and not es_rival_pendiente_pegado(jugador2) and
                    "/" not in jugador1 and "/" not in jugador2
                )

                if valid_country_pair and valid_players and not current_ignore_doubles:
                    matches.append({
                        "raw": f"{line} · {jugador1} - {jugador2}",
                        "time": line,
                        "p1_raw": jugador1,
                        "p2_raw": jugador2,
                        "surface": current_surface,
                        "torneo": current_tournament,
                        "circuito_detectado": current_circuit,
                        "odd1": None,
                        "odd2": None,
                        "quoted_side": None,
                        "quoted_odd": None,
                        "quoted_text": None
                    })
                    i = j + 4
                    continue

            # Si no encaja, avanzamos sin romper el parser completo.
            i += 1
            continue

        # Metadatos de circuito/categoría/superficie/números.
        low = normalizar_texto(line).lower().strip()
        up = normalizar_texto(line).upper().strip()
        is_circuit_or_category = (
            up in {"ATP", "WTA", "CHALLENGER", "ITF", "GRAND SLAM", "QUALIFYING", "QUALIFICATION"} or
            any(k in up for k in [
                "ATP 1000", "ATP 500", "ATP 250",
                "WTA 1000", "WTA 500", "WTA 250", "WTA 125",
                "CHALLENGER 50", "CHALLENGER 75", "CHALLENGER 100", "CHALLENGER 125", "CHALLENGER 175",
                "GRAND SLAM", "QUALIFYING", "QUALIFICATION"
            ])
        )
        is_surface = normalizar_superficie_pegada(line, default="") in ["Hard", "Clay", "Grass"]

        if is_circuit_or_category or is_surface or is_number_line_sofa_schedule(line):
            add_meta(line)
            i += 1
            continue

        # Cabecera/título de torneo. Importante para Hamburg/Geneva/Rome/Valencia/Istanbul.
        if _is_schedule_header_candidate(line, lines[i + 1:i + 7]):
            # v23.26.7: líneas sueltas tipo "Grand Slam" / "Qualifying"
            # son metadatos de categoría, no cabeceras de torneo. Si se trataban
            # como torneo, borraban ATP/WTA y luego no se leía nada.
            if up in {"GRAND SLAM", "QUALIFYING", "QUALIFICATION"}:
                add_meta(line)
                i += 1
                continue

            current_tournament = line
            # Subcabeceras tipo "Hamburg, Germany, Qualifying" o "Istanbul, Türkiye"
            # pueden venir justo antes de la hora y deben HEREDAR ATP/WTA/Challenger/superficie
            # del bloque padre. Si después vienen metadatos nuevos, sí empezamos bloque nuevo.
            if not (i + 1 < len(lines) and is_time_line_sofa(lines[i + 1])):
                current_meta = []
            refresh_context()
            i += 1
            continue

        i += 1

    return matches


def parse_sofascore_day_grouped_paste(raw_text):
    """
    Parser para pegar TODO el día desde Sofascore/Flashscore.
    Separa por encabezados ATP/WTA/Challenger, ignora dobles y descarta rivales pendientes.
    No cambia cálculos: solo devuelve lista limpia de partidos con circuito/superficie detectados.

    v23.25.6: si el pegado viene como horarios sin fecha, usa parser específico
    HORA / - / País / Jugador / País / Jugador para no romperse con cabeceras vacías.
    """
    lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    if lines and not any(is_date_line_sofa_result(x) for x in lines) and any(is_time_line_sofa(x) for x in lines):
        return parse_sofascore_schedule_no_date_paste(raw_text)
    matches = []
    current_tournament = ""
    current_meta = []
    current_circuit = "DESCONOCIDO"
    current_surface = "Clay"
    current_ignore_doubles = False

    def refresh_context():
        nonlocal current_circuit, current_surface, current_ignore_doubles
        current_circuit = clasificar_bloque_torneo_pegado(current_tournament, current_meta)
        current_ignore_doubles = current_circuit == "IGNORAR_DOBLES"
        for ml in current_meta:
            surf = normalizar_superficie_pegada(ml, default="")
            if surf in ["Hard", "Clay", "Grass"]:
                current_surface = surf

    i = 0
    while i < len(lines):
        line = lines[i]

        # Nuevo bloque de torneo. Guardamos título y unas líneas meta posteriores.
        if es_linea_torneo_principal_resultados(line) and not es_rival_pendiente_pegado(line):
            current_tournament = line
            current_meta = []
            j = i + 1
            while j < len(lines) and len(current_meta) < 6 and not is_date_line_sofa_result(lines[j]) and not is_time_line_sofa(lines[j]):
                # Si aparece otro título antes de fecha, seguimos usándolo como meta, porque Sofascore duplica nombres.
                current_meta.append(lines[j])
                j += 1
            refresh_context()
            i += 1
            continue

        # Partido: fecha / hora / país-jugador / país-jugador.
        if is_date_line_sofa_result(line):
            fecha = line
            if i + 1 >= len(lines) or not is_time_line_sofa(lines[i + 1]):
                i += 1
                continue
            hora = lines[i + 1]
            j = i + 2
            candidates = []
            pending_or_doubles = False

            while j < len(lines) and not is_date_line_sofa_result(lines[j]) and len(candidates) < 2:
                ln = lines[j].strip()

                # Si entra un nuevo torneo antes de completar 2 jugadores, el partido está incompleto.
                if es_linea_torneo_pegado(ln) and len(candidates) < 2:
                    pending_or_doubles = True
                    break

                if ln == "-" or is_country_line_sofa(ln) or is_country_like_name(ln) or is_sofa_meta_line(ln):
                    j += 1
                    continue
                if "/" in ln or es_rival_pendiente_pegado(ln):
                    pending_or_doubles = True
                    break
                if looks_like_player_line_sofa(ln) and not is_country_like_name(ln):
                    candidates.append(ln)
                j += 1

            if len(candidates) >= 2 and not pending_or_doubles and not current_ignore_doubles:
                matches.append({
                    "raw": f"{fecha} {hora} · {candidates[0]} - {candidates[1]}",
                    "date": fecha,
                    "time": hora,
                    "p1_raw": candidates[0],
                    "p2_raw": candidates[1],
                    "surface": current_surface,
                    "torneo": current_tournament,
                    "circuito_detectado": current_circuit,
                    "odd1": None,
                    "odd2": None,
                    "quoted_side": None,
                    "quoted_odd": None,
                    "quoted_text": None
                })
            i = max(j, i + 1)
            continue

        i += 1

    return matches


def is_date_line_sofa_result(x):
    return bool(re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", str(x).strip()))


def is_status_line_sofa_result(x):
    t = normalizar_texto(x).lower().strip()
    return t in {"ft", "final", "cancelado", "retirado", "walkover", "wo", "aplazado", "postponed"}


def parse_sofascore_results_paste(raw_text):
    """
    v23.12: parser Sofascore resultados con juegos reales.
    Detecta bloques como:
    fecha / FT / país / jugador1 / país / jugador2 /
    juegos por set + sets finales.

    Ejemplo 2 sets:
    6 6 4 4 2 2 0 0
    => sets reales 2-0, games reales 20

    Ejemplo con tiebreak/super tie:
    6 8 1 7 10 6 0 0 2 2
    => juegos reales estimados sumando todos los números de set antes de los 4 finales.
       Nota: si Sofascore incluye puntos de tie-break (8/10), el total de games puede inflarse.
    """
    raw_lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []
    i = 0

    while i < len(raw_lines):
        line = raw_lines[i]

        if not is_date_line_sofa_result(line):
            i += 1
            continue

        fecha = line
        if i + 1 >= len(raw_lines):
            i += 1
            continue

        status = normalizar_texto(raw_lines[i + 1]).strip()
        status_low = status.lower()
        j = i + 2

        # Saltar cancelados/retirados: no son buenos para backtest.
        if status_low not in {"ft", "final"}:
            while j < len(raw_lines) and not is_date_line_sofa_result(raw_lines[j]):
                j += 1
            i = j
            continue

        candidates = []
        numbers = []
        doubles_match = False

        while j < len(raw_lines) and not is_date_line_sofa_result(raw_lines[j]):
            ln = raw_lines[j].strip()

            if "/" in ln:
                doubles_match = True
                break

            if re.match(r"^\d+$", ln):
                numbers.append(int(ln))
                j += 1
                continue

            if ln == "-" or is_country_line_sofa(ln) or is_country_like_name(ln):
                j += 1
                continue

            if is_pending_opponent_name(ln):
                doubles_match = True
                break

            # Ignorar metadatos de torneo.
            if len(candidates) < 2 and is_sofa_meta_line(ln) and not looks_like_player_line_sofa(ln):
                j += 1
                continue

            if looks_like_player_line_sofa(ln) and len(candidates) < 2:
                candidates.append(ln)

            j += 1

        if len(candidates) >= 2 and not doubles_match:
            p1_raw, p2_raw = candidates[0], candidates[1]

            p1_sets = p2_sets = None
            winner_side = None
            actual_total_games = None
            score_games = ""

            # Sofascore suele dejar los 4 últimos números como:
            # p1_sets, p1_sets, p2_sets, p2_sets
            # y todo lo anterior como juegos/puntos visibles.
            game_nums = []
            if len(numbers) >= 6:
                p1_sets = numbers[-4]
                p2_sets = numbers[-2]
                game_nums = numbers[:-4]
            elif len(numbers) >= 4:
                # fallback antiguo: solo sets duplicados
                p1_sets = numbers[0]
                p2_sets = numbers[2]
                game_nums = []
            elif len(numbers) >= 2:
                p1_sets = numbers[0]
                p2_sets = numbers[1]
                game_nums = []

            if p1_sets is not None and p2_sets is not None:
                if p1_sets > p2_sets:
                    winner_side = 1
                elif p2_sets > p1_sets:
                    winner_side = 2

            # Calcular total games si hay juegos por set.
            # Emparejamos números de 2 en 2: p1_set1, p2_set1, p1_set2, p2_set2...
            # Si hay 6-8 / 7-10 por puntos de tie-break, esto puede inflar el total.
            if len(game_nums) >= 4 and len(game_nums) % 2 == 0:
                pairs = [(game_nums[k], game_nums[k+1]) for k in range(0, len(game_nums), 2)]
                actual_total_games = int(sum(a + b for a, b in pairs))
                score_games = " ".join([f"{a}-{b}" for a, b in pairs])

            matches.append({
                "raw": f"{fecha} · {p1_raw} - {p2_raw}",
                "date": fecha,
                "time": "",
                "status": status,
                "p1_raw": p1_raw,
                "p2_raw": p2_raw,
                "p1_sets_real": p1_sets,
                "p2_sets_real": p2_sets,
                "actual_winner_side": winner_side,
                "actual_total_games": actual_total_games,
                "score_games": score_games,
                "odd1": None,
                "odd2": None,
                "quoted_side": None,
                "quoted_odd": None,
                "quoted_text": None
            })

        while j < len(raw_lines) and not is_date_line_sofa_result(raw_lines[j]):
            j += 1
        i = max(j, i + 1)

    return matches


def parse_sofascore_results_grouped_paste(raw_text):
    """
    Parser para pegar resultados completos del día desde Sofascore/Flashscore.
    Mantiene el parser de resultados original, pero añade contexto de torneo/circuito/superficie
    para poder filtrar ATP/WTA/Challenger sin mezclar motores.
    No cambia cálculos: solo clasifica, filtra dobles y conserva ganador/games reales.
    """
    lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []
    current_tournament = ""
    current_meta = []
    current_circuit = "DESCONOCIDO"
    current_surface = "Clay"
    current_ignore_doubles = False

    def refresh_context():
        nonlocal current_circuit, current_surface, current_ignore_doubles
        current_circuit = clasificar_bloque_torneo_pegado(current_tournament, current_meta)
        current_ignore_doubles = current_circuit == "IGNORAR_DOBLES"
        for ml in current_meta:
            surf = normalizar_superficie_pegada(ml, default="")
            if surf in ["Hard", "Clay", "Grass"]:
                current_surface = surf

    i = 0
    while i < len(lines):
        line = lines[i]

        # Nuevo bloque de torneo antes de un resultado. Guardamos metadatos cercanos.
        if es_linea_torneo_pegado(line) and not es_rival_pendiente_pegado(line):
            current_tournament = line
            current_meta = []
            j = i + 1
            while j < len(lines) and len(current_meta) < 8 and not is_date_line_sofa_result(lines[j]):
                if is_status_line_sofa_result(lines[j]):
                    break
                current_meta.append(lines[j])
                j += 1
            refresh_context()
            i += 1
            continue

        if not is_date_line_sofa_result(line):
            i += 1
            continue

        fecha = line
        if i + 1 >= len(lines):
            i += 1
            continue

        status = normalizar_texto(lines[i + 1]).strip()
        status_low = status.lower()
        j = i + 2

        # Saltar cancelados/retirados/no finalizados y avanzar al siguiente resultado/bloque.
        if status_low not in {"ft", "final"}:
            while j < len(lines) and not is_date_line_sofa_result(lines[j]) and not es_linea_torneo_pegado(lines[j]):
                j += 1
            i = max(j, i + 1)
            continue

        candidates = []
        numbers = []
        invalid_match = False

        while j < len(lines) and not is_date_line_sofa_result(lines[j]) and not es_linea_torneo_pegado(lines[j]):
            ln = lines[j].strip()

            if "/" in ln:
                invalid_match = True
                break

            if re.match(r"^\d+$", ln):
                numbers.append(int(ln))
                j += 1
                continue

            if ln == "-" or is_country_line_sofa(ln) or is_country_like_name(ln):
                j += 1
                continue

            if es_rival_pendiente_pegado(ln):
                invalid_match = True
                break

            # Ignorar metadatos del torneo dentro del bloque si aparecen copiados entre líneas.
            if len(candidates) < 2 and is_sofa_meta_line(ln) and not looks_like_player_line_sofa(ln):
                j += 1
                continue

            if looks_like_player_line_sofa(ln) and len(candidates) < 2:
                candidates.append(ln)

            j += 1

        if len(candidates) >= 2 and not invalid_match and not current_ignore_doubles:
            p1_raw, p2_raw = candidates[0], candidates[1]

            p1_sets = p2_sets = None
            winner_side = None
            actual_total_games = None
            score_games = ""

            game_nums = []
            if len(numbers) >= 6:
                p1_sets = numbers[-4]
                p2_sets = numbers[-2]
                game_nums = numbers[:-4]
            elif len(numbers) >= 4:
                p1_sets = numbers[0]
                p2_sets = numbers[2]
                game_nums = []
            elif len(numbers) >= 2:
                p1_sets = numbers[0]
                p2_sets = numbers[1]
                game_nums = []

            if p1_sets is not None and p2_sets is not None:
                if p1_sets > p2_sets:
                    winner_side = 1
                elif p2_sets > p1_sets:
                    winner_side = 2

            if len(game_nums) >= 4 and len(game_nums) % 2 == 0:
                pairs = [(game_nums[k], game_nums[k+1]) for k in range(0, len(game_nums), 2)]
                actual_total_games = int(sum(a + b for a, b in pairs))
                score_games = " ".join([f"{a}-{b}" for a, b in pairs])

            matches.append({
                "raw": f"{fecha} · {p1_raw} - {p2_raw}",
                "date": fecha,
                "time": "",
                "status": status,
                "p1_raw": p1_raw,
                "p2_raw": p2_raw,
                "p1_sets_real": p1_sets,
                "p2_sets_real": p2_sets,
                "actual_winner_side": winner_side,
                "actual_total_games": actual_total_games,
                "score_games": score_games,
                "surface": current_surface,
                "torneo": current_tournament,
                "circuito_detectado": current_circuit,
                "odd1": None,
                "odd2": None,
                "quoted_side": None,
                "quoted_odd": None,
                "quoted_text": None
            })

        # Si hemos parado porque empieza un torneo, no saltarlo; el bucle lo procesará.
        i = max(j, i + 1)

    return matches


# =========================================================
# v23.25.3 Parser resultados Sofascore: hora + estado
# =========================================================
def is_live_or_nonfinal_status_sofa_result(x):
    """Estados que NO se usan para backtest/resultados cerrados."""
    t = normalizar_texto(x).lower().strip()
    return (
        t in {"-", "sin jugar", "suspendido", "interrumpido", "retirado", "cancelado", "aplazado", "postponed", "walkover", "wo"}
        or "set" in t
        or "retir" in t
        or "suspend" in t
        or "interrump" in t
    )


def is_final_status_sofa_result(x):
    t = normalizar_texto(x).lower().strip()
    return t in {"ft", "final"}


def es_linea_torneo_principal_resultados(line):
    """Título real de bloque, no metadatos sueltos tipo ATP/WTA/Challenger/ATP 1000."""
    t = normalizar_texto(line).strip()
    u = t.upper()
    if not es_linea_torneo_pegado(t):
        return False
    meta_sueltos = {
        "ATP", "WTA", "CHALLENGER", "ITF", "ATP 1000", "ATP 500", "ATP 250",
        "WTA 1000", "WTA 500", "WTA 250", "WTA 125", "CHALLENGER 50",
        "CHALLENGER 75", "CHALLENGER 100", "CHALLENGER 125", "CHALLENGER 175"
    }
    if u in meta_sueltos:
        return False
    return True


def is_match_start_sofa_result(lines, idx):
    """Inicio de partido de resultados: fecha+estado o hora+estado."""
    if idx + 1 >= len(lines):
        return False
    return (is_date_line_sofa_result(lines[idx]) or is_time_line_sofa(lines[idx])) and (
        is_final_status_sofa_result(lines[idx + 1]) or is_live_or_nonfinal_status_sofa_result(lines[idx + 1])
    )


def decodificar_games_sofascore(numbers):
    """
    v23.38.5: decodifica marcadores copiados desde SofaScore.

    Caso importante para Grand Slam / BO5:
    SofaScore, al copiar resultados, suele dejar los números así:
      jugador 1: juegos de cada set + posibles puntos de tie-break
      jugador 2: juegos de cada set + posibles puntos de tie-break
      final: sets_j1, sets_j2

    Ejemplos reales:
      6 6 6 1 3 4 3 0 -> 6-1 6-3 6-4, sets 3-0
      7 11 3 6 6 3 6 9 6 7 8 6 1 3
        -> 7-6(11-9) 3-6 6-7(6-8) 3-6, sets 1-3

    Devuelve: p1_sets, p2_sets, actual_total_games, score_games, games_ok
    """
    p1_sets = p2_sets = None
    actual_total_games = None
    score_games = ""
    games_ok = False

    if not numbers:
        return p1_sets, p2_sets, actual_total_games, score_games, games_ok

    nums = []
    for x in numbers:
        try:
            nums.append(int(x))
        except Exception:
            pass

    if len(nums) < 2:
        return p1_sets, p2_sets, actual_total_games, score_games, games_ok

    def _decode_rows(row1, row2, expected_sets=0):
        """Convierte dos filas SofaScore en pares de sets, saltando columnas de tie-break."""
        set_pairs = []
        annotated = []
        last_set_index = None

        for a, b in zip(row1, row2):
            try:
                a, b = int(a), int(b)
            except Exception:
                continue

            # Columna de tie-break copiada por SofaScore: suele aparecer justo después
            # de un set 7-6 / 6-7 y ambos valores pueden ser 7, 8, 9, 10, 11...
            if last_set_index is not None and (a >= 6 and b >= 6):
                pa, pb = set_pairs[last_set_index]
                if sorted([pa, pb]) == [6, 7]:
                    annotated[last_set_index] = f"{pa}-{pb}({a}-{b})"
                    continue

            # Columna normal de juegos de set.
            # Aceptamos sets normales: 6-x, 7-5, 7-6, 0-6, etc.
            looks_like_set = (
                (max(a, b) >= 6 and abs(a - b) >= 2) or
                (sorted([a, b]) in ([5, 7], [6, 7])) or
                (a == 0 and b >= 6) or (b == 0 and a >= 6)
            )

            if not looks_like_set:
                # Si parece columna rara de SofaScore y ya tenemos suficientes sets, se ignora.
                if expected_sets and len(set_pairs) >= expected_sets:
                    continue
                # Si no, la dejamos pasar para no perder información en copiados raros.

            set_pairs.append((a, b))
            annotated.append(f"{a}-{b}")
            last_set_index = len(set_pairs) - 1

            if expected_sets and len(set_pairs) >= expected_sets:
                # No cortamos aquí: puede venir justo después la columna de tie-break del último set.
                pass

        if expected_sets and len(set_pairs) > expected_sets:
            set_pairs = set_pairs[:expected_sets]
            annotated = annotated[:expected_sets]

        return set_pairs, annotated

    # Formato actual observado en SofaScore resultados manual:
    # [fila_j1..., fila_j2..., sets_j1, sets_j2]
    best_pairs = []
    best_annotated = []
    if len(nums) >= 4:
        cand_p1_sets, cand_p2_sets = nums[-2], nums[-1]
        cand_sets_played = cand_p1_sets + cand_p2_sets
        body = nums[:-2]

        if 0 <= cand_p1_sets <= 5 and 0 <= cand_p2_sets <= 5 and 1 <= cand_sets_played <= 5 and len(body) >= 2 and len(body) % 2 == 0:
            mid = len(body) // 2
            row1 = body[:mid]
            row2 = body[mid:]
            pairs, annotated = _decode_rows(row1, row2, expected_sets=cand_sets_played)

            if len(pairs) >= cand_sets_played:
                p1_sets, p2_sets = cand_p1_sets, cand_p2_sets
                best_pairs, best_annotated = pairs[:cand_sets_played], annotated[:cand_sets_played]

    # Fallback antiguo: algunos copiados dejaban últimos 4 números duplicados.
    if not best_pairs and len(nums) >= 6:
        cand_p1_sets, cand_p2_sets = nums[-4], nums[-2]
        cand_sets_played = cand_p1_sets + cand_p2_sets
        body = nums[:-4]
        if 0 <= cand_p1_sets <= 5 and 0 <= cand_p2_sets <= 5 and 1 <= cand_sets_played <= 5 and len(body) >= 2 and len(body) % 2 == 0:
            mid = len(body) // 2
            pairs, annotated = _decode_rows(body[:mid], body[mid:], expected_sets=cand_sets_played)
            if len(pairs) >= cand_sets_played:
                p1_sets, p2_sets = cand_p1_sets, cand_p2_sets
                best_pairs, best_annotated = pairs[:cand_sets_played], annotated[:cand_sets_played]

    # Fallback mínimo: números intercalados por parejas + sets al final.
    if not best_pairs and len(nums) >= 4:
        cand_p1_sets, cand_p2_sets = nums[-2], nums[-1]
        cand_sets_played = cand_p1_sets + cand_p2_sets
        body = nums[:-2]
        if 0 <= cand_p1_sets <= 5 and 0 <= cand_p2_sets <= 5 and 1 <= cand_sets_played <= 5 and len(body) >= cand_sets_played * 2:
            raw_pairs = list(zip(body[::2], body[1::2]))
            pairs, annotated = _decode_rows([a for a, _ in raw_pairs], [b for _, b in raw_pairs], expected_sets=cand_sets_played)
            if len(pairs) >= cand_sets_played:
                p1_sets, p2_sets = cand_p1_sets, cand_p2_sets
                best_pairs, best_annotated = pairs[:cand_sets_played], annotated[:cand_sets_played]

    if best_pairs:
        actual_total_games = int(sum(a + b for a, b in best_pairs))
        score_games = " ".join(best_annotated)
        games_ok = True

    return p1_sets, p2_sets, actual_total_games, score_games, games_ok

def _parse_finished_result_from_position(lines, i, current_tournament="", current_surface="Clay", current_circuit="DESCONOCIDO", current_ignore_doubles=False):
    """
    Lee un resultado desde lines[i]. Acepta:
      fecha / FT / pais / jugador / pais / jugador / números
      hora  / FT / pais / jugador / pais / jugador / números
    Devuelve (match_o_None, nuevo_indice).
    """
    start_line = lines[i]
    is_date_start = is_date_line_sofa_result(start_line)
    fecha = start_line if is_date_start else ""
    hora = "" if is_date_start else start_line

    if i + 1 >= len(lines):
        return None, i + 1

    status = normalizar_texto(lines[i + 1]).strip()
    j = i + 2

    # Solo analizamos partidos cerrados. Los live/sin jugar/suspendidos/retirados se saltan.
    if not is_final_status_sofa_result(status):
        while j < len(lines) and not is_match_start_sofa_result(lines, j) and not es_linea_torneo_principal_resultados(lines[j]):
            j += 1
        return None, max(j, i + 1)

    candidates = []
    numbers = []
    invalid_match = False

    while j < len(lines) and not is_match_start_sofa_result(lines, j) and not es_linea_torneo_principal_resultados(lines[j]):
        ln = lines[j].strip()

        if "/" in ln:
            invalid_match = True
            break

        if re.match(r"^\d+$", ln):
            numbers.append(int(ln))
            j += 1
            continue

        if ln == "-" or is_country_line_sofa(ln) or is_country_like_name(ln):
            j += 1
            continue

        if es_rival_pendiente_pegado(ln):
            invalid_match = True
            break

        if len(candidates) < 2 and is_sofa_meta_line(ln) and not looks_like_player_line_sofa(ln):
            j += 1
            continue

        if looks_like_player_line_sofa(ln) and len(candidates) < 2:
            candidates.append(ln)

        j += 1

    if len(candidates) < 2 or invalid_match or current_ignore_doubles:
        return None, max(j, i + 1)

    p1_raw, p2_raw = candidates[0], candidates[1]
    p1_sets = p2_sets = None
    winner_side = None
    actual_total_games = None
    score_games = ""

    p1_sets, p2_sets, actual_total_games, score_games, _games_ok = decodificar_games_sofascore(numbers)

    if p1_sets is not None and p2_sets is not None:
        if p1_sets > p2_sets:
            winner_side = 1
        elif p2_sets > p1_sets:
            winner_side = 2

    raw_prefix = (f"{fecha} {hora}" if fecha and hora else (fecha or hora)).strip()
    match = {
        "raw": f"{raw_prefix} · {p1_raw} - {p2_raw}",
        "date": fecha,
        "time": hora,
        "status": status,
        "p1_raw": p1_raw,
        "p2_raw": p2_raw,
        "p1_sets_real": p1_sets,
        "p2_sets_real": p2_sets,
        "actual_winner_side": winner_side,
        "actual_total_games": actual_total_games,
        "score_games": score_games,
        "games_leidos": bool(actual_total_games is not None),
        "surface": current_surface,
        "torneo": current_tournament,
        "circuito_detectado": current_circuit,
        "odd1": None,
        "odd2": None,
        "quoted_side": None,
        "quoted_odd": None,
        "quoted_text": None
    }
    return match, max(j, i + 1)


def parse_sofascore_results_paste(raw_text):
    """
    v23.25.3: parser resultados compatible con bloques que empiezan por fecha o por hora.
    Ignora partidos live/sin jugar/suspendidos/retirados; solo devuelve FT/final.
    """
    lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []
    i = 0
    while i < len(lines):
        if not is_match_start_sofa_result(lines, i):
            i += 1
            continue
        match, ni = _parse_finished_result_from_position(lines, i)
        if match:
            matches.append(match)
        i = max(ni, i + 1)
    return matches


def parse_sofascore_results_grouped_paste(raw_text):
    """
    v23.25.3: resultados completos del día desde Sofascore/Flashscore.
    Acepta partidos cerrados con formato hora+FT o fecha+FT.
    Ignora dobles, live, sin jugar, suspendidos y retirados. No toca cálculos.
    """
    lines = [ln.strip() for ln in str(raw_text).splitlines() if ln.strip()]
    matches = []
    current_tournament = ""
    current_meta = []
    current_circuit = "DESCONOCIDO"
    current_surface = "Clay"
    current_ignore_doubles = False

    def refresh_context():
        nonlocal current_circuit, current_surface, current_ignore_doubles
        current_circuit = clasificar_bloque_torneo_pegado(current_tournament, current_meta)
        current_ignore_doubles = current_circuit == "IGNORAR_DOBLES"
        for ml in current_meta:
            surf = normalizar_superficie_pegada(ml, default="")
            if surf in ["Hard", "Clay", "Grass"]:
                current_surface = surf

    i = 0
    while i < len(lines):
        line = lines[i]

        if es_linea_torneo_principal_resultados(line) and not es_rival_pendiente_pegado(line):
            current_tournament = line
            current_meta = []
            j = i + 1
            while j < len(lines) and len(current_meta) < 10 and not is_match_start_sofa_result(lines, j):
                # Si ya viene otro título real, dejamos que el bucle principal lo procese como nuevo bloque.
                # No cortamos por metadatos sueltos tipo ATP/WTA/Challenger/ATP 1000.
                if j != i + 1 and es_linea_torneo_principal_resultados(lines[j]):
                    break
                current_meta.append(lines[j])
                j += 1
            refresh_context()
            i += 1
            continue

        if not is_match_start_sofa_result(lines, i):
            i += 1
            continue

        match, ni = _parse_finished_result_from_position(
            lines, i,
            current_tournament=current_tournament,
            current_surface=current_surface,
            current_circuit=current_circuit,
            current_ignore_doubles=current_ignore_doubles
        )
        if match:
            matches.append(match)
        i = max(ni, i + 1)

    return matches

def find_player_in_db(name, db):
    if is_pending_opponent_name(name) or is_country_like_name(name):
        return None, 0.0

    target = normalizar_texto(name).lower()
    target_clean = limpiar(name).lower()
    best_key = None
    best_score = 0.0

    for key in db.keys():
        k1 = normalizar_texto(key).lower()
        k2 = limpiar(key).lower()

        score = max(
            SequenceMatcher(None, target, k1).ratio(),
            SequenceMatcher(None, target_clean, k2).ratio(),
            abbreviation_player_score(name, key)
        )

        # Bonus apellido+inicial en ambos sentidos.
        tparts = name_words_keep_initials(name)
        kparts = name_words_keep_initials(key)
        if len(tparts) >= 2 and len(kparts) >= 2:
            # "C Ruud" vs "Casper Ruud"
            t_initial = len(tparts[0]) == 1 and kparts[0].startswith(tparts[0])
            same_last = tparts[-1] == kparts[-1]
            if t_initial and same_last:
                score = max(score, 0.97)
            elif same_last:
                score = max(score, 0.88)

            # Apellidos compuestos.
            t_surname = " ".join(tparts[1:]) if len(tparts[0]) == 1 else " ".join(tparts[1:])
            k_surname = " ".join(kparts[1:])
            if t_initial and t_surname and (t_surname in k_surname or k_surname in t_surname):
                score = max(score, 0.96)

        # v23.20 Fix nombres: si viene "D. Chiesa", no aceptar candidatos
        # cuyo apellido no sea Chiesa aunque el fuzzy general sea alto.
        if not strict_abbrev_surname_compatible(name, key):
            score = min(score, 0.40)

        if score > best_score:
            best_score = score
            best_key = key

    return best_key, best_score

def find_player_candidates_in_db(name, db, top_n=5):
    if is_pending_opponent_name(name) or is_country_like_name(name):
        return [{"candidate": "Rival pendiente / país detectado", "score": 0.0, "reason": "no analizable"}]

    target = normalizar_texto(name).lower()
    target_clean = limpiar(name).lower()
    rows = []

    for key in db.keys():
        k1 = normalizar_texto(key).lower()
        k2 = limpiar(key).lower()
        score = max(
            SequenceMatcher(None, target, k1).ratio(),
            SequenceMatcher(None, target_clean, k2).ratio(),
            abbreviation_player_score(name, key)
        )

        tparts = name_words_keep_initials(name)
        kparts = name_words_keep_initials(key)

        reason = "fuzzy"
        if len(tparts) >= 2 and len(kparts) >= 2:
            t_initial = len(tparts[0]) == 1 and kparts[0].startswith(tparts[0])
            same_last = tparts[-1] == kparts[-1]
            t_surname = " ".join(tparts[1:]) if len(tparts[0]) == 1 else " ".join(tparts[1:])
            k_surname = " ".join(kparts[1:])

            if t_initial and same_last:
                score = max(score, 0.97)
                reason = "inicial+apellido"
            elif t_initial and t_surname and (t_surname in k_surname or k_surname in t_surname):
                score = max(score, 0.96)
                reason = "inicial+apellido compuesto"
            elif same_last:
                score = max(score, 0.88)
                reason = "apellido"

        # v23.20 Fix nombres: evita sugerencias falsas por fuzzy cuando
        # la consulta es inicial + apellido y el apellido no coincide.
        if not strict_abbrev_surname_compatible(name, key):
            score = min(score, 0.40)
            reason = "apellido no coincide - bloqueado"

        rows.append({"candidate": key, "score": score, "reason": reason})

    rows = sorted(rows, key=lambda x: x["score"], reverse=True)[:top_n]
    return rows


# =========================================================
# v23.25.8 FALLBACK DE LECTURA
# =========================================================
# Antes, si un jugador no estaba en el ELO principal, el partido completo caía
# en "No encontrado". Esto era demasiado duro para Challenger/ITF/qualy.
# Ahora: si el nombre no aparece en db, creamos un perfil estimado MUY prudente
# y lo marcamos como baja confianza. Así el partido entra en la hoja, pero con
# aviso claro de que NO debe tratarse como pick fuerte.


def elo_estimado_por_quality(q, circuito="ATP"):
    try:
        total = int(q.get("matches_total", 0) or 0)
        tq = float(q.get("tour_quality", 0.45) or 0.45)
        levels = q.get("level_counts", {}) or {}
        tour = int(levels.get("tour", 0) or 0)
        chall = int(levels.get("challenger", 0) or 0)
        itf = int(levels.get("itf", 0) or 0)
    except Exception:
        total, tq, tour, chall, itf = 0, 0.45, 0, 0, 0

    # Base prudente. No queremos regalar Elo a un desconocido.
    if circuito == "WTA":
        elo = 1485
    else:
        elo = 1500

    if total >= 80: elo += 55
    elif total >= 40: elo += 35
    elif total >= 20: elo += 20
    elif total >= 8: elo += 8
    elif total <= 2: elo -= 20

    if tour >= 10: elo += 35
    elif tour >= 3: elo += 18
    if chall >= 15: elo += 14
    if itf >= max(3, total * 0.45): elo -= 22

    elo += int((tq - 0.45) * 80)
    return float(np.clip(elo, 1360, 1625))


def crear_jugador_estimado(nombre, circuito="ATP", surface="Clay"):
    raw = normalizar_texto(nombre)
    # Intentamos sacar match count real desde históricos. Si no existe, perfil neutro.
    try:
        q = buscar_quality_directo_historicos(raw, circuito)
    except Exception:
        q = {}

    if not isinstance(q, dict):
        q = {}

    elo = elo_estimado_por_quality(q, circuito)
    rank = 999
    stats_by_surface = {}
    for sf in ["Hard", "Clay", "Grass"]:
        stats_by_surface[sf] = stats_default_por_elo(elo, rank, sf, circuito)
        stats_by_surface[sf]["match_type"] = "fallback_estimado"
        stats_by_surface[sf]["raw_name_stats"] = "ESTIMADO - sin ELO principal"

    if not q or int(q.get("matches_total", 0) or 0) == 0:
        q = {
            "matches_total": 0,
            "matches_surface": {"Hard": 0, "Clay": 0, "Grass": 0},
            "level_counts": {"tour": 0, "challenger": 0, "itf": 0, "qualy": 0, "unknown": 0},
            "tour_quality": 0.35,
            "stability": {"Hard": 0.05, "Clay": 0.05, "Grass": 0.05},
            "confidence": {"Hard": 0.28, "Clay": 0.28, "Grass": 0.28},
            "raw_names": [raw],
            "matched_name": "ESTIMADO SIN HISTÓRICO",
            "match_score": 0.0,
        }

    return {
        "Player": raw,
        "Rank": rank,
        "Hard": elo,
        "Clay": elo,
        "Grass": elo,
        "Stats": stats_by_surface.get(surface, stats_by_surface["Clay"]),
        "StatsBySurface": stats_by_surface,
        "Fatigue": {},
        "Quality": q,
        "FallbackEstimado": True,
    }


def resolver_player_batch(raw_name, db, circuito="ATP", surface="Clay", allow_fallback=True):
    key, score = find_player_in_db(raw_name, db)
    if key and score >= 0.66:
        d = db[key]
        try:
            d = d.copy()
            d["FallbackEstimado"] = False
        except Exception:
            pass
        return key, d, float(score), "OK"

    if not allow_fallback:
        return key, None, float(score), "NO_ENCONTRADO"

    # No usar fallback para países, BYE, dobles o placeholders.
    if is_pending_opponent_name(raw_name) or is_country_like_name(raw_name) or "/" in str(raw_name):
        return key, None, float(score), "NO_ANALIZABLE"

    fallback = crear_jugador_estimado(raw_name, circuito, surface)
    return normalizar_texto(raw_name), fallback, max(float(score), 0.50), "ESTIMADO"

def enrich_not_found_with_suggestions(ko_df, db):
    if ko_df is None or ko_df.empty:
        return ko_df

    out = ko_df.copy()
    suggestions = []
    scores = []
    reasons = []

    for _, row in out.iterrows():
        raw_match = str(row.get("Partido", ""))
        parts = raw_match.split(" vs ")
        raw_names = [p.strip() for p in parts if p.strip()]

        cand_bits = []
        score_bits = []
        reason_bits = []

        for nm in raw_names:
            cands = find_player_candidates_in_db(nm, db, top_n=3)
            if cands:
                best = cands[0]
                cand_bits.append(f"{nm} → {best['candidate']}")
                score_bits.append(f"{best['score']:.0%}")
                reason_bits.append(best["reason"])

        suggestions.append(" | ".join(cand_bits))
        scores.append(" | ".join(score_bits))
        reasons.append(" | ".join(reason_bits))

    out["Sugerencia"] = suggestions
    out["Score sugerencia"] = scores
    out["Tipo sugerencia"] = reasons
    return out

def batch_pick_label(sim, over17, over18, over19, over20, over22, under22, p1_name, p2_name, circuito=None, surface=None):
    p1c, p2c = sim.get("p1_cal", 0.5), sim.get("p2_cal", 0.5)
    fav_prob = max(p1c, p2c)
    fav_name = p1_name if p1c >= p2c else p2_name
    dog_name = p2_name if p1c >= p2c else p1_name

    candidates = [
        (f"ML {fav_name}", fav_prob),
        ("Over 18.5", over18),
        ("Over 19.5", over19),
        ("Over 20.5", over20),
        ("Under 22.5", under22),
        (f"{dog_name} gana al menos 1 set", sim.get("dog_wins_set", 0.0)),
    ]
    # Over 17.5 se añade solo para WTA. ATP/Challenger no se modifican.
    if circuito == "WTA":
        candidates.insert(1, ("Over 17.5", over17))

    # v23.20: etiqueta rápida selectiva para WTA Clay.
    # Solo damos bonus al Over 18.5 en los patrones concretos recuperados.
    if circuito == "WTA" and surface == "Clay":
        if fav_prob >= 0.75:
            penalty = 0.10
        elif fav_prob >= 0.72:
            penalty = 0.08
        elif fav_prob >= 0.68:
            penalty = 0.04
        else:
            penalty = 0.00

        adjusted = []
        for label, prob in candidates:
            p = prob
            if label.startswith("Over"):
                p = prob - penalty
                if label == "Over 17.5" and 0.64 <= fav_prob < 0.72 and prob >= 0.74:
                    p += 0.06
                elif label == "Over 17.5" and 0.50 <= fav_prob < 0.64 and prob >= 0.78:
                    p += 0.03
                elif label == "Over 18.5" and 0.64 < fav_prob < 0.67 and prob >= 0.68:
                    p += 0.05
                elif label == "Over 18.5" and 0.70 <= fav_prob < 0.72 and prob >= 0.66:
                    p += 0.04
            adjusted.append((label, p))
        candidates = adjusted

    candidates = [(k, v) for k, v in candidates if v is not None]

    # v23.25.2 Over Focus Priority Fix:
    # En ATP/Challenger el ML queda como contexto. Si la estructura de juegos es clara,
    # el mejor mercado visible debe ser Over 18.5 / Over 19.5 y no ML ni "gana set".
    if circuito != "WTA":
        try:
            set3_prob = float(sim.get("set3", sim.get("market_3sets", sim.get("prob_3sets", 0.0))) or 0.0)
            o18 = float(over18 or 0.0)
            o19 = float(over19 or 0.0)
            if o18 >= 0.73:
                return "Over 18.5", o18
            if o18 >= 0.70 and o19 >= 0.66:
                return "Over 19.5", o19
            if o18 >= 0.70 and set3_prob >= 0.45:
                return "Over 18.5", o18
        except Exception:
            pass

    # v23.25 WTA Over17 Priority:
    # Para WTA Clay, si el 17.5 está alto, se muestra como mejor mercado conservador.
    # ATP/Challenger no pasan por aquí porque Over 17.5 solo se inserta para WTA.
    if circuito == "WTA" and surface == "Clay" and over17 is not None:
        try:
            if float(over17) >= 0.77 and 0.50 <= float(fav_prob) < 0.75:
                return "Over 17.5", float(over17)
        except Exception:
            pass

    label, prob = max(candidates, key=lambda x: x[1])
    return label, float(prob)



def _safe_float_v23370(x, default=0.0):
    try:
        if x is None or pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def _confianza_acierto_v23370(prob):
    prob = _safe_float_v23370(prob, 0.0)
    if prob >= 0.84:
        return "🔥 Muy alta"
    if prob >= 0.78:
        return "✅ Alta"
    if prob >= 0.70:
        return "✅ Media-alta"
    if prob >= 0.64:
        return "⚖️ Media"
    return "⚠️ Baja"


def selector_mercado_maximo_acierto_v23370(sim, over17, over18, over19, over20, over22, under22,
                                           p1_name, p2_name, circuito=None, surface=None):
    """
    v23.37 Selector Maestro de Máximo Acierto.
    Objetivo: para cada partido elegir el mercado con MAYOR probabilidad estimada.
    - No usa cuotas.
    - No calcula value.
    - No modifica ninguna fórmula ni guard del motor Over.
    - Usa las probabilidades ya calculadas por la simulación: ML, gana set, overs y 3 sets.
    """
    p1c = _safe_float_v23370(sim.get("p1_cal", 0.5), 0.5)
    p2c = _safe_float_v23370(sim.get("p2_cal", 1 - p1c), 1 - p1c)

    p1_set = _safe_float_v23370(sim.get("p1_wins_set_any", 0.0), 0.0)
    p2_set = _safe_float_v23370(sim.get("p2_wins_set_any", 0.0), 0.0)

    # Compatibilidad con versiones/cachés antiguas donde aún no existían p1_wins_set_any/p2_wins_set_any.
    if p1_set <= 0.0 and p2_set <= 0.0:
        dog_set = _safe_float_v23370(sim.get("dog_wins_set", 0.0), 0.0)
        if p1c >= p2c:
            p2_set = dog_set
            p1_set = max(p1c, 1.0 - _safe_float_v23370(sim.get("p2_2_0", 0.0), 0.0))
        else:
            p1_set = dog_set
            p2_set = max(p2c, 1.0 - _safe_float_v23370(sim.get("p1_2_0", 0.0), 0.0))

    fav_is_p1 = p1c >= p2c
    fav_name = p1_name if fav_is_p1 else p2_name
    dog_name = p2_name if fav_is_p1 else p1_name
    fav_prob = p1c if fav_is_p1 else p2c
    dog_prob = p2c if fav_is_p1 else p1c
    fav_set = p1_set if fav_is_p1 else p2_set
    dog_set = p2_set if fav_is_p1 else p1_set

    # v23.39.4: RankGap para reglas WTA del Analyzer.
    rank_gap = sim.get("rank_gap", None)
    try:
        rank_gap = abs(float(rank_gap))
    except Exception:
        try:
            rank_gap = abs(float(sim.get("rank1")) - float(sim.get("rank2")))
        except Exception:
            rank_gap = None
    wta_rankgap_ideal_over17 = (str(circuito).upper().strip() == "WTA" and rank_gap is not None and 21 <= rank_gap <= 50)

    set3 = _safe_float_v23370(sim.get("set3", sim.get("market_3sets", sim.get("prob_3sets", 0.0))), 0.0)

    # v23.37.2: selector orientado SOLO a máxima tasa de acierto.
    # Ajuste tras backtest manual:
    # - Prioridad real: Over 18.5 + favorito gana set + ML muy filtrado.
    # - Under/2-0 no debe competir como mercado principal salvo caso clarísimo.
    # - Over 19.5/20.5 solo compiten si salen muy fuertes; si no, preferimos Over 18.5.
    fav20 = _safe_float_v23370(sim.get("p1_2_0", 0.0) if fav_is_p1 else sim.get("p2_2_0", 0.0), 0.0)
    over18_p = _safe_float_v23370(over18, 0.0)
    over19_p = _safe_float_v23370(over19, 0.0)
    over20_p = _safe_float_v23370(over20, 0.0)
    under_ok_principal = (fav20 >= 0.72 and fav_prob >= 0.76 and set3 <= 0.28 and over18_p < 0.70)

    candidatos = [
        {
            "mercado": f"{fav_name} gana al menos 1 set",
            "prob": fav_set,
            "tipo": "SET_FAV",
            "motivo": "Mercado prudente: favorito con margen para ganar mínimo un set"
        },
        {
            "mercado": f"{fav_name} gana",
            "prob": fav_prob,
            "tipo": "ML",
            "motivo": "Favorito por probabilidad de partido"
        },
        {
            "mercado": f"{dog_name} gana al menos 1 set",
            "prob": dog_set,
            "tipo": "SET_DOG",
            "motivo": "Underdog con opciones reales de rascar un set"
        },
        {
            "mercado": "Over 18.5",
            "prob": _safe_float_v23370(over18, 0.0),
            "tipo": "OVER",
            "motivo": "Línea baja de juegos con probabilidad superior al ML alternativo"
        },
        {
            "mercado": "Over 19.5",
            "prob": _safe_float_v23370(over19, 0.0),
            "tipo": "OVER",
            "motivo": "Partido con proyección de juegos suficiente"
        },
        {
            "mercado": "Over 20.5",
            "prob": _safe_float_v23370(over20, 0.0),
            "tipo": "OVER",
            "motivo": "Partido con proyección larga"
        },
        {
            "mercado": "+2.5 sets",
            "prob": set3,
            "tipo": "SETS3",
            "motivo": "Probabilidad de partido a tres sets"
        },
        {
            "mercado": "Under 22.5",
            "prob": _safe_float_v23370(under22, 0.0),
            "tipo": "UNDER",
            "motivo": "Proyección de partido corto / control del favorito"
        },
    ]

    if str(circuito).upper().strip() == "WTA":
        candidatos.append({
            "mercado": "Over 17.5",
            "prob": _safe_float_v23370(over17, 0.0),
            "tipo": "OVER17_WTA",
            "motivo": "Línea WTA más conservadora"
        })

    # Limpieza básica + filtros v23.37.2: no dejar señales que el backtest mostró peligrosas
    # competir como mercado principal de acierto.
    limpios = []
    for c in candidatos:
        p = _safe_float_v23370(c.get("prob", 0.0), 0.0)
        mercado_u = str(c.get("mercado", "")).upper()
        tipo_u = str(c.get("tipo", "")).upper()

        # Under/2-0 solo como principal si el partido está MUY controlado.
        if tipo_u == "UNDER" and not under_ok_principal:
            continue

        # Over 19.5/20.5 no deben quitarle el sitio al Over 18.5 salvo señal claramente superior.
        if "OVER 19.5" in mercado_u and p < 0.76:
            continue
        if "OVER 20.5" in mercado_u and p < 0.80:
            continue

        # +2.5 sets solo si es una señal fuerte; si no, es mercado de cuota, no de máximo acierto.
        if tipo_u == "SETS3" and p < 0.62:
            continue

        # ML solo si es realmente limpio. Si no, el mercado prudente es gana al menos 1 set.
        if tipo_u == "ML" and p < 0.78:
            continue

        if 0.01 <= p <= 0.995:
            c = c.copy()
            c["prob"] = float(np.clip(p, 0.0, 0.995))
            limpios.append(c)

    if not limpios:
        return {
            "mercado": "⛔ EVITAR",
            "prob": 0.0,
            "confianza": "⚠️ Baja",
            "motivo": "Sin probabilidades utilizables",
            "tipo": "EVITAR",
        }

    # v23.39.4 WTA Analyzer Rules Upgrade:
    # 1) Favorita gana al menos 1 set >=90% es mercado WTA estable.
    # 2) Over 17.5 sube cuando el Analyzer encontró la zona fuerte: RankGap 21-50 + Over17 >=80%.
    if str(circuito).upper().strip() == "WTA":
        over17_p = _safe_float_v23370(over17, 0.0)
        if fav_set >= 0.90:
            return {
                "mercado": f"{fav_name} gana al menos 1 set",
                "prob": float(np.clip(fav_set, 0.0, 0.995)),
                "confianza": _confianza_acierto_v23370(fav_set),
                "motivo": "WTA Analyzer: favorito gana al menos 1 set >=90%, mercado más estable",
                "tipo": "SET_FAV",
            }
        if over17_p >= 0.80 and wta_rankgap_ideal_over17:
            return {
                "mercado": "Over 17.5",
                "prob": float(np.clip(over17_p, 0.0, 0.995)),
                "confianza": _confianza_acierto_v23370(over17_p),
                "motivo": "WTA Analyzer: Over 17.5 >=80% con RankGap 21-50",
                "tipo": "OVER17_WTA",
            }

    # v23.37.9 híbrida:
    # Recuperamos Over 18.5 como mercado principal cuando está realmente fuerte.
    # Motivo: en los backtests manuales Over 18.5 >= 76% fue mucho más estable
    # que muchos mercados de "gana al menos 1 set". No cambia el motor Over;
    # solo cambia la prioridad visual del selector de máximo acierto.
    over18_candidate = None
    for c in limpios:
        if str(c.get("mercado", "")).upper() == "OVER 18.5":
            over18_candidate = c.copy()
            break

    if over18_candidate is not None and over18_p >= 0.76:
        # Solo mantenemos el set como primera opción si es un set market de élite:
        # probabilidad muy alta y favorito suficientemente claro.
        set_elite = (fav_set >= 0.93 and fav_prob >= 0.75)
        if not set_elite:
            over18_candidate["prob"] = float(np.clip(over18_p, 0.0, 0.995))
            over18_candidate["motivo"] = "Over 18.5 fuerte priorizado por backtest: línea baja con soporte >= 76%"
            return {
                "mercado": over18_candidate["mercado"],
                "prob": float(over18_candidate["prob"]),
                "confianza": _confianza_acierto_v23370(over18_candidate["prob"]),
                "motivo": over18_candidate["motivo"],
                "tipo": over18_candidate.get("tipo", "OVER"),
            }

    # Orden por máxima probabilidad. En empate, priorizamos el mercado más prudente.
    prioridad_tipo = {"SET_FAV": 5, "ML": 4, "OVER17_WTA": 4, "OVER": 3, "UNDER": 3, "SET_DOG": 2, "SETS3": 1}
    limpios.sort(key=lambda c: (c["prob"], prioridad_tipo.get(c["tipo"], 0)), reverse=True)
    top = limpios[0]

    # Umbral de seguridad: si nada llega al 62%, mejor no venderlo como mercado probable.
    if top["prob"] < 0.62:
        return {
            "mercado": "⛔ EVITAR",
            "prob": top["prob"],
            "confianza": "⚠️ Baja",
            "motivo": f"Ningún mercado supera zona mínima de acierto; mejor candidato: {top['mercado']}",
            "tipo": "EVITAR",
        }

    # Motivo algo más explicativo para el caso clave que estabas buscando.
    if top["tipo"] == "SET_FAV":
        if fav_prob >= 0.78:
            top["motivo"] = "Favorito claro: el mercado más conservador es que gane al menos un set"
        elif fav_prob >= 0.64:
            top["motivo"] = "Favorito con riesgo ML: se baja a gana al menos un set para maximizar acierto"
        else:
            top["motivo"] = "Partido igualado: gana al menos un set es más prudente que elegir ganador"
    elif top["tipo"] == "ML":
        if str(circuito).upper().strip() == "WTA":
            if fav_prob >= 0.80:
                top["motivo"] = "WTA ML confirmado: favorita supera filtro fuerte de confianza y datos extra"
            else:
                top["motivo"] = "WTA ML permitido: favorita supera el mínimo de seguridad sin bajar a mercado de set"
        elif fav_prob >= 0.78:
            top["motivo"] = "Superioridad suficiente para recomendar ganador sin bajar a mercado de set"
        else:
            top["motivo"] = "ML permitido por filtro de seguridad del selector"
    elif top["tipo"] in ["OVER", "OVER17_WTA"]:
        if str(top.get("mercado", "")).upper() == "OVER 17.5":
            top["motivo"] = "WTA Over 17.5: mercado conservador validado como OBSERVAR/JUGAR según RankGap"
        elif str(top.get("mercado", "")).upper() == "OVER 18.5":
            top["motivo"] = "Over 18.5 priorizado: línea baja validada mejor en el backtest"
        else:
            top["motivo"] = "Total de juegos fuerte; pasa filtro restrictivo de máximo acierto"
    elif top["tipo"] == "UNDER":
        top["motivo"] = "Under solo permitido por control muy alto: favorito 2-0 fuerte, ML alto y bajo riesgo de 3 sets"

    return {
        "mercado": top["mercado"],
        "prob": float(top["prob"]),
        "confianza": _confianza_acierto_v23370(top["prob"]),
        "motivo": top["motivo"],
        "tipo": top["tipo"],
    }



# =========================================================
# v23.38 GRAND SLAM ATP 5 SET ENGINE
# =========================================================
# Rama nueva para el selector Grand Slam ya existente. No toca fórmulas ni guards
# del Over BO3 blindado; calcula mercados BO5 desde la simulación best_of=5.

def calcular_grand_slam_5set_markets(sim, p1_name="Jugador 1", p2_name="Jugador 2"):
    games = sim.get("games", []) or []
    n = max(1, len(games))

    def prob_over(line):
        return float(sum(float(x) > line for x in games) / n) if games else 0.0

    p1c = float(sim.get("p1_cal", 0.50))
    p2c = float(sim.get("p2_cal", 0.50))
    fav_is_p1 = p1c >= p2c
    fav_name = p1_name if fav_is_p1 else p2_name
    dog_name = p2_name if fav_is_p1 else p1_name

    fav_3_0 = float(sim.get("p1_3_0", 0.0) if fav_is_p1 else sim.get("p2_3_0", 0.0))
    dog_set = float(sim.get("p2_wins_set_any", 0.0) if fav_is_p1 else sim.get("p1_wins_set_any", 0.0))
    set4 = float(sim.get("set4_bo5", 0.0))
    set5 = float(sim.get("set5_bo5", 0.0))
    four_plus = float(np.clip(set4 + set5, 0.0, 1.0))
    fav_no_3_0 = float(np.clip(1.0 - fav_3_0, 0.0, 1.0))

    markets = {
        "Over 30.5 GS": prob_over(30.5),
        "Over 32.5 GS": prob_over(32.5),
        "Over 35.5 GS": prob_over(35.5),
        "Over 37.5 GS": prob_over(37.5),
        "Over 39.5 GS": prob_over(39.5),
        f"{dog_name} gana al menos 1 set": dog_set,
        f"{fav_name} NO gana 3-0": fav_no_3_0,
        "Partido a 4+ sets": four_plus,
        "Partido a 5 sets": set5,
        f"{fav_name} gana 3-0": fav_3_0,
    }

    # Selector prudente: prioriza mercados BO5 útiles, no líneas demasiado obvias.
    candidates = []
    rules = [
        ("Over 39.5 GS", 0.57, 5),
        ("Over 37.5 GS", 0.60, 4),
        ("Over 35.5 GS", 0.64, 3),
        ("Over 32.5 GS", 0.70, 2),
        (f"{dog_name} gana al menos 1 set", 0.62, 4),
        (f"{fav_name} NO gana 3-0", 0.63, 4),
        ("Partido a 4+ sets", 0.58, 3),
        ("Partido a 5 sets", 0.30, 2),
        (f"{fav_name} gana 3-0", 0.62, 2),
    ]
    for label, threshold, weight in rules:
        pr = float(markets.get(label, 0.0))
        if pr >= threshold:
            candidates.append((pr + weight * 0.012, pr, label, threshold))

    if candidates:
        candidates.sort(reverse=True)
        _, best_prob, best_label, best_threshold = candidates[0]
        if best_prob >= max(0.70, best_threshold + 0.06):
            action = "✅ JUGAR"
            confidence = "🔥 Alta GS"
        elif best_prob >= best_threshold:
            action = "👀 OBSERVAR"
            confidence = "✅ Apta GS"
        else:
            action = "⚠️ EVITAR"
            confidence = "⚠️ Baja GS"
    else:
        best_label = "Sin mercado GS claro"
        best_prob = 0.0
        action = "⚠️ EVITAR"
        confidence = "⚠️ Sin señal GS"

    return {
        "markets": markets,
        "best_label": best_label,
        "best_prob": float(best_prob),
        "action": action,
        "confidence": confidence,
        "fav_name": fav_name,
        "dog_name": dog_name,
        "avg_games": float(np.mean(games)) if games else 0.0,
        "median_games": float(np.median(games)) if games else 0.0,
        "set4": set4,
        "set5": set5,
        "four_plus": four_plus,
        "fav_3_0": fav_3_0,
        "fav_no_3_0": fav_no_3_0,
        "dog_set": dog_set,
        "note": "Motor Grand Slam BO5 activo: usa líneas 30.5-39.5, 4+ sets, 5 sets y 3-0/no 3-0. Over BO3 queda intacto."
    }


def _fmt_pct_gs(markets, label):
    try:
        return f"{float(markets.get(label, 0.0)):.1%}"
    except Exception:
        return ""

def analyze_batch_matches(parsed_matches, db, circuito, surface, best_of, sims, progress_callback=None):
    rows = []
    total = len(parsed_matches)
    for idx, m in enumerate(parsed_matches, start=1):
        if progress_callback:
            progress_callback(idx-1, total, f"Emparejando {m.get('p1_raw')} vs {m.get('p2_raw')}")

        match_surface = m.get("surface", surface) if m.get("surface", surface) in ["Hard", "Clay", "Grass"] else surface
        lookup_circuit = circuito_lookup_para_match(m, circuito)
        circuito_calc = circuito_sim_para_lookup(lookup_circuit, circuito)

        p1_key, d1, p1_score, p1_status = resolver_player_batch(m["p1_raw"], db, lookup_circuit, match_surface, allow_fallback=True)
        p2_key, d2, p2_score, p2_status = resolver_player_batch(m["p2_raw"], db, lookup_circuit, match_surface, allow_fallback=True)

        # Solo queda como No encontrado si es rival pendiente/país/dobles o imposible de analizar.
        if d1 is None or d2 is None:
            rows.append({
                "Fecha": m.get("date", ""),
                "Hora": m.get("time", ""),
                "Partido": f"{m['p1_raw']} vs {m['p2_raw']}",
                "Estado": "No encontrado",
                "Jugador 1 encontrado": p1_key or "N/A",
                "Score J1": f"{p1_score:.0%}",
                "Jugador 2 encontrado": p2_key or "N/A",
                "Score J2": f"{p2_score:.0%}",
            })
            continue
        sim = sim_match(d1, d2, match_surface, circuito_calc, best_of, sims, context_row={})
        games = sim.get("games", [])
        avg_games = float(np.mean(games)) if len(games) else 0.0
        games_p1 = sim.get("games_p1", [])
        games_p2 = sim.get("games_p2", [])
        avg_g1 = float(np.mean(games_p1)) if len(games_p1) else 0.0
        avg_g2 = float(np.mean(games_p2)) if len(games_p2) else 0.0

        # v23.48.0 Player Games Markets: no toca motor Over;
        # solo lee la distribución de juegos ya simulada para cada jugador.
        p1_games_over_10_5 = float(sum(float(x) > 10.5 for x in games_p1) / max(1, len(games_p1))) if len(games_p1) else 0.0
        p1_games_over_11_5 = float(sum(float(x) > 11.5 for x in games_p1) / max(1, len(games_p1))) if len(games_p1) else 0.0
        p2_games_over_10_5 = float(sum(float(x) > 10.5 for x in games_p2) / max(1, len(games_p2))) if len(games_p2) else 0.0
        p2_games_over_11_5 = float(sum(float(x) > 11.5 for x in games_p2) / max(1, len(games_p2))) if len(games_p2) else 0.0
        player_game_candidates = [
            (f"{p1_key} +10.5 juegos", p1_games_over_10_5, p1_key, "+10.5"),
            (f"{p1_key} +11.5 juegos", p1_games_over_11_5, p1_key, "+11.5"),
            (f"{p2_key} +10.5 juegos", p2_games_over_10_5, p2_key, "+10.5"),
            (f"{p2_key} +11.5 juegos", p2_games_over_11_5, p2_key, "+11.5"),
        ]
        player_games_best_label, player_games_best_prob, player_games_best_player, player_games_best_line = max(player_game_candidates, key=lambda x: x[1])
        if player_games_best_prob >= 0.86:
            player_games_tier = "🔥 Muy alta"
        elif player_games_best_prob >= 0.80:
            player_games_tier = "✅ Alta"
        elif player_games_best_prob >= 0.74:
            player_games_tier = "👀 Observar"
        else:
            player_games_tier = "⚠️ Baja"

        over17_raw = sum(x > 17.5 for x in games)/sims if sims else 0
        over18_raw = sum(x > 18.5 for x in games)/sims if sims else 0
        over19_raw = sum(x > 19.5 for x in games)/sims if sims else 0
        over20_raw = sum(x > 20.5 for x in games)/sims if sims else 0
        over22_raw = sum(x > 22.5 for x in games)/sims if sims else 0
        caps = aplicar_market_sanity_caps(sim, circuito_calc, match_surface, over18_raw, over19_raw, over20_raw, over22_raw)
        over18, over19, over20, over22 = caps["over18"], caps["over19"], caps["over20"], caps["over22"]
        under22 = 1 - over22
        # Over 17.5 solo se usa como mercado operativo WTA. En ATP/Challenger queda calculado pero no se muestra como señal.
        over17 = over17_raw if circuito_calc == "WTA" else 0.0
        sim["market_over17"] = over17
        sim["market_over18"] = over18
        sim["market_over19"] = over19
        sim["market_over20"] = over20
        sim["market_over22"] = over22
        sim["market_cap_notes"] = caps.get("notes", [])

        p1c, p2c = sim["p1_cal"], sim["p2_cal"]
        fav_name = p1_key if p1c >= p2c else p2_key
        fav_prob = max(p1c, p2c)
        dog_name = p2_key if p1c >= p2c else p1_key

        best_label, best_prob = batch_pick_label(sim, over17, over18, over19, over20, over22, under22, p1_key, p2_key, circuito_calc, match_surface)

        # v23.37: selector independiente de máxima probabilidad de acierto.
        # Es una capa visible nueva: no usa cuotas y no toca el motor Over.
        max_acierto = selector_mercado_maximo_acierto_v23370(
            sim, over17, over18, over19, over20, over22, under22,
            p1_key, p2_key, circuito_calc, match_surface
        )

        gs5 = calcular_grand_slam_5set_markets(sim, p1_key, p2_key) if (best_of == 5 and str(circuito_calc).upper() in {"ATP", "CHALLENGER"}) else None
        if gs5:
            # En BO5 no dejamos que el selector oficial muestre Over 18.5/19.5 como mercado principal.
            max_acierto = {
                "mercado": gs5.get("best_label", ""),
                "prob": gs5.get("best_prob", 0.0),
                "confianza": gs5.get("confidence", ""),
                "motivo": gs5.get("note", ""),
            }
            best_label, best_prob = gs5.get("best_label", best_label), gs5.get("best_prob", best_prob)

        wta_watchlist = wta_over_watchlist_reason(circuito_calc, match_surface, fav_prob, over18, over17)

        filters = betting_filter_engine(circuito_calc, match_surface, sim, p1_key, p2_key)
        trust = filters.get("status", "")
        rationale = " · ".join(filters.get("reasons", [])[:2]) if isinstance(filters, dict) else ""
        if p1_status != "OK" or p2_status != "OK":
            rationale = (rationale + " · " if rationale else "") + "fallback estimado: datos incompletos"
            trust = "⚠️ DATOS PARCIALES"
            # v23.37.6: NO contaminamos el nombre del mercado con "OBSERVAR".
            # La columna 🎯 Acción final decide si es JUGAR/OBSERVAR/EVITAR.
            if isinstance(max_acierto, dict) and max_acierto.get("mercado"):
                max_acierto = max_acierto.copy()
                max_acierto["confianza"] = "⚠️ Datos parciales"
                max_acierto["motivo"] = str(max_acierto.get("motivo", "")) + " · fallback estimado: no tratar como pick fuerte"

        # Cuota pegada: puede ser de un jugador concreto.
        quoted_player = None
        quoted_odd = m.get("quoted_odd")
        quoted_prob_model = None
        edge = None
        fair_odd = None

        if m.get("odd1") is not None:
            # si hay dos cuotas limpias, usar cuota del favorito modelo
            quoted_player = fav_name
            quoted_odd = m["odd1"] if fav_name == p1_key else m["odd2"]
            quoted_prob_model = fav_prob
        elif quoted_odd is not None and m.get("quoted_side") in [1, 2]:
            quoted_player = p1_key if m["quoted_side"] == 1 else p2_key
            quoted_prob_model = p1c if m["quoted_side"] == 1 else p2c

        if quoted_odd and quoted_prob_model:
            fair_odd = 1 / max(0.01, quoted_prob_model)
            edge = quoted_odd * quoted_prob_model - 1

        # v23.38.2: métricas reales Grand Slam BO5 para validación/export.
        fav_side_model = 1 if fav_name == p1_key else 2 if fav_name == p2_key else None
        gs_reales = calcular_gs_bo5_reales(m, fav_side_model) if gs5 else {}
        gs_market_label = gs5.get("best_label", "") if gs5 else ""
        gs_acierto = acierta_mercado_gs_bo5(gs_market_label, gs_reales) if gs5 else ""

        rows.append({
            "Versión app": APP_VERSION,
            "Fecha": m.get("date", ""),
            "Hora": m.get("time", ""),
            "Circuito fuente": m.get("circuito_detectado", ""),
            "Circuito datos": str(lookup_circuit).upper(),
            "Circuito cálculo": str(circuito_calc).upper(),
            "Torneo": m.get("torneo", ""),
            "Superficie": match_surface,
            "Partido": f"{p1_key} vs {p2_key}",
            "Estado": "OK" if p1_status == "OK" and p2_status == "OK" else "OK con jugador estimado",
            "Lectura J1": p1_status,
            "Lectura J2": p2_status,
            "Aviso datos": (
                "⚠️ Fallback estimado: usar solo como lectura orientativa / NO pick fuerte"
                if p1_status != "OK" or p2_status != "OK" else ""
            ),
            "Favorito modelo": fav_name,
            "ML favorito": f"{fav_prob:.1%}",
            "🎯 Mercado más probable": max_acierto.get("mercado", ""),
            "🎯 Prob máxima": f"{max_acierto.get('prob', 0.0):.1%}" if isinstance(max_acierto, dict) else "",
            "🎯 Confianza acierto": max_acierto.get("confianza", "") if isinstance(max_acierto, dict) else "",
            "🎯 Motivo acierto": max_acierto.get("motivo", "") if isinstance(max_acierto, dict) else "",
            "Mejor señal": best_label,
            "Prob señal": f"{best_prob:.1%}",
            "Mejor mercado Over Focus": best_label if any(x in str(best_label) for x in ["Over", "3 sets", "Partido"]) else "",
            "Over Focus Label": over_focus_label(lookup_circuit, best_label, best_prob, sim.get("set3", 0.0), over17, over18, over19),
            "WTA Watchlist": wta_watchlist,
            "Mejor mercado WTA": (best_label if circuito_calc == "WTA" else ""),
            "WTA Over17 Priority": ("Sí" if circuito_calc == "WTA" and best_label == "Over 17.5" and over17 >= 0.77 else ""),
            "Signal Trust": trust,
            "Over Quality Guard": filters.get("over_guard_label", ""),
            "Motivos Over Guard": " · ".join(filters.get("over_quality_reasons", [])[:4]) if isinstance(filters, dict) else "",
            "Under 2.5 Rescue": filters.get("under25_label", ""),
            "Under 2.5 ajustado": f"{filters.get('under25_adjusted', 0.0):.1%}" if isinstance(filters, dict) else "",
            "Confianza mínima": f"{filters.get('min_confidence', 1.0):.0%}" if isinstance(filters, dict) else "",
            "Mín. partidos superficie": filters.get("min_surface_matches", "") if isinstance(filters, dict) else "",
            "Gap Elo/modelo": f"{filters.get('elo_pure_gap', 0.0):.1%}" if isinstance(filters, dict) else "",
            "Over 17.5": f"{over17:.1%}" if circuito_calc == "WTA" else "",
            "Over 18.5": f"{over18:.1%}",
            "Over 19.5": f"{over19:.1%}",
            "Under 22.5": f"{under22:.1%}",
            "🏆 GS 5 sets activo": "Sí" if gs5 else "",
            "🏆 Mercado GS 5 sets": gs5.get("best_label", "") if gs5 else "",
            "🏆 Prob GS 5 sets": f"{gs5.get('best_prob', 0.0):.1%}" if gs5 else "",
            "🏆 Acción GS": gs5.get("action", "") if gs5 else "",
            "🏆 Confianza GS": gs5.get("confidence", "") if gs5 else "",
            "GS Over 30.5": _fmt_pct_gs(gs5.get("markets", {}), "Over 30.5 GS") if gs5 else "",
            "GS Over 32.5": _fmt_pct_gs(gs5.get("markets", {}), "Over 32.5 GS") if gs5 else "",
            "GS Over 35.5": _fmt_pct_gs(gs5.get("markets", {}), "Over 35.5 GS") if gs5 else "",
            "GS Over 37.5": _fmt_pct_gs(gs5.get("markets", {}), "Over 37.5 GS") if gs5 else "",
            "GS Over 39.5": _fmt_pct_gs(gs5.get("markets", {}), "Over 39.5 GS") if gs5 else "",
            "GS Partido 4+ sets": f"{gs5.get('four_plus', 0.0):.1%}" if gs5 else "",
            "GS Partido 5 sets": f"{gs5.get('set5', 0.0):.1%}" if gs5 else "",
            "GS Favorito 3-0": f"{gs5.get('fav_3_0', 0.0):.1%}" if gs5 else "",
            "GS Favorito NO 3-0": f"{gs5.get('fav_no_3_0', 0.0):.1%}" if gs5 else "",
            "GS Underdog gana set": f"{gs5.get('dog_set', 0.0):.1%}" if gs5 else "",
            "Jugador gana set": dog_name,
            "Prob gana set": f"{sim.get('dog_wins_set', 0):.1%}",
            f"{p1_key} gana al menos 1 set": f"{sim.get('p1_wins_set_any', 0):.1%}",
            f"{p2_key} gana al menos 1 set": f"{sim.get('p2_wins_set_any', 0):.1%}",
            "Favorito gana al menos 1 set": f"{(sim.get('p1_wins_set_any', 0) if p1c >= p2c else sim.get('p2_wins_set_any', 0)):.1%}",
            "Partido a 3 sets": f"{sim.get('set3', 0):.1%}",
            "Favorito 2-0": f"{sim.get('fav_2_0', 0):.1%}",
            "Juegos J1": f"{avg_g1:.1f}",
            "Juegos J2": f"{avg_g2:.1f}",
            "Total games": f"{avg_games:.1f}",
            "J1 +10.5 juegos": f"{p1_games_over_10_5:.1%}",
            "J1 +11.5 juegos": f"{p1_games_over_11_5:.1%}",
            "J2 +10.5 juegos": f"{p2_games_over_10_5:.1%}",
            "J2 +11.5 juegos": f"{p2_games_over_11_5:.1%}",
            "Mejor juegos jugador": player_games_best_label,
            "Prob juegos jugador": f"{player_games_best_prob:.1%}",
            "Confianza juegos jugador": player_games_tier,
            "Jugador juegos premium": player_games_best_player,
            "Línea juegos premium": player_games_best_line,
            "Cuota pegada": quoted_odd if quoted_odd else "",
            "Jugador cuota": quoted_player or "",
            "Cuota justa": f"{fair_odd:.2f}" if fair_odd else "",
            "Edge": f"{edge:.1%}" if edge is not None else "",
            "Ganador real": (
                p1_key if m.get("actual_winner_side") == 1 else
                p2_key if m.get("actual_winner_side") == 2 else ""
            ),
            "Resultado sets": (
                f"{m.get('p1_sets_real')}-{m.get('p2_sets_real')}"
                if m.get("p1_sets_real") is not None and m.get("p2_sets_real") is not None else ""
            ),
            "Marcador games": m.get("score_games", ""),
            "Total games real": m.get("actual_total_games", ""),
            "GS Juegos reales": gs_reales.get("gs_total_games_real", "") if gs5 else "",
            "GS Sets reales totales": gs_reales.get("gs_sets_total_real", "") if gs5 else "",
            "GS Over 30.5 real": _gs_bool_to_si_no(gs_reales.get("gs_over_30_5_real_bool"), gs_reales.get("available_games", False)) if gs5 else "",
            "GS Over 32.5 real": _gs_bool_to_si_no(gs_reales.get("gs_over_32_5_real_bool"), gs_reales.get("available_games", False)) if gs5 else "",
            "GS Over 35.5 real": _gs_bool_to_si_no(gs_reales.get("gs_over_35_5_real_bool"), gs_reales.get("available_games", False)) if gs5 else "",
            "GS Over 37.5 real": _gs_bool_to_si_no(gs_reales.get("gs_over_37_5_real_bool"), gs_reales.get("available_games", False)) if gs5 else "",
            "GS Over 39.5 real": _gs_bool_to_si_no(gs_reales.get("gs_over_39_5_real_bool"), gs_reales.get("available_games", False)) if gs5 else "",
            "GS Partido 4+ sets real": _gs_bool_to_si_no(gs_reales.get("gs_partido_4_plus_real_bool"), gs_reales.get("available_sets", False)) if gs5 else "",
            "GS Partido 5 sets real": _gs_bool_to_si_no(gs_reales.get("gs_partido_5_sets_real_bool"), gs_reales.get("available_sets", False)) if gs5 else "",
            "GS Favorito 3-0 real": _gs_bool_to_si_no(gs_reales.get("gs_fav_3_0_real_bool"), gs_reales.get("available_sets", False)) if gs5 else "",
            "GS Favorito NO 3-0 real": _gs_bool_to_si_no(gs_reales.get("gs_fav_no_3_0_real_bool"), gs_reales.get("available_sets", False)) if gs5 else "",
            "GS Underdog gana set real": _gs_bool_to_si_no(gs_reales.get("gs_dog_set_real_bool"), gs_reales.get("available_sets", False)) if gs5 else "",
            "Acierta mercado GS": gs_acierto if gs5 else "",
            "Acierta ML modelo": (
                "Sí" if (
                    (m.get("actual_winner_side") == 1 and fav_name == p1_key) or
                    (m.get("actual_winner_side") == 2 and fav_name == p2_key)
                ) else "No" if m.get("actual_winner_side") in [1,2] else ""
            ),
            "Over 17.5 real": (
                "Sí" if circuito_calc == "WTA" and m.get("actual_total_games", None) is not None and float(m.get("actual_total_games")) > 17.5
                else "No" if circuito_calc == "WTA" and m.get("actual_total_games", None) is not None
                else "N/D" if circuito_calc == "WTA" else ""
            ),
            "Acierta Over 17.5": (
                "Sí" if circuito_calc == "WTA" and m.get("actual_total_games", None) is not None and ((over17 >= 0.50) == (float(m.get("actual_total_games")) > 17.5))
                else "No" if circuito_calc == "WTA" and m.get("actual_total_games", None) is not None and ((over17 >= 0.50) != (float(m.get("actual_total_games")) > 17.5))
                else "N/D" if circuito_calc == "WTA" else ""
            ),
            "Over 18.5 real": (
                "Sí" if m.get("actual_total_games", None) is not None and float(m.get("actual_total_games")) > 18.5
                else "No" if m.get("actual_total_games", None) is not None
                else "N/D"
            ),
            "3 sets real": (
                "Sí" if m.get("p1_sets_real") is not None and m.get("p2_sets_real") is not None and int(m.get("p1_sets_real")) + int(m.get("p2_sets_real")) == 3
                else "No" if m.get("p1_sets_real") is not None and m.get("p2_sets_real") is not None
                else "N/D"
            ),
            "Acierta Over 18.5": (
                "Sí" if m.get("actual_total_games", None) is not None and ((over18 >= 0.50) == (float(m.get("actual_total_games")) > 18.5))
                else "No" if m.get("actual_total_games", None) is not None and ((over18 >= 0.50) != (float(m.get("actual_total_games")) > 18.5))
                else "N/D"
            ),
            "Acierta 3 sets": (
                "Sí" if m.get("p1_sets_real") is not None and m.get("p2_sets_real") is not None and ((sim.get("set3", 0) >= 0.50) == (int(m.get("p1_sets_real")) + int(m.get("p2_sets_real")) == 3))
                else "No" if m.get("p1_sets_real") is not None and m.get("p2_sets_real") is not None and ((sim.get("set3", 0) >= 0.50) != (int(m.get("p1_sets_real")) + int(m.get("p2_sets_real")) == 3))
                else "N/D"
            ),
            "Resultado válido": (
                "Sí" if (
                    (best_of == 5 and m.get("p1_sets_real") in [0,1,2,3] and m.get("p2_sets_real") in [0,1,2,3] and m.get("p1_sets_real") != m.get("p2_sets_real") and max(m.get("p1_sets_real"), m.get("p2_sets_real")) == 3) or
                    (best_of != 5 and m.get("p1_sets_real") in [0,1,2] and m.get("p2_sets_real") in [0,1,2] and m.get("p1_sets_real") != m.get("p2_sets_real"))
                ) else "Revisar"
            ),
            "Riesgos": rationale,
            "Match J1": f"{p1_score:.0%}",
            "Match J2": f"{p2_score:.0%}",
        })

        # v23.10: liberar arrays de simulación en lotes largos.
        try:
            del sim, games, games_p1, games_p2
        except Exception:
            pass
        if idx % 3 == 0:
            gc.collect()

        if progress_callback:
            progress_callback(idx, total, f"Analizado {p1_key} vs {p2_key}")

    gc.collect()
    return pd.DataFrame(rows)


def _pct_to_float(x):
    try:
        s = str(x).replace("%", "").replace(",", ".").strip()
        if not s:
            return None
        return float(s) / 100.0
    except Exception:
        return None


# =========================================================
# v23.26.8 COMBI SAFE ENGINE
# Constructor de combinadas para evitar el "fallo por uno"
# =========================================================

def _combi_float(x, default=None):
    try:
        if x is None:
            return default
        s = str(x).replace("%", "").replace(",", ".").strip()
        if s == "" or s.lower() in ["nan", "none", "n/d"]:
            return default
        return float(s)
    except Exception:
        return default


def _combi_pct(x, default=0.0):
    v = _combi_float(x, None)
    if v is None:
        return default
    if v > 1:
        v = v / 100.0
    return float(np.clip(v, 0.0, 1.0))


def _combi_tipo_mercado(market):
    m = str(market or "").upper().replace(",", ".")
    if "NO BET" in m:
        return "no_bet"
    if "OVER 17.5" in m:
        return "over17"
    if "OVER 18.5" in m:
        return "over18"
    if "OVER 19.5" in m:
        return "over19"
    if "OVER 20.5" in m:
        return "over20"
    if "UNDER 22.5" in m:
        return "under22"
    if "+2.5" in m or "+3.5" in m or "MÁS DE 3.5" in m or "MAS DE 3.5" in m or "MÁS DE 3,5" in m or "MAS DE 3,5" in m or "NÚMERO DE SETS" in m or "NUMERO DE SETS" in m or "3 SET" in m or "TRES SET" in m:
        return "sets3"
    if "2-0" in m or "UNDER 2.5 SET" in m:
        return "fav20"
    if "GANA AL MENOS 1 SET" in m or "GANA SET" in m:
        return "dog_set"
    if "ML" in m or "GANADOR" in m or "FAVORITO" in m:
        return "ml"
    return "otro"


COMBI_SAFE_PROFILES_V23268 = {
    # v23.40: perfiles mucho más estrictos tras revisar combinadas falladas por una selección.
    # La idea es NO convertir picks buenos en combinadas largas: primero rentabilidad, luego cuota.
    "🔒 Conservador": {
        "ml": 0.82, "over17": 0.84, "over18": 0.83, "over19": 0.86, "over20": 0.90,
        "under22": 0.76, "sets3": 0.66, "fav20": 0.76, "dog_set": 0.93, "otro": 0.82,
        "max_picks": 2,
    },
    "⚖️ Normal": {
        "ml": 0.79, "over17": 0.81, "over18": 0.81, "over19": 0.84, "over20": 0.88,
        "under22": 0.74, "sets3": 0.62, "fav20": 0.73, "dog_set": 0.90, "otro": 0.79,
        "max_picks": 2,
    },
    "🔥 Agresivo": {
        "ml": 0.76, "over17": 0.79, "over18": 0.79, "over19": 0.82, "over20": 0.86,
        "under22": 0.72, "sets3": 0.58, "fav20": 0.70, "dog_set": 0.88, "otro": 0.76,
        "max_picks": 3,
    },
}


def _combi_prob_from_row(row, market):
    """Usa Prob mercado recomendado y, si falta, busca la columna específica."""
    p = _combi_pct(row.get("Prob mercado recomendado", ""), None)
    if p is not None and p > 0:
        return p

    tipo = _combi_tipo_mercado(market)
    col_map = {
        "ml": "ML favorito",
        "over17": "Over 17.5",
        "over18": "Over 18.5",
        "over19": "Over 19.5",
        "over20": "Over 20.5",
        "under22": "Under 22.5",
        "sets3": "Partido a 3 sets",
        "fav20": "Favorito 2-0",
        "dog_set": "Prob gana set",
    }
    col = col_map.get(tipo)
    if col and col in row.index:
        return _combi_pct(row.get(col, ""), 0.0)
    return _combi_pct(row.get("Prob señal", ""), 0.0)


def _combi_cuota_from_row(row, prob):
    """Prioriza cuota pegada. Si no existe, usa cuota justa/estimada para poder ordenar."""
    q = _combi_float(row.get("Cuota pegada", ""), None)
    if q is not None and q > 1.0:
        return float(q), "pegada"

    q = _combi_float(row.get("Cuota justa", ""), None)
    if q is not None and q > 1.0:
        return float(q), "justa ML"

    if prob and prob > 0:
        # Estimación conservadora de cuota posible. Solo orienta si no se han pegado cuotas.
        return float(np.clip((1.0 / prob) * 0.94, 1.01, 3.50)), "estimada"
    return 1.01, "estimada"


def clasificar_combi_safe_row_v23268(row, profile_name="⚖️ Normal"):
    profile = COMBI_SAFE_PROFILES_V23268.get(profile_name, COMBI_SAFE_PROFILES_V23268["⚖️ Normal"])

    market = str(row.get("Mercado recomendado", "") or "").strip()
    rec = str(row.get("Recomendación", "") or "").strip()
    trust = str(row.get("Signal Trust", "") or "").strip()
    risks = str(row.get("Riesgos", "") or "").strip()
    estado = str(row.get("Estado", "") or "").strip()
    circuito = " ".join([
        str(row.get("Circuito fuente", "") or ""),
        str(row.get("Circuito datos", "") or ""),
        str(row.get("Circuito cálculo", "") or ""),
        str(row.get("Torneo", "") or ""),
    ]).upper()
    surface = str(row.get("Superficie", "") or "")
    partido = str(row.get("Partido", "") or "").strip()

    tipo = _combi_tipo_mercado(market)
    prob = _combi_prob_from_row(row, market)
    cuota, cuota_tipo = _combi_cuota_from_row(row, prob)

    reasons = []
    min_prob = float(profile.get(tipo, profile.get("otro", 0.75)))

    blocked_words = ["NO BET", "WATCH", "OBSERVAR", "SOLO CONTEXTO", "NO COMBI"]
    all_signal = f"{market} {rec} {trust}".upper()
    hard_block = any(w in all_signal for w in blocked_words)
    # v23.29: si el Over Guard bloqueó, no entra en combinada salvo que el mercado recomendado sea Under 2.5.
    if _row_over_guard_active(row) and "UNDER 2.5" not in market.upper():
        hard_block = True

    if tipo == "no_bet" or hard_block:
        return {
            "Partido": partido, "Mercado": market, "Prob": prob, "Cuota": cuota,
            "Cuota tipo": cuota_tipo, "Tipo": tipo, "Min": min_prob,
            "Etiqueta": "❌ NO COMBI", "Combi Safe": False, "Score": 0.0,
            "Motivos": "recomendación/watch/no bet: no entra en combinada",
        }

    # v23.40 Combi Guard: reglas duras vistas en tickets reales.
    # Cuotas muy bajas aportan poco a la combinada y siguen pudiendo romperla.
    if cuota_tipo == "pegada" and cuota < 1.25 and tipo in ["ml", "dog_set", "fav20"]:
        return {
            "Partido": partido, "Mercado": market, "Prob": prob, "Cuota": cuota,
            "Cuota tipo": cuota_tipo, "Tipo": tipo, "Min": min_prob,
            "Etiqueta": "🔥 FUERTE SIMPLE", "Combi Safe": False, "Score": float(prob * 100.0 - 8),
            "Motivos": "cuota <1.25: aporta poco y puede romper la combinada; mejor simple o fuera",
        }

    # El mercado gana al menos 1 set solo entra en combi si es realmente premium.
    if tipo == "dog_set" and prob < max(min_prob, 0.90):
        return {
            "Partido": partido, "Mercado": market, "Prob": prob, "Cuota": cuota,
            "Cuota tipo": cuota_tipo, "Tipo": tipo, "Min": max(min_prob, 0.90),
            "Etiqueta": "🔥 FUERTE SIMPLE" if prob >= 0.84 else "❌ NO COMBI", "Combi Safe": False,
            "Score": float(prob * 100.0 - 6),
            "Motivos": "gana al menos 1 set no alcanza umbral premium para combinar",
        }

    # Over 19.5/20.5 suelen fallar por paliza: se aceptan solo como simples salvo probabilidad muy alta.
    if tipo in ["over19", "over20"] and prob < min_prob:
        return {
            "Partido": partido, "Mercado": market, "Prob": prob, "Cuota": cuota,
            "Cuota tipo": cuota_tipo, "Tipo": tipo, "Min": min_prob,
            "Etiqueta": "🔥 FUERTE SIMPLE" if prob >= min_prob - 0.04 else "❌ NO COMBI", "Combi Safe": False,
            "Score": float(prob * 100.0 - 5),
            "Motivos": "Over alto: riesgo de paliza; no entra en combinada si no supera umbral premium",
        }

    if "CHALLENGER" in circuito or "CHALL" in circuito:
        min_prob += 0.04
        reasons.append("Challenger: sube exigencia")

    if "QUAL" in circuito or "QUALY" in circuito or "PREVIA" in circuito:
        min_prob += 0.03
        reasons.append("Qualy/previa: sube exigencia")

    if "WTA" in circuito and tipo == "ml":
        min_prob += 0.03
        reasons.append("WTA ML: sube exigencia")

    if "OK CON JUGADOR ESTIMADO" in estado.upper() or "FALLBACK" in f"{risks} {row.get('Aviso datos','')}".upper():
        min_prob += 0.05
        reasons.append("Datos parciales/fallback: no forzar combinada")

    if any(x in f"{risks} {trust}".upper() for x in ["UPSET", "VULNERABLE", "RIESGO", "PARCIALES"]):
        min_prob += 0.02
        reasons.append("Riesgo detectado por el analista")

    if tipo in ["sets3", "dog_set"]:
        reasons.append("Mercado de más varianza: mejor simple salvo probabilidad muy alta")

    if tipo in ["over17", "over18"]:
        reasons.append("Mercado prioritario para combinadas si supera umbral")

    # Score interno para ordenar combinadas.
    score = prob * 100.0
    if prob >= 0.85:
        score += 10
    elif prob >= 0.80:
        score += 7
    elif prob >= 0.75:
        score += 4
    else:
        score -= 8

    if tipo in ["over17", "over18"]:
        score += 5
    elif tipo == "over19":
        score += 3
    elif tipo in ["sets3", "dog_set"]:
        score -= 5
    elif tipo == "ml":
        score += 0
    elif tipo in ["fav20", "under22"]:
        score -= 1
    else:
        score -= 2

    if "CHALLENGER" in circuito or "CHALL" in circuito:
        score -= 6
    if "QUAL" in circuito or "QUALY" in circuito:
        score -= 5
    if "WTA" in circuito and tipo == "ml":
        score -= 5
    if "FALLBACK" in f"{risks} {row.get('Aviso datos','')}".upper():
        score -= 8

    if prob >= min_prob:
        etiqueta = "🧱 COMBI SAFE"
        safe = True
    elif prob >= min_prob - 0.04:
        etiqueta = "🔥 FUERTE SIMPLE"
        safe = False
        reasons.append("bueno para simple, corto para combinada")
    else:
        etiqueta = "❌ NO COMBI"
        safe = False
        reasons.append("no supera el mínimo de combinada")

    return {
        "Partido": partido,
        "Mercado": market,
        "Prob": float(prob),
        "Cuota": float(cuota),
        "Cuota tipo": cuota_tipo,
        "Tipo": tipo,
        "Min": float(min_prob),
        "Etiqueta": etiqueta,
        "Combi Safe": bool(safe),
        "Score": float(score),
        "Motivos": " · ".join(reasons) if reasons else "sin alerta adicional",
        # v23.41.1: conservar probabilidades alternativas para comparar cuotas reales de otras líneas.
        "Prob Over 17.5": _row_pct(row, "Over 17.5", 0.0),
        "Prob Over 18.5": _row_pct(row, "Over 18.5", 0.0),
        "Prob Over 19.5": _row_pct(row, "Over 19.5", 0.0),
        "Prob Over 20.5": _row_pct(row, "Over 20.5", 0.0),
        "Prob Over 22.5": _row_pct(row, "Over 22.5", 0.0),
        "Prob +2.5 sets": _row_pct(row, "Partido a 3 sets", 0.0),
        "Prob gana set alt": _row_pct(row, "Prob gana set", 0.0),
    }


def construir_combinadas_v23268(ok_df, profile_name="⚖️ Normal", cuota_min=1.60, cuota_max=1.80, min_picks=2, max_picks=3):
    if ok_df is None or ok_df.empty:
        return pd.DataFrame(), []

    picks = []
    for _, row in ok_df.iterrows():
        p = clasificar_combi_safe_row_v23268(row, profile_name=profile_name)
        if p.get("Partido") and p.get("Mercado"):
            picks.append(p)

    safe_picks = [p for p in picks if p.get("Combi Safe")]
    combos = []

    for n in range(int(min_picks), int(max_picks) + 1):
        for combo in combinations(safe_picks, n):
            partidos = [p["Partido"] for p in combo]
            if len(set(partidos)) != len(partidos):
                continue

            tipos_combo = [str(p.get("Tipo", "")) for p in combo]
            # v23.40: evitar el fallo por uno acumulando mercados frágiles.
            if sum(t in ["sets3", "dog_set"] for t in tipos_combo) > 1:
                continue
            if sum(t in ["over19", "over20"] for t in tipos_combo) > 1:
                continue
            if any(float(p.get("Cuota", 1.0) or 1.0) < 1.25 for p in combo):
                continue

            cuota_total = 1.0
            confianza = 1.0
            score_medio = 0.0
            for p in combo:
                cuota_total *= float(p["Cuota"])
                confianza *= float(p["Prob"])
                score_medio += float(p["Score"])
            score_medio /= max(1, len(combo))

            min_confianza_combo = 0.56 if n == 2 else 0.46
            if float(cuota_min) <= cuota_total <= float(cuota_max) and confianza >= min_confianza_combo:
                weak = min(combo, key=lambda x: x["Prob"])
                combos.append({
                    "Nº picks": n,
                    "Cuota total": cuota_total,
                    "Confianza global": confianza,
                    "Score medio": score_medio,
                    "Pick más débil": f"{weak['Mercado']} — {weak['Partido']} ({weak['Prob']:.1%})",
                    "Picks": list(combo),
                })

    combos = sorted(combos, key=lambda x: (x["Score medio"], x["Confianza global"], -x["Nº picks"]), reverse=True)
    return pd.DataFrame(picks), combos



def construir_combinadas_plan_b_v23270(picks_df, cuota_min=1.60, cuota_max=1.80, min_picks=2, max_picks=3):
    """
    Plan B: NO marca como segura. Solo propone candidatas controladas si el modo normal/conservador
    no encuentra nada. Usa picks COMBI SAFE + FUERTE SIMPLE, excluye NO COMBI y mantiene partidos únicos.
    """
    if picks_df is None or picks_df.empty:
        return []

    df = picks_df.copy()
    if "Etiqueta" not in df.columns:
        return []

    # Solo picks con cierta calidad. No rescatamos NO COMBI para no volver al fallo por uno.
    df = df[df["Etiqueta"].isin(["🧱 COMBI SAFE", "🔥 FUERTE SIMPLE"])].copy()
    if df.empty:
        return []

    # Evita mercados de varianza si son solo fuertes simples.
    def usable(row):
        tipo = str(row.get("Tipo", ""))
        etiqueta = str(row.get("Etiqueta", ""))
        prob = float(row.get("Prob", 0.0) or 0.0)
        min_req = float(row.get("Min", 0.0) or 0.0)
        if etiqueta == "🧱 COMBI SAFE":
            return True
        if tipo in ["sets3", "dog_set"]:
            return False
        # Fuerte simple pero muy cerca del mínimo.
        return prob >= (min_req - 0.055)

    df = df[df.apply(usable, axis=1)].copy()
    if df.empty:
        return []

    picks = df.to_dict(orient="records")
    combos = []

    for n in range(int(min_picks), int(max_picks) + 1):
        for combo in combinations(picks, n):
            partidos = [p.get("Partido", "") for p in combo]
            if len(set(partidos)) != len(partidos):
                continue

            cuota_total = 1.0
            confianza = 1.0
            score_medio = 0.0
            gaps = []
            labels = []

            for p in combo:
                cuota_total *= float(p.get("Cuota", 1.0) or 1.0)
                confianza *= float(p.get("Prob", 0.0) or 0.0)
                score_medio += float(p.get("Score", 0.0) or 0.0)
                gaps.append(float(p.get("Prob", 0.0) or 0.0) - float(p.get("Min", 0.0) or 0.0))
                labels.append(str(p.get("Etiqueta", "")))

            score_medio /= max(1, len(combo))

            if float(cuota_min) <= cuota_total <= float(cuota_max):
                weak = min(combo, key=lambda x: float(x.get("Prob", 0.0) or 0.0))
                safe_count = sum(1 for lab in labels if lab == "🧱 COMBI SAFE")
                strong_simple_count = sum(1 for lab in labels if lab == "🔥 FUERTE SIMPLE")
                min_gap = min(gaps) if gaps else -1.0

                combos.append({
                    "Nº picks": n,
                    "Cuota total": cuota_total,
                    "Confianza global": confianza,
                    "Score medio": score_medio,
                    "Pick más débil": f"{weak['Mercado']} — {weak['Partido']} ({float(weak['Prob']):.1%})",
                    "Picks": list(combo),
                    "Safe": safe_count,
                    "Fuerte simple": strong_simple_count,
                    "Gap mínimo": min_gap,
                })

    combos = sorted(
        combos,
        key=lambda x: (x["Safe"], x["Gap mínimo"], x["Score medio"], x["Confianza global"], -x["Nº picks"]),
        reverse=True
    )
    return combos



# =========================================================
# v23.40.1 BET PLAN + STAKE ENGINE
# Convierte picks en plan operativo: simples, dobles y stake.
# No toca el predictor ni el motor Over; solo gestiona riesgo.
# =========================================================

def _stake_profile_params_v23401(mode):
    mode = str(mode or "⚖️ Normal")
    if "Conservador" in mode:
        return {"simple_cap_u": 1.00, "combo_u": 0.25, "kelly_frac": 0.12, "max_singles": 4, "daily_pct": 3.0}
    if "Agresivo" in mode:
        return {"simple_cap_u": 1.50, "combo_u": 0.60, "kelly_frac": 0.25, "max_singles": 6, "daily_pct": 5.0}
    return {"simple_cap_u": 1.25, "combo_u": 0.40, "kelly_frac": 0.18, "max_singles": 5, "daily_pct": 4.0}


def _edge_kelly_v23401(prob, cuota):
    try:
        p = float(prob or 0.0)
        q = float(cuota or 0.0)
        if p <= 0 or q <= 1.0:
            return None, None, None
        implied = 1.0 / q
        edge = p - implied
        b = q - 1.0
        kelly = ((b * p) - (1.0 - p)) / b if b > 0 else 0.0
        return float(implied), float(edge), float(max(0.0, kelly))
    except Exception:
        return None, None, None


def _stake_units_for_pick_v23401(row, stake_mode="⚖️ Normal"):
    params = _stake_profile_params_v23401(stake_mode)
    etiqueta = str(row.get("Etiqueta", ""))
    tipo = str(row.get("Tipo", ""))
    prob = float(row.get("Prob", 0.0) or 0.0)
    cuota = float(row.get("Cuota", 1.0) or 1.0)
    cuota_tipo = str(row.get("Cuota tipo", ""))

    # Base por calidad del pick.
    if etiqueta == "🧱 COMBI SAFE":
        units = 1.00
    elif etiqueta == "🔥 FUERTE SIMPLE":
        units = 0.60
    else:
        return 0.0, "PASAR", "no es pick jugable para plan"

    # Ajustes por probabilidad.
    if prob >= 0.92:
        units += 0.20
    elif prob >= 0.86:
        units += 0.10
    elif prob < 0.76:
        units -= 0.20

    # Ajustes por mercado: algunos picks son buenos, pero frágiles en staking.
    if tipo in ["sets3", "dog_set"]:
        units *= 0.60
    elif tipo in ["over19", "over20"]:
        units *= 0.55
    elif tipo in ["over17", "over18"]:
        units *= 1.05
    elif tipo == "ml":
        units *= 0.90

    # Cuotas muy bajas: mejor no cargarlas, aunque sean verdes.
    if cuota_tipo == "pegada" and cuota < 1.25:
        units *= 0.45

    implied, edge, kelly = _edge_kelly_v23401(prob, cuota)
    nota_value = ""
    if cuota_tipo == "pegada" and implied is not None:
        if edge <= 0:
            return 0.0, "PASAR", f"sin value con cuota {cuota:.2f}: prob modelo {prob:.1%} vs implícita {implied:.1%}"
        if edge < 0.025:
            units *= 0.50
            nota_value = f"edge pequeño +{edge:.1%}: stake reducido"
        elif edge >= 0.08:
            units *= 1.10
            nota_value = f"edge bueno +{edge:.1%}"
        else:
            nota_value = f"edge positivo +{edge:.1%}"

        # Kelly fraccionado como techo, no como orden agresiva.
        if kelly is not None and kelly > 0:
            # Convertimos Kelly % en unidades aproximadas asumiendo 1u≈1% bankroll.
            kelly_units_cap = max(0.20, min(params["simple_cap_u"], kelly * params["kelly_frac"] * 100.0))
            units = min(units, kelly_units_cap)

    units = float(np.clip(units, 0.0, params["simple_cap_u"]))
    if units <= 0:
        return 0.0, "PASAR", nota_value or "stake cero por filtros"
    label = "SIMPLE" if etiqueta == "🧱 COMBI SAFE" else "SIMPLE BAJO"
    return units, label, nota_value or "stake por confianza/mercado"


def construir_plan_apuestas_v23401(picks_df, combos, bankroll=100.0, unit_amount=1.0, max_daily_pct=4.0, stake_mode="⚖️ Normal"):
    if picks_df is None or picks_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), 0.0, 0.0

    params = _stake_profile_params_v23401(stake_mode)
    try:
        bankroll = float(bankroll or 0.0)
        unit_amount = float(unit_amount or 0.0)
        max_daily_pct = float(max_daily_pct or params["daily_pct"])
    except Exception:
        bankroll, unit_amount, max_daily_pct = 100.0, 1.0, params["daily_pct"]

    daily_cap = max(0.0, bankroll * max_daily_pct / 100.0)
    rows = []
    df = picks_df.copy()
    df = df.sort_values(["Etiqueta", "Score", "Prob"], ascending=[True, False, False])

    for _, r in df.iterrows():
        units, play_type, note = _stake_units_for_pick_v23401(r, stake_mode=stake_mode)
        if units <= 0:
            continue
        stake = units * unit_amount
        rows.append({
            "Juego": play_type,
            "Partido": r.get("Partido", ""),
            "Mercado": r.get("Mercado", ""),
            "Prob %": round(float(r.get("Prob", 0.0) or 0.0) * 100.0, 1),
            "Cuota": round(float(r.get("Cuota", 1.0) or 1.0), 2),
            "Cuota tipo": r.get("Cuota tipo", ""),
            "Stake u": round(units, 2),
            "Stake €": round(stake, 2),
            "Motivo stake": note,
            "Etiqueta": r.get("Etiqueta", ""),
            "Tipo": r.get("Tipo", ""),
            "Score": float(r.get("Score", 0.0) or 0.0),
        })

    singles = pd.DataFrame(rows)
    if not singles.empty:
        max_singles = int(params.get("max_singles", 5))
        singles = singles.sort_values(["Stake u", "Score", "Prob %"], ascending=[False, False, False]).head(max_singles).copy()

    combo_rows = []
    # Solo primeras 2 combinadas oficiales. La combinada es complemento, no base.
    for i, combo in enumerate((combos or [])[:2], start=1):
        combo_units = float(params.get("combo_u", 0.40))
        if float(combo.get("Nº picks", 2)) >= 3:
            combo_units *= 0.60
        stake = combo_units * unit_amount
        combo_rows.append({
            "Juego": f"DOBLE/TRIPLE #{i}",
            "Nº picks": combo.get("Nº picks", ""),
            "Cuota total": round(float(combo.get("Cuota total", 1.0) or 1.0), 2),
            "Confianza global": f"{float(combo.get('Confianza global', 0.0) or 0.0):.1%}",
            "Stake u": round(combo_units, 2),
            "Stake €": round(stake, 2),
            "Pick más débil": combo.get("Pick más débil", ""),
            "Picks": " + ".join([str(p.get("Mercado", "")) + " — " + str(p.get("Partido", "")) for p in combo.get("Picks", [])]),
            "Motivo stake": "combinada controlada; stake pequeño para evitar fallo por una",
        })
    combo_plan = pd.DataFrame(combo_rows)

    total_stake = 0.0
    if not singles.empty:
        total_stake += float(singles["Stake €"].sum())
    if not combo_plan.empty:
        total_stake += float(combo_plan["Stake €"].sum())

    # Escalado por límite diario.
    if daily_cap > 0 and total_stake > daily_cap:
        factor = daily_cap / total_stake
        if not singles.empty:
            singles["Stake €"] = (singles["Stake €"] * factor).round(2)
            singles["Stake u"] = (singles["Stake u"] * factor).round(2)
            singles["Motivo stake"] = singles["Motivo stake"].astype(str) + f" · escalado por límite diario {max_daily_pct:.1f}%"
        if not combo_plan.empty:
            combo_plan["Stake €"] = (combo_plan["Stake €"] * factor).round(2)
            combo_plan["Stake u"] = (combo_plan["Stake u"] * factor).round(2)
            combo_plan["Motivo stake"] = combo_plan["Motivo stake"].astype(str) + f" · escalado por límite diario {max_daily_pct:.1f}%"
        total_stake = daily_cap

    no_play = picks_df[picks_df["Etiqueta"].astype(str).str.contains("NO COMBI", na=False)].copy() if "Etiqueta" in picks_df.columns else pd.DataFrame()
    if not no_play.empty:
        no_play = no_play.sort_values("Score", ascending=False).head(8).copy()
        no_play["Prob %"] = (no_play["Prob"].astype(float) * 100).round(1)
        no_play["Cuota"] = no_play["Cuota"].astype(float).round(2)

    return singles, combo_plan, no_play, float(total_stake), float(daily_cap)



# =========================================================
# v23.41.0 PLAN GLOBAL DEL DÍA
# =========================================================

GLOBAL_BET_PLAN_KEY_V23410 = "global_bet_plan_picks_v23410"

def _global_pick_key_v23410(row):
    return "|".join([
        limpiar(str(row.get("Partido", ""))),
        limpiar(str(row.get("Mercado", ""))),
    ])


def _normalizar_picks_global_v23410(df, fuente="Tanda actual"):
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for c in ["Partido", "Mercado", "Etiqueta", "Tipo", "Motivos", "Cuota tipo"]:
        if c not in out.columns:
            out[c] = ""
    for c in ["Prob", "Cuota", "Min", "Score"]:
        if c not in out.columns:
            out[c] = 0.0
    out["Fuente análisis"] = str(fuente)
    out["Añadido"] = time.strftime("%H:%M:%S")
    out["GlobalKey"] = out.apply(_global_pick_key_v23410, axis=1)
    out = out[out["GlobalKey"].astype(str).str.len() > 2].copy()
    return out


def _get_global_picks_v23410():
    data = st.session_state.get(GLOBAL_BET_PLAN_KEY_V23410)
    if isinstance(data, pd.DataFrame):
        return data.copy()
    return pd.DataFrame()


def _set_global_picks_v23410(df):
    if df is None or df.empty:
        st.session_state[GLOBAL_BET_PLAN_KEY_V23410] = pd.DataFrame()
        return
    out = df.copy()
    if "GlobalKey" not in out.columns:
        out["GlobalKey"] = out.apply(_global_pick_key_v23410, axis=1)
    # Si el mismo partido+mercado aparece varias veces, conserva el de mayor Score/Prob.
    sort_cols = [c for c in ["Score", "Prob"] if c in out.columns]
    if sort_cols:
        out = out.sort_values(sort_cols, ascending=[False] * len(sort_cols)).copy()
    out = out.drop_duplicates("GlobalKey", keep="first").reset_index(drop=True)
    st.session_state[GLOBAL_BET_PLAN_KEY_V23410] = out



# =========================================================
# v23.41.1 CUOTAS REALES + MERCADOS ALTERNATIVOS
# Permite comparar, por ejemplo, Over 18.5 recomendado vs Over 19.5 real.
# No cambia el predictor; solo decide qué línea tiene mejor value/stake.
# =========================================================

ALT_ODDS_STATE_KEY_V23411 = "global_alt_odds_editor_v23411"


def _prob_stake_cap_v23411(prob, tipo="otro"):
    """Cap de probabilidad para staking: nunca tratamos un pick como 100% real."""
    try:
        p = float(prob or 0.0)
    except Exception:
        return 0.0
    tipo = str(tipo or "otro")
    if tipo in ["over17", "over18"]:
        cap = 0.92
    elif tipo in ["over19", "over20"]:
        cap = 0.89
    elif tipo in ["dog_set", "fav20"]:
        cap = 0.96
    elif tipo == "sets3":
        cap = 0.78
    elif tipo == "ml":
        cap = 0.95
    else:
        cap = 0.92
    return float(np.clip(p, 0.0, cap))


def _edge_alt_value_v23411(prob, cuota, tipo="otro"):
    try:
        q = float(str(cuota).replace(",", "."))
        if q <= 1.0:
            return None
        p_raw = float(prob or 0.0)
        p = _prob_stake_cap_v23411(p_raw, tipo)
        implied = 1.0 / q
        edge = p - implied
        ev = (p * q) - 1.0
        return {"prob_raw": p_raw, "prob_used": p, "cuota": q, "implied": implied, "edge": edge, "ev": ev}
    except Exception:
        return None


def _alt_candidates_from_row_v23411(row):
    """Devuelve mercados alternativos disponibles para una fila acumulada."""
    base = []
    partido = str(row.get("Partido", "") or "")
    actual_market = str(row.get("Mercado", "") or "")
    try:
        actual_prob = float(str(row.get("Prob", 0.0) or 0.0).replace(",", "."))
    except Exception:
        actual_prob = 0.0
    actual_tipo = str(row.get("Tipo", "") or _combi_tipo_mercado(actual_market))
    base.append({"Partido": partido, "Mercado": actual_market, "Prob": actual_prob, "Tipo": actual_tipo, "odds_col": "Cuota real actual", "kind": "actual"})

    for line, col, tipo in [
        ("17.5", "Prob Over 17.5", "over17"),
        ("18.5", "Prob Over 18.5", "over18"),
        ("19.5", "Prob Over 19.5", "over19"),
        ("20.5", "Prob Over 20.5", "over20"),
        ("22.5", "Prob Over 22.5", "over22"),
    ]:
        try:
            pr = float(row.get(col, 0.0) or 0.0)
        except Exception:
            pr = 0.0
        if pr > 0:
            base.append({"Partido": partido, "Mercado": f"OVER {line}", "Prob": pr, "Tipo": tipo, "odds_col": f"Cuota Over {line}", "kind": "alt"})

    # Eliminar duplicados por mercado conservando el primero.
    seen = set()
    out = []
    for c in base:
        k = limpiar(c.get("Mercado", ""))
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(c)
    return out


def _preparar_editor_cuotas_v23411(global_df):
    if global_df is None or global_df.empty:
        return pd.DataFrame()
    df = global_df.copy().reset_index(drop=True)
    if "GlobalKey" not in df.columns:
        df["GlobalKey"] = df.apply(_global_pick_key_v23410, axis=1)
    for c in ["Cuota real actual", "Cuota Over 17.5", "Cuota Over 18.5", "Cuota Over 19.5", "Cuota Over 20.5", "Cuota Over 22.5"]:
        if c not in df.columns:
            df[c] = ""
    show_cols = [
        "GlobalKey", "Etiqueta", "Partido", "Mercado", "Prob", "Cuota", "Cuota tipo",
        "Cuota real actual", "Cuota Over 17.5", "Cuota Over 18.5", "Cuota Over 19.5", "Cuota Over 20.5", "Cuota Over 22.5",
        "Prob Over 17.5", "Prob Over 18.5", "Prob Over 19.5", "Prob Over 20.5", "Prob Over 22.5"
    ]
    return df[[c for c in show_cols if c in df.columns]].copy()


def _aplicar_cuotas_alternativas_v23411(global_df, odds_editor_df):
    """Elige el mejor mercado por edge positivo entre mercado actual y líneas alternativas con cuota real."""
    if global_df is None or global_df.empty or odds_editor_df is None or odds_editor_df.empty:
        return global_df.copy() if isinstance(global_df, pd.DataFrame) else pd.DataFrame(), pd.DataFrame()

    df = global_df.copy().reset_index(drop=True)
    if "GlobalKey" not in df.columns:
        df["GlobalKey"] = df.apply(_global_pick_key_v23410, axis=1)
    odds = odds_editor_df.copy()
    if "GlobalKey" not in odds.columns:
        return df, pd.DataFrame()
    odds = odds.drop_duplicates("GlobalKey", keep="last").set_index("GlobalKey")

    decisions = []
    updated_rows = []
    for _, row in df.iterrows():
        r = row.copy()
        gk = r.get("GlobalKey")
        odds_row = odds.loc[gk] if gk in odds.index else None
        best = None
        evals = []
        for cand in _alt_candidates_from_row_v23411(r):
            q = None
            if odds_row is not None and cand["odds_col"] in odds_row.index:
                q = odds_row.get(cand["odds_col"], "")
            # Para el mercado actual, si no se rellenó cuota real, mantenemos cuota pegada/estimada solo como referencia.
            if (q is None or str(q).strip() == "") and cand["kind"] == "actual":
                q = r.get("Cuota", "") if str(r.get("Cuota tipo", "")).lower() == "pegada" else ""
            val = _edge_alt_value_v23411(cand.get("Prob"), q, cand.get("Tipo"))
            if val is None:
                continue
            item = {**cand, **val}
            evals.append(item)
            # Exigimos edge positivo real. Para alternativas más altas pedimos un poco más.
            min_edge = 0.015 if cand.get("kind") == "actual" else 0.020
            if item["edge"] >= min_edge:
                if best is None or (item["ev"], item["edge"], item["cuota"]) > (best["ev"], best["edge"], best["cuota"]):
                    best = item

        if best is not None:
            r["Mercado"] = best["Mercado"]
            r["Prob"] = float(best["prob_used"])
            r["Prob modelo original"] = float(best["prob_raw"])
            r["Cuota"] = float(best["cuota"])
            r["Cuota tipo"] = "real alternativa" if best.get("kind") == "alt" else "real"
            r["Tipo"] = best["Tipo"]
            r["Edge real"] = float(best["edge"])
            r["EV real"] = float(best["ev"])
            r["Mercado elegido value"] = best["Mercado"]
            if str(r.get("Etiqueta", "")) == "❌ NO COMBI" and best["prob_used"] >= float(r.get("Min", 0.80) or 0.80):
                r["Etiqueta"] = "🔥 FUERTE SIMPLE"
            decisions.append({
                "Partido": r.get("Partido", ""),
                "Mercado elegido": best["Mercado"],
                "Prob usada %": round(best["prob_used"] * 100, 1),
                "Prob modelo %": round(best["prob_raw"] * 100, 1),
                "Cuota real": round(best["cuota"], 2),
                "Implícita %": round(best["implied"] * 100, 1),
                "Edge %": round(best["edge"] * 100, 1),
                "EV %": round(best["ev"] * 100, 1),
                "Tipo cuota": r.get("Cuota tipo", ""),
            })
        else:
            # Si se han rellenado cuotas pero ninguna tiene value, bloqueamos para stake automático.
            if evals:
                r["Etiqueta"] = "❌ NO COMBI"
                r["Combi Safe"] = False
                r["Motivos"] = str(r.get("Motivos", "")) + " · sin value real en cuotas introducidas"
                best_no = max(evals, key=lambda x: x["ev"])
                decisions.append({
                    "Partido": r.get("Partido", ""),
                    "Mercado elegido": "PASAR",
                    "Prob usada %": round(best_no["prob_used"] * 100, 1),
                    "Prob modelo %": round(best_no["prob_raw"] * 100, 1),
                    "Cuota real": round(best_no["cuota"], 2),
                    "Implícita %": round(best_no["implied"] * 100, 1),
                    "Edge %": round(best_no["edge"] * 100, 1),
                    "EV %": round(best_no["ev"] * 100, 1),
                    "Tipo cuota": "sin value",
                })
        updated_rows.append(r)

    out = pd.DataFrame(updated_rows)
    if "Combi Safe" not in out.columns:
        out["Combi Safe"] = out.get("Etiqueta", "").astype(str).str.contains("COMBI SAFE", na=False)
    else:
        out["Combi Safe"] = out.get("Etiqueta", "").astype(str).str.contains("COMBI SAFE", na=False)
    return out, pd.DataFrame(decisions)


def construir_combinadas_desde_picks_v23410(picks_df, cuota_min=1.70, cuota_max=2.40, min_picks=2, max_picks=2):
    """Construye combinadas oficiales desde picks ya clasificados en la bolsa global."""
    if picks_df is None or picks_df.empty:
        return []
    df = picks_df.copy()
    if "Combi Safe" not in df.columns:
        df["Combi Safe"] = df.get("Etiqueta", "").astype(str).str.contains("COMBI SAFE", na=False)
    safe = df[df["Combi Safe"].astype(bool)].copy()
    if safe.empty:
        return []

    picks = safe.to_dict(orient="records")
    combos = []
    for n in range(int(min_picks), int(max_picks) + 1):
        for combo in combinations(picks, n):
            partidos = [str(p.get("Partido", "")) for p in combo]
            if len(set(partidos)) != len(partidos):
                continue

            tipos_combo = [str(p.get("Tipo", "")) for p in combo]
            if sum(t in ["sets3", "dog_set"] for t in tipos_combo) > 1:
                continue
            if sum(t in ["over19", "over20"] for t in tipos_combo) > 1:
                continue
            if any(float(p.get("Cuota", 1.0) or 1.0) < 1.25 for p in combo):
                continue

            cuota_total = 1.0
            confianza = 1.0
            score_medio = 0.0
            for p in combo:
                cuota_total *= float(p.get("Cuota", 1.0) or 1.0)
                confianza *= float(p.get("Prob", 0.0) or 0.0)
                score_medio += float(p.get("Score", 0.0) or 0.0)
            score_medio /= max(1, len(combo))

            min_confianza_combo = 0.56 if n == 2 else 0.46
            if float(cuota_min) <= cuota_total <= float(cuota_max) and confianza >= min_confianza_combo:
                weak = min(combo, key=lambda x: float(x.get("Prob", 0.0) or 0.0))
                combos.append({
                    "Nº picks": n,
                    "Cuota total": cuota_total,
                    "Confianza global": confianza,
                    "Score medio": score_medio,
                    "Pick más débil": f"{weak.get('Mercado','')} — {weak.get('Partido','')} ({float(weak.get('Prob',0.0) or 0.0):.1%})",
                    "Picks": list(combo),
                })
    return sorted(combos, key=lambda x: (x["Score medio"], x["Confianza global"], -x["Nº picks"]), reverse=True)


def render_plan_global_dia_v23410(picks_df_actual, cuota_min, cuota_max, max_picks, bankroll_plan, unit_plan, daily_pct_plan, stake_mode):
    st.markdown("### 📦 Plan Global del Día")
    st.caption("Usa esto cuando analices Challenger, ATP Grand Slam y WTA por separado. Añade cada tanda y al final genera un único stake global.")

    b1, b2, b3 = st.columns([1.2, 1.0, 2.2])
    with b1:
        if st.button("➕ Añadir picks actuales al plan global", key="add_global_picks_v23410", use_container_width=True):
            current = _normalizar_picks_global_v23410(picks_df_actual, fuente="Tanda actual")
            previous = _get_global_picks_v23410()
            merged = pd.concat([previous, current], ignore_index=True) if not previous.empty else current
            _set_global_picks_v23410(merged)
            st.success(f"Añadidos {len(current)} picks actuales al plan global. Duplicados eliminados automáticamente.")
    with b2:
        if st.button("🧹 Limpiar plan global", key="clear_global_picks_v23410", use_container_width=True):
            _set_global_picks_v23410(pd.DataFrame())
            st.warning("Plan global limpiado.")
    with b3:
        st.info("Flujo: analiza Challenger → añadir · analiza ATP GS → añadir · analiza WTA → añadir · mira este plan global.")

    global_df = _get_global_picks_v23410()

    # v23.42.1: Auto perfiles visible incluso antes de tener plan global acumulado.
    # Así puedes analizar una tanda, ver faltantes y buscarlos sin tener que descubrir el bloque por error.
    try:
        render_auto_profile_finder_v23420(current_df=picks_df_actual, global_df=global_df)
    except Exception as e:
        st.warning("No se pudo mostrar el bloque de Auto perfiles, pero el resto del Plan Global sigue funcionando.")
        st.caption(f"Detalle técnico Auto perfiles: {type(e).__name__}: {e}")

    if global_df.empty:
        st.warning("Todavía no hay picks acumulados. Para que puedas meter cuotas sin perder tiempo, uso la tanda actual como VISTA PREVIA. Para el stake global real, pulsa primero ➕ Añadir picks actuales al plan global.")
        try:
            preview_current = _normalizar_picks_global_v23410(picks_df_actual, fuente="Tanda actual · preview cuotas")
        except Exception:
            preview_current = pd.DataFrame()
        if preview_current is None or preview_current.empty:
            st.warning("No hay picks actuales suficientes para mostrar cuotas. Analiza una lista o añade picks al plan global.")
            return
        global_df = preview_current

    safe_count = int(global_df.get("Combi Safe", pd.Series(dtype=bool)).astype(bool).sum()) if "Combi Safe" in global_df.columns else 0
    fuerte_count = int(global_df.get("Etiqueta", pd.Series(dtype=str)).astype(str).str.contains("FUERTE SIMPLE", na=False).sum()) if "Etiqueta" in global_df.columns else 0
    g1, g2, g3, g4 = st.columns(4)
    g1.metric("Picks acumulados", len(global_df))
    g2.metric("🧱 Combi Safe", safe_count)
    g3.metric("🔥 Fuertes simple", fuerte_count)
    g4.metric("Límite diario", f"{float(bankroll_plan) * float(daily_pct_plan) / 100.0:.2f} €")

    show_global = global_df.copy()
    if "Prob" in show_global.columns:
        show_global["Prob %"] = (show_global["Prob"].astype(float) * 100).round(1)
    if "Min" in show_global.columns:
        show_global["Mínimo %"] = (show_global["Min"].astype(float) * 100).round(1)
    if "Cuota" in show_global.columns:
        show_global["Cuota"] = pd.to_numeric(show_global["Cuota"].astype(str).str.replace(",", ".", regex=False), errors="coerce").round(2)

    with st.expander("Ver picks acumulados en el plan global", expanded=False):
        cols = ["Etiqueta", "Partido", "Mercado", "Prob %", "Mínimo %", "Cuota", "Cuota tipo", "Fuente análisis", "Motivos"]
        st.dataframe(show_global[[c for c in cols if c in show_global.columns]], width='stretch', hide_index=True)

    st.markdown("#### 💸 Cuotas reales y mercados alternativos")
    st.caption("Rellena la cuota real del mercado recomendado y, si existe, la cuota de otra línea del mismo partido. Ejemplo: pick Over 18.5, puedes poner también cuota Over 19.5 y la app elegirá la opción con más value real.")

    editor_base = _preparar_editor_cuotas_v23411(global_df)
    with st.expander("Editar cuotas reales / alternativas", expanded=True):
        st.info("Modo seguro móvil: para evitar caídas de Streamlit en el móvil, las cuotas se introducen en texto. Usa punto o coma decimal.")

        tabla_ref = editor_base.copy().reset_index(drop=True)
        tabla_ref["Nº"] = range(1, len(tabla_ref) + 1)
        if "Prob" in tabla_ref.columns:
            tabla_ref["Prob %"] = (pd.to_numeric(tabla_ref["Prob"], errors="coerce") * 100).round(1)
        ref_cols = ["Nº", "Partido", "Mercado", "Prob %", "Cuota", "Cuota tipo"]
        st.dataframe(tabla_ref[[c for c in ref_cols if c in tabla_ref.columns]], width='stretch', hide_index=True)

        # v23.43.0: Plantilla rápida para buscar y pegar cuotas sin ir editando celda a celda.
        try:
            template_lines = []
            search_lines = []
            for _, rr in tabla_ref.iterrows():
                n = int(rr.get("Nº", 0))
                partido_txt = str(rr.get("Partido", "")).strip()
                mercado_txt = str(rr.get("Mercado", "")).strip()
                prob_txt = rr.get("Prob %", "")
                try:
                    prob_txt = f"{float(prob_txt):.1f}%"
                except Exception:
                    prob_txt = ""
                search_lines.append(f"{n}. {partido_txt} | {mercado_txt} | Modelo {prob_txt}")
                # Formato: Nº; cuota real; over17.5; over18.5; over19.5; over20.5; over22.5
                template_lines.append(f"{n};;;;;;  # {partido_txt} | {mercado_txt}")

            with st.expander("📋 Plantilla rápida para buscar/pegar cuotas", expanded=False):
                st.caption("Copia la primera lista para buscar cuotas en la casa. Luego rellena la segunda plantilla con las cuotas reales que encuentres.")
                st.text_area(
                    "Lista de picks para buscar cuotas",
                    value="\n".join(search_lines),
                    height=180,
                    key="quick_odds_search_template_v23430",
                )
                st.text_area(
                    "Plantilla para pegar cuotas",
                    value="\n".join(template_lines),
                    height=180,
                    key="quick_odds_input_template_v23430",
                )
                st.caption("Borra los comentarios si quieres, o deja las líneas tal cual: la app solo lee lo que está antes del #.")
        except Exception as e:
            st.caption(f"No se pudo crear la plantilla rápida de cuotas: {type(e).__name__}")

        st.caption("Formato: Nº; cuota real; over17.5; over18.5; over19.5; over20.5; over22.5")
        st.caption("Ejemplo: 3;1.42;;1.42;1.65;1.90;  → rellena la fila 3 con Over18/19/20.")
        odds_text = st.text_area(
            "Cuotas reales y alternativas",
            value=st.session_state.get("odds_text_global_v23413", ""),
            key="odds_text_global_v23413",
            height=140,
            placeholder="1;1.38;;1.38;1.55;1.78;\n2;1,25;;;;;",
        )

        odds_editor = editor_base.copy()
        try:
            if str(odds_text).strip():
                col_map = ["Cuota real actual", "Cuota Over 17.5", "Cuota Over 18.5", "Cuota Over 19.5", "Cuota Over 20.5", "Cuota Over 22.5"]
                for raw_line in str(odds_text).splitlines():
                    line = raw_line.strip()
                    if "#" in line:
                        line = line.split("#", 1)[0].strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = [x.strip().replace(",", ".") for x in line.split(";")]
                    try:
                        idx = int(float(parts[0])) - 1
                    except Exception:
                        continue
                    if idx < 0 or idx >= len(odds_editor):
                        continue
                    for j, col in enumerate(col_map, start=1):
                        if j < len(parts) and parts[j] != "":
                            odds_editor.loc[idx, col] = parts[j]
                st.success("Cuotas aplicadas al plan global.")
            else:
                st.warning("Sin cuotas reales introducidas: la app usará cuotas estimadas solo como orientación.")
        except Exception as e:
            st.error("No se pudieron leer las cuotas introducidas. Revisa el formato de las líneas.")
            st.caption(f"Detalle técnico: {type(e).__name__}: {e}")
            odds_editor = editor_base.copy()

    try:
        global_df_value, value_decisions = _aplicar_cuotas_alternativas_v23411(global_df, odds_editor)
    except Exception as e:
        st.error("No se pudo recalcular value con cuotas reales. Uso el plan global sin cuotas alternativas.")
        st.caption(f"Detalle técnico: {type(e).__name__}: {e}")
        global_df_value, value_decisions = global_df, pd.DataFrame()
    if not value_decisions.empty:
        st.markdown("##### 🧮 Mejor mercado por value real")
        st.dataframe(value_decisions, width='stretch', hide_index=True)
        st.success("El plan de stake de abajo ya usa estas cuotas reales y el mercado alternativo elegido cuando tiene más value.")
    else:
        global_df_value = global_df

    global_combos = construir_combinadas_desde_picks_v23410(
        global_df_value,
        cuota_min=cuota_min,
        cuota_max=cuota_max,
        min_picks=2,
        max_picks=max_picks,
    )
    singles_g, combos_g, no_play_g, total_g, cap_g = construir_plan_apuestas_v23401(
        global_df_value,
        global_combos,
        bankroll=bankroll_plan,
        unit_amount=unit_plan,
        max_daily_pct=daily_pct_plan,
        stake_mode=stake_mode,
    )

    st.markdown("#### 🎯 Plan GLOBAL recomendado")
    pg1, pg2, pg3 = st.columns(3)
    pg1.metric("Stake total global", f"{total_g:.2f} €")
    pg2.metric("Límite diario", f"{cap_g:.2f} €")
    pg3.metric("Combinadas globales", len(global_combos))

    if len(global_combos) == 0 and not singles_g.empty:
        st.warning("🔴 GLOBAL: no hay combinada oficial limpia. Jugar solo las mejores simples acumuladas.")
    elif len(global_combos) > 0:
        st.success("🟢 GLOBAL: hay simples y combinada pequeña permitida dentro del límite diario.")
    else:
        st.error("GLOBAL: no hay forma limpia de jugar con estos picks. Mejor pasar o esperar más partidos.")

    if not singles_g.empty:
        st.markdown("##### 🎾 Simples globales")
        st.dataframe(singles_g[["Juego", "Partido", "Mercado", "Prob %", "Cuota", "Cuota tipo", "Stake u", "Stake €", "Motivo stake"]], width='stretch', hide_index=True)
    if not combos_g.empty:
        st.markdown("##### 🔗 Combinadas globales pequeñas")
        st.dataframe(combos_g[["Juego", "Nº picks", "Cuota total", "Confianza global", "Stake u", "Stake €", "Pick más débil", "Motivo stake"]], width='stretch', hide_index=True)
        with st.expander("Ver composición de combinadas globales"):
            st.dataframe(combos_g[["Juego", "Picks"]], width='stretch', hide_index=True)


def render_constructor_combinadas_v23268(ok_df):
    st.divider()
    st.subheader("💰 Plan de apuestas + 🧱 Combi Guard")
    st.caption("Objetivo: rentabilidad y evitar el fallo por una. Prioriza simples, dobles limpias y bloquea cuotas bajas/mercados frágiles.")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        profile_name = st.selectbox("Modo combi", list(COMBI_SAFE_PROFILES_V23268.keys()), index=1, key="combi_profile_v23268")
    with c2:
        cuota_min = st.number_input("Cuota mínima", min_value=1.01, max_value=10.0, value=1.70, step=0.05, key="combi_cuota_min_v23268")
    with c3:
        cuota_max = st.number_input("Cuota máxima", min_value=1.01, max_value=10.0, value=2.40, step=0.05, key="combi_cuota_max_v23268")
    with c4:
        max_default = COMBI_SAFE_PROFILES_V23268.get(profile_name, {}).get("max_picks", 3)
        max_picks = st.slider("Máx picks", 2, 3, int(max_default), key="combi_max_picks_v23268")

    min_picks = 2
    picks_df, combos = construir_combinadas_v23268(
        ok_df,
        profile_name=profile_name,
        cuota_min=cuota_min,
        cuota_max=cuota_max,
        min_picks=min_picks,
        max_picks=max_picks,
    )

    st.markdown("### 🎯 Cómo jugarlos + stake")
    s1, s2, s3, s4 = st.columns(4)
    with s1:
        bankroll_plan = st.number_input("Bankroll €", min_value=1.0, max_value=100000.0, value=100.0, step=10.0, key="stake_bankroll_v23401")
    with s2:
        unit_plan = st.number_input("Unidad base €", min_value=0.10, max_value=1000.0, value=1.0, step=0.10, key="stake_unit_v23401")
    with s3:
        default_daily = _stake_profile_params_v23401(profile_name).get("daily_pct", 4.0)
        daily_pct_plan = st.number_input("Riesgo máx diario %", min_value=0.1, max_value=20.0, value=float(default_daily), step=0.25, key="stake_daily_pct_v23401")
    with s4:
        stake_mode = st.selectbox("Modo stake", ["🔒 Conservador", "⚖️ Normal", "🔥 Agresivo"], index=1, key="stake_mode_v23401")

    if picks_df.empty:
        st.warning("No hay picks suficientes para construir combinadas.")
        return

    render_plan_global_dia_v23410(
        picks_df,
        cuota_min=cuota_min,
        cuota_max=cuota_max,
        max_picks=max_picks,
        bankroll_plan=bankroll_plan,
        unit_plan=unit_plan,
        daily_pct_plan=daily_pct_plan,
        stake_mode=stake_mode,
    )
    st.divider()
    st.markdown("### 📍 Plan solo de la tanda actual")
    st.caption("Este bloque mira únicamente los picks que acabas de analizar. Para controlar el stake real del día, usa el Plan Global de arriba.")

    show = picks_df.copy()
    show["Prob %"] = show["Prob"].apply(lambda x: round(float(x) * 100, 1))
    show["Mínimo %"] = show["Min"].apply(lambda x: round(float(x) * 100, 1))
    show["Cuota"] = show["Cuota"].apply(lambda x: round(float(x), 2))
    show["Score"] = show["Score"].apply(lambda x: round(float(x), 1))
    cols = ["Etiqueta", "Partido", "Mercado", "Prob %", "Mínimo %", "Cuota", "Cuota tipo", "Score", "Motivos"]

    safe_count = int(picks_df["Combi Safe"].sum()) if "Combi Safe" in picks_df.columns else 0
    m1, m2, m3 = st.columns(3)
    m1.metric("Picks analizados", len(picks_df))
    m2.metric("🧱 Combi Safe", safe_count)
    m3.metric("Combinadas encontradas", len(combos))

    with st.expander("Ver clasificación combi de todos los picks", expanded=False):
        st.dataframe(show[[c for c in cols if c in show.columns]], width='stretch', hide_index=True)

    singles_plan, combo_stake_plan, no_play_plan, total_stake_plan, daily_cap_plan = construir_plan_apuestas_v23401(
        picks_df,
        combos,
        bankroll=bankroll_plan,
        unit_amount=unit_plan,
        max_daily_pct=daily_pct_plan,
        stake_mode=stake_mode,
    )

    st.markdown("### ✅ Plan recomendado de juego")
    p1, p2, p3 = st.columns(3)
    p1.metric("Stake total plan", f"{total_stake_plan:.2f} €")
    p2.metric("Límite diario", f"{daily_cap_plan:.2f} €")
    p3.metric("% bankroll usado", f"{(total_stake_plan / bankroll_plan * 100.0) if bankroll_plan else 0:.2f}%")

    # v23.40.2: decisión operativa clara. Evita que el usuario lea "0 combis" como fallo.
    if len(combos) == 0 and safe_count == 0 and not singles_plan.empty:
        st.error("🔴 DECISIÓN OFICIAL: NO HACER COMBINADA")
        st.success("✅ Día apto solo para SIMPLES con stake controlado. La app ha encontrado picks jugables, pero ninguno suficientemente limpio para combinada.")
    elif len(combos) == 0 and not singles_plan.empty:
        st.warning("🟡 DECISIÓN OFICIAL: jugar simples; combinada solo si ajustas cuotas/rango y aceptas más riesgo.")
    elif len(combos) > 0:
        st.success("🟢 DECISIÓN OFICIAL: simples + combinada pequeña permitida por el Combi Guard.")

    if singles_plan.empty and combo_stake_plan.empty:
        st.warning("Hoy la app no encuentra una forma limpia de jugar estos picks con stake. Mejor no forzar.")
    else:
        if not singles_plan.empty:
            st.markdown("#### 🎾 Simples recomendadas")
            st.caption("Estas son las apuestas que la app sí jugaría hoy. Si no hay combinada segura, este bloque es el plan principal.")
            st.dataframe(singles_plan[["Juego", "Partido", "Mercado", "Prob %", "Cuota", "Cuota tipo", "Stake u", "Stake €", "Motivo stake"]], width='stretch', hide_index=True)
        if not combo_stake_plan.empty:
            st.markdown("#### 🔗 Combinadas pequeñas recomendadas")
            st.caption("Complemento de stake bajo. Nunca debe pesar más que las simples.")
            st.dataframe(combo_stake_plan[["Juego", "Nº picks", "Cuota total", "Confianza global", "Stake u", "Stake €", "Pick más débil", "Motivo stake"]], width='stretch', hide_index=True)
            with st.expander("Ver composición de combinadas"):
                st.dataframe(combo_stake_plan[["Juego", "Picks"]], width='stretch', hide_index=True)

    if not no_play_plan.empty:
        with st.expander("❌ Picks que NO usaría para combinada"):
            st.dataframe(no_play_plan[[c for c in ["Partido", "Mercado", "Prob %", "Cuota", "Motivos"] if c in no_play_plan.columns]], width='stretch', hide_index=True)

    if not combos:
        st.error("❌ No hay combinada segura dentro del rango de cuota objetivo. Esto NO es un fallo: es una señal para proteger bankroll.")
        if safe_count == 0:
            st.warning("Hoy no hay ningún pick 🧱 COMBI SAFE. Plan recomendado: simples pequeñas o pasar; no forzar combinada oficial.")
        elif safe_count == 1:
            st.warning("Solo hay 1 pick 🧱 COMBI SAFE. Mejor simple o esperar.")
        else:
            st.info("Hay picks seguros, pero no encajan en la cuota objetivo. Revisa si pegaste cuotas reales o baja un poco la cuota mínima.")

        st.markdown("### 🟡 Plan B: candidatas controladas, no oficiales")
        st.caption("Este bloque NO sustituye al filtro seguro. Solo muestra combinadas con picks 🧱 COMBI SAFE + 🔥 FUERTE SIMPLE muy cercanos al mínimo, para que veas la opción menos mala sin forzar a ciegas.")

        plan_b = construir_combinadas_plan_b_v23270(
            picks_df,
            cuota_min=cuota_min,
            cuota_max=cuota_max,
            min_picks=2,
            max_picks=max_picks,
        )

        if not plan_b:
            st.warning("Tampoco hay una candidata controlada dentro del rango. Señal fuerte de NO combinar hoy en este modo.")
            top = picks_df[picks_df["Etiqueta"].isin(["🧱 COMBI SAFE", "🔥 FUERTE SIMPLE"])].copy()
            if not top.empty:
                top["Prob %"] = top["Prob"].apply(lambda x: round(float(x) * 100, 1))
                top["Mínimo %"] = top["Min"].apply(lambda x: round(float(x) * 100, 1))
                top["Falta/sobra %"] = ((top["Prob"] - top["Min"]) * 100).round(1)
                top["Cuota"] = top["Cuota"].apply(lambda x: round(float(x), 2))
                st.markdown("#### Mejores picks sueltos para simple")
                st.dataframe(
                    top.sort_values(["Etiqueta", "Score"], ascending=[True, False])[["Etiqueta", "Partido", "Mercado", "Prob %", "Mínimo %", "Falta/sobra %", "Cuota", "Cuota tipo", "Motivos"]].head(10),
                    width='stretch',
                    hide_index=True,
                )
            return

        st.warning("Estas combinadas son PLAN B: se pueden estudiar, pero no son 🧱 COMBI SAFE puras.")
        for i, combo in enumerate(plan_b[:3], start=1):
            st.markdown(f"#### 🟡 Candidata controlada #{i}")
            a, b, c, d = st.columns(4)
            a.metric("Cuota total", f"{combo['Cuota total']:.2f}")
            b.metric("Confianza global", f"{combo['Confianza global']:.1%}")
            c.metric("Nº picks", combo["Nº picks"])
            d.metric("Fuerte simple", combo["Fuerte simple"])

            combo_df = pd.DataFrame(combo["Picks"]).copy()
            combo_df["Prob %"] = combo_df["Prob"].apply(lambda x: round(float(x) * 100, 1))
            combo_df["Mínimo %"] = combo_df["Min"].apply(lambda x: round(float(x) * 100, 1))
            combo_df["Falta/sobra %"] = ((combo_df["Prob"] - combo_df["Min"]) * 100).round(1)
            combo_df["Cuota"] = combo_df["Cuota"].apply(lambda x: round(float(x), 2))
            st.dataframe(
                combo_df[["Etiqueta", "Partido", "Mercado", "Prob %", "Mínimo %", "Falta/sobra %", "Cuota", "Cuota tipo"]],
                width='stretch',
                hide_index=True,
            )
            st.warning(f"⚠️ Pick más débil: {combo['Pick más débil']}")
            st.info("Lectura: si quieres reducir el fallo por uno, esta candidata debería jugarse con stake menor o quedarse como referencia; la oficial sigue siendo NO COMBI SAFE.")
            st.divider()
        return

    st.success(f"✅ Combinadas recomendadas encontradas: {len(combos)}")

    for i, combo in enumerate(combos[:5], start=1):
        st.markdown(f"### 🧱 Combinada recomendada #{i}")
        a, b, c, d = st.columns(4)
        a.metric("Cuota total", f"{combo['Cuota total']:.2f}")
        b.metric("Confianza global", f"{combo['Confianza global']:.1%}")
        c.metric("Nº picks", combo["Nº picks"])
        d.metric("Score medio", f"{combo['Score medio']:.1f}")

        combo_df = pd.DataFrame(combo["Picks"]).copy()
        combo_df["Prob %"] = combo_df["Prob"].apply(lambda x: round(float(x) * 100, 1))
        combo_df["Cuota"] = combo_df["Cuota"].apply(lambda x: round(float(x), 2))
        st.dataframe(
            combo_df[["Partido", "Mercado", "Prob %", "Cuota", "Cuota tipo", "Etiqueta"]],
            width='stretch',
            hide_index=True
        )
        st.warning(f"⚠️ Pick más débil: {combo['Pick más débil']}")

        if combo["Nº picks"] >= 4:
            st.info("Para evitar el fallo por uno, intenta llegar a cuota parecida con 2-3 picks si es posible.")
        elif combo["Confianza global"] >= 0.50:
            st.success("Estructura limpia: pocos picks y todos superan filtro COMBI SAFE.")
        else:
            st.info("Apta, pero con riesgo acumulado. No subiría más picks.")


def wta_over_watchlist_reason(circuito, surface, fav_prob, over18, over17=None):
    """
    v23.20 WTA Over Watchlist Tight.
    Over 17.5 solo para WTA. El patrón flojo Over 18.5 72% con favorita 50-60
    deja de ser watchlist directa; se observa mejor por Over 17.5.
    """
    try:
        fav_prob = float(fav_prob or 0)
        over18 = float(over18 or 0)
        over17 = float(over17 or 0) if over17 is not None else 0.0
    except Exception:
        return ""

    if str(circuito).upper().strip() != "WTA" or str(surface).strip() != "Clay":
        return ""

    if over17 >= 0.78 and 0.50 <= fav_prob < 0.64:
        return "👀 OBSERVAR OVER 17.5: mercado WTA alto con partido igualado"
    if over17 >= 0.74 and 0.64 <= fav_prob < 0.72:
        return "👀 OBSERVAR OVER 17.5: zona ideal favorita 64-72%"
    if over18 >= 0.68 and 0.64 <= fav_prob < 0.67:
        return "👀 OBSERVAR OVER 18.5: favorita 64-67%"
    if over18 >= 0.66 and 0.68 <= fav_prob < 0.71:
        return "👀 OBSERVAR OVER 18.5: favorita 68-71%"
    return ""



def _is_challenger_context(value):
    """Devuelve True para Challenger/Qualy aunque venga escrito de varias formas."""
    c = str(value or "").upper().replace(" ", "_")
    return any(x in c for x in ["CHALLENGER", "CHALL", "QUAL_CHALL", "ITF_ATP", "QUALIFYING", "QUALY"])


def _es_entorno_challenger_match(m, circuito_ui):
    """v23.26.2: detector operativo de Challenger/Qualy.
    En la prueba v23.26.1 algunos partidos Challenger entraban como ATP y
    por eso un Over 18.5 de 73-75% todavía salía como FUERTE.
    Aquí usamos fuente + torneo + texto bruto para activar filtros prudentes.
    """
    ui = str(circuito_ui or "").upper().strip()
    if ui != "ATP":
        return False
    txt = " ".join([
        str((m or {}).get("circuito_detectado", "")),
        str((m or {}).get("torneo", "")),
        str((m or {}).get("raw_block", "")),
        str((m or {}).get("source_text", "")),
    ]).upper()
    # Challenger, ITF o previas: tratamos como entorno de alta varianza.
    return any(x in txt for x in [
        "CHALLENGER", "CHALL", " ITF", "ITF_", "QUALIFYING", "QUALY", " Q1", " Q2", " Q3"
    ])


def _row_pct(row, col, default=0.0):
    """Lee porcentajes en formato 72.5%, 0.725 o vacío."""
    try:
        v = row.get(col, default)
        if v is None or str(v).strip() == "":
            return float(default)
        s = str(v).replace("%", "").replace(",", ".").strip()
        n = float(s)
        return n / 100.0 if n > 1 else n
    except Exception:
        return float(default)


def _row_over_guard_active(row):
    """True solo cuando el Over está bloqueado.
    v23.29.1: WATCH ya no bloquea automáticamente; solo avisa/no combi visual.
    v23.36.4: en Challenger, un bloqueo por baja muestra no debe matar
    automáticamente un Over muy alto si la línea 19.5 también acompaña.
    No cambia el cálculo del Over; solo afina el selector.
    """
    label = str(row.get("Over Quality Guard", "") or "").upper()
    if "BLOQUEADO" not in label:
        return False
    circuito_txt = " ".join([
        str(row.get("Circuito fuente", "") or ""),
        str(row.get("Circuito datos", "") or ""),
        str(row.get("Circuito cálculo", "") or ""),
        str(row.get("Torneo", "") or ""),
    ])
    is_chall = _is_challenger_context(circuito_txt)
    over18 = _row_pct(row, "Over 18.5", 0.0)
    over19 = _row_pct(row, "Over 19.5", 0.0)
    motivos = str(row.get("Motivos Over Guard", "") or "").lower()
    solo_baja_muestra = (
        "confianza" in motivos
        or "muestra" in motivos
        or "rating sanity" in motivos
    ) and "favorito 2-0" not in motivos and "perfil 2-0" not in motivos
    if is_chall and solo_baja_muestra and over18 >= 0.76 and over19 >= 0.66:
        return False
    return True

def _row_over_guard_watch(row):
    label = str(row.get("Over Quality Guard", "") or "").upper()
    return "WATCH" in label

def _row_under25_adjusted(row, default=0.0):
    return _row_pct(row, "Under 2.5 ajustado", default)


def _downgrade_strong_to_apto(label, extra=""):
    txt = str(label or "")
    txt = txt.replace("🔥", "✅").replace("FUERTE", "APTO")
    return (txt + (" · " + extra if extra else "")).strip()

def over_focus_label(circuito, best_label, best_prob, set3, over17, over18, over19):
    """Etiqueta visual v23.25 para priorizar Over/3 sets sobre ML."""
    label = str(best_label or "")
    try:
        best_prob = float(best_prob or 0)
        set3 = float(set3 or 0)
        over17 = float(over17 or 0)
        over18 = float(over18 or 0)
        over19 = float(over19 or 0)
    except Exception:
        return ""

    c = str(circuito).upper().strip()
    is_chall = _is_challenger_context(c)
    if "Over 17.5" in label:
        if c == "WTA" and over17 >= 0.78:
            return "🔥 OVER 17.5 FUERTE"
        return "✅ OVER 17.5 APTO"
    if "Over 18.5" in label:
        # v23.26.1: Challenger más prudente. Antes 73% ya era FUERTE; en el backtest salieron falsos fuertes.
        if is_chall:
            if over18 >= 0.80:
                return "🔥 OVER 18.5 FUERTE"
            if over18 >= 0.76:
                return "✅ OVER 18.5 APTO"
            if over18 >= 0.70:
                return "👀 OVER 18.5 WATCH"
            if over18 >= 0.65:
                return "👀 OVER 18.5 WATCH / NO BET"
            return "ML SOLO CONTEXTO"
        if c == "ATP" and over18 >= 0.73:
            return "🔥 OVER 18.5 FUERTE"
        if over18 >= 0.68:
            return "✅ OVER 18.5 APTO"
        return "👀 OVER 18.5 WATCH"
    if "Over 19.5" in label:
        if over18 >= 0.70 and over19 >= 0.66:
            return "✅ OVER 19.5 APTO"
        if over19 >= 0.68:
            return "✅ OVER 19.5 APTO"
        return "👀 OVER 19.5 WATCH"
    if "3 sets" in label or "Partido a 3" in label:
        if set3 >= 0.48:
            return "🎯 3 SETS WATCH FUERTE"
        return "🎯 3 SETS WATCH"
    return "ML SOLO CONTEXTO" if "ML" in label else ""



def market_selector_v23263(row):
    """
    v23.26.6 Market Selector + Visual Coherence.
    No fuerza siempre Over 18.5. Clasifica el perfil del partido y propone:
    - Over 18.5 / Over 19.5
    - +2.5 sets
    - Underdog gana set
    - Favorito 2-0 / Under 2.5 sets
    - NO BET
    """
    circuito_txt = " ".join([
        str(row.get("Circuito cálculo", "")),
        str(row.get("Circuito datos", "")),
        str(row.get("Circuito fuente", "")),
        str(row.get("Torneo", "")),
        str(row.get("Estado", "")),
    ])
    is_chall = _is_challenger_context(circuito_txt)
    estado = str(row.get("Estado", ""))
    trust = str(row.get("Signal Trust", ""))
    risks = str(row.get("Riesgos", "")).lower()
    partial = ("estimado" in estado.lower()) or ("parcial" in trust.lower()) or ("fallback" in risks)

    fav = _row_pct(row, "ML favorito", 0.0)
    over18 = _row_pct(row, "Over 18.5", 0.0)
    over19 = _row_pct(row, "Over 19.5", 0.0)
    under22 = _row_pct(row, "Under 22.5", 0.0)
    set3 = _row_pct(row, "Partido a 3 sets", 0.0)
    dog_set = _row_pct(row, "Prob gana set", 0.0)
    fav20 = _row_pct(row, "Favorito 2-0", 0.0)

    favorito = str(row.get("Favorito modelo", "Favorito"))
    dog = str(row.get("Jugador gana set", "Underdog"))

    def out(market, prob, motivo):
        try:
            prob_txt = f"{float(prob):.1%}"
        except Exception:
            prob_txt = ""
        return pd.Series({
            "Mercado recomendado": market,
            "Prob mercado recomendado": prob_txt,
            "Motivo Market Selector": motivo,
        })

    # v23.29: si el Over Guard está activo, el selector no puede recomendar Over.
    over_guard_active = _row_over_guard_active(row)
    under25_adj = _row_under25_adjusted(row, 1 - set3)
    if over_guard_active:
        # v23.29.3: máximo acierto. Under 2.5 queda como WATCH hasta tener más validación.
        if under25_adj >= 0.63:
            return out("👀 WATCH UNDER 2.5 SETS", under25_adj, "Over Quality Guard activo: estudiar Under 2.5, no apuesta oficial")
        return out("❌ NO OVER / NO BET", over18, "Over Quality Guard activo: bloquear Over 18.5/19.5")

    # Fuera de Challenger, conservamos lógica prudente: el selector ayuda, no sustituye todo.
    if not is_chall:
        if over18 >= 0.78 and fav20 <= 0.54:
            return out("OVER 18.5", over18, "Over alto sin riesgo fuerte de 2-0")
        if set3 >= 0.46 and fav <= 0.64:
            return out("+2.5 SETS", set3, "partido igualado con probabilidad alta de 3 sets")
        return out("NO BET / ML SOLO CONTEXTO", fav, "sin patrón suficientemente claro")

    # 1) Bloqueo de zona mala detectada en tus backtests: Over 70-75.9% no es apuesta automática.
    zona_over_watch = 0.70 <= over18 < 0.76

    # 2) Partido dominado: mejor mirar 2-0/Under que Over.
    # v23.26.4: hacemos este detector más útil porque los fallos de Over venían de 2-0 cortos.
    perfil_2_0_fuerte = (fav >= 0.70 and fav20 >= 0.58 and (under22 >= 0.56 or over18 < 0.76))
    perfil_2_0_moderado = (fav >= 0.64 and fav20 >= 0.52 and over18 < 0.76 and set3 <= 0.45)

    if perfil_2_0_fuerte or perfil_2_0_moderado:
        if perfil_2_0_fuerte:
            market = f"✅ {favorito} 2-0 / UNDER 2.5 SETS"
            motivo = "favorito con perfil claro de 2-0; evitar Over 18.5 por riesgo de marcador corto"
        else:
            market = f"👀 WATCH {favorito} 2-0 / UNDER 2.5 SETS"
            motivo = "perfil moderado de 2-0; no forzar Over 18.5"

        if partial:
            market = market.replace("✅ ", "👀 OBSERVAR ")
            if not market.startswith("👀"):
                market = "👀 OBSERVAR " + market
            motivo += " · datos parciales"
        return out(market, fav20, motivo)

    # 3) Partido realmente largo: Over solo si es elite o si Over 19.5 acompaña.
    if over18 >= 0.80 and fav20 <= 0.52:
        return out("🔥 OVER 18.5", over18, "Over elite y riesgo de 2-0 controlado")
    if over18 >= 0.76 and over19 >= 0.62 and fav20 <= 0.54:
        return out("✅ OVER 18.5 / OVER 19.5", min(over18, over19), "Over 18.5 apto con confirmación en línea 19.5")

    # 4) Tres sets: mejor que ML si hay igualdad real.
    if fav <= 0.64 and set3 >= 0.44 and over19 >= 0.58:
        market = "+2.5 SETS"
        motivo = "igualdad real: favorito no dominante, 3 sets y Over 19.5 acompañan"
        if set3 < 0.48:
            market = "👀 WATCH +2.5 SETS"
        return out(market, set3, motivo)

    # 5) Underdog gana set: cuando el modelo detecta resistencia pero no queremos dog ML.
    if 0.38 <= (1 - fav) <= 0.48 and dog_set >= 0.46 and fav20 <= 0.52:
        market = f"{dog} GANA AL MENOS 1 SET"
        motivo = "underdog vivo: mejor buscar set ganado que ML"
        if dog_set < 0.50 or partial:
            market = "👀 WATCH " + market
            if partial:
                motivo += " · datos parciales"
        return out(market, dog_set, motivo)

    # 6) Zona Over watch: la dejamos como observación, no apuesta.
    if zona_over_watch:
        return out("👀 WATCH OVER 18.5 / NO BET", over18, "zona 70-75.9%: en tus tests falló mucho por 2-0 corto")

    # 7) Si nada destaca, no forzar apuesta.
    return out("NO BET", max(fav, over18, set3, dog_set, fav20), "sin mercado con ventaja clara según selector")


def limpiar_signal_trust_v23263(row):
    """Evita que Signal Trust muestre fuego cuando la recomendación final/selector no es fuerte."""
    trust = str(row.get("Signal Trust", ""))
    rec = str(row.get("Recomendación", ""))
    market = str(row.get("Mercado recomendado", ""))
    is_strong = ("🔥" in rec) or ("🔥" in market) or ("FUERTE" in rec.upper()) or ("FUERTE" in market.upper())
    if "🔥" in trust and not is_strong:
        return trust.replace("🔥", "👀").replace("SPOT FUERTE", "SPOT WATCH").replace("FUERTE", "WATCH")
    return trust


def alinear_market_selector_v23266(row):
    """
    v23.26.6 Visual Coherence.
    La recomendación final manda siempre. El mercado recomendado se reconstruye
    desde la propia recomendación para evitar contradicciones tipo:
      Recomendación = ✅ OVER 19.5 APTO
      Mercado recomendado = +2.5 SETS
    No modifica probabilidades ni filtros; solo limpia la señal visual/export.
    """
    rec = str(row.get("Recomendación", "")).strip()
    market = str(row.get("Mercado recomendado", "")).strip()
    prob = str(row.get("Prob mercado recomendado", "")).strip()
    motivo = str(row.get("Motivo Market Selector", "")).strip()
    rec_u = rec.upper()
    trust_u = str(row.get("Signal Trust", "")).upper()
    risk_u = str(row.get("Riesgos", "")).upper()
    force_watch_signal = (
        "SPOT WATCH" in trust_u
        or "DATOS PARCIALES" in trust_u
        or "FALLBACK" in risk_u
        or "DATOS INCOMPLETOS" in risk_u
    )

    # v23.36.4 Challenger Over Selector Fix
    # Problema detectado en los Excel de hoy: muchos Challenger venían con
    # Confianza mínima=25% y Mín. partidos superficie=0, lo que convertía casi
    # todos los Over APTO en WATCH aunque la lectura de juegos fuera alta.
    # No se toca el motor Over ni sus probabilidades: solo impedimos que el
    # aviso genérico de baja muestra degrade automáticamente un Over Challenger
    # realmente alto. El Over Guard BLOQUEADO sigue mandando.
    circuito_txt_fix = " ".join([
        str(row.get("Circuito fuente", "") or ""),
        str(row.get("Circuito datos", "") or ""),
        str(row.get("Circuito cálculo", "") or ""),
        str(row.get("Torneo", "") or ""),
    ])
    is_chall_fix = _is_challenger_context(circuito_txt_fix)
    over18_fix = _row_pct(row, "Over 18.5", 0.0)
    over19_fix = _row_pct(row, "Over 19.5", 0.0)
    challenger_over_release = (
        is_chall_fix
        and not _row_over_guard_active(row)
        and (
            over18_fix >= 0.78
            or (over18_fix >= 0.76 and over19_fix >= 0.68)
        )
    )

    def fmt_prob(col):
        p = _row_pct(row, col, None)
        try:
            if p is None or pd.isna(p):
                return prob
            return f"{float(p):.1%}"
        except Exception:
            return prob

    def out(m, p="", extra=""):
        base = motivo
        if extra:
            base = (base + " · " if base else "") + extra
        return pd.Series({
            "Mercado recomendado": m,
            "Prob mercado recomendado": p,
            "Motivo Market Selector": base,
        })

    # v23.29: Over Guard manda también en coherencia visual.
    if _row_over_guard_active(row):
        under25_adj = _row_under25_adjusted(row, _row_pct(row, "Under 2.5 ajustado", 0.0))
        # v23.29.3: Under 2.5 no es pick oficial todavía.
        if under25_adj >= 0.63:
            return out("👀 WATCH UNDER 2.5 SETS", f"{under25_adj:.1%}", "Over Quality Guard activo: watch, no apuesta oficial")
        return out("❌ NO OVER / NO BET", "", "Over Quality Guard activo")

    # 1) Bloqueos finales: no puede aparecer ningún mercado con ✅/🔥.
    if "NO BET" in rec_u:
        return out("❌ NO BET", "", "alineado con recomendación final")

    if "ML SOLO CONTEXTO" in rec_u or "SOLO CONTEXTO" in rec_u:
        return out("👀 SOLO CONTEXTO / OBSERVAR", "", "alineado con recomendación final")

    # 2) WATCH/OBSERVAR: el mercado recomendado debe copiar el tipo de watch de la recomendación.
    if "WATCH" in rec_u or "OBSERVAR" in rec_u or rec.startswith("👀"):
        if "OVER 19.5" in rec_u:
            return out("👀 WATCH OVER 19.5", fmt_prob("Over 19.5"), "alineado como watch, no apuesta")
        if "OVER 18.5" in rec_u:
            return out("👀 WATCH OVER 18.5", fmt_prob("Over 18.5"), "alineado como watch, no apuesta")
        if "OVER 17.5" in rec_u:
            return out("👀 WATCH OVER 17.5", fmt_prob("Over 17.5"), "alineado como watch, no apuesta")
        if "2-0" in rec_u or "UNDER 2.5" in rec_u:
            favorito = str(row.get("Favorito modelo", "Favorito")) or "Favorito"
            return out(f"👀 WATCH {favorito} 2-0 / UNDER 2.5 SETS", fmt_prob("Favorito 2-0"), "alineado como watch, no apuesta")
        if "3 SET" in rec_u or "3 SETS" in rec_u or "+2.5" in rec_u:
            return out("👀 WATCH +2.5 SETS", fmt_prob("Partido a 3 sets"), "alineado como watch, no apuesta")
        return out("👀 SOLO OBSERVAR", "", "alineado como watch, no apuesta")

    # 3) Recomendaciones positivas: Mercado recomendado debe ser exactamente el mismo tipo de mercado.
    emoji = "🔥" if (rec.startswith("🔥") or "FUERTE" in rec_u) else "✅"

    if "OVER 19.5" in rec_u:
        if force_watch_signal and not challenger_over_release:
            return out("👀 WATCH OVER 19.5", fmt_prob("Over 19.5"), "Signal Trust en watch/datos parciales: no apuesta fuerte")
        return out(f"{emoji} OVER 19.5", fmt_prob("Over 19.5"), "mercado alineado con recomendación final" + (" · v23.36.4 release Challenger Over" if challenger_over_release else ""))

    if "OVER 18.5" in rec_u:
        if force_watch_signal and not challenger_over_release:
            return out("👀 WATCH OVER 18.5", fmt_prob("Over 18.5"), "Signal Trust en watch/datos parciales: no apuesta fuerte")
        return out(f"{emoji} OVER 18.5", fmt_prob("Over 18.5"), "mercado alineado con recomendación final" + (" · v23.36.4 release Challenger Over" if challenger_over_release else ""))

    if "OVER 17.5" in rec_u:
        return out(f"{emoji} OVER 17.5", fmt_prob("Over 17.5"), "mercado alineado con recomendación final")

    if "2-0" in rec_u or "UNDER 2.5" in rec_u:
        favorito = str(row.get("Favorito modelo", "Favorito")) or "Favorito"
        fav20_now = _row_pct(row, "Favorito 2-0", 0.0)
        over18_now = _row_pct(row, "Over 18.5", 0.0)
        if fav20_now >= 0.70 and over18_now <= 0.68 and emoji == "🔥":
            return out(f"{emoji} {favorito} 2-0 / UNDER 2.5 SETS", fmt_prob("Favorito 2-0"), "mercado alineado con recomendación final")
        return out(f"👀 WATCH {favorito} 2-0 / UNDER 2.5 SETS", fmt_prob("Favorito 2-0"), "Under 2.5 prudente: watch hasta validar")

    if "3 SET" in rec_u or "3 SETS" in rec_u or "+2.5" in rec_u:
        return out(f"{emoji} +2.5 SETS", fmt_prob("Partido a 3 sets"), "mercado alineado con recomendación final")

    # 4) Fallback: si no sabemos mapearlo, quitamos contradicciones obvias.
    if not market:
        market = rec if rec else "NO BET"
    return out(market, prob, "revisión de coherencia v23.26.6")


# =========================================================
# v23.30 ML QUALITY GUARD
# Objetivo: añadir ML solo cuando el favorito sea realmente fiable.
# Mantiene Over Guard como prioridad: un Over oficial limpio no se sustituye.
# =========================================================

def _is_official_over_market_v23300(market):
    mu = str(market or "").upper()
    return (mu.strip().startswith("✅") or mu.strip().startswith("🔥")) and "OVER" in mu and "WATCH" not in mu and "NO COMBI" not in mu


def ml_quality_guard_v23300(row):
    fav = _row_pct(row, "ML favorito", 0.0)
    fav20 = _row_pct(row, "Favorito 2-0", 0.0)
    dog_set = _row_pct(row, "Prob gana set", 0.0)
    set3 = _row_pct(row, "Partido a 3 sets", 0.0)
    min_conf = _row_pct(row, "Confianza mínima", 0.0)
    gap = _row_pct(row, "Gap Elo/modelo", 0.0)
    try:
        min_surface = int(float(str(row.get("Mín. partidos superficie", "0") or 0).replace(",", ".")))
    except Exception:
        min_surface = 0

    favorito = str(row.get("Favorito modelo", "Favorito") or "Favorito").strip()
    estado = str(row.get("Estado", "") or "")
    trust = str(row.get("Signal Trust", "") or "").upper()
    risks = str(row.get("Riesgos", "") or "").upper()
    aviso = str(row.get("Aviso datos", "") or "").upper()
    surface = str(row.get("Superficie", "") or "").upper()
    circuito_txt = " ".join([
        str(row.get("Circuito fuente", "") or ""),
        str(row.get("Circuito datos", "") or ""),
        str(row.get("Circuito cálculo", "") or ""),
        str(row.get("Torneo", "") or ""),
    ]).upper()

    is_wta = "WTA" in circuito_txt
    is_chall = "CHALL" in circuito_txt
    is_qualy = "QUAL" in circuito_txt or "PREVIA" in circuito_txt
    is_partial = (
        "ESTIMADO" in estado.upper()
        or "FALLBACK" in risks
        or "FALLBACK" in aviso
        or "DATOS PARCIALES" in trust
        or "DATOS INCOMPLETOS" in risks
    )

    # Umbrales base. Son duros porque el ML entra para mejorar combinadas, no para generar volumen.
    min_ml = 0.72
    min_conf_req = 0.65
    min_surface_req = 10

    if is_wta:
        min_ml = 0.76
        min_conf_req = 0.68
        min_surface_req = 12
    if is_qualy:
        min_ml = max(min_ml, 0.75)
        min_conf_req = max(min_conf_req, 0.68)
        min_surface_req = max(min_surface_req, 12)
    if is_chall:
        min_ml = max(min_ml, 0.78)
        min_conf_req = max(min_conf_req, 0.70)
        min_surface_req = max(min_surface_req, 15)

    reasons = []
    block = False
    watch = False

    if fav <= 0:
        return {
            "label": "",
            "market": "",
            "prob": fav,
            "reasons": "sin probabilidad ML",
            "official": False,
            "watch": False,
        }

    if fav < min_ml:
        watch = fav >= (min_ml - 0.04)
        reasons.append(f"ML {fav:.1%} por debajo del mínimo {min_ml:.1%}")
        if not watch:
            block = True

    if min_conf < min_conf_req:
        reasons.append(f"confianza {min_conf:.0%} < {min_conf_req:.0%}")
        if min_conf < 0.50:
            block = True
        else:
            watch = True

    if min_surface < min_surface_req:
        reasons.append(f"muestra superficie {min_surface} < {min_surface_req}")
        if min_surface < 5:
            block = True
        else:
            watch = True

    if gap >= 0.18:
        reasons.append("gap Elo/modelo >=18%")
        block = True
    elif gap >= 0.12:
        reasons.append("gap Elo/modelo elevado")
        watch = True

    if is_partial:
        reasons.append("datos parciales/fallback")
        block = True

    # Rival con mucha probabilidad de set = favorito menos cómodo. No siempre bloquea, pero sí evita oficial si el ML no es alto.
    if dog_set >= 0.65 and fav < 0.76:
        reasons.append("underdog gana set alto")
        watch = True
    if set3 >= 0.48 and fav < 0.76:
        reasons.append("partido a 3 sets elevado")
        watch = True

    # ML puede ser oficial si el favorito tiene perfil de 2-0 o control, pero no exigimos 2-0.
    control_bonus = fav20 >= 0.58 or dog_set <= 0.38 or set3 <= 0.40
    if control_bonus:
        reasons.append("perfil de control del favorito")

    if block:
        return {
            "label": "🚫 ML BLOQUEADO",
            "market": f"🚫 NO ML {favorito}",
            "prob": fav,
            "reasons": " · ".join(reasons),
            "official": False,
            "watch": False,
        }

    if fav >= min_ml and not watch:
        return {
            "label": "✅ ML OFICIAL",
            "market": f"✅ ML {favorito}",
            "prob": fav,
            "reasons": " · ".join(reasons) if reasons else "favorito fiable: probabilidad, datos y Elo coherentes",
            "official": True,
            "watch": False,
        }

    if fav >= (min_ml - 0.04):
        return {
            "label": "👀 WATCH ML",
            "market": f"👀 WATCH ML {favorito}",
            "prob": fav,
            "reasons": " · ".join(reasons) if reasons else "cerca de ML oficial, pero no suficientemente limpio",
            "official": False,
            "watch": True,
        }

    return {
        "label": "❌ NO ML",
        "market": f"❌ NO ML {favorito}",
        "prob": fav,
        "reasons": " · ".join(reasons) if reasons else "ML insuficiente",
        "official": False,
        "watch": False,
    }




def pick_oficial_v23301(row):
    """
    v23.30.1: columna visual para pantalla/export.
    Devuelve solo picks oficiales reales: Over oficial o ML oficial.
    Los WATCH/NO BET/NO OVER/NO ML quedan en blanco para no confundir.
    """
    market = str(row.get("Mercado recomendado", "") or "").strip()
    prob = str(row.get("Prob mercado recomendado", "") or "").strip()
    if not market:
        return ""

    mu = market.upper()
    bad_tokens = [
        "WATCH", "NO BET", "NO OVER", "NO ML", "BLOQUEADO",
        "SOLO CONTEXTO", "OBSERVAR", "REVISAR", "DUDOS"
    ]
    if any(tok in mu for tok in bad_tokens):
        return ""

    is_positive = market.startswith("✅") or market.startswith("🔥") or market.startswith("🧱")
    is_supported_market = ("OVER" in mu) or ("ML" in mu) or ("GANADOR" in mu)

    if is_positive and is_supported_market:
        return f"{market} ({prob})" if prob else market

    return ""


# =========================================================
# v23.32 TELEGRAM PICKS SENDER
# Lee TELEGRAM_BOT_TOKEN y TELEGRAM_CHAT_ID desde .streamlit/secrets.toml
# y permite enviar picks oficiales + recomendados/watch sin poner datos en pantalla.
# =========================================================

def _telegram_get_secret(key, default=""):
    try:
        return str(st.secrets.get(key, default) or "").strip()
    except Exception:
        return default


def _telegram_html_escape(x):
    s = str(x if x is not None else "")
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
    )


def _telegram_str(row, col, default=""):
    try:
        v = row.get(col, default)
    except Exception:
        v = default
    if pd.isna(v):
        return default
    return str(v).strip()


def _telegram_pct_text(row, col):
    raw = _telegram_str(row, col, "")
    if raw == "":
        return ""
    try:
        val = float(str(raw).replace("%", "").replace(",", "."))
        if 0 <= val <= 1:
            val *= 100
        return f"{val:.1f}%"
    except Exception:
        return raw


def _telegram_is_watch_recomendado(row):
    """Devuelve True para WATCH/PREMIUM relevantes y False para NO BET/NO OVER/etc."""
    market = _telegram_str(row, "Mercado recomendado", "")
    pick = _telegram_str(row, "Pick oficial", "")
    if pick:
        return False
    if not market:
        return False

    txt = " ".join([
        market,
        _telegram_str(row, "Motivo Market Selector", ""),
        _telegram_str(row, "WTA Over17 Official Guard", ""),
        _telegram_str(row, "Over Quality Guard", ""),
        _telegram_str(row, "ML Quality Guard", ""),
    ]).upper()

    hard_bad = ["NO BET", "NO OVER", "NO ML", "BLOQUEADO", "DESCART", "SOLO CONTEXTO"]
    if any(x in txt for x in hard_bad):
        # Excepción: si explícitamente es Premium Watch, sí interesa verlo.
        if "PREMIUM WATCH" not in txt:
            return False

    good_watch = ["WATCH", "PREMIUM", "RECOMEND", "OVER 17.5"]
    return any(x in txt for x in good_watch)


def _telegram_row_line(row, idx=None, oficial=True):
    partido = _telegram_str(row, "Partido", "Partido sin nombre")
    hora = _telegram_str(row, "Hora", "")
    if oficial:
        pick = _telegram_str(row, "Pick oficial", "")
        prob = _telegram_pct_text(row, "Prob mercado recomendado")
    else:
        pick = _telegram_str(row, "Mercado recomendado", "")
        prob = _telegram_pct_text(row, "Prob mercado recomendado")
    conf = _telegram_pct_text(row, "Confianza mínima")
    sup = _telegram_str(row, "Mín. partidos superficie", "")
    motivo = _telegram_str(row, "Motivo Market Selector", "")
    guard_wta = _telegram_str(row, "WTA Over17 Official Guard", "")
    guard_over = _telegram_str(row, "Over Quality Guard", "")
    guard_ml = _telegram_str(row, "ML Quality Guard", "")

    prefix = f"{idx}. " if idx is not None else ""
    lineas = []
    title = f"{prefix}<b>{_telegram_html_escape(partido)}</b>"
    if hora:
        title += f" · {_telegram_html_escape(hora)}"
    lineas.append(title)

    if pick:
        lineas.append(f"Pick: {_telegram_html_escape(pick)}")
    if prob:
        lineas.append(f"Prob: {_telegram_html_escape(prob)}")
    extras = []
    if conf:
        extras.append(f"Conf {_telegram_html_escape(conf)}")
    if sup:
        extras.append(f"Sup {_telegram_html_escape(sup)}")
    if extras:
        lineas.append(" · ".join(extras))

    if not oficial:
        visible_motivos = " · ".join([x for x in [guard_wta, guard_over, guard_ml, motivo] if x])
        if visible_motivos:
            visible_motivos = visible_motivos[:220]
            lineas.append(f"Motivo: {_telegram_html_escape(visible_motivos)}")

    return "\n".join(lineas)


def _telegram_text_norm(x):
    try:
        if x is None or pd.isna(x):
            return ""
    except Exception:
        pass
    return re.sub(r"[^A-Z0-9]+", "", str(x).upper())


def _telegram_prob_max_text(row):
    for col in ["🎯 Prob máxima", "Prob.", "Prob"]:
        txt = _telegram_pct_text(row, col)
        if txt:
            return txt
    return ""


def _telegram_is_nuevo_jugar(row):
    """Nuevo selector de máximo acierto: filas marcadas como ✅ JUGAR."""
    accion = _telegram_str(row, "🎯 Acción final", "") or _telegram_str(row, "Acción", "")
    mercado = _telegram_str(row, "🎯 Mercado más probable", "") or _telegram_str(row, "Mercado", "")
    if "JUGAR" not in accion.upper():
        return False
    if not mercado:
        return False
    if "EVITAR" in mercado.upper() or mercado.startswith("⛔"):
        return False
    return True


def _telegram_es_duplicado_oficial(row):
    """Evita repetir un nuevo pick si ya coincide con el Pick oficial del mismo partido."""
    oficial = _telegram_str(row, "Pick oficial", "")
    nuevo = _telegram_str(row, "🎯 Mercado más probable", "") or _telegram_str(row, "Mercado", "")
    if not oficial or not nuevo:
        return False
    no = _telegram_text_norm(oficial)
    nn = _telegram_text_norm(nuevo)
    if not no or not nn:
        return False
    return (no in nn) or (nn in no) or ("OVER185" in no and "OVER185" in nn)


def _telegram_nuevo_row_line(row, idx=None):
    partido = _telegram_str(row, "Partido", "Partido sin nombre")
    hora = _telegram_str(row, "Hora", "")
    mercado = _telegram_str(row, "🎯 Mercado más probable", "") or _telegram_str(row, "Mercado", "")
    prob = _telegram_prob_max_text(row)
    decision = _telegram_str(row, "🎯 Decisión acierto", "") or _telegram_str(row, "Confianza", "")
    estado_extra = _telegram_str(row, "📥 Estado Datos extra", "") or _telegram_str(row, "Estado Datos extra", "")
    ajuste_extra = _telegram_str(row, "📥 Ajuste Datos extra", "") or _telegram_str(row, "Ajuste Datos extra", "")
    motivo = _telegram_str(row, "🎯 Motivo acierto", "") or _telegram_str(row, "Motivo", "")
    motivo_extra = _telegram_str(row, "📥 Motivo Datos extra", "") or _telegram_str(row, "Motivo Datos extra", "")

    prefix = f"{idx}. " if idx is not None else ""
    lineas = []
    title = f"{prefix}<b>{_telegram_html_escape(partido)}</b>"
    if hora:
        title += f" · {_telegram_html_escape(hora)}"
    lineas.append(title)
    if mercado:
        lineas.append(f"Pick nuevo: {_telegram_html_escape(mercado)}")
    if prob:
        lineas.append(f"Prob: {_telegram_html_escape(prob)}")
    extras = [x for x in [decision, estado_extra, ajuste_extra] if x]
    if extras:
        lineas.append(_telegram_html_escape(" · ".join(extras))[:260])
    visible_motivos = " · ".join([x for x in [motivo_extra, motivo] if x])
    if visible_motivos:
        lineas.append(f"Motivo: {_telegram_html_escape(visible_motivos[:240])}")
    return "\n".join(lineas)


def _telegram_get_oficiales_y_nuevos(ok_df):
    df = ok_df.copy() if ok_df is not None else pd.DataFrame()
    if df.empty:
        return df, df
    if "Pick oficial" not in df.columns:
        try:
            df["Pick oficial"] = df.apply(pick_oficial_v23301, axis=1)
        except Exception:
            df["Pick oficial"] = ""
    oficiales = df[df["Pick oficial"].astype(str).str.strip() != ""].copy()
    nuevos = df[df.apply(_telegram_is_nuevo_jugar, axis=1)].copy()
    if not nuevos.empty:
        nuevos = nuevos[~nuevos.apply(_telegram_es_duplicado_oficial, axis=1)].copy()
    return oficiales, nuevos


def construir_mensaje_telegram_picks(ok_df, incluir_watch=False, max_watch=12, incluir_nuevos=False, solo_nuevos=False):
    if ok_df is None or ok_df.empty:
        return "🎾 <b>Tennis IA</b>\nNo hay análisis disponible."

    df = ok_df.copy()
    oficiales, nuevos = _telegram_get_oficiales_y_nuevos(df)
    watch = df[df.apply(_telegram_is_watch_recomendado, axis=1)].copy() if incluir_watch else pd.DataFrame()

    partes = []
    partes.append("🎾 <b>Tennis IA — Picks</b>")

    if not solo_nuevos:
        partes.append("\n🎯 <b>PICKS OFICIALES</b>")
        if oficiales.empty:
            partes.append("No hay picks oficiales. No forzar combinada.")
        else:
            for i, (_, row) in enumerate(oficiales.iterrows(), start=1):
                partes.append(_telegram_row_line(row, i, oficial=True))

    if incluir_nuevos or solo_nuevos:
        partes.append("\n🧠 <b>NUEVOS — MÁXIMO ACIERTO / TA</b>")
        if nuevos.empty:
            partes.append("No hay picks nuevos marcados como ✅ JUGAR.")
        else:
            for i, (_, row) in enumerate(nuevos.iterrows(), start=1):
                partes.append(_telegram_nuevo_row_line(row, i))

    if incluir_watch and not solo_nuevos:
        partes.append("\n👀 <b>RECOMENDADOS / WATCH</b>")
        if watch.empty:
            partes.append("No hay watch/recomendados relevantes.")
        else:
            for i, (_, row) in enumerate(watch.head(max_watch).iterrows(), start=1):
                partes.append(_telegram_row_line(row, i, oficial=False))
            if len(watch) > max_watch:
                partes.append(f"… y {len(watch) - max_watch} watch más en la app/Excel.")

    return "\n\n".join(partes)


def enviar_telegram_mensaje(mensaje):
    token = _telegram_get_secret("TELEGRAM_BOT_TOKEN", "")
    chat_id = _telegram_get_secret("TELEGRAM_CHAT_ID", "")

    if not token or not chat_id:
        return False, "Falta TELEGRAM_BOT_TOKEN o TELEGRAM_CHAT_ID en .streamlit/secrets.toml"

    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": mensaje,
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }

    try:
        r = requests.post(url, data=payload, timeout=12)
        if r.status_code == 200:
            return True, "✅ Mensaje enviado a Telegram."
        return False, f"❌ Error Telegram {r.status_code}: {r.text[:400]}"
    except Exception as e:
        return False, f"❌ Error enviando Telegram: {e}"


def render_telegram_sender_panel(ok_df):
    st.markdown("### 📲 Telegram")

    token_ok = bool(_telegram_get_secret("TELEGRAM_BOT_TOKEN", ""))
    chat_ok = bool(_telegram_get_secret("TELEGRAM_CHAT_ID", ""))

    c1, c2, c3 = st.columns(3)
    c1.metric("Bot token", "OK" if token_ok else "Falta")
    c2.metric("Chat ID", "OK" if chat_ok else "Falta")
    c3.metric("Estado", "Listo" if token_ok and chat_ok else "Configurar")

    if not (token_ok and chat_ok):
        st.info('Añade TELEGRAM_BOT_TOKEN y TELEGRAM_CHAT_ID en `.streamlit/secrets.toml`.')
        return

    oficiales_df, nuevos_df = _telegram_get_oficiales_y_nuevos(ok_df)
    c4, c5 = st.columns(2)
    c4.metric("Oficiales", len(oficiales_df))
    c5.metric("Nuevos ✅ JUGAR", len(nuevos_df))

    incluir_nuevos_preview = st.toggle("Incluir nuevos ✅ JUGAR en vista previa", value=True, key="tg_preview_nuevos")
    incluir_watch_preview = st.toggle("Incluir watch/recomendados en vista previa", value=False, key="tg_preview_watch")
    mensaje_preview = construir_mensaje_telegram_picks(
        ok_df,
        incluir_watch=incluir_watch_preview,
        max_watch=12,
        incluir_nuevos=incluir_nuevos_preview,
    )
    with st.expander("👀 Vista previa del mensaje", expanded=True):
        st.text_area("Mensaje que se enviará", value=mensaje_preview, height=320, key="tg_preview_text", disabled=True)

    b1, b2, b3, b4 = st.columns(4)
    with b1:
        if st.button("📲 Enviar prueba", key="tg_test_btn"):
            ok, msg = enviar_telegram_mensaje("✅ <b>Prueba Tennis IA</b>\nTelegram está conectado correctamente.")
            st.success(msg) if ok else st.error(msg)

    with b2:
        if st.button("🎯 Enviar oficiales", key="tg_official_btn"):
            mensaje = construir_mensaje_telegram_picks(ok_df, incluir_watch=False, incluir_nuevos=False)
            ok, msg = enviar_telegram_mensaje(mensaje)
            st.success(msg) if ok else st.error(msg)

    with b3:
        if st.button("🧠 Enviar nuevos", key="tg_new_btn"):
            mensaje = construir_mensaje_telegram_picks(ok_df, solo_nuevos=True, incluir_nuevos=True)
            ok, msg = enviar_telegram_mensaje(mensaje)
            st.success(msg) if ok else st.error(msg)

    with b4:
        if st.button("🎯🧠 Oficiales + nuevos", key="tg_official_new_btn"):
            mensaje = construir_mensaje_telegram_picks(ok_df, incluir_watch=False, incluir_nuevos=True)
            ok, msg = enviar_telegram_mensaje(mensaje)
            st.success(msg) if ok else st.error(msg)

    if st.button("🎯🧠👀 Oficiales + nuevos + watch", key="tg_all_btn"):
        mensaje = construir_mensaje_telegram_picks(ok_df, incluir_watch=True, max_watch=12, incluir_nuevos=True)
        ok, msg = enviar_telegram_mensaje(mensaje)
        st.success(msg) if ok else st.error(msg)


# =========================================================
# v23.31.2 WTA OVER 17.5 PREMIUM WATCH GUARD
# Objetivo: WTA no usa ML ni Over 18.5 como mercado principal.
# El Over 17.5 puede subir a PREMIUM WATCH si es >=82% y limpio,
# pero NO se convierte todavía en pick oficial.
# =========================================================

def aplicar_wta_over17_oficial_v23310(row):
    current_market = str(row.get("Mercado recomendado", "") or "")
    current_prob = str(row.get("Prob mercado recomendado", "") or "")
    current_motivo = str(row.get("Motivo Market Selector", "") or "")

    circuito_txt = " ".join([
        str(row.get("Circuito cálculo", "") or ""),
        str(row.get("Circuito datos", "") or ""),
        str(row.get("Circuito fuente", "") or ""),
        str(row.get("Torneo", "") or ""),
        str(row.get("Estado", "") or ""),
    ]).upper()

    is_wta = "WTA" in circuito_txt
    over17 = _row_pct(row, "Over 17.5", 0.0)
    over18 = _row_pct(row, "Over 18.5", 0.0)
    fav = _row_pct(row, "ML favorito", 0.0)
    fav20 = _row_pct(row, "Favorito 2-0", 0.0)
    set3 = _row_pct(row, "Partido a 3 sets", 0.0)
    dog_set = _row_pct(row, "Prob gana set", 0.0)
    min_conf = _row_pct(row, "Confianza mínima", 0.0)
    gap = _row_pct(row, "Gap Elo/modelo", 0.0)
    try:
        min_surface = int(float(str(row.get("Mín. partidos superficie", "0") or 0).replace(",", ".")))
    except Exception:
        min_surface = 0

    trust = str(row.get("Signal Trust", "") or "").upper()
    risks = str(row.get("Riesgos", "") or "").upper()
    aviso = str(row.get("Aviso datos", "") or "").upper()
    estado = str(row.get("Estado", "") or "").upper()

    def out(market, prob=None, extra="", wta_label=""):
        prob_txt = current_prob
        if prob is not None:
            try:
                prob_txt = f"{float(prob):.1%}"
            except Exception:
                pass
        motivo = current_motivo
        if extra:
            motivo = (motivo + " · " if motivo else "") + extra
        return pd.Series({
            "Mercado recomendado": market,
            "Prob mercado recomendado": prob_txt,
            "Motivo Market Selector": motivo,
            "WTA Over17 Official Guard": wta_label,
        })

    if not is_wta:
        return out(current_market, None, "", "")

    # Si ya hay un Over oficial ATP/WTA muy claro, no lo pisamos salvo que sea WTA Over 18.5:
    # en WTA preferimos la línea 17.5 como mercado principal cuando cumple filtro.
    mu = current_market.upper()
    motivo_u = current_motivo.upper()

    # v23.31.1 prudente:
    # No convertir en pick oficial una señal que el propio selector/Trust ya marcó como WATCH o DUDOSA.
    # Esto evita que un Over17 alto, pero nacido de una lectura débil, aparezca como apuesta oficial.
    señal_base_dudosa = (
        "WATCH" in mu
        or "WATCH" in motivo_u
        or "SPOT DUDOSO" in trust
        or "DUDOSO" in trust
        or "DUDOSA" in trust
        or "SPOT WATCH" in trust
        or "NO COMBI" in mu
        or "NO COMBI" in motivo_u
    )
    if señal_base_dudosa:
        if over17 >= 0.77:
            return out("👀 WATCH OVER 17.5", over17, "WTA v23.31.1: Over17 alto, pero señal base WATCH/DUDOSA; no oficial", "👀 WTA O17 WATCH")
        return out(current_market, None, "WTA v23.31.1: sin Over17 oficial por señal base dudosa", "")

    # Bloqueos duros: si Over Guard está activo o hay datos claramente pobres, no oficializamos.
    if _row_over_guard_active(row):
        return out(current_market, None, "WTA v23.31: Over17 no oficial por Over Quality Guard", "🚫 WTA O17 BLOQUEADO")

    datos_malos = (
        min_conf < 0.55
        or min_surface < 6
        or gap >= 0.18
        or "FALLBACK" in risks
        or "FALLBACK" in aviso
        or "DATOS PARCIALES" in trust
        or "DATOS INCOMPLETOS" in risks
        or "ESTIMADO" in estado
    )
    if datos_malos:
        if over17 >= 0.77:
            return out("👀 WATCH OVER 17.5", over17, "WTA v23.31: Over17 alto pero datos insuficientes", "👀 WTA O17 WATCH")
        return out(current_market, None, "WTA v23.31: sin Over17 oficial por datos", "")

    # Riesgo de 2-0 dominante: en WTA el 17.5 aguanta más que el 18.5, pero no queremos
    # oficializar si el favorito tiene perfil de paseo y no hay apoyo de partido largo.
    perfil_largo = (set3 >= 0.45) or (dog_set >= 0.50) or (over18 >= 0.66)
    riesgo_2_0 = (fav >= 0.70 and fav20 >= 0.58 and not perfil_largo) or (fav20 >= 0.64 and set3 < 0.42)
    if riesgo_2_0:
        if over17 >= 0.80:
            return out("👀 WATCH OVER 17.5 / RIESGO 2-0", over17, "WTA v23.31: Over17 alto pero riesgo de 2-0 dominante", "👀 WTA O17 WATCH")
        return out(current_market, None, "WTA v23.31: riesgo 2-0, sin Over17 oficial", "")

    # v23.31.2: todavía no oficializamos WTA.
    # Separación visual: Over17 limpio >=82% pasa a PREMIUM WATCH.
    if over17 >= 0.82:
        return out("🟢 PREMIUM WATCH OVER 17.5", over17, "WTA v23.31.2: Over17 >=82% con datos limpios; premium watch, no oficial", "🟢 WTA O17 PREMIUM WATCH")
    if over17 >= 0.80:
        return out("👀 WATCH OVER 17.5", over17, "WTA v23.31.2: Over17 80%-81.9%; watch normal, no oficial", "👀 WTA O17 WATCH")
    if over17 >= 0.77:
        return out("👀 WATCH OVER 17.5", over17, "WTA v23.31.2: Over17 watch 77%-79.9%", "👀 WTA O17 WATCH")

    return out(current_market, None, "WTA v23.31: Over17 sin umbral oficial", "")

def aplicar_ml_quality_guard_v23300(row):
    """Capa final: añade ML oficial/watch sin pisar Over oficial limpio."""
    current_market = str(row.get("Mercado recomendado", "") or "")
    current_prob = str(row.get("Prob mercado recomendado", "") or "")
    current_motivo = str(row.get("Motivo Market Selector", "") or "")
    ml = ml_quality_guard_v23300(row)

    def out(market, prob=None, motivo_extra=""):
        prob_txt = current_prob
        if prob is not None:
            try:
                prob_txt = f"{float(prob):.1%}"
            except Exception:
                pass
        motivo = current_motivo
        if motivo_extra:
            motivo = (motivo + " · " if motivo else "") + motivo_extra
        return pd.Series({
            "Mercado recomendado": market,
            "Prob mercado recomendado": prob_txt,
            "Motivo Market Selector": motivo,
            "ML Quality Guard": ml.get("label", ""),
            "Motivos ML Guard": ml.get("reasons", ""),
        })

    # Si hay Over oficial limpio, lo mantenemos. El ML queda visible en sus columnas.
    if _is_official_over_market_v23300(current_market):
        return out(current_market, None, "ML guard evaluado; se mantiene Over oficial")

    mu = current_market.upper()

    # v23.31.2: WTA Over17 Premium Watch debe verse en pantalla/export,
    # no lo sustituimos por ML aunque el ML Guard detecte algo.
    if "PREMIUM WATCH OVER 17.5" in mu:
        return out(current_market, None, "ML guard evaluado; se mantiene WTA Over17 Premium Watch")
    current_is_official = (current_market.strip().startswith("✅") or current_market.strip().startswith("🔥")) and "WATCH" not in mu

    # Si no hay pick oficial y el ML sí es oficial, lo proponemos.
    if ml.get("official"):
        return out(ml.get("market"), ml.get("prob"), "v23.30: ML oficial añadido por ML Quality Guard")

    # Si la app no tiene apuesta clara, podemos enseñar Watch ML.
    no_pick_context = any(x in mu for x in ["NO BET", "SOLO CONTEXTO", "OBSERVAR", "WATCH", "NO OVER"])
    if ml.get("watch") and no_pick_context and not current_is_official:
        return out(ml.get("market"), ml.get("prob"), "v23.30: ML solo watch, no combinada")

    return out(current_market, None, "ML guard evaluado")

def aplicar_max_acierto_v23294(row):
    """
    v23.30.5 Máximo acierto tuned:
    - Si Signal Trust es WATCH, ningún Over queda como pick oficial.
    - Over 19.5 oficial solo si probabilidad >=69%; 66%-68.9% pasa a watch/no combi.
    - Si un Over oficial tiene Favorito 2-0 >=60% y no hay perfil largo fuerte, baja a WATCH.
    - Perfil 2-0 moderado con favorito >=64% y Fav 2-0 alto pasa a watch.
    - Under 2.5 nunca queda como ✅ oficial; solo watch hasta validar más muestras.
    """
    market = str(row.get("Mercado recomendado", "") or "")
    motivo = str(row.get("Motivo Market Selector", "") or "")
    trust = str(row.get("Signal Trust", "") or "").upper()
    fav = _row_pct(row, "ML favorito", 0.0)
    fav20 = _row_pct(row, "Favorito 2-0", 0.0)
    over18 = _row_pct(row, "Over 18.5", 0.0)
    over19 = _row_pct(row, "Over 19.5", 0.0)
    set3 = _row_pct(row, "Partido a 3 sets", 0.0)
    dog_set = _row_pct(row, "Prob gana set", 0.0)

    def out(m, p=None, extra=""):
        prob_txt = str(row.get("Prob mercado recomendado", "") or "")
        if p is not None:
            try:
                prob_txt = f"{float(p):.1%}"
            except Exception:
                pass
        base = motivo
        if extra:
            base = (base + " · " if base else "") + extra
        return pd.Series({
            "Mercado recomendado": m,
            "Prob mercado recomendado": prob_txt,
            "Motivo Market Selector": base,
        })

    mu = market.upper()

    if "UNDER 2.5" in mu and market.strip().startswith("✅"):
        return out(market.replace("✅", "👀 WATCH"), None, "v23.29.4: Under 2.5 solo watch hasta validar")

    if "OVER" in mu and ("SPOT WATCH" in trust or "WATCH" in trust):
        if "19.5" in mu:
            return out("👀 WATCH OVER 19.5", over19, "v23.29.4: Signal Trust en watch; no apuesta oficial")
        if "18.5" in mu:
            return out("👀 WATCH OVER 18.5", over18, "v23.29.4: Signal Trust en watch; no apuesta oficial")
        return out("👀 WATCH OVER", None, "v23.29.4: Signal Trust en watch; no apuesta oficial")

    # v23.29.4: el último backtest mostró que Over 19.5 al 68% seguía colándose
    # como oficial. Para máximo acierto, 19.5 necesita al menos 69%.
    if "OVER 19.5" in mu and over19 < 0.69:
        return out("👀 WATCH OVER 19.5 / NO COMBI", over19, "v23.29.4: Over 19.5 <69%; baja a watch/no combi")

    # v23.30.5: si el propio modelo da Favorito 2-0 >=60%, no dejamos que el Over
    # sea pick oficial salvo que haya una señal larga muy clara. Esta regla nace del
    # fallo Buse vs Gaston: Over 18.5 oficial con Fav 2-0 63.2% terminó corto.
    official_over = ("OVER" in mu) and (market.strip().startswith("✅") or market.strip().startswith("🔥"))
    perfil_largo_fuerte = (set3 >= 0.48) or (over19 >= 0.70 and dog_set >= 0.55)
    if official_over and fav20 >= 0.60 and not perfil_largo_fuerte:
        if "19.5" in mu:
            return out("👀 WATCH OVER 19.5 / RIESGO 2-0", over19, "v23.30.5: Favorito 2-0 >=60%; no hay perfil largo fuerte")
        return out("👀 WATCH OVER 18.5 / RIESGO 2-0", over18, "v23.30.5: Favorito 2-0 >=60%; no hay perfil largo fuerte")

    if "OVER" in mu and fav >= 0.64 and fav20 >= 0.48 and over18 < 0.78:
        return out("👀 WATCH OVER 18.5 / RIESGO 2-0", over18, "v23.29.4: favorito + 2-0 moderado; no apuesta oficial")

    return pd.Series({
        "Mercado recomendado": row.get("Mercado recomendado", ""),
        "Prob mercado recomendado": row.get("Prob mercado recomendado", ""),
        "Motivo Market Selector": row.get("Motivo Market Selector", ""),
    })


def batch_recommendation(row):
    trust = str(row.get("Signal Trust", "")).upper()
    signal = str(row.get("Mejor señal", ""))
    focus = str(row.get("Over Focus Label", "")).strip()
    watchlist = str(row.get("WTA Watchlist", "")).strip()
    risks = str(row.get("Riesgos", "")).lower()

    circuito_txt = " ".join([
        str(row.get("Circuito cálculo", "")),
        str(row.get("Circuito datos", "")),
        str(row.get("Circuito fuente", "")),
        str(row.get("Torneo", "")),
        str(row.get("Estado", "")),
    ])
    is_chall = _is_challenger_context(circuito_txt)
    estado = str(row.get("Estado", ""))
    partial_data = ("estimado" in estado.lower()) or ("parcial" in trust.lower()) or ("fallback" in risks)
    fav_prob = _row_pct(row, "ML favorito", 0.0)
    fav20 = _row_pct(row, "Favorito 2-0", 0.0)
    over18 = _row_pct(row, "Over 18.5", 0.0)

    def apply_guards(rec):
        rec = str(rec or "")

        # v23.29: Over Quality Guard manda por encima de señales antiguas.
        if _row_over_guard_active(row):
            under25_adj = _row_under25_adjusted(row, 1 - _row_pct(row, "Partido a 3 sets", 0.0))
            # v23.29.3: no damos Under 2.5 como pick oficial; solo vigilancia.
            if under25_adj >= 0.63:
                return "👀 WATCH UNDER 2.5 SETS"
            return "🚫 OVER BLOQUEADO / NO BET"

        # v23.29.3: si Signal Trust está en WATCH, ningún Over puede ser apuesta oficial.
        if "WATCH" in trust and "OVER" in rec.upper():
            return rec.replace("🔥", "👀").replace("✅", "👀").replace("FUERTE", "WATCH").replace("APTO", "WATCH / NO COMBI")

        # v23.26.1: en Challenger ningún pick con datos parciales puede salir como FUERTE.
        if partial_data and "FUERTE" in rec:
            rec = _downgrade_strong_to_apto(rec, "datos parciales")

        if is_chall:
            # v23.26.4: si el perfil apunta a 2-0/Under, no dejamos que la recomendación final sea Over.
            if fav_prob >= 0.64 and fav20 >= 0.52 and over18 < 0.76:
                if fav_prob >= 0.70 and fav20 >= 0.58:
                    return "✅ MIRAR FAVORITO 2-0 / UNDER 2.5 SETS"
                return "👀 WATCH FAVORITO 2-0 / UNDER 2.5 SETS"

            # v23.26.2: re-etiquetado final de Over Challenger por umbral real,
            # aunque llegue desde una etiqueta antigua ATP.
            if "OVER 18.5" in rec:
                if over18 >= 0.80:
                    rec = "🔥 OVER 18.5 FUERTE"
                elif over18 >= 0.76:
                    rec = "✅ OVER 18.5 APTO"
                elif over18 >= 0.70:
                    rec = "👀 OVER 18.5 WATCH"
                elif over18 >= 0.65:
                    rec = "👀 OVER 18.5 WATCH / NO BET"
                else:
                    rec = "ML SOLO CONTEXTO"

            # Challenger ML: nunca lo vendemos como apuesta fuerte desde el resumen automático.
            if "ML" in rec and "SOLO CONTEXTO" not in rec:
                if fav_prob < 0.77:
                    return "ML SOLO CONTEXTO / NO BET CHALLENGER"
                return "👀 ML OBSERVAR CHALLENGER"

            # Filtro anti 2-0 corto: favorito claro + 2-0 alto + Over no elite.
            if "OVER 18.5" in rec and fav_prob >= 0.72 and fav20 >= 0.55 and over18 < 0.80:
                rec = _downgrade_strong_to_apto(rec, "riesgo 2-0 corto") if "FUERTE" in rec else rec + " · riesgo 2-0 corto"

            # Si Challenger Over 18.5 no llega a 70%, no se etiqueta como apuesta.
            if "OVER 18.5" in rec and over18 < 0.70:
                return "👀 OVER 18.5 WATCH / NO BET"

        return rec

    if focus:
        if "FUERTE" in trust and "OVER" in focus:
            return apply_guards(focus)
        if "APTO" in trust and "OVER" in focus:
            return apply_guards(focus)
        if "DUDOSO" in trust and "OVER" in focus:
            return apply_guards(focus.replace("✅", "👀").replace("FUERTE", "WATCH"))
        if "3 SETS" in focus and ("DUDOSO" in trust or "APTO" in trust or "FUERTE" in trust):
            return apply_guards(focus)

    if "NO BET" in trust:
        if watchlist:
            if "17.5" in watchlist:
                return "OBSERVAR OVER 17.5"
            if "18.5" in watchlist:
                return "OBSERVAR OVER 18.5"
            return "OBSERVAR OVER"
        return "NO BET"

    if "OVER" in signal.upper():
        if "FUERTE" in trust:
            return apply_guards("🔥 OVER FUERTE")
        if "APTO" in trust:
            return apply_guards("✅ OVER APTO")
        if "DUDOSO" in trust:
            return apply_guards("👀 OVER WATCHLIST")

    if "3 SETS" in signal.upper() or "PARTIDO A 3" in signal.upper():
        return "🎯 3 SETS WATCH"

    if "upset" in risks or "riesgo" in risks:
        return "NO BET / REVISAR"

    return "ML SOLO CONTEXTO"



# =========================================================
# v23.37.1 CLARIDAD SELECTOR MÁXIMO ACIERTO
# =========================================================
def _pct_text_to_float_v23371(x, default=0.0):
    """Convierte '84.5%' o 0.845 a float 0-1 sin romper la tabla."""
    try:
        if x is None or pd.isna(x):
            return default
        s = str(x).replace('%', '').replace(',', '.').strip()
        if s == '' or s.lower() in {'nan', 'none'}:
            return default
        v = float(s)
        if v > 1:
            v /= 100.0
        return float(np.clip(v, 0.0, 1.0))
    except Exception:
        return default





# =========================================================
# v23.38.2 GRAND SLAM BO5 REAL RESULT VALIDATOR / EXPORT
# =========================================================
# Solo se usa en modo Grand Slam (5 sets). No modifica el motor Over BO3.
# Añade lectura real de juegos/sets BO5 y validación del mercado GS seleccionado.

def _gs_bool_to_si_no(value, available=True):
    if not available:
        return "N/D"
    return "Sí" if bool(value) else "No"


def _gs_float_or_none(x):
    try:
        if x is None or pd.isna(x):
            return None
        s = str(x).replace(',', '.').strip()
        if s == '' or s.lower() in {'nan', 'none', 'n/d'}:
            return None
        return float(s)
    except Exception:
        return None


def calcular_gs_bo5_reales(m, fav_side=None):
    """Devuelve métricas reales BO5 desde el resultado SofaScore ya parseado.

    fav_side: 1 si el favorito modelo es jugador 1, 2 si es jugador 2.
    """
    total_games = _gs_float_or_none((m or {}).get('actual_total_games'))
    p1_sets = (m or {}).get('p1_sets_real')
    p2_sets = (m or {}).get('p2_sets_real')
    try:
        p1_sets = int(p1_sets) if p1_sets is not None and not pd.isna(p1_sets) else None
        p2_sets = int(p2_sets) if p2_sets is not None and not pd.isna(p2_sets) else None
    except Exception:
        p1_sets, p2_sets = None, None

    sets_available = p1_sets is not None and p2_sets is not None
    games_available = total_games is not None
    total_sets = (p1_sets + p2_sets) if sets_available else None

    if fav_side == 1:
        fav_sets, dog_sets = p1_sets, p2_sets
    elif fav_side == 2:
        fav_sets, dog_sets = p2_sets, p1_sets
    else:
        fav_sets, dog_sets = None, None

    fav_known = sets_available and fav_side in [1, 2]
    fav_3_0 = bool(fav_known and fav_sets == 3 and dog_sets == 0)

    return {
        'available_games': games_available,
        'available_sets': sets_available,
        'gs_total_games_real': total_games if games_available else '',
        'gs_sets_total_real': total_sets if sets_available else '',
        'gs_over_30_5_real_bool': bool(games_available and total_games > 30.5),
        'gs_over_32_5_real_bool': bool(games_available and total_games > 32.5),
        'gs_over_35_5_real_bool': bool(games_available and total_games > 35.5),
        'gs_over_37_5_real_bool': bool(games_available and total_games > 37.5),
        'gs_over_39_5_real_bool': bool(games_available and total_games > 39.5),
        'gs_partido_4_plus_real_bool': bool(sets_available and total_sets >= 4),
        'gs_partido_5_sets_real_bool': bool(sets_available and total_sets == 5),
        'gs_fav_3_0_real_bool': fav_3_0,
        'gs_fav_no_3_0_real_bool': bool(fav_known and not fav_3_0),
        'gs_dog_set_real_bool': bool(fav_known and dog_sets is not None and dog_sets >= 1),
    }


def acierta_mercado_gs_bo5(label, reales):
    """Valida si el mercado GS seleccionado se cumplió. Devuelve Sí/No/N/D."""
    if not label or reales is None:
        return "N/D"
    lab = str(label).upper()

    # Over de juegos: requiere total games real.
    if 'OVER 30.5' in lab:
        return _gs_bool_to_si_no(reales.get('gs_over_30_5_real_bool'), reales.get('available_games', False))
    if 'OVER 32.5' in lab:
        return _gs_bool_to_si_no(reales.get('gs_over_32_5_real_bool'), reales.get('available_games', False))
    if 'OVER 35.5' in lab:
        return _gs_bool_to_si_no(reales.get('gs_over_35_5_real_bool'), reales.get('available_games', False))
    if 'OVER 37.5' in lab:
        return _gs_bool_to_si_no(reales.get('gs_over_37_5_real_bool'), reales.get('available_games', False))
    if 'OVER 39.5' in lab:
        return _gs_bool_to_si_no(reales.get('gs_over_39_5_real_bool'), reales.get('available_games', False))

    # Mercados de sets: requieren sets reales.
    if '4+ SET' in lab or '4+ SETS' in lab or 'PARTIDO A 4' in lab:
        return _gs_bool_to_si_no(reales.get('gs_partido_4_plus_real_bool'), reales.get('available_sets', False))
    if '5 SET' in lab or '5 SETS' in lab or 'PARTIDO A 5' in lab:
        return _gs_bool_to_si_no(reales.get('gs_partido_5_sets_real_bool'), reales.get('available_sets', False))
    if 'NO GANA 3-0' in lab or 'NO 3-0' in lab:
        return _gs_bool_to_si_no(reales.get('gs_fav_no_3_0_real_bool'), reales.get('available_sets', False))
    if 'GANA 3-0' in lab:
        return _gs_bool_to_si_no(reales.get('gs_fav_3_0_real_bool'), reales.get('available_sets', False))
    if 'GANA AL MENOS 1 SET' in lab or 'UNDERDOG GANA SET' in lab:
        return _gs_bool_to_si_no(reales.get('gs_dog_set_real_bool'), reales.get('available_sets', False))

    return "N/D"


# =========================================================
# v23.38.1 GRAND SLAM BO5 FINAL SELECTOR LOCK
# =========================================================
# Capa final visual/export: si Grand Slam 5 sets está activo, el mercado oficial
# de máxima probabilidad NO puede volver a Over BO3 (18.5/19.5/20.5/22.5).
# No cambia el motor Over blindado; solo corrige la elección final mostrada.

GS_BO3_BLOCKED_MARKETS = ["OVER 18.5", "OVER 19.5", "OVER 20.5", "OVER 22.5", "OVER 17.5"]


def _gs_bo5_activo_row(row):
    return str(row.get("🏆 GS 5 sets activo", "")).strip().lower() in {"sí", "si", "yes", "true", "1"}


def _gs_bo5_prob_float(row):
    try:
        return _pct_text_to_float_v23371(row.get("🏆 Prob GS 5 sets", 0), 0.0)
    except Exception:
        return 0.0


def _gs_bo5_datos_pobres(row):
    txt = " ".join(str(row.get(c, "")) for c in [
        "Estado", "Aviso datos", "Signal Trust", "Lectura J1", "Lectura J2", "🎯 Motivo acierto"
    ]).lower()
    try:
        ms_raw = str(row.get('Mín. partidos superficie', '')).replace(',', '.').strip()
        ms = float(ms_raw) if ms_raw not in {'', 'nan', 'None'} else 99.0
    except Exception:
        ms = 99.0
    try:
        min_conf = _pct_text_to_float_v23371(row.get('Confianza mínima', ''), 1.0)
    except Exception:
        min_conf = 1.0
    return ('fallback' in txt or 'estimado' in txt or 'datos parciales' in txt or ms <= 3 or min_conf < 0.45)




def _gs_pct_row(row, col, default=0.0):
    try:
        return _pct_text_to_float_v23371(row.get(col, ''), default)
    except Exception:
        return default


def _gs_partes_partido(row):
    txt = str(row.get("Partido", ""))
    if " vs " in txt:
        a, b = txt.split(" vs ", 1)
        return a.strip(), b.strip()
    return "", ""


def _gs_fav_lado(row):
    p1, p2 = _gs_partes_partido(row)
    fav = str(row.get("Favorito modelo", "")).strip()
    if fav and p1 and limpiar(fav) == limpiar(p1):
        return "J1"
    if fav and p2 and limpiar(fav) == limpiar(p2):
        return "J2"
    return ""


def _gs_refinar_mercado_por_ch(row, gs_market, gs_prob):
    """v23.38.6: refinamiento SOLO visual/export Grand Slam BO5.

    No toca el motor Over. Usa las columnas CH ya calculadas para evitar tres
    patrones que fallaron en el primer backtest:
    - 3-0 favorito cuando el rival tiene alta señal de ganar set.
    - gana-set con riesgo de paliza >= 25%.
    - NO gana 3-0 cuando el total GS Over 30.5 es más estable.
    """
    market = str(gs_market or "").strip()
    upper = market.upper()
    notes = []
    action_override = None

    blowout = _gs_pct_row(row, "🧠 CH Riesgo paliza", 0.0)
    ch_set_j1 = _gs_pct_row(row, "🧠 CH Set J1", 0.0)
    ch_set_j2 = _gs_pct_row(row, "🧠 CH Set J2", 0.0)
    over30 = _gs_pct_row(row, "GS Over 30.5", 0.0)
    over32 = _gs_pct_row(row, "GS Over 32.5", 0.0)

    fav_lado = _gs_fav_lado(row)
    rival_set = ch_set_j2 if fav_lado == "J1" else ch_set_j1 if fav_lado == "J2" else max(ch_set_j1, ch_set_j2)

    # 1) Bloquear 3-0 si el rival tiene señal fuerte de ganar set.
    if "GANA 3-0" in upper and "NO GANA 3-0" not in upper:
        if rival_set >= 0.80:
            notes.append(f"GS v6 bloquea 3-0: rival con CH gana-set {rival_set:.0%}.")
            # Si el propio mercado NO 3-0/Over tiene soporte, reconducimos a una línea más segura.
            if over30 >= 0.72:
                market = "Over 30.5 GS"
                gs_prob = over30
                notes.append("Se reconduce a Over 30.5 GS por soporte de juegos.")
            else:
                action_override = "👀 OBSERVAR"

    # 2) Gana-set + riesgo de paliza alto: no tratar como JUGAR.
    if "GANA AL MENOS 1 SET" in upper and blowout >= 0.25:
        notes.append(f"GS v6 baja gana-set: riesgo paliza CH {blowout:.0%}.")
        action_override = "👀 OBSERVAR"

    # 3) Cuando NO gana 3-0 compite con Over 30.5 alto, preferimos la línea de juegos.
    # Evita casos tipo 7-6 7-6 6-4 donde hay juegos, pero sí hay 3-0.
    if "NO GANA 3-0" in upper:
        if over30 >= 0.72 and over30 >= gs_prob - 0.08:
            market = "Over 30.5 GS"
            gs_prob = over30
            notes.append("GS v6 prioriza Over 30.5 frente a NO 3-0 cuando el total de juegos es más robusto.")
        elif over32 >= 0.70 and over32 >= gs_prob - 0.05:
            market = "Over 32.5 GS"
            gs_prob = over32
            notes.append("GS v6 prioriza Over 32.5 frente a NO 3-0 por proyección de juegos.")

    return market, float(gs_prob or 0.0), action_override, " | ".join(notes)



def _gs_parse_resultado_sets_row(row):
    """Devuelve (p1_sets, p2_sets) desde la columna Resultado sets, si existe."""
    try:
        raw = str(row.get("Resultado sets", "")).strip()
        m = re.search(r"(\d+)\s*-\s*(\d+)", raw)
        if not m:
            return None, None
        return int(m.group(1)), int(m.group(2))
    except Exception:
        return None, None


def _gs_player_side_from_label(row, label):
    """Detecta si el jugador mencionado al inicio del mercado es J1 o J2."""
    p1, p2 = _gs_partes_partido(row)
    lab = str(label or "").strip()
    if not lab:
        return None
    lab_clean = limpiar(lab)
    if p1 and limpiar(p1) and limpiar(p1) in lab_clean:
        return 1
    if p2 and limpiar(p2) and limpiar(p2) in lab_clean:
        return 2
    # Fallback: comparar por tokens/apellidos para etiquetas abreviadas.
    try:
        if p1 and similitud_nombre(lab, p1) >= 0.72:
            return 1
        if p2 and similitud_nombre(lab, p2) >= 0.72:
            return 2
    except Exception:
        pass
    return None


def _gs_real_col_to_result(row, col):
    v = str(row.get(col, "")).strip()
    if v in {"Sí", "Si", "SI", "SÍ", "sí", "si"}:
        return "Sí"
    if v.lower() in {"no"}:
        return "No"
    if v.upper() in {"N/D", "ND"}:
        return "N/D"
    return ""


def _gs_acierto_final_desde_row(label, row):
    """v23.38.7: valida EXACTAMENTE el mercado final visible.

    Antes se validaba el mercado GS original, pero el refinamiento visual podía
    cambiarlo a Over 30.5/32.5 o bajar/cambiar NO 3-0. Esta función sincroniza:
    PICK LIMPIO = DETALLE TÉCNICO = Acierta mercado GS.
    """
    lab = str(label or "").strip()
    upper = lab.upper()
    if not lab or "SIN MERCADO GS" in upper:
        return ""

    # Totales GS.
    if "OVER 30.5" in upper:
        return _gs_real_col_to_result(row, "GS Over 30.5 real")
    if "OVER 32.5" in upper:
        return _gs_real_col_to_result(row, "GS Over 32.5 real")
    if "OVER 35.5" in upper:
        return _gs_real_col_to_result(row, "GS Over 35.5 real")
    if "OVER 37.5" in upper:
        return _gs_real_col_to_result(row, "GS Over 37.5 real")
    if "OVER 39.5" in upper:
        return _gs_real_col_to_result(row, "GS Over 39.5 real")

    if "4+" in upper or "4 +" in upper:
        return _gs_real_col_to_result(row, "GS Partido 4+ sets real")
    if "5 SET" in upper:
        return _gs_real_col_to_result(row, "GS Partido 5 sets real")

    p1_sets, p2_sets = _gs_parse_resultado_sets_row(row)
    side = _gs_player_side_from_label(row, lab)

    # Mercados de jugador: validamos por el jugador visible en el texto, no por el mercado GS original.
    if "NO GANA 3-0" in upper:
        if p1_sets is not None and p2_sets is not None and side in {1, 2}:
            s, other = (p1_sets, p2_sets) if side == 1 else (p2_sets, p1_sets)
            return "No" if (s == 3 and other == 0) else "Sí"
        return _gs_real_col_to_result(row, "GS Favorito NO 3-0 real")

    if "GANA 3-0" in upper:
        if p1_sets is not None and p2_sets is not None and side in {1, 2}:
            s, other = (p1_sets, p2_sets) if side == 1 else (p2_sets, p1_sets)
            return "Sí" if (s == 3 and other == 0) else "No"
        return _gs_real_col_to_result(row, "GS Favorito 3-0 real")

    if "GANA AL MENOS 1 SET" in upper:
        if p1_sets is not None and p2_sets is not None and side in {1, 2}:
            s = p1_sets if side == 1 else p2_sets
            return "Sí" if s >= 1 else "No"
        return _gs_real_col_to_result(row, "GS Underdog gana set real")

    return ""

def aplicar_grand_slam_bo5_selector_final(df):
    """Aplica el candado final BO5 sobre tabla ya calculada.

    Reglas:
    - Con GS activo, 🎯 Mercado más probable debe salir de 🏆 Mercado GS 5 sets.
    - Over 18.5/19.5/20.5/22.5/17.5 quedan bloqueados como mercado final.
    - Si no hay señal GS clara, acción final = 👀 OBSERVAR, nunca ✅ JUGAR.
    - Si hay fallback/datos pobres, acción final = 👀 OBSERVAR aunque GS diga JUGAR.
    """
    if df is None or getattr(df, "empty", True) or "🏆 GS 5 sets activo" not in df.columns:
        return df

    out = df.copy()
    needed = [
        "🎯 Mercado más probable", "🎯 Prob máxima", "🎯 Confianza acierto", "🎯 Motivo acierto",
        "🎯 Acción final", "🎯 Decisión acierto", "🎯 Aviso acierto", "Mejor señal", "Prob señal"
    ]
    for c in needed:
        if c not in out.columns:
            out[c] = ""

    for idx, row in out.iterrows():
        if not _gs_bo5_activo_row(row):
            continue

        current_market = str(row.get("🎯 Mercado más probable", "")).upper()
        gs_market = str(row.get("🏆 Mercado GS 5 sets", "")).strip()
        gs_prob = _gs_bo5_prob_float(row)
        gs_action = str(row.get("🏆 Acción GS", "")).strip()
        gs_conf = str(row.get("🏆 Confianza GS", row.get("🎯 Confianza acierto", ""))).strip()
        gs_market, gs_prob, gs_action_override, gs_refine_note = _gs_refinar_mercado_por_ch(row, gs_market, gs_prob)

        blocked_current = any(x in current_market for x in GS_BO3_BLOCKED_MARKETS)
        no_gs_signal = (not gs_market or "SIN MERCADO GS" in gs_market.upper() or gs_prob <= 0.0)

        if no_gs_signal:
            out.at[idx, "🎯 Mercado más probable"] = "Sin mercado GS claro"
            out.at[idx, "🎯 Prob máxima"] = "0.0%"
            out.at[idx, "🎯 Confianza acierto"] = "⚠️ Sin señal GS"
            out.at[idx, "🎯 Motivo acierto"] = "Grand Slam BO5 activo: sin mercado 5 sets suficientemente claro. Over BO3 bloqueado como mercado final."
            out.at[idx, "Mejor señal"] = "Sin mercado GS claro"
            out.at[idx, "Prob señal"] = "0.0%"
            out.at[idx, "🏆 Mercado GS 5 sets"] = "Sin mercado GS claro"
            out.at[idx, "🏆 Prob GS 5 sets"] = "0.0%"
            out.at[idx, "Acierta mercado GS"] = ""
            out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
            out.at[idx, "🎯 Decisión acierto"] = "👀 Observar / sin señal GS"
            out.at[idx, "🎯 Aviso acierto"] = "Grand Slam 5 sets: no hay señal BO5 clara; no convertir Over 18.5/19.5 en pick."
            continue

        # En GS, siempre mandan las columnas GS aunque alguna capa posterior haya recuperado Over 18.5.
        if blocked_current or True:
            out.at[idx, "🎯 Mercado más probable"] = gs_market
            out.at[idx, "🎯 Prob máxima"] = f"{gs_prob:.1%}"
            out.at[idx, "🎯 Confianza acierto"] = gs_conf if gs_conf else str(row.get("🏆 Acción GS", ""))
            base_motivo_gs = "Grand Slam BO5 activo: selector final bloquea Over BO3 y usa solo mercados 30.5-39.5, 4+ sets, 5 sets o 3-0/no 3-0."
            out.at[idx, "🎯 Motivo acierto"] = base_motivo_gs + ((" | " + gs_refine_note) if gs_refine_note else "")
            out.at[idx, "Mejor señal"] = gs_market
            out.at[idx, "Prob señal"] = f"{gs_prob:.1%}"
            # v23.38.7: sincronizar detalle técnico y validación con el mercado visible.
            out.at[idx, "🏆 Mercado GS 5 sets"] = gs_market
            out.at[idx, "🏆 Prob GS 5 sets"] = f"{gs_prob:.1%}"
            out.at[idx, "Acierta mercado GS"] = _gs_acierto_final_desde_row(gs_market, out.loc[idx])

        datos_pobres = _gs_bo5_datos_pobres(row)
        if gs_action_override:
            out.at[idx, "🎯 Acción final"] = gs_action_override
            out.at[idx, "🎯 Decisión acierto"] = "👀 Observar / GS v6 filtro histórico"
            out.at[idx, "🎯 Aviso acierto"] = gs_refine_note or "Grand Slam BO5: bajado por filtro histórico v6."
        elif datos_pobres:
            out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
            out.at[idx, "🎯 Decisión acierto"] = "👀 Observar / GS datos parciales"
            out.at[idx, "🎯 Aviso acierto"] = "Grand Slam BO5: señal útil, pero con fallback/muestra baja; no tratar como pick fuerte."
        elif "JUGAR" in gs_action.upper() and gs_prob >= 0.70:
            out.at[idx, "🎯 Acción final"] = "✅ JUGAR"
            out.at[idx, "🎯 Decisión acierto"] = "🔥 Alto acierto GS"
            out.at[idx, "🎯 Aviso acierto"] = "OK: mercado Grand Slam BO5 válido; Over BO3 bloqueado."
        elif "OBSERVAR" in gs_action.upper() or gs_prob >= 0.58:
            out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
            out.at[idx, "🎯 Decisión acierto"] = "👀 Observar / GS apto"
            out.at[idx, "🎯 Aviso acierto"] = "Mercado GS interesante, pero no llega a zona JUGAR."
        else:
            out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
            out.at[idx, "🎯 Decisión acierto"] = "👀 Observar / sin señal GS"
            out.at[idx, "🎯 Aviso acierto"] = "Grand Slam BO5: señal insuficiente; no jugar."

    return out

def decision_acierto_v23371(row):
    """
    Etiqueta final para el objetivo actual del proyecto:
    máxima tasa de acierto por partido, sin cuotas ni value.

    No cambia ningún cálculo. Solo traduce el selector 🎯 a una decisión visual.
    """
    mercado = str(row.get('🎯 Mercado más probable', '')).strip()
    motivo = str(row.get('🎯 Motivo acierto', '')).lower()
    estado = str(row.get('Estado', '')).lower()
    aviso = str(row.get('Aviso datos', '')).lower()
    trust = str(row.get('Signal Trust', '')).lower()
    lectura1 = str(row.get('Lectura J1', '')).lower()
    lectura2 = str(row.get('Lectura J2', '')).lower()
    prob = _pct_text_to_float_v23371(row.get('🎯 Prob máxima', 0), 0.0)
    min_surface = _pct_text_to_float_v23371(row.get('Mín. partidos superficie', ''), -1.0)
    # Mín. partidos superficie suele venir como entero/string. Lo parseamos aparte.
    try:
        ms_raw = str(row.get('Mín. partidos superficie', '')).replace(',', '.').strip()
        ms = float(ms_raw) if ms_raw not in {'', 'nan', 'None'} else 99.0
    except Exception:
        ms = 99.0
    min_conf = _pct_text_to_float_v23371(row.get('Confianza mínima', ''), 1.0)

    datos_pobres = (
        'fallback' in motivo or 'fallback' in aviso or 'estimado' in estado or
        'estimado' in lectura1 or 'estimado' in lectura2 or 'datos parciales' in trust or
        ms <= 3 or min_conf < 0.45
    )

    if not mercado or mercado.startswith('⛔') or 'evitar' in mercado.lower() or prob < 0.62:
        return pd.Series({
            '🎯 Acción final': '⛔ EVITAR',
            '🎯 Decisión acierto': '⛔ Evitar',
            '🎯 Aviso acierto': 'Ningún mercado supera el mínimo de seguridad.'
        })

    if datos_pobres:
        return pd.Series({
            '🎯 Acción final': '👀 OBSERVAR',
            '🎯 Decisión acierto': '👀 Observar / datos pobres',
            '🎯 Aviso acierto': 'Probabilidad estimada por fallback o muestra baja: no tratar como pick fuerte.'
        })

    mercado_u = mercado.upper()
    ml_fav = _pct_text_to_float_v23371(row.get('ML favorito', 0), 0.0)
    over18 = _pct_text_to_float_v23371(row.get('Over 18.5', 0), 0.0)
    is_set_market = 'GANA AL MENOS 1 SET' in mercado_u
    is_under_market = ('UNDER' in mercado_u or '2-0' in mercado_u)
    spot_dudoso = 'dudoso' in trust
    oficial_no_bet = 'no bet' in str(row.get('Mercado recomendado', '')).lower() or 'no bet' in str(row.get('Recomendación', '')).lower()

    if is_under_market:
        # Tras los backtests, Under/2-0 queda capado visualmente aunque pase el filtro estricto.
        dec = '👀 Observar / under restrictivo'
        aviso_ok = 'Under/2-0 capado: revisar solo si el favorito 2-0 y el ML son muy claros.'
    elif is_set_market:
        # v23.37.29 Analyzer Threshold Selector:
        # El backtest masivo confirmó que gana-set es útil solo en zona élite
        # o con confirmación externa. La decisión base NO lo juega salvo >=94%.
        set_elite_ok = (prob >= 0.94 and ml_fav >= 0.68 and not oficial_no_bet)

        if set_elite_ok:
            dec = '🔥 Alto acierto'
            aviso_ok = 'OK: gana-set en zona élite según backtest masivo; revisar si histórico/TA no contradice.'
        elif prob >= 0.84:
            dec = '👀 Observar / gana-set alto no jugable'
            aviso_ok = 'Gana-set alto, pero el analyzer exige >=94% o confirmación histórica/TA para JUGAR.'
        elif prob >= 0.78:
            dec = '👀 Observar / set con contexto débil'
            aviso_ok = 'Probabilidad útil, pero gana-set queda bloqueado por backtest irregular.'
        else:
            dec = '👀 Observar'
            aviso_ok = 'Set market por debajo de zona jugable estricta.'
    elif 'OVER 18.5' in mercado_u and over18 >= 0.80 and not oficial_no_bet:
        dec = '🔥 Alto acierto'
        aviso_ok = 'OK: Over 18.5 >=80%, umbral validado por analyzer masivo.'
    elif ('OVER 19.5' in mercado_u or 'OVER 20.5' in mercado_u) and prob >= 0.74 and not oficial_no_bet:
        dec = '👀 Observar / over alto secundario'
        aviso_ok = 'Over superior requiere confirmación histórica/TA; no es JUGAR base.'
    elif (' GANA' in (' ' + mercado_u) and 'SET' not in mercado_u and 'OVER' not in mercado_u and ml_fav >= 0.70 and not oficial_no_bet):
        dec = '🔥 Alto acierto'
        aviso_ok = 'OK: ML favorito >=70%, zona fuerte validada por analyzer masivo.'
    elif prob >= 0.84 and not oficial_no_bet:
        dec = '👀 Observar / alto no validado'
        aviso_ok = 'Probabilidad alta, pero no encaja en mercados validados por analyzer para JUGAR.'
    elif prob >= 0.78 and not oficial_no_bet:
        if 'OVER 18.5' in mercado_u and (over18 >= 0.76 or 'fuerte' in trust or 'apto' in trust):
            dec = '👀 Observar / over watch'
            aviso_ok = 'Over 18.5 interesante, pero analyzer recomienda JUGAR desde 80%.'
        else:
            dec = '👀 Observar / buen acierto no jugable'
            aviso_ok = 'Buen acierto, pero no pasa umbral analyzer de JUGAR.'
    elif prob >= 0.70:
        dec = '⚖️ Apto prudente'
        aviso_ok = 'OK, pero no tratar como pick fuerte.'
    else:
        dec = '👀 Observar'
        aviso_ok = 'Probabilidad útil, pero por debajo de zona fuerte.'

    # v23.37.9: acción final híbrida; 🔥 alto + Over 18.5 fuerte pueden ser JUGAR.
    # ✅ JUGAR solo si es 🔥 Alto acierto, o ✅ Buen acierto cuando viene de Over 18.5 fuerte/apto.
    # Los mercados de set en zona ✅ Buen acierto pasan a OBSERVAR por backtest irregular.
    if dec == '🔥 Alto acierto' or (dec == '✅ Buen acierto' and 'OVER 18.5' in mercado_u):
        accion = '✅ JUGAR'
    elif dec.startswith('⛔'):
        accion = '⛔ EVITAR'
    else:
        accion = '👀 OBSERVAR'

    return pd.Series({
        '🎯 Acción final': accion,
        '🎯 Decisión acierto': dec,
        '🎯 Aviso acierto': aviso_ok
    })




def radar_datos_extra_v233710(row):
    """
    v23.37.10 Radar Datos extra.
    No cambia probabilidades ni motores. Solo decide qué partidos merecen revisión extra
    en Stats Center de Datos extra para no revisar 40 partidos a mano.
    """
    def pct(x, default=0.0):
        try:
            if x is None or pd.isna(x):
                return default
            t = str(x).replace('%','').replace(',','.').strip()
            if t == '' or t.lower() in {'nan','none'}:
                return default
            v = float(t)
            if v > 1:
                v /= 100.0
            return float(np.clip(v, 0, 1))
        except Exception:
            return default

    def num(x, default=99.0):
        try:
            if x is None or pd.isna(x):
                return default
            t = str(x).replace(',','.').strip()
            if t == '' or t.lower() in {'nan','none'}:
                return default
            return float(t)
        except Exception:
            return default

    accion = str(row.get('🎯 Acción final', '')).strip()
    mercado = str(row.get('🎯 Mercado más probable', '')).strip()
    mercado_u = mercado.upper()
    decision = str(row.get('🎯 Decisión acierto', '')).lower()
    aviso = str(row.get('🎯 Aviso acierto', '')).lower()
    motivo = str(row.get('🎯 Motivo acierto', '')).lower()
    estado = str(row.get('Estado', '')).lower()
    trust = str(row.get('Signal Trust', '')).lower()
    over_guard = str(row.get('Over Quality Guard', '')).lower()
    motivos_over = str(row.get('Motivos Over Guard', '')).lower()
    reco = str(row.get('Recomendación', '')).lower()
    oficial = str(row.get('Pick oficial', '')).lower()

    prob = pct(row.get('🎯 Prob máxima', 0), 0.0)
    ml_fav = pct(row.get('ML favorito', 0), 0.0)
    over18 = pct(row.get('Over 18.5', 0), 0.0)
    over19 = pct(row.get('Over 19.5', 0), 0.0)
    set3 = pct(row.get('Partido a 3 sets', 0), 0.0)
    min_conf = pct(row.get('Confianza mínima', ''), 1.0)
    min_surface = num(row.get('Mín. partidos superficie', ''), 99.0)

    is_set = 'GANA AL MENOS 1 SET' in mercado_u
    is_over18 = 'OVER 18.5' in mercado_u
    is_under = 'UNDER' in mercado_u or '2-0' in mercado_u
    datos_pobres = (
        'fallback' in aviso or 'fallback' in motivo or 'estimado' in estado or
        'datos parciales' in trust or min_surface <= 8 or min_conf < 0.58
    )
    over_fuerte = over18 >= 0.76 or 'fuerte' in trust or 'apto' in trust or 'watch' in reco or 'over' in oficial

    # Por defecto no pedir nada: ya es claro, malo o no merece perder tiempo.
    necesita = 'NO'
    prioridad = ''
    mirar = ''
    why = 'No necesita datos extra: decisión suficientemente clara o sin potencial jugable.'

    if accion == '⛔ EVITAR':
        return pd.Series({
            '📥 Necesita Datos extra': 'NO',
            '📥 Prioridad Datos extra': '',
            '📥 Qué mirar Datos extra': '',
            '📥 Motivo Datos extra': 'Evitar: no merece revisión manual.'
        })

    # 1) Over fuerte en observar: principal candidato para subir a JUGAR si Datos extra confirma marcadores largos.
    if accion == '👀 OBSERVAR' and over18 >= 0.76 and not is_under:
        necesita = 'SÍ'
        prioridad = 'ALTA' if over18 >= 0.79 or over_fuerte else 'MEDIA'
        mirar = 'Últimos 10 de ambos; marcadores 20+ juegos; 7-6/7-5; partidos a 3 sets; derrotas/palizas 2-0.'
        why = f'Over 18.5 alto ({over18:.1%}) pero la app lo dejó en OBSERVAR: Datos extra puede confirmar si subirlo.'

    # 2) JUGAR con set market y datos pobres: confirmar que no viene de perder fácil.
    elif accion == '✅ JUGAR' and is_set and (datos_pobres or ml_fav < 0.70):
        necesita = 'SÍ'
        prioridad = 'ALTA'
        mirar = 'Forma últimos 10; si el jugador elegido gana sets con frecuencia; derrotas 2-0; marcadores recientes del elegido.'
        why = 'Es JUGAR por gana-set, pero conviene confirmar forma reciente y que no pierda fácil 2-0.'

    # 3) Set muy alto en observar: puede ser útil, pero necesita confirmación.
    elif accion == '👀 OBSERVAR' and is_set and prob >= 0.84 and 0.58 <= ml_fav <= 0.72:
        necesita = 'SÍ'
        prioridad = 'MEDIA'
        mirar = 'Forma últimos 10; derrotas 2-0 del jugador elegido; si compite sets ante rivales similares.'
        why = f'Gana-set alto ({prob:.1%}) con ML no totalmente claro ({ml_fav:.1%}); necesita confirmación de forma.'

    # 4) Datos pobres con probabilidad muy alta: revisar solo si hay potencial real.
    elif accion in {'✅ JUGAR', '👀 OBSERVAR'} and datos_pobres and prob >= 0.84:
        necesita = 'SÍ'
        prioridad = 'MEDIA'
        mirar = 'Forma últimos 10; superficie; marcadores recientes; si hay mucha retirada/paliza o partidos largos.'
        why = 'Probabilidad alta, pero basada en muestra baja/fallback: confirmar en Datos extra.'

    # 5) Under/2-0 no lo priorizamos para el objetivo actual.
    elif is_under:
        necesita = 'NO'
        prioridad = ''
        mirar = ''
        why = 'Under/2-0 no se prioriza por backtest irregular.'

    return pd.Series({
        '📥 Necesita Datos extra': necesita,
        '📥 Prioridad Datos extra': prioridad,
        '📥 Qué mirar Datos extra': mirar,
        '📥 Motivo Datos extra': why
    })

def prepare_batch_display_table(ok_df):
    if ok_df is None or ok_df.empty:
        return ok_df

    df = ok_df.copy()

    # Unificar columnas dinámicas del tipo "Jugador set".
    set_cols = [c for c in df.columns if c.endswith(" set")]
    if set_cols:
        jugador_set = []
        prob_set = []
        for _, row in df.iterrows():
            found_name = ""
            found_prob = ""
            for c in set_cols:
                val = row.get(c, "")
                if str(val).strip():
                    found_name = c[:-4]
                    found_prob = val
                    break
            jugador_set.append(found_name)
            prob_set.append(found_prob)
        df["Jugador gana set"] = jugador_set
        df["Prob gana set"] = prob_set
        df = df.drop(columns=set_cols, errors="ignore")

    df["Recomendación"] = df.apply(batch_recommendation, axis=1)
    selector_cols = df.apply(market_selector_v23263, axis=1)
    df = pd.concat([df, selector_cols], axis=1)

    # v23.26.6: la recomendación final manda; el mercado recomendado no puede contradecirla.
    aligned_selector_cols = df.apply(alinear_market_selector_v23266, axis=1)
    for _col in ["Mercado recomendado", "Prob mercado recomendado", "Motivo Market Selector"]:
        df[_col] = aligned_selector_cols[_col]

    df["Signal Trust"] = df.apply(limpiar_signal_trust_v23263, axis=1)

    # v23.29.3: capa final de máximo acierto después de limpiar Signal Trust.
    final_prudence_cols = df.apply(aplicar_max_acierto_v23294, axis=1)
    for _col in ["Mercado recomendado", "Prob mercado recomendado", "Motivo Market Selector"]:
        df[_col] = final_prudence_cols[_col]

    # v23.31.2: WTA Over 17.5 Premium Watch específico. Se aplica antes del ML Guard
    # para separar los mejores WTA Over17 sin convertirlos aún en oficiales.
    wta_o17_cols = df.apply(aplicar_wta_over17_oficial_v23310, axis=1)
    for _col in ["Mercado recomendado", "Prob mercado recomendado", "Motivo Market Selector", "WTA Over17 Official Guard"]:
        df[_col] = wta_o17_cols[_col]

    # v23.30: ML Quality Guard. Añade ML oficial/watch solo si supera filtro duro.
    ml_guard_cols = df.apply(aplicar_ml_quality_guard_v23300, axis=1)
    for _col in ["Mercado recomendado", "Prob mercado recomendado", "Motivo Market Selector", "ML Quality Guard", "Motivos ML Guard"]:
        df[_col] = ml_guard_cols[_col]

    # v23.30.1: columna visible con el pick oficial real para revisar antes de descargar Excel.
    df["Pick oficial"] = df.apply(pick_oficial_v23301, axis=1)

    # v23.37.5: decisión visual + acción final específica para máximo acierto, sin cuotas.
    if "🎯 Mercado más probable" in df.columns and "🎯 Prob máxima" in df.columns:
        decision_cols = df.apply(decision_acierto_v23371, axis=1)
        df = pd.concat([df, decision_cols], axis=1)

        # v23.37.10: Radar Datos extra. Solo dice qué partidos merece revisar manualmente.
        radar_cols = df.apply(radar_datos_extra_v233710, axis=1)
        df = pd.concat([df, radar_cols], axis=1)

        # v23.38.1: último candado del selector Grand Slam BO5.
        df = aplicar_grand_slam_bo5_selector_final(df)

    preferred = [
        "Versión app",
        "Recomendación",
        "Pick oficial",
        "Mercado recomendado",
        "Prob mercado recomendado",
        "Motivo Market Selector",
        "Fecha",
        "Hora",
        "Partido",
        "🎯 Acción final",
        "🎯 Decisión acierto",
        "🎯 Mercado más probable",
        "🎯 Prob máxima",
        "🎯 Confianza acierto",
        "🎯 Aviso acierto",
        "🎯 Motivo acierto",
        "📥 Necesita Datos extra",
        "📥 Prioridad Datos extra",
        "📥 Qué mirar Datos extra",
        "📥 Motivo Datos extra",
        "Favorito modelo",
        "ML favorito",
        "Mejor señal",
        "Prob señal",
        "Mejor mercado Over Focus",
        "Over Focus Label",
        "WTA Watchlist",
        "Mejor mercado WTA",
        "WTA Over17 Priority",
        "WTA Over17 Official Guard",
        "Signal Trust",
        "Over Quality Guard",
        "Motivos Over Guard",
        "Under 2.5 Rescue",
        "Under 2.5 ajustado",
        "ML Quality Guard",
        "Motivos ML Guard",
        "Confianza mínima",
        "Mín. partidos superficie",
        "Gap Elo/modelo",
        "Cuota pegada",
        "Jugador cuota",
        "Cuota justa",
        "Edge",
        "Ganador real",
        "Resultado sets",
        "Marcador games",
        "Total games real",
        "🏆 GS 5 sets activo",
        "🏆 Mercado GS 5 sets",
        "🏆 Prob GS 5 sets",
        "🏆 Acción GS",
        "🏆 Confianza GS",
        "GS Over 30.5",
        "GS Over 32.5",
        "GS Over 35.5",
        "GS Over 37.5",
        "GS Over 39.5",
        "GS Partido 4+ sets",
        "GS Partido 5 sets",
        "GS Favorito 3-0",
        "GS Favorito NO 3-0",
        "GS Underdog gana set",
        "GS Juegos reales",
        "GS Sets reales totales",
        "GS Over 30.5 real",
        "GS Over 32.5 real",
        "GS Over 35.5 real",
        "GS Over 37.5 real",
        "GS Over 39.5 real",
        "GS Partido 4+ sets real",
        "GS Partido 5 sets real",
        "GS Favorito 3-0 real",
        "GS Favorito NO 3-0 real",
        "GS Underdog gana set real",
        "Acierta mercado GS",
        "Acierta ML modelo",
        "Over 17.5",
        "Over 17.5 real",
        "Acierta Over 17.5",
        "Over 18.5",
        "Over 18.5 real",
        "Acierta Over 18.5",
        "3 sets real",
        "Acierta 3 sets",
        "Over 19.5",
        "Under 22.5",
        "Jugador gana set",
        "Prob gana set",
        "Favorito gana al menos 1 set",
        "J1 +10.5 juegos",
        "J1 +11.5 juegos",
        "J2 +10.5 juegos",
        "J2 +11.5 juegos",
        "Mejor juegos jugador",
        "Prob juegos jugador",
        "Confianza juegos jugador",
        "Juegos J1",
        "Juegos J2",
        "Total games",
        "Riesgos",
        "Match J1",
        "Match J2",
    ]

    # Si no hay WTA, ocultar columnas Over 17.5 para no tocar ATP/Challenger visualmente.
    over17_cols = ["Over 17.5", "Over 17.5 real", "Acierta Over 17.5"]
    has_over17 = str(globals().get("circuito", "")).upper().strip() == "WTA"
    for c in over17_cols:
        if c in df.columns and df[c].astype(str).str.strip().replace("nan", "").ne("").any():
            has_over17 = True
            break
    if not has_over17:
        df = df.drop(columns=over17_cols, errors="ignore")
        preferred = [c for c in preferred if c not in over17_cols]

    # v23.48.0: pantalla/Excel limpios. Las cuotas/stake/value ya no forman parte
    # del flujo operativo; se decide fuera de la app.
    odds_cols = ["Cuota pegada", "Jugador cuota", "Cuota justa", "Edge"]
    df = df.drop(columns=odds_cols, errors="ignore")
    preferred = [c for c in preferred if c not in odds_cols]

    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    return df[cols]





# =========================================================
# v23.37.12 REANÁLISIS DATOS EXTRA MANUAL
# =========================================================

def _datos_extra_match_key(row):
    return str(row.get("Partido", "")).strip()


def _match_key_ui(row, idx=""):
    """Clave por partido para que Streamlit no reutilice fichas de otro análisis."""
    try:
        raw = f"{idx}_{row.get('Fecha','')}_{row.get('Hora','')}_{row.get('Torneo','')}_{row.get('Partido','')}"
    except Exception:
        raw = f"{idx}"
    key = limpiar(raw)[:90]
    return key or str(idx)


def _nombre_ficha_extra(d):
    if not isinstance(d, dict):
        return ""
    ta = d.get("tennisabstract")
    fs = d.get("flashscore")
    if isinstance(ta, dict):
        return str(ta.get("player", "") or "")
    if isinstance(fs, dict):
        return str(fs.get("player", "") or "")
    return ""


def _ficha_corresponde_jugador(d, jugador_esperado):
    """Comprueba que la ficha pegada pertenece al jugador de esa fila."""
    nombre = _nombre_ficha_extra(d)
    if not nombre:
        return True, ""
    esperado = normalizar_texto(jugador_esperado or "")
    if not esperado:
        return True, ""
    sc, reason = nombre_score_quality_direct(esperado, nombre)
    exp_surname = surname_key(esperado)
    ficha_surname = surname_key(nombre)
    exp_clean = limpiar(esperado)
    ficha_clean = limpiar(nombre)
    ok = False
    if sc >= 0.84:
        ok = True
    if exp_surname and ficha_surname and exp_surname == ficha_surname:
        ok = True
    if exp_clean and ficha_clean and (exp_clean in ficha_clean or ficha_clean in exp_clean):
        ok = True
    if ok:
        return True, f"Ficha coincide con {esperado}: {nombre}"
    return False, f"❌ Ficha no coincide: esperábamos {esperado} pero la ficha es de {nombre}. No se aplica al reanálisis."


def _invalidar_lectura_si_no_coincide(d, jugador_esperado):
    if not isinstance(d, dict):
        return d
    ok, msg = _ficha_corresponde_jugador(d, jugador_esperado)
    if ok:
        if msg:
            d["name_match_ok"] = msg
        return d
    d = d.copy()
    d["mismatch"] = msg
    d.pop("tennisabstract", None)
    d.pop("flashscore", None)
    d["largos"] = ""
    d["palizas"] = ""
    d["nota"] = (str(d.get("nota", "")) + " | " + msg).strip(" |")
    return d



TA_PROFILE_MAX_AGE_DAYS = 7  # v23.47.3: refrescar fichas TA antiguas

# =========================================================
# v23.37.24 TENNIS ABSTRACT PROFILE CACHE
# Guarda fichas TA/Flashscore por jugador con fecha de subida para no pegarlas cada día.
# =========================================================

def _ta_profile_cache_path():
    try:
        os.makedirs("datos", exist_ok=True)
        return os.path.join("datos", "ta_profile_cache.json")
    except Exception:
        return "ta_profile_cache.json"


def _ta_profile_cache_load():
    import json
    path = _ta_profile_cache_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _ta_profile_cache_save(cache):
    import json
    path = _ta_profile_cache_path()
    try:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache or {}, f, ensure_ascii=False, indent=2)
        return True, path
    except Exception as e:
        return False, str(e)


def _ta_profile_cache_key(name):
    return limpiar(name or "")


def _ta_profile_cache_today():
    from datetime import date
    return date.today().isoformat()


def _ta_profile_cache_age_days(uploaded_at):
    from datetime import date, datetime
    try:
        d = datetime.fromisoformat(str(uploaded_at)[:10]).date()
        return max(0, (date.today() - d).days)
    except Exception:
        return None


def _ta_profile_cache_age_label(uploaded_at):
    age = _ta_profile_cache_age_days(uploaded_at)
    if age is None:
        return "fecha desconocida"
    if age == 0:
        return "subido hoy"
    if age == 1:
        return "subido ayer"
    return f"subido hace {age} días"




def _ta_profile_cache_is_ta_real_permanent(item):
    """v23.49.2: TA Real/Manual queda permanente.
    Se puede marcar como antigua, pero nunca se elimina ni se sustituye por Auto Recent.
    """
    if not isinstance(item, dict):
        return False
    src = str(item.get("source", "") or "")
    raw = str(item.get("raw_text", "") or "")[:1200]
    low = (src + "\n" + raw).lower()
    try:
        if int(item.get("ta_full_score", 0) or 0) >= 3:
            return True
    except Exception:
        pass
    if item.get("permanent_ta") is True:
        return True
    if "manual" in low and ("tennis abstract" in low or "tennisabstract" in low):
        return True
    if "tennis abstract" in low and "auto recent" not in low and "auto challenger recent" not in low:
        return True
    return False


def _ta_profile_cache_source_priority(item):
    """Mayor = más valiosa. Evita pisar TA real/manual con perfiles automáticos."""
    if not isinstance(item, dict):
        return 0
    src = str(item.get("source", "") or "")
    raw = str(item.get("raw_text", "") or "")[:1200]
    low = (src + "\n" + raw).lower()
    try:
        full_score = int(item.get("ta_full_score", 0) or 0)
    except Exception:
        full_score = 0
    if _ta_profile_cache_is_ta_real_permanent(item) and full_score >= 3:
        return 100
    if _ta_profile_cache_is_ta_real_permanent(item):
        return 90
    if "manual" in low and "auto recent" not in low:
        return 80
    if "flashscore" in low:
        return 60
    if "auto recent" in low or "auto challenger recent" in low or "histórico local" in low or "historico local" in low:
        return 20
    return 40

def _ta_profile_cache_is_stale(item, max_age_days=TA_PROFILE_MAX_AGE_DAYS):
    """v23.47.3: considera caducada una ficha TA si tiene más de max_age_days.
    Así la app la vuelve a pedir/buscar, pero no la borra: solo la marca para refrescar.
    """
    if not isinstance(item, dict):
        return True
    age = _ta_profile_cache_age_days(item.get("uploaded_at", ""))
    if age is None:
        return True
    try:
        return int(age) > int(max_age_days)
    except Exception:
        return True


def _ta_profile_cache_find(player_name, allow_stale=False):
    """Busca ficha guardada por nombre, usando exacto y fuzzy apellido+inicial."""
    cache = _ta_profile_cache_load()
    if not player_name:
        return None
    key = _ta_profile_cache_key(player_name)
    if key and key in cache:
        item = cache[key]
        if isinstance(item, dict):
            item = item.copy()
            item["cache_key"] = key
            item["match_score"] = 1.0
            item["stale"] = _ta_profile_cache_is_stale(item)
            if item["stale"] and not allow_stale and not _ta_profile_cache_is_ta_real_permanent(item):
                return None
            return item

    best_key, best_item, best_score = "", None, 0.0
    for k, item in cache.items():
        if isinstance(item, dict) and _ta_profile_cache_is_stale(item) and not allow_stale:
            continue
        pname = item.get("player", k) if isinstance(item, dict) else k
        sc, _reason = nombre_score_quality_direct(player_name, pname)
        if sc > best_score:
            best_key, best_item, best_score = k, item, sc
    if isinstance(best_item, dict) and best_score >= 0.94:
        out = best_item.copy()
        out["cache_key"] = best_key
        out["match_score"] = float(best_score)
        out["stale"] = _ta_profile_cache_is_stale(out)
        if out["stale"] and not allow_stale and not _ta_profile_cache_is_ta_real_permanent(out):
            return None
        return out
    return None


def _ta_profile_cache_num_matches(item):
    """Extrae nº de partidos de un perfil guardado/parseado de forma robusta."""
    try:
        if not isinstance(item, dict):
            return 0
        for k in ["matches", "n"]:
            v = item.get(k, "")
            if isinstance(v, (int, float)) and not pd.isna(v):
                return int(v)
            m = re.search(r"\d+", str(v or ""))
            if m:
                return int(m.group(0))
        summary = str(item.get("summary", "") or "")
        m = re.search(r"n\s*=\s*(\d+)", summary, flags=re.I) or re.search(r"(\d+)\s+partidos", summary, flags=re.I)
        return int(m.group(1)) if m else 0
    except Exception:
        return 0


def _ta_profile_cache_should_replace(old_item, new_item, force_manual=True):
    """Decide si una ficha nueva debe pisar la guardada.
    v23.46.9: si el usuario pega una ficha mejor, debe reemplazar perfiles parciales n=0.
    """
    if not isinstance(old_item, dict):
        return True
    old_n = _ta_profile_cache_num_matches(old_item)
    new_n = _ta_profile_cache_num_matches(new_item)
    old_raw_len = len(str(old_item.get("raw_text", "") or ""))
    new_raw_len = len(str(new_item.get("raw_text", "") or ""))

    # v23.49.2: una TA Real/Manual permanente nunca puede ser pisada por Auto Recent.
    old_pr = _ta_profile_cache_source_priority(old_item)
    new_pr = _ta_profile_cache_source_priority(new_item)
    if old_pr >= 80 and new_pr < old_pr:
        return False

    # v23.49.1: una ficha TennisAbstract completa siempre pisa Auto Recent o fichas antiguas.
    try:
        if int((new_item or {}).get("ta_full_score", 0) or 0) >= 3:
            return True
    except Exception:
        pass

    # Regla principal: una ficha con más partidos siempre mejora una ficha parcial.
    if new_n > old_n:
        return True
    # Si la guardada no tenía partidos, permite reemplazar con texto más completo aunque el parser dé igual.
    if old_n <= 0 and new_raw_len >= old_raw_len:
        return True
    # Si el usuario pega ficha manual, permitimos actualizar el texto mientras no sea claramente peor.
    if force_manual and new_n >= old_n and new_raw_len >= max(80, int(old_raw_len * 0.75)):
        return True
    return False


def _ta_profile_cache_upsert_from_raw(expected_player, raw_text, parsed=None):
    """Guarda ficha válida con fecha y sobrescribe perfiles parciales con fichas mejores.
    v23.46.9: antes podía quedar guardado un perfil n=0 y no pisarse al pegar una ficha completa.
    Ahora guarda también bajo la clave del jugador esperado, para que la búsqueda exacta recupere la ficha buena.
    """
    raw_text = str(raw_text or "").strip()
    if len(raw_text) < 80:
        return False, "sin texto suficiente"
    d = _extraer_datos_extra_desde_texto(raw_text, None)
    if not isinstance(d, dict):
        return False, "no interpretado"
    # Si el texto no es TA ni Flashscore, no lo guardamos como perfil persistente.
    if not isinstance(d.get("tennisabstract"), dict) and not isinstance(d.get("flashscore"), dict):
        return False, "no es ficha TA/Flashscore reconocida"
    ok, msg = _ficha_corresponde_jugador(d, expected_player)
    if not ok:
        return False, msg

    parsed_obj = d.get("tennisabstract") or d.get("flashscore") or {}
    player = str(parsed_obj.get("player") or expected_player or "").strip()
    expected_player = str(expected_player or player or "").strip()

    # Claves donde guardamos: nombre esperado del partido y nombre detectado de la ficha.
    # Esto evita que quede una entrada antigua exacta (n=0) tapando una nueva buena con país/formato distinto.
    keys = []
    for nm in [expected_player, player]:
        k = _ta_profile_cache_key(nm)
        if k and k not in keys:
            keys.append(k)
    if not keys:
        return False, "sin nombre"

    new_item = {
        "player": player or expected_player,
        "expected_player": expected_player,
        "source": parsed_obj.get("source", "Tennis Abstract/Flashscore"),
        "uploaded_at": _ta_profile_cache_today(),
        "raw_text": raw_text,
        "summary": parsed_obj.get("summary", ""),
        "matches": parsed_obj.get("matches", ""),
        "over_rate": parsed_obj.get("over_rate", ""),
        "set_rate": parsed_obj.get("set_rate", ""),
        "three_rate": parsed_obj.get("three_rate", ""),
        "elo": parsed_obj.get("elo", ""),
        # v23.48.3 auditoría: guardamos también qué datos reales venían en la ficha.
        # OJO: matches = partidos recientes usados por el modelo (máx. 12), no histórico total.
        "recent_used": parsed_obj.get("matches", ""),
        "challenger_2026_matches": (parsed_obj.get("challenger_2026", {}) or {}).get("matches", ""),
        "ta_player_detected": parsed_obj.get("player", ""),
        "ta_full_score": parsed_obj.get("ta_full_score", 0),
        "ta_full": parsed_obj.get("ta_full", {}),
        "ta_best_hard": parsed_obj.get("ta_best_hard", {}),
        "ta_best_clay": parsed_obj.get("ta_best_clay", {}),
        "ta_best_grass": parsed_obj.get("ta_best_grass", {}),
        "ta_year_current": (parsed_obj.get("ta_full", {}) or {}).get("year_end_current", {}),
        "tour_2026": (parsed_obj.get("ta_full", {}) or {}).get("tour_seasons", {}).get("2026", {}),
        "challenger_2026_full": (parsed_obj.get("ta_full", {}) or {}).get("challenger_seasons", {}).get("2026", {}),
        "career_stats": (parsed_obj.get("ta_full", {}) or {}).get("career_splits", {}),
        "last52_stats": (parsed_obj.get("ta_full", {}) or {}).get("last52_splits", {}),
        "profile_source_type": "TA Real Completo" if int(parsed_obj.get("ta_full_score", 0) or 0) >= 3 else parsed_obj.get("source", ""),
        "permanent_ta": True if (int(parsed_obj.get("ta_full_score", 0) or 0) >= 3 or "Tennis Abstract" in str(parsed_obj.get("source", ""))) else False,
        "expires_policy": "permanent_use_stale_only_warning",
    }

    cache = _ta_profile_cache_load()
    replaced = 0
    skipped = 0
    for key in keys:
        old_item = cache.get(key)
        if _ta_profile_cache_should_replace(old_item, new_item, force_manual=True):
            cache[key] = new_item.copy()
            replaced += 1
        else:
            skipped += 1

    ok_save, info = _ta_profile_cache_save(cache)
    if not ok_save:
        return False, f"no guardado: {info}"

    n = _ta_profile_cache_num_matches(new_item)
    extra = f" · n={n}" if n else " · perfil parcial"
    if skipped and not replaced:
        return True, f"perfil ya era igual/mejor para {player or expected_player}{extra} ({new_item['uploaded_at']})"
    return True, f"perfil guardado/actualizado para {player or expected_player}{extra} ({new_item['uploaded_at']})"


def _ta_profile_cache_status_text(player_name):
    item = _ta_profile_cache_find(_ta_alias_resolve(player_name), allow_stale=True) or _ta_profile_cache_find(player_name, allow_stale=True)
    if not item:
        return "", None
    uploaded = item.get("uploaded_at", "")
    src = item.get("source", "perfil")
    player = item.get("player", player_name)
    age = _ta_profile_cache_age_label(uploaded)
    stale = _ta_profile_cache_is_stale(item)
    summary = str(item.get("summary", ""))[:180]
    icon = "♻️" if stale else "💾"
    txt = f"{icon} Perfil guardado · {player} · {src} · {uploaded or 'sin fecha'} ({age})"
    if stale:
        if _ta_profile_cache_is_ta_real_permanent(item):
            txt += f" · ANTIGUA >{TA_PROFILE_MAX_AGE_DAYS}d: se puede refrescar, pero se sigue usando y NO se borra"
        else:
            txt += f" · CADUCADO >{TA_PROFILE_MAX_AGE_DAYS}d: se volverá a pedir/buscar"
    if summary:
        txt += f" · {summary}"
    return txt, item


# =========================================================
# v23.43.2 AUTO TA MATCH REINFORCER
# Descarga de forma controlada fichas TennisAbstract para el bloque
# "Datos extra" de picks Over. No toca el motor Over: solo rellena
# la caché de fichas que antes se pegaban a mano.
# =========================================================

def _ta_match_fetch_url_v23432(url, timeout_sec=22):
    """Fetch sin cachear errores.
    v23.43.3: la versión anterior usaba st.cache_data y podía dejar cacheado
    un timeout durante 24h. Ahora reintenta y no arrastra fallos antiguos.
    """
    last = ""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 TennisIA/AutoTAReinforcer",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
        "Connection": "close",
    }
    for attempt in range(2):
        try:
            resp = requests.get(str(url), timeout=max(8, int(timeout_sec)), headers=headers)
            if resp.status_code != 200:
                last = f"HTTP {resp.status_code}"
                continue
            txt = resp.text or ""
            if not txt.strip():
                last = "respuesta vacía"
                continue
            return txt, "OK"
        except Exception as e:
            last = f"{type(e).__name__}: {e}"
            time.sleep(0.4)
    return "", last or "sin respuesta"


def _ta_decode_jsfrag_to_htmlish_v234915(js_text):
    """v23.49.15: TennisAbstract carga casi todas las tablas desde /jsfrags/*.js.
    El HTML principal sí descarga bien, pero solo trae cabecera y scripts. Esta función
    convierte los fragmentos JS escapados a HTML/texto para que el parser pueda ver tablas.
    No toca cálculos: solo recuperación de datos TA.
    """
    raw = str(js_text or "")
    if not raw.strip():
        return ""
    variants = [raw]
    try:
        # Muchos jsfrags incluyen HTML escapado dentro de strings JS.
        variants.append(bytes(raw, "utf-8").decode("unicode_escape", errors="ignore"))
    except Exception:
        pass
    out = []
    for txt in variants:
        try:
            t = txt
            t = t.replace('\\/', '/')
            t = t.replace('\\n', '\n').replace('\\t', '\t').replace('\\r', '\n')
            t = t.replace('\x3c', '<').replace('\x3e', '>').replace('\u003c', '<').replace('\u003e', '>')
            t = t.replace("\\'", "'").replace('\\"', '"')
            out.append(t)
            # Extrae strings largos que contengan HTML o títulos de secciones.
            for m in re.finditer(r"(['\"])(.*?)(?<!\\)\1", t, flags=re.S):
                frag = m.group(2)
                if len(frag) < 20:
                    continue
                low = frag.lower()
                if any(k in low for k in ['<table', '<tr', '<td', 'tour-level', 'challenger', 'last 52', 'year-end', 'recent results', 'match-results']):
                    out.append(frag)
        except Exception:
            continue
    return "\n".join(out)


def _ta_fetch_jsfrags_from_html_v234915(html_text, timeout_sec=18):
    """Descarga scripts /jsfrags/ enlazados por la página TA.
    Es la pieza que faltaba: player.cgi contiene poco texto y las tablas van en JS externo.
    """
    html_raw = str(html_text or "")
    urls = []
    for src in re.findall(r'<script[^>]+src=["\']([^"\']+)["\']', html_raw, flags=re.I):
        src = src.strip()
        if not src:
            continue
        if 'jsfrags/' not in src and '/cgi-bin/jsmatches/' not in src:
            continue
        if src.startswith('//'):
            src = 'https:' + src
        elif src.startswith('/'):
            src = 'https://www.tennisabstract.com' + src
        elif not src.startswith('http'):
            src = 'https://www.tennisabstract.com/' + src.lstrip('/')
        if src not in urls:
            urls.append(src)
    pieces = []
    debug = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 TennisIA/TAJSFrag",
        "Accept": "*/*",
        "Referer": "https://www.tennisabstract.com/",
        "Connection": "close",
    }
    for u in urls[:6]:
        try:
            r = requests.get(u, timeout=max(8, int(timeout_sec)), headers=headers)
            status = f"HTTP {getattr(r, 'status_code', '')}"
            txt = r.text or ""
            debug.append(f"{os.path.basename(u)}:{status}:chars={len(txt)}")
            if getattr(r, 'status_code', 0) == 200 and txt.strip():
                pieces.append(f"\n<!-- TA_JSFRAG {u} -->\n" + _ta_decode_jsfrag_to_htmlish_v234915(txt))
                time.sleep(0.25)
        except Exception as e:
            debug.append(f"{os.path.basename(u)}:{type(e).__name__}")
    if debug:
        pieces.append("\nTA_JSFRAG_DEBUG\n" + "\n".join(debug) + "\n")
    return "\n".join(pieces)


def _ta_match_html_to_profile_text_v23432(html_text, expected_player=""):
    """Convierte HTML de TennisAbstract a texto parseable.
    v23.49.15: además descarga y añade /jsfrags/*.js, porque las tablas actuales
    no están en el HTML principal sino en scripts externos.
    """
    try:
        import html as _html
        raw0 = str(html_text or "")
        js_extra = _ta_fetch_jsfrags_from_html_v234915(raw0)
        raw = raw0 + "\n" + js_extra

        def _table_to_tsv(table_html):
            try:
                rows = []
                for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", table_html, flags=re.I | re.S):
                    cells = []
                    for cell in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, flags=re.I | re.S):
                        c = re.sub(r"<br\s*/?>", " ", cell, flags=re.I)
                        c = re.sub(r"<[^>]+>", " ", c)
                        c = _html.unescape(c)
                        c = re.sub(r"\s+", " ", c).strip()
                        cells.append(c)
                    if cells:
                        rows.append("\t".join(cells))
                return "\n".join(rows)
            except Exception:
                return ""

        # Antes de eliminar scripts, ya hemos añadido el contenido de jsfrags.
        raw = re.sub(r"<table[^>]*>.*?</table>", lambda m: "\n" + _table_to_tsv(m.group(0)) + "\n", raw, flags=re.I | re.S)
        txt = raw
        txt = re.sub(r"<style[^>]*>.*?</style>", " ", txt, flags=re.I | re.S)
        # Quita solo scripts del HTML original; los JSFRAGS ya están fuera en texto/htmlish.
        txt = re.sub(r"<script[^>]*>.*?</script>", " ", txt, flags=re.I | re.S)
        txt = re.sub(r"</t[dh]>\s*<t[dh][^>]*>", "\t", txt, flags=re.I)
        txt = re.sub(r"<t[dh][^>]*>", "", txt, flags=re.I)
        txt = re.sub(r"</t[dh]>", "\t", txt, flags=re.I)
        txt = re.sub(r"</tr>", "\n", txt, flags=re.I)
        txt = re.sub(r"<br\s*/?>", "\n", txt, flags=re.I)
        txt = re.sub(r"</p>|</div>|</h\d>|</li>", "\n", txt, flags=re.I)
        txt = re.sub(r"<[^>]+>", " ", txt)
        txt = _html.unescape(txt)
        txt = txt.replace('\\n', '\n').replace('\\t', '\t')
        lines = []
        for line in txt.splitlines():
            parts = [re.sub(r"\s+", " ", x).strip() for x in line.split("\t")]
            line2 = "\t".join([x for x in parts if x])
            line2 = re.sub(r"[ \u00a0]+", " ", line2).strip()
            if line2:
                lines.append(line2)
        body = "\n".join(lines)
        header = normalizar_texto(expected_player or "").strip()
        if header:
            return f"{header}\nTennis Abstract\n{body}"
        return f"Tennis Abstract\n{body}"
    except Exception:
        return str(html_text or "")


def _ta_match_profile_urls_v23432(name):
    """URLs candidatas de ficha TA.
    v23.49.3: primero intenta resolver la URL real desde el leaderboard Elo de TA,
    luego prueba variantes directas con/sin www. Esto arregla jugadores conocidos
    que antes salían como Sin ficha pese a existir en TennisAbstract.
    """
    urls = []
    # 0) Alias manual primero.
    name0 = normalizar_texto(name)
    name1 = _ta_alias_resolve(name0)
    for nm in [name1, name0]:
        for u in _ta_match_urls_from_elo_leaderboard_v23493(nm, circuito="ATP", min_score=0.88):
            if u and u not in urls:
                urls.append(u)

    # 1) Generador existente.
    for nm in [name1, name0]:
        try:
            for u in _auto_profile_direct_urls_v23431(nm):
                if u and u not in urls:
                    urls.append(u)
        except Exception:
            pass

    # 2) Variantes directas robustas.
    for nm in [name1, name0]:
        toks = tokens(nm)
        if not toks:
            continue
        title_tokens = [str(t).title() for t in toks]
        variants = []
        variants.append("".join(title_tokens))              # AlexanderZverev
        variants.append("-".join(title_tokens))             # Alexander-Zverev
        variants.append("".join(toks).title())              # NZverev / NKyrgios si abreviado
        # Si viene "Apellido Inicial" o "Inicial Apellido", probamos apellido+inicial, pero no como única opción.
        if len(toks) >= 2:
            variants.append(toks[-1].title() + toks[0][0].title())
            variants.append(toks[0][0].title() + toks[-1].title())
        for v in variants:
            if not v:
                continue
            for host in ["https://www.tennisabstract.com", "https://tennisabstract.com"]:
                u = f"{host}/cgi-bin/player.cgi?p={v}"
                if u not in urls:
                    urls.append(u)
    return urls[:14]


def _ta_match_auto_cache_one_player_v23432(player_name, timeout_sec=22):
    """Busca la ficha TA y, si TA no responde/no parsea, crea ficha de refuerzo
    desde histórico local Challenger/WTA/ATP disponible. Esto evita tener que pegar
    20 fichas manuales cuando la web se bloquea o tarda.
    """
    name = normalizar_texto(player_name)
    if not name:
        return False, "sin nombre"
    existing = _ta_profile_cache_find(name, allow_stale=True)
    existing_priority = _ta_profile_cache_source_priority(existing) if isinstance(existing, dict) else 0
    # Si ya tenemos TA Real/Manual permanente, no la tocamos.
    if existing and str(existing.get("raw_text", "")).strip() and existing_priority >= 80:
        return True, f"ya estaba en caché TA Real/Manual permanente: {existing.get('player', name)}"

    errors = []
    if existing and str(existing.get("raw_text", "")).strip() and existing_priority < 80:
        errors.append("existía Auto Recent; se intenta mejorar a TA Real")
    # 1) Intento real TennisAbstract.
    for url in _ta_match_profile_urls_v23432(name):
        html_text, status = _ta_match_fetch_url_v23432(url, timeout_sec=timeout_sec)
        if not html_text:
            errors.append(f"{url.split('=')[-1]}: {status}")
            continue
        low = html_text.lower()
        if "player not found" in low or "no player found" in low:
            errors.append("TA: jugador no encontrado")
            continue
        profile_text = _ta_match_html_to_profile_text_v23432(html_text, name)
        parsed = _parse_tennisabstract_player_profile(profile_text)
        if parsed and (int(parsed.get("matches", 0) or 0) >= 1 or int(parsed.get("ta_full_score", 0) or 0) >= 3):
            # Guardamos el texto completo descargado y dejamos que el upsert decida si pisa Auto Recent.
            ok, msg = _ta_profile_cache_upsert_from_raw(name, profile_text, parsed=parsed)
            if ok:
                full = int(parsed.get("ta_full_score", 0) or 0)
                return True, f"{name}: TA Real guardada desde URL ({parsed.get('matches', 0)} recientes, full={full})"
            errors.append(f"TA leída pero no guardada: {msg}")
        elif parsed:
            ok, msg = _ta_profile_cache_upsert_from_raw(name, profile_text, parsed=parsed)
            if ok:
                return True, f"{name}: ficha TA básica guardada"
            errors.append(f"TA básica no guardada: {msg}")
        else:
            # Si descargó página pero el formato no trae Recent Results parseable,
            # no la damos por buena para Over. Pasamos a fallback local.
            errors.append("TA descargada pero sin Recent Results parseables")

    # 2) Fallback local: usa los históricos que ya tiene la app para generar una ficha automática.
    # Esto NO toca el motor Over; solo rellena la caja de refuerzo con evidencia local.
    try:
        # Intento por superficie genérica; el propio buscador elige últimos partidos.
        st_ch = _ch_recent_stats(name, surface=None, limit=12)
        if isinstance(st_ch, dict) and st_ch.get("found") and int(st_ch.get("n", 0) or 0) >= 3:
            raw = (
                f"Auto Challenger Recent Form\n"
                f"Player: {st_ch.get('player', name)}\n"
                f"n={int(st_ch.get('n', 0) or 0)}; wins={int(st_ch.get('wins', 0) or 0)}; "
                f"over18={int(st_ch.get('over18', 0) or 0)}; over19={int(st_ch.get('over19', 0) or 0)}; "
                f"three_sets={int(st_ch.get('three_sets', 0) or 0)}; set_won={int(st_ch.get('set_won', 0) or 0)}; "
                f"lost02={int(st_ch.get('lost_02', 0) or 0)}; short={int(st_ch.get('short', 0) or 0)}; "
                f"avg_games={float(st_ch.get('avg_games', 0.0) or 0.0):.1f}\n"
                f"Summary: {st_ch.get('summary', '')}\n"
                f"Nota: TennisAbstract no respondió o no fue parseable; se usa histórico local como refuerzo provisional."
            )
            parsed_local = _parse_auto_recent_profile_v23433(raw) or {}
            ok, msg = _ta_profile_cache_upsert_auto_recent_v23433(name, parsed_local, raw)
            if ok:
                return True, f"{name}: {msg} (TA falló; fallback local n={st_ch.get('n')})"
            errors.append(msg)
    except Exception as e:
        errors.append(f"fallback local error {type(e).__name__}: {e}")

    detail = " | ".join(errors[-3:]) if errors else "sin coincidencia"
    return False, f"{name}: no extraída ({detail})"


def _ta_match_players_from_extra_candidates_v23432(cand_df):
    names = []
    try:
        for _, row in (cand_df if cand_df is not None else pd.DataFrame()).iterrows():
            partido = str(row.get("Partido", ""))
            if " vs " in partido:
                p1, p2 = [x.strip() for x in partido.split(" vs ", 1)]
                names.extend([p1, p2])
    except Exception:
        pass
    # únicos conservando orden
    out, seen = [], set()
    for n in names:
        k = limpiar(n)
        if k and k not in seen:
            out.append(n); seen.add(k)
    return out


def _ta_match_cache_status_counts_v23432(players):
    ok, missing = [], []
    for n in players or []:
        canon = _ta_alias_resolve(n)
        item = _ta_profile_cache_find(canon) or _ta_profile_cache_find(n)
        if item and str(item.get("raw_text", "")).strip():
            ok.append(n)
        else:
            missing.append(n)
    return ok, missing




def _ta_profile_source_kind_v23485(item):
    """Clasifica una ficha guardada para no confundir TA real con Auto Recent.
    No toca motor Over; solo auditoría y guardado manual.
    """
    if not isinstance(item, dict) or not str(item.get("raw_text", "") or "").strip():
        return "sin", "🔴 Sin ficha"
    src = str(item.get("source", "") or "")
    raw = str(item.get("raw_text", "") or "")
    low = (src + "\n" + raw[:500]).lower()
    if int(item.get("ta_full_score", 0) or 0) >= 3 and ("tennis abstract" in low or "manual" in low):
        return "ta", "🟢 TA Real Completa 🔒"
    if "manual" in low and ("tennis abstract" in low or "flashscore" in low):
        return "manual", "📋 Manual"
    if "auto recent" in low or "auto challenger recent form" in low or "histórico local" in low or "historico local" in low:
        return "auto", "🟠 Auto Recent"
    if "tennis abstract" in low or "tennisabstract" in low or "recent results" in low or "elo rank" in low:
        return "ta", "🟢 TA Real"
    if "flashscore" in low:
        return "manual", "📋 Manual/Flashscore"
    return "otro", "⚪ Perfil guardado"


def _ta_profile_cache_upsert_manual_v23485(expected_player, raw_text):
    """Guarda ficha pegada manualmente y la marca como Manual/TA si se reconoce.
    Mantiene la validación de jugador de _ta_profile_cache_upsert_from_raw.
    """
    ok, msg = _ta_profile_cache_upsert_from_raw(expected_player, raw_text)
    if not ok:
        return ok, msg
    try:
        cache = _ta_profile_cache_load()
        for nm in [expected_player]:
            k = _ta_profile_cache_key(nm)
            if k in cache and isinstance(cache[k], dict):
                item = cache[k].copy()
                src_old = str(item.get("source", "") or "")
                if "Auto Recent" not in src_old:
                    if "manual" not in src_old.lower():
                        item["source"] = "Manual " + (src_old or "TennisAbstract")
                item["manual_saved_at"] = _ta_profile_cache_today()
                item["permanent_ta"] = True
                item["expires_policy"] = "manual_permanent"
                cache[k] = item
        _ta_profile_cache_save(cache)
    except Exception:
        pass
    return ok, msg + " · marcado como manual"

def _ta_profile_audit_detail_v23483(item):
    """v23.48.3: auditoría honesta de fuente.
    - matches/recent_used = últimos partidos parseados y usados por el modelo (máx. 12).
    - challenger_2026_matches = total temporada Challenger 2026 si la ficha TA lo trae.
    - No inventa histórico total si la ficha no lo aporta.
    """
    detail = {
        "recent_used": 0,
        "season_total": 0,
        "reference_n": 0,
        "source_label": "Sin fuente",
        "quality": "⚪ Sin datos",
        "note": "",
    }
    if not isinstance(item, dict):
        return detail

    raw = str(item.get("raw_text", "") or "")
    src = str(item.get("source", "") or "")

    def _to_int(v):
        try:
            if v is None or v == "":
                return 0
            if isinstance(v, (int, float)) and not pd.isna(v):
                return int(v)
            m = re.search(r"\d+", str(v))
            return int(m.group(0)) if m else 0
        except Exception:
            return 0

    recent_used = _to_int(item.get("recent_used", item.get("matches", item.get("n", 0))))
    season_total = _to_int(item.get("challenger_2026_matches", 0))

    # Si la entrada antigua no guardaba estos campos, intentamos reconstruir desde raw_text.
    parsed = None
    try:
        if raw and ("Tennis Abstract" in raw or "Recent Results" in raw or "Elo rank" in raw):
            parsed = _parse_tennisabstract_player_profile(raw)
        elif raw and ("Auto Challenger Recent Form" in raw or "Auto Recent Form" in raw):
            parsed = _parse_auto_recent_profile_v23433(raw)
    except Exception:
        parsed = None

    if isinstance(parsed, dict):
        recent_used = max(recent_used, _to_int(parsed.get("matches", parsed.get("n", 0))))
        ch = parsed.get("challenger_2026", {}) or {}
        season_total = max(season_total, _to_int(ch.get("matches", 0)))

    reference_n = max(season_total, recent_used)

    if "Auto Recent" in src or "local" in src.lower() or "Auto Challenger Recent Form" in raw:
        source_label = "Histórico local reciente"
        note = "No es TennisAbstract completo; es refuerzo local reciente."
    elif season_total > 0:
        source_label = "TennisAbstract temporada 2026"
        note = f"Usa {recent_used} recientes para señales; TA informa {season_total} partidos 2026."
    elif recent_used > 0:
        source_label = "TennisAbstract Recent Results"
        note = "TA solo aporta/parsea recientes para el modelo; no hay total temporada disponible."
    else:
        source_label = src or "Perfil sin partidos"
        note = "Ficha guardada, pero sin partidos parseables."

    if reference_n >= 50:
        quality = "🟢 Sólida"
    elif reference_n >= 20:
        quality = "🟡 Media"
    elif reference_n >= 8:
        quality = "🟠 Reciente limitada"
    elif reference_n >= 1:
        quality = "🔴 Muy limitada"
    else:
        quality = "⚪ Sin datos"

    detail.update({
        "recent_used": int(recent_used),
        "season_total": int(season_total),
        "reference_n": int(reference_n),
        "source_label": source_label,
        "quality": quality,
        "note": note,
    })
    return detail

def _ta_match_cache_audit_rows_v23481(players):
    """Auditoría visible de caché TA por jugador detectado.
    v23.48.3: deja claro si los números son recientes usados por el modelo o total temporada TA.
    No toca motor Over; solo visualización y diagnóstico.
    """
    rows = []
    for n in players or []:
        detected = normalizar_texto(n)
        canon = _ta_alias_resolve(detected)
        alias_used = canon != detected
        item = _ta_profile_cache_find(canon, allow_stale=True) or _ta_profile_cache_find(detected, allow_stale=True)
        if isinstance(item, dict) and str(item.get("raw_text", "")).strip():
            stale = bool(item.get("stale", _ta_profile_cache_is_stale(item)))
            detail = _ta_profile_audit_detail_v23483(item)
            source_kind, source_label_visible = _ta_profile_source_kind_v23485(item)
            recent_used = int(detail.get("recent_used", 0) or 0)
            season_total = int(detail.get("season_total", 0) or 0)
            ref_n = int(detail.get("reference_n", 0) or 0)
            if stale:
                status = "♻️ Caducada >7d"
            elif ref_n >= 8:
                status = "✅ OK"
            elif ref_n >= 1:
                status = "⚠️ Parcial"
            else:
                status = "⚠️ Sin partidos"
            rows.append({
                "Jugador detectado": detected,
                "Nombre usado": canon,
                "Alias": "Sí" if alias_used else "No",
                "Estado ficha": status,
                "Fuente perfil": source_label_visible,
                "Tipo fuente": source_kind,
                "Calidad": detail.get("quality", ""),
                "Recientes usados": recent_used,
                "Total temporada TA": season_total,
                "Referencia datos": ref_n,
                "Fuente conteo": detail.get("source_label", ""),
                "Edad": _ta_profile_cache_age_label(item.get("uploaded_at", "")),
                "Fecha ficha": item.get("uploaded_at", ""),
                "Jugador ficha": item.get("player", ""),
                "Nota": detail.get("note", ""),
                "Cache key": item.get("cache_key", ""),
            })
        else:
            rows.append({
                "Jugador detectado": detected,
                "Nombre usado": canon,
                "Alias": "Sí" if alias_used else "No",
                "Estado ficha": "❌ Pendiente",
                "Fuente perfil": "🔴 Sin ficha",
                "Tipo fuente": "sin",
                "Calidad": "⚪ Sin datos",
                "Recientes usados": 0,
                "Total temporada TA": 0,
                "Referencia datos": 0,
                "Fuente conteo": "Sin ficha",
                "Edad": "",
                "Fecha ficha": "",
                "Jugador ficha": "",
                "Nota": "No hay ficha en caché para este nombre/alias.",
                "Cache key": "",
            })
    return rows


def _ta_match_cache_audit_summary_v23481(players):
    rows = _ta_match_cache_audit_rows_v23481(players)
    total = len(rows)
    ok = sum(1 for r in rows if str(r.get("Estado ficha", "")).startswith("✅"))
    partial = sum(1 for r in rows if str(r.get("Estado ficha", "")).startswith("⚠️"))
    stale = sum(1 for r in rows if "Caducada" in str(r.get("Estado ficha", "")))
    missing = sum(1 for r in rows if str(r.get("Estado ficha", "")).startswith("❌"))
    aliases = sum(1 for r in rows if r.get("Alias") == "Sí")
    ta_real = sum(1 for r in rows if r.get("Tipo fuente") == "ta")
    auto_recent = sum(1 for r in rows if r.get("Tipo fuente") == "auto")
    manual = sum(1 for r in rows if r.get("Tipo fuente") == "manual")
    otros = sum(1 for r in rows if r.get("Tipo fuente") == "otro")
    cobertura = (100.0 * (total - missing) / total) if total else 0.0
    return rows, {
        "total": total, "ok": ok, "partial": partial, "stale": stale,
        "missing": missing, "aliases": aliases, "cobertura": cobertura,
        "ta_real": ta_real, "auto_recent": auto_recent, "manual": manual, "otros": otros
    }


def _inferir_jugador_mercado(row):
    """Intenta saber qué jugador es el elegido en mercados tipo 'X gana al menos 1 set'."""
    mercado = str(row.get("🎯 Mercado más probable", ""))
    partido = str(row.get("Partido", ""))
    if " vs " in partido:
        p1, p2 = [x.strip() for x in partido.split(" vs ", 1)]
    else:
        p1, p2 = "", ""
    m_clean = limpiar(mercado)
    p1_clean = limpiar(p1)
    p2_clean = limpiar(p2)
    if p1_clean and p1_clean in m_clean:
        return "J1", p1
    if p2_clean and p2_clean in m_clean:
        return "J2", p2
    return "", ""


def _evaluar_datos_extra_real(row, data):
    """Convierte campos reales visibles de Datos extra en confirmar/neutral/descartar.
    No usa cuotas. Solo refuerza o frena la acción final.
    """
    mercado = str(row.get("🎯 Mercado más probable", "")).upper()
    accion_old = str(row.get("🎯 Acción final", ""))
    try:
        prob = leer_porcentaje(row.get("🎯 Prob máxima", 0), 0)
    except Exception:
        prob = 0

    j1_v = data.get("j1_v", None)
    j2_v = data.get("j2_v", None)
    largos = str(data.get("largos", ""))
    palizas = str(data.get("palizas", ""))
    h2h = str(data.get("h2h", ""))
    manual = str(data.get("manual", "Auto"))

    if manual.startswith("✅"):
        return "confirmar", "Confirmado manualmente por revisión Datos extra"
    if manual.startswith("❌"):
        return "descartar", "Descartado manualmente por revisión Datos extra"
    if manual.startswith("⚠️"):
        return "neutral", "Datos extra neutral manual"

    # Tennis Abstract All-Markets: si hay dos fichas TA, puede cambiar el mercado candidato,
    # no solo confirmar el Over inicial.
    ta_sig = data.get("ta_all_markets") if isinstance(data, dict) else None
    if ta_sig:
        estado = ta_sig.get("estado", "neutral")
        motivo = ta_sig.get("motivo", "Tennis Abstract revisado")
        if estado == "confirmar":
            return "confirmar", "Tennis Abstract confirma mejor mercado: " + motivo
        return "neutral", "Tennis Abstract revisa todos los mercados: " + motivo

    # Si no hay datos suficientes, no tocar.
    if j1_v is None and j2_v is None and not largos and not palizas and not h2h:
        return "sin_revisar", "Sin datos Datos extra"

    # Mercado Over: lo que más nos interesa es que haya partidos largos y pocas palizas.
    if "OVER 18.5" in mercado or "OVER18" in mercado:
        if largos.startswith("Muchos") and not palizas.startswith("Muchas"):
            return "confirmar", "Datos extra confirma patrón de partidos largos"
        if largos.startswith("Algunos") and prob >= 0.78 and not palizas.startswith("Muchas"):
            return "confirmar", "Datos extra apoya Over: algunos largos + probabilidad alta"
        if palizas.startswith("Muchas") or largos.startswith("Pocos"):
            return "descartar", "Datos extra no apoya Over: palizas o pocos marcadores largos"
        return "neutral", "Datos extra no confirma ni descarta claramente el Over"

    # Mercado gana set: mirar forma del jugador elegido y si pierde fácil.
    lado, nombre = _inferir_jugador_mercado(row)
    chosen_wins = None
    if lado == "J1":
        chosen_wins = j1_v
    elif lado == "J2":
        chosen_wins = j2_v

    if "GANA AL MENOS 1 SET" in mercado or "GANA SET" in mercado:
        if palizas.startswith("Jugador elegido pierde fácil"):
            return "descartar", "Datos extra muestra riesgo: el jugador elegido pierde fácil 2-0"
        if chosen_wins is not None:
            if chosen_wins >= 6 and prob >= 0.88:
                return "confirmar", f"Datos extra confirma forma del elegido: {chosen_wins}/10"
            if chosen_wins <= 2:
                return "descartar", f"Datos extra no confirma: elegido solo {chosen_wins}/10"
        if largos.startswith("Muchos") and prob >= 0.90:
            return "confirmar", "Datos extra apoya que compita set: marcadores largos + probabilidad alta"
        return "neutral", "Datos extra no confirma suficiente el gana-set"

    # Resto mercados: solo manual o neutral.
    return "neutral", "Datos extra revisado, sin señal automática fuerte"


def aplicar_reanalisis_datos_extra_manual(df, ajustes):
    """Aplica datos reales visibles de Datos extra a la tabla ya analizada.
    No recalcula motores. Solo sube/baja la acción final y deja trazabilidad.
    ajustes: dict index -> dict(j1_v, j2_v, largos, palizas, h2h, manual, nota)
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    for c in ["📥 Estado Datos extra", "📥 Ajuste Datos extra"]:
        if c not in out.columns:
            out[c] = ""

    for idx, data in (ajustes or {}).items():
        if idx not in out.index:
            continue

        estado, motivo_estado = _evaluar_datos_extra_real(out.loc[idx], data)
        if estado == "sin_revisar":
            continue

        accion_old = str(out.at[idx, "🎯 Acción final"] if "🎯 Acción final" in out.columns else "")
        ajuste = [motivo_estado]

        if estado == "confirmar":
            out.at[idx, "📥 Estado Datos extra"] = "✅ Confirmado"
            out.at[idx, "🎯 Acción final"] = "✅ JUGAR"
            ta_sig = data.get("ta_all_markets") if isinstance(data, dict) else None
            if ta_sig and ta_sig.get("nuevo_mercado"):
                if "🎯 Mercado más probable" in out.columns:
                    out.at[idx, "🎯 Mercado más probable"] = ta_sig.get("nuevo_mercado")
                if "🎯 Prob máxima" in out.columns and ta_sig.get("nueva_prob") is not None:
                    out.at[idx, "🎯 Prob máxima"] = f"{float(ta_sig.get('nueva_prob')):.1%}"
            if "🎯 Decisión acierto" in out.columns:
                out.at[idx, "🎯 Decisión acierto"] = (data.get("ta_all_markets", {}) or {}).get("decision", "🔥 Alto acierto + Datos extra")
            if "🎯 Aviso acierto" in out.columns:
                out.at[idx, "🎯 Aviso acierto"] = "Datos extra/Tennis Abstract confirma datos recientes: subido o mantenido como JUGAR."
            ajuste.append(f"Sube/mantiene JUGAR desde {accion_old}")
        elif estado == "descartar":
            out.at[idx, "📥 Estado Datos extra"] = "❌ Descartado"
            out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
            if "🎯 Aviso acierto" in out.columns:
                out.at[idx, "🎯 Aviso acierto"] = "Datos extra no confirma: bajado a OBSERVAR."
            ajuste.append(f"Baja a OBSERVAR desde {accion_old}")
        else:
            out.at[idx, "📥 Estado Datos extra"] = "⚠️ Neutral"
            ta_sig = data.get("ta_all_markets") if isinstance(data, dict) else None
            if ta_sig and ta_sig.get("nuevo_mercado"):
                if "🎯 Mercado más probable" in out.columns:
                    out.at[idx, "🎯 Mercado más probable"] = ta_sig.get("nuevo_mercado")
                if "🎯 Prob máxima" in out.columns and ta_sig.get("nueva_prob") is not None:
                    out.at[idx, "🎯 Prob máxima"] = f"{float(ta_sig.get('nueva_prob')):.1%}"
                if "🎯 Decisión acierto" in out.columns:
                    out.at[idx, "🎯 Decisión acierto"] = ta_sig.get("decision", "👀 Observar + TA")
            # v23.37.28: si Datos extra/TA fue revisado pero NO confirma, nunca puede quedar como JUGAR.
            # Evita casos donde la app mantenía ✅ JUGAR aunque TA dijera neutral/no confirma.
            if "🎯 Acción final" in out.columns:
                out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
            if "🎯 Aviso acierto" in out.columns:
                out.at[idx, "🎯 Aviso acierto"] = "Datos extra/Tennis Abstract neutral: no confirma mercado jugable; bajado a OBSERVAR."
            ajuste.append("Baja/mantiene OBSERVAR: Datos extra no confirma lo suficiente para JUGAR")

        for label, key in [
            ("J1 últimos 10", "j1_v"), ("J2 últimos 10", "j2_v"),
            ("Marcadores largos", "largos"), ("Palizas 2-0", "palizas"),
            ("H2H", "h2h"), ("Nota", "nota")
        ]:
            val = data.get(key, "")
            if val not in [None, ""]:
                ajuste.append(f"{label}: {val}")
        out.at[idx, "📥 Ajuste Datos extra"] = " | ".join(map(str, ajuste))

    # Reaplicar candado Grand Slam BO5 después de Datos extra/TA por si intentó recuperar Over 18.5.
    out = aplicar_grand_slam_bo5_selector_final(out)
    return out


def _ocr_datos_extra_image(uploaded_file):
    """Lee una captura Datos extra con OCR opcional.
    Si pytesseract/tesseract no está instalado, devuelve aviso sin romper la app.
    """
    if uploaded_file is None:
        return "", ""
    try:
        from PIL import Image, ImageOps, ImageEnhance
        import pytesseract
    except Exception as e:
        return "", "OCR no disponible. Instala: pip install pytesseract pillow y Tesseract OCR en Windows."

    try:
        img = Image.open(uploaded_file).convert("RGB")
        # Preprocesado simple para capturas oscuras de Datos extra.
        gray = ImageOps.grayscale(img)
        gray = ImageEnhance.Contrast(gray).enhance(1.8)
        gray = ImageEnhance.Sharpness(gray).enhance(1.4)
        # Reescalar ayuda mucho con texto pequeño.
        w, h = gray.size
        if w < 1600:
            scale = 1600 / max(w, 1)
            gray = gray.resize((int(w * scale), int(h * scale)))
        txt = pytesseract.image_to_string(gray, lang="eng")
        return txt.strip(), ""
    except Exception as e:
        return "", f"No he podido leer la captura con OCR: {e}"





def _score_pairs_from_ta_score(score):
    """Parsea scores tipo '4-6 6-3 6-3' o '7-6(4) 7-5' en parejas de juegos.
    Devuelve solo juegos, ignora tie-break entre paréntesis.
    """
    pairs = []
    for part in str(score or '').split():
        part = re.sub(r"\([^)]*\)", "", part.strip())
        m = re.match(r"^(\d{1,2})-(\d{1,2})$", part)
        if not m:
            continue
        a, b = int(m.group(1)), int(m.group(2))
        if 0 <= a <= 7 and 0 <= b <= 7 and max(a, b) >= 4:
            pairs.append((a, b))
    return pairs

def _ta_valid_elo_rating(x):
    """Valida ratings Elo de Tennis Abstract.
    Importante: no debe confundir años 2019/2024/2025 con Elo.
    Solo se usa como rating si viene de un campo claramente identificado.
    """
    try:
        x = int(float(x))
    except Exception:
        return None
    return x if 800 <= x <= 2400 else None


def _ta_parse_current_elo_line(line):
    """Lee la fila Current de Year-End Rankings de Tennis Abstract.

    Formato esperado tras 'Current (fecha)':
    ATP Rank, Points, Elo Rank, Elo, hElo Rank, hElo, cElo Rank, cElo, gElo Rank, gElo.

    El bug anterior filtraba todos los números 1200-2300 y podía tomar años
    como 2024/2025/2019 como si fueran Elo. Aquí usamos posiciones, no rango bruto.
    """
    ln = normalizar_texto(line or "").strip()
    if not ln.startswith("Current ("):
        return None, None, None, None

    # Elimina la fecha Current (2026-05-25) para que 2026/05/25 no entren en la secuencia.
    ln2 = re.sub(r"^Current\s*\([^)]*\)", "", ln).strip()
    if not ln2:
        return None, None, None, None

    # Con pegado por tabs, Tennis Abstract conserva cada celda. Si no, también vale con espacios.
    toks = [t.strip() for t in re.split(r"\t+|\s{2,}", ln2) if str(t).strip()]
    nums = []
    for t in toks:
        # Celdas tipo '-' o vacías se saltan. Solo número entero puro.
        if re.fullmatch(r"\d{1,4}", t):
            nums.append(int(t))

    # Fallback si el navegador pegó todo con espacios simples.
    if len(nums) < 4:
        nums = [int(x) for x in re.findall(r"\b\d{1,4}\b", ln2)]

    # Posiciones: 0 ATP Rank, 1 Points, 2 Elo Rank, 3 Elo, 4 hElo Rank, 5 hElo, ...
    elo = _ta_valid_elo_rating(nums[3]) if len(nums) > 3 else None
    helo = _ta_valid_elo_rating(nums[5]) if len(nums) > 5 else None
    celo = _ta_valid_elo_rating(nums[7]) if len(nums) > 7 else None
    gelo = _ta_valid_elo_rating(nums[9]) if len(nums) > 9 else None

    return elo, helo, celo, gelo


def _ta_elo_is_suspicious_year(x):
    try:
        x = int(float(x))
    except Exception:
        return False
    return 1900 <= x <= 2035





# =========================================================
# v23.49.0 TENNISABSTRACT FULL PROFILE PARSER
# Lee secciones completas de TA (no solo Recent Results):
# Tour-Level Seasons, Challenger Seasons, Splits y Elo por superficie.
# No toca motor Over: solo enriquece caché/ficha para auditoría y futuras lecturas.
# =========================================================

def _ta_cell_pct_v23490(x, default=None):
    try:
        t = normalizar_texto(x).replace('%','').replace(',','.').strip()
        if t in ['', '-', 'nan']:
            return default
        return float(t)
    except Exception:
        return default


def _ta_cell_int_v23490(x, default=0):
    try:
        t = normalizar_texto(x).strip()
        if t in ['', '-', 'nan']:
            return default
        m = re.search(r"-?\d+", t)
        return int(m.group(0)) if m else default
    except Exception:
        return default


def _ta_split_cells_v23490(line):
    line = normalizar_texto(line or '').strip()
    if not line:
        return []
    if '\t' in line:
        return [c.strip() for c in line.split('\t') if str(c).strip()]
    # Fallback conservador: separa por 2+ espacios. No rompe nombres como "Best of 3" si hay tabs.
    return [c.strip() for c in re.split(r"\s{2,}", line) if str(c).strip()]


def _ta_parse_stats_row_v23490(cells, row_label_type='season'):
    """Parsea filas TA con columnas:
    Label/Year, M,W,L,Win%,Set W-L,Set%,Game W-L,Game%,TB W-L,TB%,MS,Hld%,Brk%,A%,DF%,1stIn,1st%,2nd%,SPW,RPW,TPW,DR,...
    """
    try:
        if not cells or len(cells) < 22:
            return None
        label = cells[0]
        out = {
            'label': label,
            'matches': _ta_cell_int_v23490(cells[1]),
            'wins': _ta_cell_int_v23490(cells[2]),
            'losses': _ta_cell_int_v23490(cells[3]),
            'win_pct': _ta_cell_pct_v23490(cells[4]),
            'set_wl': cells[5] if len(cells) > 5 else '',
            'set_pct': _ta_cell_pct_v23490(cells[6] if len(cells) > 6 else None),
            'game_wl': cells[7] if len(cells) > 7 else '',
            'game_pct': _ta_cell_pct_v23490(cells[8] if len(cells) > 8 else None),
            'tb_wl': cells[9] if len(cells) > 9 else '',
            'tb_pct': _ta_cell_pct_v23490(cells[10] if len(cells) > 10 else None),
            'ms': _ta_cell_int_v23490(cells[11] if len(cells) > 11 else None),
            'hold_pct': _ta_cell_pct_v23490(cells[12] if len(cells) > 12 else None),
            'break_pct': _ta_cell_pct_v23490(cells[13] if len(cells) > 13 else None),
            'ace_pct': _ta_cell_pct_v23490(cells[14] if len(cells) > 14 else None),
            'df_pct': _ta_cell_pct_v23490(cells[15] if len(cells) > 15 else None),
            'first_in_pct': _ta_cell_pct_v23490(cells[16] if len(cells) > 16 else None),
            'first_won_pct': _ta_cell_pct_v23490(cells[17] if len(cells) > 17 else None),
            'second_won_pct': _ta_cell_pct_v23490(cells[18] if len(cells) > 18 else None),
            'spw_pct': _ta_cell_pct_v23490(cells[19] if len(cells) > 19 else None),
            'rpw_pct': _ta_cell_pct_v23490(cells[20] if len(cells) > 20 else None),
            'tpw_pct': _ta_cell_pct_v23490(cells[21] if len(cells) > 21 else None),
            'dr': _ta_cell_pct_v23490(cells[22] if len(cells) > 22 else None),
        }
        return out if out['matches'] > 0 else None
    except Exception:
        return None


def _ta_parse_year_end_current_v23490(lines):
    out = {}
    for ln in lines or []:
        if not str(ln).startswith('Current ('):
            continue
        # Current (date) ATP Rank Points Elo Rank Elo hElo Rank hElo cElo Rank cElo gElo Rank gElo
        ln2 = re.sub(r"^Current\s*\([^)]*\)", "", normalizar_texto(ln)).strip()
        cells = _ta_split_cells_v23490(ln2)
        nums = []
        for c in cells:
            if re.fullmatch(r"\d{1,5}", str(c).strip()):
                nums.append(int(c))
        if len(nums) < 10:
            nums = [int(x) for x in re.findall(r"\b\d{1,5}\b", ln2)]
        if len(nums) >= 4:
            out['rank'] = nums[0]
            out['points'] = nums[1] if len(nums) > 1 else None
            out['elo_rank'] = nums[2] if len(nums) > 2 else None
            out['elo'] = _ta_valid_elo_rating(nums[3])
            out['hElo_rank'] = nums[4] if len(nums) > 4 else None
            out['hElo'] = _ta_valid_elo_rating(nums[5]) if len(nums) > 5 else None
            out['cElo_rank'] = nums[6] if len(nums) > 6 else None
            out['cElo'] = _ta_valid_elo_rating(nums[7]) if len(nums) > 7 else None
            out['gElo_rank'] = nums[8] if len(nums) > 8 else None
            out['gElo'] = _ta_valid_elo_rating(nums[9]) if len(nums) > 9 else None
            return out
    return out


def _ta_parse_full_sections_v23490(text, lines):
    """Devuelve dict TA completo con secciones estructuradas."""
    result = {
        'tour_seasons': {},
        'challenger_seasons': {},
        'career_splits': {},
        'last52_splits': {},
        'year_end_current': {},
        'full_score': 0,
    }
    section = None
    for ln in lines or []:
        if ln.startswith('Tour-Level Seasons'):
            section = 'tour'; continue
        if ln.startswith('Challenger Seasons'):
            section = 'challenger'; continue
        if ln.startswith('Career Tour-Level Splits') or ln.startswith('Career Challenger-Level Splits'):
            section = 'career_splits'; continue
        if ln.startswith('Last 52 Weeks Tour-Level Splits'):
            section = 'last52_splits'; continue
        if ln.startswith('Year-End Rankings'):
            section = 'year_end'; continue
        # cortes principales
        if any(ln.startswith(x) for x in ['Recent Titles', 'Major and Recent Events', 'Winners and Unforced Errors', 'Key Points', 'Match Charting', 'Most Frequent Head-to-Heads']):
            if section in ['tour','challenger','career_splits','last52_splits','year_end']:
                section = None
        cells = _ta_split_cells_v23490(ln)
        if section in ['tour','challenger'] and cells:
            label = cells[0]
            if label == 'Career' or re.fullmatch(r"20\d{2}", label):
                row = _ta_parse_stats_row_v23490(cells)
                if row:
                    result['tour_seasons' if section == 'tour' else 'challenger_seasons'][label] = row
        elif section in ['career_splits','last52_splits'] and cells:
            label = cells[0]
            valid_labels = {'Hard','Clay','Grass','Grand Slams','Masters','Other Tours','Best of 5','Best of 3','vs Righties','vs Lefties','vs Top 10','Finals','Semi-finals','Quarter-finals'}
            if label in valid_labels or label.startswith('vs '):
                row = _ta_parse_stats_row_v23490(cells)
                if row:
                    result[section][label] = row
    result['year_end_current'] = _ta_parse_year_end_current_v23490(lines)
    # puntuación de riqueza: sirve para clasificar TA Real Completo.
    score = 0
    score += 2 if result['tour_seasons'] else 0
    score += 2 if result['challenger_seasons'] else 0
    score += 2 if result['career_splits'] else 0
    score += 2 if result['last52_splits'] else 0
    score += 1 if result['year_end_current'] else 0
    result['full_score'] = score
    return result


def _ta_best_surface_stats_v23490(full, surface='Hard'):
    """Elige stats TA útiles para integrar/mostrar por superficie sin tocar motor."""
    if not isinstance(full, dict):
        return {}
    surf = {'Hard':'Hard','Clay':'Clay','Grass':'Grass'}.get(str(surface), str(surface))
    # Prioridad: Last52 superficie > Career superficie > Tour 2026 > Challenger 2026 > Career tour/challenger.
    for path in [
        ('last52_splits', surf), ('career_splits', surf),
        ('tour_seasons', '2026'), ('challenger_seasons', '2026'),
        ('tour_seasons', 'Career'), ('challenger_seasons', 'Career')
    ]:
        row = (full.get(path[0], {}) or {}).get(path[1])
        if isinstance(row, dict) and row.get('matches',0):
            return row
    return {}

def _parse_tennisabstract_player_profile(raw_text):
    """Extrae señales desde una ficha copiada de Tennis Abstract.
    Usa Recent Results para calcular Over18.5, Over19.5, 3 sets, set ganado,
    derrotas 0-2 y palizas cortas. También intenta leer Elo/hElo/cElo/gElo.
    """
    text = normalizar_texto(raw_text or "")
    if "Tennis Abstract" not in text and "Recent Results" not in text and "Elo rank" not in text:
        return None
    lines = [normalizar_texto(x).strip() for x in text.splitlines() if normalizar_texto(x).strip()]
    if not lines:
        return None

    # Nombre del jugador: primera línea normalmente "Luca Nardi [ITA]".
    player_line = lines[0]
    player_name = re.sub(r"\[[A-Z]{2,3}\].*$", "", player_line).strip()
    player_clean = limpiar(player_name)
    player_surname = surname_key(player_name)

    # Elo general de cabecera.
    # FIX v23.37.25: solo aceptamos el rating si viene de la etiqueta explícita
    # "Elo rank: ... (rating: XXXX)". Así evitamos confundir años con Elo.
    elo = None
    m = re.search(r"Elo rank:\s*\d+\s*\(rating:\s*(\d+)\)", text)
    if m:
        elo = _ta_valid_elo_rating(m.group(1))

    # Línea Current de Year-End Rankings: ATP Rank Points Elo Rank Elo hElo Rank hElo cElo Rank cElo gElo Rank gElo...
    # FIX v23.37.25: usar posiciones reales de columnas, no "todos los números 1200-2300".
    helo = celo = gelo = None
    for ln in lines:
        if ln.startswith("Current (") and re.search(r"\b\d{2,4}\b", ln):
            cur_elo, cur_helo, cur_celo, cur_gelo = _ta_parse_current_elo_line(ln)
            if elo is None and cur_elo is not None:
                elo = cur_elo
            helo = cur_helo
            celo = cur_celo
            gelo = cur_gelo
            break

    # Si por cualquier formato raro se detecta un año como Elo y no venía de cabecera, lo anulamos.
    # En Challenger esos valores 2019-2026 eran el síntoma principal del bug.
    if elo is not None and _ta_elo_is_suspicious_year(elo) and not re.search(r"Elo rank:\s*\d+\s*\(rating:\s*" + str(int(elo)) + r"\)", text):
        elo = None

    # Challenger 2026 resumen si aparece.
    ch_2026 = {}
    in_ch = False
    for ln in lines:
        if ln.startswith("Challenger Seasons"):
            in_ch = True
            continue
        if in_ch and ln.startswith("2026"):
            parts = re.split(r"\s+", ln)
            try:
                # Year M W L Win% Set W-L ... Game W-L ...
                ch_2026 = {"matches": int(parts[1]), "wins": int(parts[2]), "losses": int(parts[3])}
            except Exception:
                pass
            break
        if in_ch and ln.startswith("Recent Titles"):
            break

    ta_full = _ta_parse_full_sections_v23490(text, lines)
    year_current = ta_full.get('year_end_current', {}) if isinstance(ta_full, dict) else {}
    if elo is None and year_current.get('elo') is not None:
        elo = year_current.get('elo')
    if helo is None and year_current.get('hElo') is not None:
        helo = year_current.get('hElo')
    if celo is None and year_current.get('cElo') is not None:
        celo = year_current.get('cElo')
    if gelo is None and year_current.get('gElo') is not None:
        gelo = year_current.get('gElo')

    matches = []
    in_recent = False
    for ln in lines:
        if ln.startswith("Recent Results"):
            in_recent = True
            continue
        if in_recent and ln.startswith("Tour-Level Seasons"):
            break
        if not in_recent:
            continue
        if not re.match(r"^\d{2}-[A-Za-z]{3}-\d{4}\t", ln):
            continue
        cols = ln.split("\t")
        if len(cols) < 8:
            continue
        try:
            date, tournament, surface, rd = cols[0], cols[1], cols[2], cols[3]
            desc = cols[6] if len(cols) > 6 else ""
            score = cols[7] if len(cols) > 7 else ""
        except Exception:
            continue
        if " d. " not in desc or not re.search(r"\d+-\d+", score):
            continue
        pairs = _score_pairs_from_ta_score(score)
        if len(pairs) < 2:
            continue
        left, right = desc.split(" d. ", 1)
        left_clean, right_clean = limpiar(left), limpiar(right)
        # Detecta si el jugador de la ficha ganó o perdió.
        player_on_left = player_clean and (player_clean in left_clean or (player_surname and player_surname in left_clean))
        player_on_right = player_clean and (player_clean in right_clean or (player_surname and player_surname in right_clean))
        if not (player_on_left or player_on_right):
            # Si no lo detecta, saltar para no orientar mal set ganado.
            continue
        win = bool(player_on_left)
        games = sum(a+b for a,b in pairs)
        three_sets = len(pairs) >= 3
        # Score está desde el ganador. Si el jugador ganó, sus juegos son a; si perdió, sus juegos son b.
        player_sets_won = 0
        opp_sets_won = 0
        for a,b in pairs:
            if win:
                player_sets_won += int(a > b)
                opp_sets_won += int(b > a)
            else:
                player_sets_won += int(b > a)
                opp_sets_won += int(a > b)
        set_won = player_sets_won >= 1
        lost_0_2 = (not win and player_sets_won == 0 and len(pairs) == 2)
        lost_0_2_easy = lost_0_2 and games <= 17
        easy_short = games <= 17 or any(sorted([a,b]) in ([0,6],[1,6],[2,6]) for a,b in pairs)
        long_sets = sum(1 for a,b in pairs if a+b >= 10 or 7 in (a,b))
        tb_sets = str(score).count("(")
        matches.append({
            "date": date, "surface": surface, "tournament": tournament,
            "win": win, "pairs": pairs, "games": games,
            "over18": games >= 19, "over19": games >= 20,
            "three_sets": three_sets, "set_won": set_won,
            "lost_0_2": lost_0_2, "lost_0_2_easy": lost_0_2_easy,
            "easy_short": easy_short, "long_sets": long_sets, "tb_sets": tb_sets,
        })
        if len(matches) >= 12:
            break

    if not matches and elo is None and not ch_2026 and not (isinstance(ta_full, dict) and ta_full.get('full_score', 0) > 0):
        return None
    n = len(matches)
    wins = sum(m["win"] for m in matches)
    overs18 = sum(m["over18"] for m in matches)
    overs19 = sum(m["over19"] for m in matches)
    threes = sum(m["three_sets"] for m in matches)
    set_won = sum(m["set_won"] for m in matches)
    lost02 = sum(m["lost_0_2"] for m in matches)
    lost02easy = sum(m["lost_0_2_easy"] for m in matches)
    easy_short = sum(m["easy_short"] for m in matches)
    long_sets = sum(m["long_sets"] for m in matches)
    tb_sets = sum(m["tb_sets"] for m in matches)
    avg_games = (sum(m["games"] for m in matches) / n) if n else 0.0
    hard_n = sum(1 for m in matches if str(m["surface"]).lower().startswith("hard"))
    clay_n = sum(1 for m in matches if str(m["surface"]).lower().startswith("clay"))
    hard_over = sum(1 for m in matches if str(m["surface"]).lower().startswith("hard") and m["over18"])
    clay_over = sum(1 for m in matches if str(m["surface"]).lower().startswith("clay") and m["over18"])

    over_rate = overs18 / max(n, 1)
    over19_rate = overs19 / max(n, 1)
    set_rate = set_won / max(n, 1)
    three_rate = threes / max(n, 1)
    lost02_rate = lost02 / max(n, 1)
    easy_rate = easy_short / max(n, 1)

    if over_rate >= 0.70 or avg_games >= 22 or threes >= 3 or long_sets >= 5:
        largos = "Muchos: Tennis Abstract muestra Over/3 sets/7-6/7-5"
    elif over_rate >= 0.50 or avg_games >= 20.0 or threes >= 1 or long_sets >= 2:
        largos = "Algunos"
    else:
        largos = "Pocos: pocos Overs recientes en Tennis Abstract"

    if easy_rate >= 0.45 or lost02easy >= 3:
        palizas = "Muchas palizas en general"
    elif easy_rate >= 0.20 or lost02 >= 2:
        palizas = "No se ve claro"
    else:
        palizas = "No se ven muchas"

    summary = (
        f"TA {player_name}: n={n}, W={wins}/{n}, Over18={overs18}/{n}, Over19={overs19}/{n}, "
        f"3sets={threes}/{n}, set ganado={set_won}/{n}, perdió 0-2={lost02}/{n}, "
        f"media juegos={avg_games:.1f}, Elo={elo or 'N/A'}"
    )
    return {
        "source": "Tennis Abstract", "player": player_name, "matches": n, "wins": wins,
        "overs": overs18, "over_rate": over_rate, "overs19": overs19, "over19_rate": over19_rate,
        "three_sets": threes, "three_rate": three_rate,
        "set_won": set_won, "set_rate": set_rate,
        "lost02": lost02, "lost02_rate": lost02_rate, "lost02_easy": lost02easy,
        "easy_short": easy_short, "easy_rate": easy_rate,
        "long_sets": long_sets, "tb_sets": tb_sets, "avg_games": avg_games,
        "hard_n": hard_n, "hard_over": hard_over, "clay_n": clay_n, "clay_over": clay_over,
        "elo": elo, "hElo": helo, "cElo": celo, "gElo": gelo, "challenger_2026": ch_2026,
        "ta_full": ta_full, "ta_full_score": (ta_full or {}).get("full_score", 0),
        "ta_best_grass": _ta_best_surface_stats_v23490(ta_full, "Grass"),
        "ta_best_clay": _ta_best_surface_stats_v23490(ta_full, "Clay"),
        "ta_best_hard": _ta_best_surface_stats_v23490(ta_full, "Hard"),
        "largos": largos, "palizas": palizas, "summary": summary,
    }



def _parse_auto_recent_profile_v23433(raw_text):
    """Lee fichas auto generadas desde históricos locales cuando TA no responde.
    No sustituye TennisAbstract si existe; solo evita pedir pegado manual cuando hay
    histórico Challenger/WTA suficiente en los CSV locales.
    """
    text = normalizar_texto(raw_text or "")
    if "Auto Challenger Recent Form" not in text and "Auto Recent Form" not in text:
        return None
    def grab(pattern, default=0.0, as_int=False):
        m = re.search(pattern, text, flags=re.I)
        if not m:
            return default
        try:
            v = float(str(m.group(1)).replace(",", "."))
            return int(v) if as_int else v
        except Exception:
            return default
    player = ""
    m_player = re.search(r"Player\s*:\s*([^\n]+)", text, flags=re.I)
    if m_player:
        player = normalizar_texto(m_player.group(1)).strip()
    n = grab(r"n\s*=\s*(\d+)", 0, True)
    wins = grab(r"wins\s*=\s*(\d+)", 0, True)
    over18 = grab(r"over18\s*=\s*(\d+)", 0, True)
    over19 = grab(r"over19\s*=\s*(\d+)", 0, True)
    threes = grab(r"three_sets\s*=\s*(\d+)", 0, True)
    setwon = grab(r"set_won\s*=\s*(\d+)", 0, True)
    lost02 = grab(r"lost02\s*=\s*(\d+)", 0, True)
    short = grab(r"short\s*=\s*(\d+)", 0, True)
    avg_games = grab(r"avg_games\s*=\s*([0-9]+(?:[\.,][0-9]+)?)", 0.0, False)
    if n <= 0:
        return None
    over_rate = over18 / max(n, 1)
    over19_rate = over19 / max(n, 1)
    three_rate = threes / max(n, 1)
    set_rate = setwon / max(n, 1)
    lost02_rate = lost02 / max(n, 1)
    easy_rate = short / max(n, 1)
    if over_rate >= 0.70 or avg_games >= 22 or threes >= 3:
        largos = "Muchos: histórico local muestra Over/3 sets"
    elif over_rate >= 0.50 or avg_games >= 20 or threes >= 1:
        largos = "Algunos"
    else:
        largos = "Pocos: pocos Overs recientes en histórico local"
    if easy_rate >= 0.45 or lost02 >= 3:
        palizas = "Muchas palizas en general"
    elif easy_rate >= 0.20 or lost02 >= 2:
        palizas = "No se ve claro"
    else:
        palizas = "No se ven muchas"
    summary = f"AutoRecent {player}: n={n}, W={wins}/{n}, Over18={over18}/{n}, Over19={over19}/{n}, 3sets={threes}/{n}, set ganado={setwon}/{n}, perdió 0-2={lost02}/{n}, media juegos={avg_games:.1f}"
    return {
        "source": "Auto Recent Form", "player": player, "matches": n, "wins": wins,
        "overs": over18, "over_rate": over_rate, "overs19": over19, "over19_rate": over19_rate,
        "three_sets": threes, "three_rate": three_rate, "set_won": setwon, "set_rate": set_rate,
        "lost02": lost02, "lost02_rate": lost02_rate, "easy_short": short, "easy_rate": easy_rate,
        "avg_games": avg_games, "largos": largos, "palizas": palizas, "summary": summary,
    }


def _ta_profile_cache_upsert_auto_recent_v23433(expected_player, stats, raw_text):
    try:
        player = normalizar_texto((stats or {}).get("player") or expected_player)
        key = _ta_profile_cache_key(player or expected_player)
        if not key:
            return False, "sin nombre"
        cache = _ta_profile_cache_load()
        old_item = cache.get(key)
        if _ta_profile_cache_is_ta_real_permanent(old_item):
            return True, f"TA Real permanente conservada para {player}; no se sustituye por Auto Recent"
        cache[key] = {
            "player": player,
            "expected_player": expected_player,
            "source": "Auto Recent Form local",
            "uploaded_at": _ta_profile_cache_today(),
            "raw_text": str(raw_text or ""),
            "summary": (stats or {}).get("summary", ""),
            "matches": (stats or {}).get("matches", (stats or {}).get("n", "")),
            "over_rate": (stats or {}).get("over_rate", (stats or {}).get("over18_rate", "")),
            "set_rate": (stats or {}).get("set_rate", (stats or {}).get("set_won_rate", "")),
            "three_rate": (stats or {}).get("three_rate", ""),
            "elo": "",
            "recent_used": (stats or {}).get("matches", (stats or {}).get("n", "")),
            "challenger_2026_matches": "",
            "ta_player_detected": player,
        }
        ok_save, info = _ta_profile_cache_save(cache)
        if not ok_save:
            return False, f"no guardado: {info}"
        return True, f"perfil local guardado para {player}"
    except Exception as e:
        return False, f"error guardando local: {type(e).__name__}: {e}"

def _safe_pct_from_row(row, col, default=0.0):
    try:
        return float(leer_porcentaje(row.get(col, default), default))
    except Exception:
        return default


def _tennisabstract_all_markets_signal(row, ta1, ta2):
    """Con dos fichas TA, reelige el mejor mercado entre Over18/Over19/gana-set/ML/+2.5.
    No usa cuotas. Devuelve selección y señal de confirmación.
    """
    if not ta1 and not ta2:
        return None
    if row is None:
        row = {}
    partido = str(row.get("Partido", ""))
    if " vs " in partido:
        p1, p2 = [x.strip() for x in partido.split(" vs ", 1)]
    else:
        p1, p2 = "Jugador 1", "Jugador 2"
    base_over18 = _safe_pct_from_row(row, "Over 18.5", 0.0)
    base_over19 = _safe_pct_from_row(row, "Over 19.5", 0.0)
    base_ml = _safe_pct_from_row(row, "ML favorito", 0.0)
    old_market = str(row.get("🎯 Mercado más probable", ""))
    old_prob = _safe_pct_from_row(row, "🎯 Prob máxima", 0.0)

    def g(d, k, default=0.0):
        try:
            if not d: return default
            return float(d.get(k, default) or default)
        except Exception:
            return default
    n1, n2 = g(ta1,"matches"), g(ta2,"matches")
    avg_over = np.mean([g(ta1,"over_rate",0.0), g(ta2,"over_rate",0.0)]) if (ta1 and ta2) else max(g(ta1,"over_rate",0), g(ta2,"over_rate",0))
    avg_over19 = np.mean([g(ta1,"over19_rate",0.0), g(ta2,"over19_rate",0.0)]) if (ta1 and ta2) else max(g(ta1,"over19_rate",0), g(ta2,"over19_rate",0))
    avg_three = np.mean([g(ta1,"three_rate",0.0), g(ta2,"three_rate",0.0)]) if (ta1 and ta2) else max(g(ta1,"three_rate",0), g(ta2,"three_rate",0))
    avg_easy = np.mean([g(ta1,"easy_rate",0.0), g(ta2,"easy_rate",0.0)]) if (ta1 and ta2) else max(g(ta1,"easy_rate",0), g(ta2,"easy_rate",0))
    avg_games = np.mean([g(ta1,"avg_games",0.0), g(ta2,"avg_games",0.0)]) if (ta1 and ta2) else max(g(ta1,"avg_games",0), g(ta2,"avg_games",0))
    set1, set2 = g(ta1,"set_rate",0.0), g(ta2,"set_rate",0.0)
    lost1, lost2 = g(ta1,"lost02_rate",0.0), g(ta2,"lost02_rate",0.0)
    elo1, elo2 = g(ta1,"elo",0.0), g(ta2,"elo",0.0)

    candidates = []
    reasons = []
    enough_both = (n1 >= 5 and n2 >= 5)
    enough_one = (max(n1, n2) >= 6)

    # Over18 confirmado por TA.
    if base_over18 >= 0.70 and enough_one:
        ta_boost = 0.06 if avg_over >= 0.70 else 0.035 if avg_over >= 0.60 else 0.0
        three_boost = 0.025 if avg_three >= 0.30 else 0.0
        easy_cut = 0.06 if avg_easy >= 0.45 else 0.03 if avg_easy >= 0.30 else 0.0
        p = float(np.clip(max(base_over18, 0.67 + avg_over*0.18 + avg_three*0.06 + min(avg_games,26)/260) + ta_boost + three_boost - easy_cut, 0.55, 0.88))
        if p >= 0.76 and avg_over >= 0.55 and avg_easy < 0.45:
            candidates.append((p, "Over 18.5", "OVER18", f"TA confirma Over: over_rate combinado {avg_over:.0%}, 3 sets {avg_three:.0%}, media juegos {avg_games:.1f}"))

    # Over19 exige más.
    if base_over19 >= 0.68 and enough_one and avg_over19 >= 0.50 and avg_easy < 0.35:
        p = float(np.clip(max(base_over19, 0.62 + avg_over19*0.20 + avg_three*0.05), 0.55, 0.82))
        if p >= 0.72:
            candidates.append((p, "Over 19.5", "OVER19", f"TA apoya Over19: over19 combinado {avg_over19:.0%}, 3 sets {avg_three:.0%}"))

    # Gana al menos 1 set por jugador.
    # v23.37.28 STRICT SET GUARD:
    # Este mercado fue el principal foco de fallos en backtest.
    # Solo puede competir como JUGAR si hay datos de LOS DOS jugadores y el elegido
    # muestra una frecuencia muy alta de ganar set con muy pocas derrotas 0-2.
    set_samples_ok = (n1 >= 6 and n2 >= 6)
    if ta1 and set_samples_ok and set1 >= 0.90 and lost1 <= 0.10:
        rival_paliza_risk = lost2 >= 0.30 or g(ta2, "easy_rate", 0.0) >= 0.40
        p = float(np.clip(0.78 + set1*0.10 - lost1*0.18 - (0.04 if rival_paliza_risk else 0.0), 0.65, 0.91))
        if p >= 0.88 and not rival_paliza_risk:
            candidates.append((p, f"{p1} gana al menos 1 set", "SET_J1", f"TA set MUY confirmado {set1:.0%}, derrotas 0-2 {lost1:.0%}; muestras {int(n1)}/{int(n2)}"))
    if ta2 and set_samples_ok and set2 >= 0.90 and lost2 <= 0.10:
        rival_paliza_risk = lost1 >= 0.30 or g(ta1, "easy_rate", 0.0) >= 0.40
        p = float(np.clip(0.78 + set2*0.10 - lost2*0.18 - (0.04 if rival_paliza_risk else 0.0), 0.65, 0.91))
        if p >= 0.88 and not rival_paliza_risk:
            candidates.append((p, f"{p2} gana al menos 1 set", "SET_J2", f"TA set MUY confirmado {set2:.0%}, derrotas 0-2 {lost2:.0%}; muestras {int(n2)}/{int(n1)}"))

    # ML por Elo si ambos tienen TA y diferencia clara.
    if elo1 and elo2 and abs(elo1-elo2) >= 120:
        p1_ml = float(elo_prob(elo1, elo2))
        fav_name = p1 if p1_ml >= 0.5 else p2
        fav_p = max(p1_ml, 1-p1_ml)
        # No sobreagresivo: requiere forma y evita que TA sustituya al modelo sin soporte.
        fav_ta = ta1 if p1_ml >= 0.5 else ta2
        fav_n = g(fav_ta,"matches",0)
        fav_w = g(fav_ta,"wins",0) / max(fav_n,1)
        if fav_p >= 0.68 and fav_w >= 0.45 and base_ml >= 0.62:
            p = float(np.clip((fav_p*0.55 + base_ml*0.45), 0.55, 0.84))
            if p >= 0.72:
                candidates.append((p, f"{fav_name} gana", "ML", f"TA Elo diferencial {abs(elo1-elo2):.0f}, forma favorita {fav_w:.0%}"))

    # +2.5 sets: útil como señal, rara vez JUGAR por probabilidad.
    if enough_both and avg_three >= 0.35 and avg_over >= 0.65 and elo1 and elo2 and abs(elo1-elo2) <= 90:
        p = float(np.clip(0.50 + avg_three*0.20 + avg_over*0.08, 0.45, 0.66))
        candidates.append((p, "+2.5 sets", "3SETS", f"TA igualdad + 3 sets frecuentes: 3sets {avg_three:.0%}, Elo gap {abs(elo1-elo2):.0f}"))

    if not candidates:
        return {
            "estado": "neutral",
            "motivo": f"TA revisado, pero no confirma mercado jugable. Over {avg_over:.0%}, set J1 {set1:.0%}, set J2 {set2:.0%}, palizas {avg_easy:.0%}",
            "nuevo_mercado": "", "nueva_prob": None, "decision": "👀 OBSERVAR"
        }

    candidates.sort(key=lambda x: x[0], reverse=True)
    p, label, typ, why = candidates[0]
    # Umbrales jugables por tipo.
    jugar = False
    if typ == "OVER18" and p >= 0.80:
        jugar = True
    elif typ.startswith("SET") and p >= 0.94:
        jugar = True
    elif typ == "ML" and p >= 0.72:
        jugar = True
    elif typ == "OVER19" and p >= 0.78:
        # Over 19.5 solo entra si TA lo confirma muy alto.
        jugar = True
    # 3SETS no lo subimos automáticamente salvo prob imposible; lo dejamos observación.
    estado = "confirmar" if jugar else "neutral"
    decision = "🔥 Alto acierto + TA" if jugar else "👀 Observar + TA"
    return {
        "estado": estado,
        "motivo": why,
        "nuevo_mercado": label,
        "nueva_prob": p,
        "decision": decision,
        "candidates": candidates[:5],
    }

def _parse_flashscore_player_profile(raw_text):
    """Extrae señales recientes desde una ficha copiada de Flashscore.
    Funciona con texto pegado donde los marcadores vienen en líneas separadas:
    2\n1\n3\n6\n6\n1\n6\n3\nP
    Ignora bloques con Retirada y calcula Over 18.5, 3 sets y marcadores largos.
    """
    text = normalizar_texto(raw_text or "")
    lines = [normalizar_texto(x).strip() for x in text.splitlines()]
    completed = []
    cur = []

    def is_result_line(x):
        return x.strip().upper() in {"G", "P", "W", "L"}

    for line in lines:
        if not line:
            continue
        cur.append(line)
        if is_result_line(line):
            block = cur[-45:]
            cur = []
            joined = " ".join(block).lower()
            if "retirada" in joined or "retired" in joined or "walkover" in joined or "w/o" in joined:
                continue
            nums = []
            for z in block:
                zz = z.strip()
                # Solo líneas que son números de marcador, no fechas/horas/rankings.
                if re.fullmatch(r"\d{1,2}", zz):
                    n = int(zz)
                    if 0 <= n <= 8:
                        nums.append(n)
            if len(nums) < 6:
                continue
            match_sets = nums[:2]
            score_nums = nums[2:]
            # Flashscore a veces incluye puntos de tie-break extra; limitamos a 6 números de sets si encaja.
            pairs = []
            i = 0
            while i + 1 < len(score_nums):
                a, b = score_nums[i], score_nums[i + 1]
                # Si aparece un 8 normalmente es tie-break points; no lo usamos como juegos.
                if a <= 7 and b <= 7 and max(a, b) >= 3:
                    pairs.append((a, b))
                i += 2
            if len(pairs) < 2:
                continue
            games = sum(a + b for a, b in pairs[:3])
            three_sets = (len(pairs) >= 3) or (len(match_sets) >= 2 and sorted(match_sets[:2]) == [1, 2])
            long_sets = sum(1 for a, b in pairs if (a + b >= 10 or 7 in (a, b)))
            easy_sets = sum(1 for a, b in pairs if sorted([a, b]) in ([0, 6], [1, 6], [2, 6]))
            result = block[-1].strip().upper()
            win = result in {"G", "W"}
            completed.append({
                "win": win,
                "pairs": pairs[:3],
                "games": games,
                "over18": games >= 19,
                "three_sets": three_sets,
                "long_sets": long_sets,
                "easy_sets": easy_sets,
            })
            if len(completed) >= 10:
                break

    # Récord de temporada/superficie si aparece al final. No es imprescindible, solo informativo.
    hard_record = ""
    m = re.search(r"2026\s+\d+\s+\d+\s+\d+\s*:\s*\d+\s+(\d+\s*:\s*\d+)", text)
    if m:
        hard_record = m.group(1).replace(" ", "")

    if not completed:
        return None
    n = len(completed)
    wins = sum(1 for m in completed if m["win"])
    overs = sum(1 for m in completed if m["over18"])
    threes = sum(1 for m in completed if m["three_sets"])
    long_sets = sum(m["long_sets"] for m in completed)
    easy_sets = sum(m["easy_sets"] for m in completed)
    under_easy = sum(1 for m in completed if m["games"] <= 17)

    over_rate = overs / max(n, 1)
    if over_rate >= 0.70 or threes >= 3 or long_sets >= 5:
        largos = "Muchos: Flashscore muestra varios Over/3 sets/7-6/7-5"
    elif over_rate >= 0.50 or threes >= 1 or long_sets >= 2:
        largos = "Algunos"
    else:
        largos = "Pocos: pocos Overs recientes en Flashscore"

    if under_easy >= max(3, n * 0.45) or easy_sets >= max(5, n):
        palizas = "Muchas palizas en general"
    elif under_easy >= 1 or easy_sets >= 2:
        palizas = "No se ve claro"
    else:
        palizas = "No se ven muchas"

    return {
        "wins": wins,
        "matches": n,
        "overs": overs,
        "over_rate": over_rate,
        "three_sets": threes,
        "long_sets": long_sets,
        "easy_sets": easy_sets,
        "under_easy": under_easy,
        "largos": largos,
        "palizas": palizas,
        "hard_record": hard_record,
        "summary": f"Flashscore: {wins}/{n} victorias, Over18.5 {overs}/{n}, 3 sets {threes}/{n}, sets largos {long_sets}, palizas/under corto {under_easy}/{n}" + (f", dura 2026 {hard_record}" if hard_record else ""),
    }

def _extraer_datos_extra_desde_texto(raw_text, row=None):
    """Convierte texto OCR/copypaste de Datos extra en los campos que usa el reanálisis.
    No usa cuotas. Busca balance últimos 10 y patrones de marcadores.
    """
    text = normalizar_texto(raw_text or "")
    low = text.lower()
    data = {
        "j1_v": None,
        "j2_v": None,
        "largos": "",
        "palizas": "",
        "h2h": "",
        "manual": "Auto",
        "nota": "",
        "ocr_text": text[:1200]
    }
    if not text.strip():
        return data

    # 0) Ficha Tennis Abstract pegada: nivel real + resultados recientes detallados.
    ta = _parse_tennisabstract_player_profile(text)
    if ta:
        data["j1_v"] = int(ta.get("wins", 0))
        data["largos"] = ta.get("largos", "")
        data["palizas"] = ta.get("palizas", "")
        data["h2h"] = "No aparece"
        data["nota"] = ta.get("summary", "")
        data["tennisabstract"] = ta
        return data

    # 0b) Ficha auto generada desde históricos locales cuando TA no responde.
    # IMPORTANTE v23.49.9:
    # Auto Recent NO es TennisAbstract real. Antes se guardaba en data["tennisabstract"]
    # y la pantalla lo pintaba como "✅ Tennis Abstract leído". Eso era el bug.
    # Lo mantenemos como refuerzo local, pero separado para no disfrazarlo de TA real.
    ar = _parse_auto_recent_profile_v23433(text)
    if ar:
        data["j1_v"] = int(ar.get("wins", 0))
        data["largos"] = ar.get("largos", "")
        data["palizas"] = ar.get("palizas", "")
        data["h2h"] = "No aparece"
        data["nota"] = ar.get("summary", "")
        data["auto_recent"] = ar
        data["is_ta_real"] = False
        return data

    # 1) Ficha Flashscore pegada: resultados recientes y marcadores.
    fs = _parse_flashscore_player_profile(text)
    if fs:
        data["j1_v"] = int(fs.get("wins", 0))
        data["largos"] = fs.get("largos", "")
        data["palizas"] = fs.get("palizas", "")
        data["h2h"] = "No aparece"
        data["nota"] = fs.get("summary", "")
        data["flashscore"] = fs
        return data

    # 1) Últimos 10: capturas/datos extra suelen mostrar "5 victorias / 5 derrotas".
    wins = []
    for m in re.finditer(r"(\d{1,2})\s*(?:victoria|victorias|win|wins|v)", low):
        try:
            n = int(m.group(1))
            if 0 <= n <= 10:
                wins.append(n)
        except Exception:
            pass
    # Fallback: líneas tipo "5 5" cerca de últimos 10 no es fiable; no inventamos.
    if len(wins) >= 1:
        data["j1_v"] = wins[0]
    if len(wins) >= 2:
        data["j2_v"] = wins[1]

    # 2) Marcadores: 7-6, 6:4, etc.
    score_pairs = []
    for a, b in re.findall(r"\b([0-7])\s*[-:]\s*([0-7])\b", text):
        try:
            x, y = int(a), int(b)
            # Evita horas tipo 17:00, pero acepta sets plausibles.
            if 0 <= x <= 7 and 0 <= y <= 7 and max(x, y) >= 4:
                score_pairs.append((x, y))
        except Exception:
            pass

    long_sets = 0
    easy_sets = 0
    for x, y in score_pairs:
        if sorted([x, y]) in ([6, 7], [5, 7]) or (x + y >= 10):
            long_sets += 1
        if sorted([x, y]) in ([0, 6], [1, 6], [2, 6]):
            easy_sets += 1

    # 3 sets si hay muchas parejas de marcador o texto explícito.
    three_set_words = any(k in low for k in ["3 sets", "tres sets", "set decisivo", "deciding set"])
    if long_sets >= 3 or three_set_words:
        data["largos"] = "Muchos: varios 7-6/7-5/3 sets"
    elif long_sets >= 1:
        data["largos"] = "Algunos"
    elif len(score_pairs) >= 3:
        data["largos"] = "Pocos: muchos 6-2/6-3"
    else:
        data["largos"] = "No se ve claro"

    if easy_sets >= 3:
        data["palizas"] = "Muchas palizas en general"
    elif easy_sets >= 1:
        data["palizas"] = "No se ve claro"
    else:
        data["palizas"] = "No se ven muchas"

    if any(k in low for k in ["cara a cara", "head to head", "h2h"]):
        data["h2h"] = "No se ve claro"
    else:
        data["h2h"] = "No aparece"

    data["nota"] = f"OCR: victorias detectadas={wins[:2]}; sets detectados={len(score_pairs)}; sets largos={long_sets}; sets fáciles={easy_sets}"
    return data


def _combinar_lecturas_datos_extra_jugadores(text_j1, text_j2, row=None):
    """Combina dos capturas separadas de Datos extra: una del jugador 1 y otra del jugador 2.
    Así evitamos que el OCR mezcle balances/partidos de ambos jugadores.
    """
    d1 = _extraer_datos_extra_desde_texto(text_j1 or "", row)
    d2 = _extraer_datos_extra_desde_texto(text_j2 or "", row)
    try:
        partido_tmp = str(row.get("Partido", "")) if row is not None else ""
    except Exception:
        partido_tmp = ""
    if " vs " in partido_tmp:
        expected_p1, expected_p2 = [x.strip() for x in partido_tmp.split(" vs ", 1)]
        d1 = _invalidar_lectura_si_no_coincide(d1, expected_p1)
        d2 = _invalidar_lectura_si_no_coincide(d2, expected_p2)

    def first_win(d):
        v = d.get("j1_v", None)
        if v is None:
            v = d.get("j2_v", None)
        return v

    def rank_largos(v):
        s = str(v or "")
        if s.startswith("Muchos"):
            return 3
        if s.startswith("Algunos"):
            return 2
        if s.startswith("Pocos"):
            return 1
        return 0

    def rank_palizas(v):
        s = str(v or "")
        if s.startswith("Muchas"):
            return 3
        if s.startswith("No se ve claro"):
            return 1
        return 0

    largos_vals = [d1.get("largos", ""), d2.get("largos", "")]
    palizas_vals = [d1.get("palizas", ""), d2.get("palizas", "")]
    best_largos = max(largos_vals, key=rank_largos) if largos_vals else ""
    worst_palizas = max(palizas_vals, key=rank_palizas) if palizas_vals else ""

    combined_text = ("[JUGADOR 1]\n" + str(text_j1 or "") + "\n\n[JUGADOR 2]\n" + str(text_j2 or "")).strip()
    notas = []
    if d1.get("nota"):
        notas.append("J1: " + str(d1.get("nota")))
    if d2.get("nota"):
        notas.append("J2: " + str(d2.get("nota")))

    combined = {
        "j1_v": first_win(d1),
        "j2_v": first_win(d2),
        "largos": best_largos or "No se ve claro",
        "palizas": worst_palizas or "No se ven muchas",
        "h2h": "No aparece",
        "manual": "Auto",
        "nota": " | ".join(notas),
        "ocr_text": combined_text[:1600],
        "ocr_j1": str(text_j1 or "")[:900],
        "ocr_j2": str(text_j2 or "")[:900],
        "lectura_j1": d1,
        "lectura_j2": d2,
    }
    ta1 = d1.get("tennisabstract") if isinstance(d1, dict) else None
    ta2 = d2.get("tennisabstract") if isinstance(d2, dict) else None
    if ta1 or ta2:
        sig = _tennisabstract_all_markets_signal(row if row is not None else {}, ta1, ta2)
        if sig:
            combined["ta_all_markets"] = sig
            extra = f"TA All Markets: {sig.get('motivo','')}"
            combined["nota"] = (combined.get("nota", "") + " | " + extra).strip(" |")
    return combined




def _estado_lectura_ficha_datos_extra(d):
    """Devuelve estado legible para mostrar si una ficha pegada se ha leído bien.

    v23.49.9: Auto Recent se muestra como Auto Recent, nunca como TennisAbstract.
    """
    if not isinstance(d, dict):
        return "❌ No leída", "No se pudo interpretar el bloque pegado."
    if d.get("mismatch"):
        return "❌ Ficha no coincide", str(d.get("mismatch"))

    raw = str(d.get("ocr_text", "") or "")
    raw_is_auto_recent = any(x in raw for x in [
        "Auto Challenger Recent Form",
        "Auto Recent Form",
        "AutoRecent "
    ])

    ar = d.get("auto_recent")
    if isinstance(ar, dict):
        player = ar.get("player") or "jugador"
        n = int(ar.get("matches", 0) or 0)
        over = int(ar.get("overs", 0) or 0)
        setwon = int(ar.get("set_won", 0) or 0)
        threes = int(ar.get("three_sets", 0) or 0)
        avg = ar.get("avg_games", 0) or 0
        return "🟠 Auto Recent detectado", f"{player}: refuerzo local reciente · {n} partidos · Over18 {over}/{n if n else 1} · Set ganado {setwon}/{n if n else 1} · 3 sets {threes}/{n if n else 1} · media {avg:.1f} juegos. NO es TennisAbstract completo."

    ta = d.get("tennisabstract")
    fs = d.get("flashscore")
    if isinstance(ta, dict):
        # Defensa extra: si por una entrada antigua llega un Auto Recent dentro de tennisabstract,
        # no lo pintamos en verde.
        src = str(ta.get("source", "") or "")
        if raw_is_auto_recent or "Auto Recent" in src or "Auto" in src:
            player = ta.get("player") or "jugador"
            n = int(ta.get("matches", 0) or 0)
            over = int(ta.get("overs", 0) or 0)
            setwon = int(ta.get("set_won", 0) or 0)
            threes = int(ta.get("three_sets", 0) or 0)
            avg = ta.get("avg_games", 0) or 0
            return "🟠 Auto Recent detectado", f"{player}: refuerzo local reciente · {n} partidos · Over18 {over}/{n if n else 1} · Set ganado {setwon}/{n if n else 1} · 3 sets {threes}/{n if n else 1} · media {avg:.1f} juegos. NO es TennisAbstract completo."

        player = ta.get("player") or "jugador"
        n = int(ta.get("matches", 0) or 0)
        over = int(ta.get("overs", 0) or 0)
        setwon = int(ta.get("set_won", 0) or 0)
        threes = int(ta.get("three_sets", 0) or 0)
        elo = ta.get("elo") or "N/A"
        avg = ta.get("avg_games", 0) or 0
        # Para pintar verde exigimos marcadores reales de TA, no solo n=12.
        ta_markers = any(k in ta for k in ["tour_2026", "challenger_2026", "career_challenger", "last52", "hElo", "cElo", "gElo", "helo", "celo", "gelo"]) or any(x in raw for x in ["Tour-Level Seasons", "Challenger Seasons", "Last 52 Weeks", "Career Splits", "Year-End Rankings", "Elo rank:"])
        if n >= 5 and ta_markers:
            return "✅ Tennis Abstract leído", f"{player}: {n} partidos · Over18 {over}/{n} · Set ganado {setwon}/{n} · 3 sets {threes}/{n} · media {avg:.1f} juegos · Elo {elo}"
        return "⚠️ Tennis Abstract parcial", f"{player}: {n} partidos detectados, pero sin marcadores completos TA. Elo {elo}."
    if isinstance(fs, dict):
        player = fs.get("player") or "jugador"
        n = int(fs.get("matches", 0) or 0)
        wins = int(fs.get("wins", 0) or 0)
        overs = int(fs.get("overs", 0) or 0)
        threes = int(fs.get("three_sets", 0) or 0)
        hard_record = fs.get("hard_record") or ""
        extra = f" · dura {hard_record}" if hard_record else ""
        if n >= 5:
            return "✅ Flashscore leído", f"{player}: {n} partidos · W {wins}/{n} · Over18 {overs}/{n} · 3 sets {threes}/{n}{extra}"
        return "⚠️ Flashscore parcial", f"{player}: solo {n} partidos detectados{extra}. Sirve, pero con menos confianza."
    raw = str(d.get("ocr_text", "") or "").strip()
    largos = str(d.get("largos", "") or "")
    nota = str(d.get("nota", "") or "")
    # OCR/texto genérico: se considera parcial si al menos detectó marcadores o alguna señal.
    if raw and ("sets detectados=" in nota or largos in ["Muchos: varios 7-6/7-5/3 sets", "Algunos", "Pocos: muchos 6-2/6-3"]):
        return "⚠️ Texto parcial leído", f"No reconoce ficha completa TA/Flashscore, pero detecta señales: {largos or 'sin largos claros'} · {nota[:140]}"
    if raw:
        return "❌ Ficha no reconocida", "Hay texto pegado, pero no se reconoció como Tennis Abstract/Flashscore ni se detectaron marcadores fiables."
    return "⬜ Sin ficha", "No has pegado texto ni subido captura para este jugador."

def render_datos_extra_reanalysis_panel(ok_saved):
    """Panel paso 2 por capturas Datos extra separadas por jugador.
    Para cada partido permite subir una captura del jugador 1 y otra del jugador 2.
    """
    if ok_saved is None or ok_saved.empty:
        return
    if "📥 Necesita Datos extra" not in ok_saved.columns:
        return

    cand = ok_saved[ok_saved["📥 Necesita Datos extra"].astype(str).str.upper().str.contains("SÍ", na=False)].copy()
    if cand.empty:
        st.info("📥 Radar datos extra: no hay partidos que necesiten revisión extra.")
        return

    st.subheader("📥 Paso 2 · Reanálisis con Tennis Abstract / Flashscore")
    st.caption("Pega la ficha de Tennis Abstract o Flashscore de cada jugador. Con Tennis Abstract la app revisa todos los mercados posibles, no solo el Over. No usa cuotas.")

    include_media = st.checkbox("Incluir también prioridad MEDIA", value=False, key="wm_include_media")
    if not include_media and "📥 Prioridad Datos extra" in cand.columns:
        cand = cand[cand["📥 Prioridad Datos extra"].astype(str).str.upper().eq("ALTA")].copy()

    cand = cand.head(10)
    if cand.empty:
        st.info("No hay prioridad ALTA. Activa MEDIA si quieres revisar más partidos.")
        return

    show_cols = [c for c in ["Hora", "Partido", "🎯 Acción final", "🎯 Mercado más probable", "🎯 Prob máxima", "📥 Prioridad Datos extra", "📥 Qué mirar Datos extra"] if c in cand.columns]
    st.dataframe(cand[show_cols], width='stretch', hide_index=True)

    st.info("Uso recomendado: Tennis Abstract → ficha del jugador completa, o Flashscore → jugador → Resultados/Partidos. Pega una ficha por jugador. Con los dos jugadores el análisis es mucho más fiable.")

    # v23.43.2: antes de pedir pegado manual, intenta rellenar las fichas TA en bloque.
    try:
        players_extra = _ta_match_players_from_extra_candidates_v23432(cand)
        ok_cached, missing_cached = _ta_match_cache_status_counts_v23432(players_extra)
        cta1, cta2, cta3 = st.columns(3)
        cta1.metric("Jugadores a reforzar", len(players_extra))
        cta2.metric("Fichas TA en caché", len(ok_cached))
        cta3.metric("Fichas TA pendientes", len(missing_cached))

        # v23.48.1: panel de auditoría general para saber si las fichas están realmente entrando.
        try:
            audit_rows, audit_sum = _ta_match_cache_audit_summary_v23481(players_extra)
            with st.expander("📦 Estado TennisAbstract / caché de estos picks", expanded=True):
                a1, a2, a3, a4, a5 = st.columns(5)
                a1.metric("Detectados", audit_sum.get("total", 0))
                a2.metric("OK", audit_sum.get("ok", 0))
                a3.metric("Parciales/caducadas", audit_sum.get("partial", 0) + audit_sum.get("stale", 0))
                a4.metric("Pendientes", audit_sum.get("missing", 0))
                a5.metric("Alias usados", audit_sum.get("aliases", 0))
                b1, b2, b3, b4 = st.columns(4)
                b1.metric("🟢 TA Real", audit_sum.get("ta_real", 0))
                b2.metric("🟠 Auto Recent", audit_sum.get("auto_recent", 0))
                b3.metric("📋 Manual", audit_sum.get("manual", 0))
                b4.metric("🔴 Sin ficha", audit_sum.get("missing", 0))
                st.caption(f"Cobertura de perfiles: {audit_sum.get('cobertura', 0):.1f}%. Ahora se distingue TA Real de Auto Recent. Auto Recent es útil como refuerzo local, pero NO equivale a ficha TennisAbstract completa. Si ves Auto Recent y quieres TA real, pega la ficha manual.")
                if audit_rows:
                    audit_df = pd.DataFrame(audit_rows)
                    st.dataframe(audit_df, width='stretch', hide_index=True)

                    # v23.48.5: pegado manual directo desde el auditor.
                    with st.expander("📋 Pegar ficha TennisAbstract manual", expanded=bool(audit_sum.get("missing", 0))):
                        st.caption("Pega aquí la ficha completa copiada de TennisAbstract para sustituir Auto Recent o rellenar pendientes. Se guarda en caché y se prioriza sobre el perfil automático.")
                        opciones_manual = []
                        for r in audit_rows:
                            label = f"{r.get('Jugador detectado','')} · {r.get('Fuente perfil','')} · {r.get('Estado ficha','')}"
                            opciones_manual.append((label, r.get('Nombre usado') or r.get('Jugador detectado')))
                        if opciones_manual:
                            labels_manual = [x[0] for x in opciones_manual]
                            chosen_label_manual = st.selectbox("Jugador", labels_manual, key="ta_manual_profile_player_v23485")
                            chosen_player_manual = opciones_manual[labels_manual.index(chosen_label_manual)][1]
                            raw_manual = st.text_area("Ficha completa TennisAbstract", height=220, key="ta_manual_profile_text_v23485", placeholder="Copia la ficha del jugador desde TennisAbstract y pégala aquí...")
                            if st.button("💾 Guardar ficha manual TA", key="ta_manual_profile_save_v23485", use_container_width=True, disabled=not bool(str(raw_manual).strip())):
                                ok_m, msg_m = _ta_profile_cache_upsert_manual_v23485(chosen_player_manual, raw_manual)
                                if ok_m:
                                    st.success(msg_m)
                                    try:
                                        st.cache_data.clear()
                                    except Exception:
                                        pass
                                    try:
                                        st.rerun()
                                    except Exception:
                                        pass
                                else:
                                    st.error(msg_m)
                        else:
                            st.info("No hay jugadores en este bloque para guardar ficha manual.")

                    # v23.48.4: inspector de ficha completa TA/cache.
                    # Permite comprobar si la caché guarda solo recientes o también texto/ficha completa.
                    try:
                        st.markdown("#### 👁️ Ver ficha completa guardada")
                        opciones_ficha = []
                        for r in audit_rows:
                            estado = str(r.get("Estado ficha", ""))
                            if not estado.startswith("❌"):
                                label = f"{r.get('Jugador detectado','')} → {r.get('Jugador ficha','') or r.get('Nombre usado','')} · {r.get('Estado ficha','')} · {r.get('Edad','')}"
                                opciones_ficha.append((label, r))

                        if opciones_ficha:
                            labels = [x[0] for x in opciones_ficha]
                            elegido_label = st.selectbox("Jugador a inspeccionar", labels, key="ta_raw_profile_viewer_select_v23484")
                            elegido = dict(opciones_ficha[labels.index(elegido_label)][1])

                            cache_item = None
                            for candidate_name in [elegido.get("Nombre usado", ""), elegido.get("Jugador detectado", ""), elegido.get("Jugador ficha", "")]:
                                if str(candidate_name).strip():
                                    cache_item = _ta_profile_cache_find(candidate_name, allow_stale=True)
                                    if isinstance(cache_item, dict) and str(cache_item.get("raw_text", "")).strip():
                                        break

                            if isinstance(cache_item, dict) and str(cache_item.get("raw_text", "")).strip():
                                raw_txt = str(cache_item.get("raw_text", ""))
                                meta_cols = st.columns(5)
                                meta_cols[0].metric("Caracteres ficha", len(raw_txt))
                                meta_cols[1].metric("Edad", _ta_profile_cache_age_label(cache_item.get("uploaded_at", "")))
                                meta_cols[2].metric("Player", str(cache_item.get("player", ""))[:22])
                                meta_cols[3].metric("Source", str(cache_item.get("source", ""))[:22])
                                meta_cols[4].metric("Stale", "Sí" if bool(cache_item.get("stale", False)) else "No")

                                parsed_info = {}
                                try:
                                    if "Auto Challenger Recent Form" in raw_txt or "Auto Recent Form" in raw_txt:
                                        parsed_info = _parse_auto_recent_profile_v23433(raw_txt) or {}
                                    else:
                                        parsed_info = _parse_tennisabstract_player_profile(raw_txt) or {}
                                except Exception as e_parse:
                                    parsed_info = {"parse_error": f"{type(e_parse).__name__}: {e_parse}"}

                                with st.expander("Resumen parseado de la ficha", expanded=True):
                                    st.json(parsed_info if parsed_info else {"info": "La ficha existe, pero no se ha podido parsear en campos estructurados."})

                                with st.expander("Texto completo guardado en caché", expanded=False):
                                    st.text_area("raw_text", raw_txt, height=360, key="ta_raw_profile_viewer_text_v23484")

                                try:
                                    st.download_button(
                                        "⬇️ Descargar ficha cache JSON",
                                        data=json.dumps(cache_item, ensure_ascii=False, indent=2),
                                        file_name=f"ta_cache_{limpiar(elegido.get('Nombre usado','jugador'))}.json",
                                        mime="application/json",
                                        key="ta_raw_profile_viewer_download_v23484",
                                        use_container_width=True,
                                    )
                                except Exception:
                                    pass
                            else:
                                st.warning("No he encontrado raw_text guardado para este jugador. Puede estar OK por histórico local, no por ficha TA completa.")
                        else:
                            st.info("No hay fichas OK para inspeccionar en este bloque.")
                    except Exception as e_viewer:
                        st.warning(f"No se pudo abrir el visor de ficha TA: {type(e_viewer).__name__}: {e_viewer}")
        except Exception as e:
            st.warning(f"No se pudo mostrar auditoría TA: {type(e).__name__}: {e}")

        with st.expander("🔎 Auto extraer fichas TennisAbstract para estos picks", expanded=bool(missing_cached)):
            st.caption("Esto intenta leer la ficha TA de los jugadores de estos picks Over y guardarla en caché. Si funciona, no tienes que pegarla a mano.")
            if missing_cached:
                st.dataframe(pd.DataFrame({"Jugador pendiente": missing_cached}), width='stretch', hide_index=True)
            else:
                st.success("Todas las fichas de estos picks ya están en caché.")
            timeout_auto = st.slider("Timeout por ficha TA", 12, 35, 22, 1, key="ta_auto_timeout_v23432")
            max_auto = st.slider("Máximo fichas a intentar ahora", 2, 30, min(20, max(2, len(missing_cached) if missing_cached else 2)), 1, key="ta_auto_max_v23432")
            if st.button("🔎 Extraer fichas TA pendientes y guardar caché", key="ta_auto_fetch_v23432", use_container_width=True, disabled=not bool(missing_cached)):
                results = []
                with st.spinner("Extrayendo fichas TennisAbstract en bloque..."):
                    for name in missing_cached[:int(max_auto)]:
                        ok, msg = _ta_match_auto_cache_one_player_v23432(name, timeout_sec=int(timeout_auto))
                        results.append({"Jugador": name, "OK": "Sí" if ok else "No", "Detalle": msg})
                if results:
                    st.dataframe(pd.DataFrame(results), width='stretch', hide_index=True)
                    saved_n = sum(1 for r in results if r.get("OK") == "Sí")
                    if saved_n:
                        st.success(f"Guardadas {saved_n} fichas. Pulsa Rerun o vuelve a analizar para que se apliquen al reanálisis.")
                    else:
                        st.warning("No se pudo guardar ninguna ficha automática. Puede ser timeout, nombre distinto o página TA sin resultados recientes parseables.")
                try:
                    st.rerun()
                except Exception:
                    pass
    except Exception as e:
        st.warning(f"No se pudo mostrar Auto TA reinforcer: {type(e).__name__}: {e}")

    with st.form("wm_reanalysis_ocr_form_2_players"):
        ajustes = {}
        for idx, row in cand.iterrows():
            partido = str(row.get("Partido", ""))
            mercado = str(row.get("🎯 Mercado más probable", ""))
            prob = row.get("🎯 Prob máxima", "")
            if " vs " in partido:
                p1, p2 = [x.strip() for x in partido.split(" vs ", 1)]
            else:
                p1, p2 = "Jugador 1", "Jugador 2"
            key_base = _match_key_ui(row, idx)

            st.markdown(f"**{partido}** · {mercado} · {prob}")
            c_up1, c_up2 = st.columns(2)
            with c_up1:
                up1 = st.file_uploader(
                    f"Captura opcional · {p1}",
                    type=["png", "jpg", "jpeg"],
                    key=f"wm_img_j1_{key_base}"
                )
            with c_up2:
                up2 = st.file_uploader(
                    f"Captura opcional · {p2}",
                    type=["png", "jpg", "jpeg"],
                    key=f"wm_img_j2_{key_base}"
                )

            ocr1, err1 = _ocr_datos_extra_image(up1) if up1 is not None else ("", "")
            ocr2, err2 = _ocr_datos_extra_image(up2) if up2 is not None else ("", "")
            if err1:
                st.warning(f"OCR {p1}: {err1}")
            if err2:
                st.warning(f"OCR {p2}: {err2}")

            cache_msg1, cache_item1 = _ta_profile_cache_status_text(p1)
            cache_msg2, cache_item2 = _ta_profile_cache_status_text(p2)
            default_raw1 = ocr1 or ((cache_item1 or {}).get("raw_text", "") if cache_item1 else "")
            default_raw2 = ocr2 or ((cache_item2 or {}).get("raw_text", "") if cache_item2 else "")

            t1, t2 = st.columns(2)
            with t1:
                if cache_msg1:
                    st.info(cache_msg1)
                raw1 = st.text_area(
                    f"Pega ficha Tennis Abstract / Flashscore o texto detectado · {p1}",
                    value=default_raw1,
                    height=120,
                    key=f"wm_raw_j1_{key_base}",
                    placeholder="Pega aquí la ficha Tennis Abstract o Flashscore del jugador 1..."
                )
            with t2:
                if cache_msg2:
                    st.info(cache_msg2)
                raw2 = st.text_area(
                    f"Pega ficha Tennis Abstract / Flashscore o texto detectado · {p2}",
                    value=default_raw2,
                    height=120,
                    key=f"wm_raw_j2_{key_base}",
                    placeholder="Pega aquí la ficha Tennis Abstract o Flashscore del jugador 2..."
                )

            auto_data = _combinar_lecturas_datos_extra_jugadores(raw1, raw2, row)
            d1 = auto_data.get("lectura_j1", {}) or {}
            d2 = auto_data.get("lectura_j2", {}) or {}
            estado1, detalle1 = _estado_lectura_ficha_datos_extra(d1)
            estado2, detalle2 = _estado_lectura_ficha_datos_extra(d2)
            chk1, chk2 = st.columns(2)
            with chk1:
                if estado1.startswith("✅"):
                    st.success(f"{estado1} · {p1}\n\n{detalle1}")
                elif estado1.startswith("⚠️"):
                    st.warning(f"{estado1} · {p1}\n\n{detalle1}")
                elif estado1.startswith("⬜"):
                    st.info(f"{estado1} · {p1}\n\n{detalle1}")
                else:
                    st.error(f"{estado1} · {p1}\n\n{detalle1}")
            with chk2:
                if estado2.startswith("✅"):
                    st.success(f"{estado2} · {p2}\n\n{detalle2}")
                elif estado2.startswith("⚠️"):
                    st.warning(f"{estado2} · {p2}\n\n{detalle2}")
                elif estado2.startswith("⬜"):
                    st.info(f"{estado2} · {p2}\n\n{detalle2}")
                else:
                    st.error(f"{estado2} · {p2}\n\n{detalle2}")

            ta_sig = auto_data.get("ta_all_markets", {}) if isinstance(auto_data, dict) else {}
            if ta_sig:
                st.info(
                    f"🧠 Reanálisis TA: {ta_sig.get('mercado', '')} · "
                    f"prob. {float(ta_sig.get('nueva_prob', 0) or 0)*100:.1f}% · "
                    f"acción sugerida {ta_sig.get('accion', '')} · {ta_sig.get('motivo', '')}"
                )
            else:
                st.caption(
                    f"Lectura conjunta: largos={auto_data.get('largos')}, palizas={auto_data.get('palizas')}"
                )

            manual = st.selectbox(
                "Forzar valoración",
                ["Auto", "✅ Confirmar", "⚠️ Neutral", "❌ Descartar"],
                key=f"wm_manual_ocr2_{key_base}"
            )
            nota_extra = st.text_input("Nota opcional", key=f"wm_nota_ocr2_{key_base}", placeholder="Ej: ambos muestran 7-6/3 sets; o uno pierde fácil 2-0")

            auto_data["manual"] = manual
            if nota_extra:
                auto_data["nota"] = str(auto_data.get("nota", "")) + " | " + nota_extra
            ajustes[idx] = auto_data
            st.divider()

        submitted = st.form_submit_button("🔁 Reanalizar con Tennis Abstract / datos extra")

    if submitted:
        # Guardar automáticamente fichas válidas con fecha para no pegarlas todos los días.
        saved_msgs = []
        for idx, data in (ajustes or {}).items():
            try:
                row = ok_saved.loc[idx]
                partido = str(row.get("Partido", ""))
                if " vs " in partido:
                    p1_save, p2_save = [x.strip() for x in partido.split(" vs ", 1)]
                else:
                    p1_save, p2_save = "", ""
                ok1, msg1 = _ta_profile_cache_upsert_from_raw(p1_save, data.get("ocr_j1", ""))
                ok2, msg2 = _ta_profile_cache_upsert_from_raw(p2_save, data.get("ocr_j2", ""))
                if ok1: saved_msgs.append(msg1)
                if ok2: saved_msgs.append(msg2)
            except Exception:
                pass

        updated = aplicar_reanalisis_datos_extra_manual(ok_saved, ajustes)
        st.session_state["batch_ok_df"] = updated
        st.success("Reanálisis con Tennis Abstract/datos extra aplicado. La tabla y el Excel se actualizan con la nueva acción final.")
        if saved_msgs:
            st.info("💾 Perfiles guardados/actualizados: " + " | ".join(saved_msgs[:8]))
        try:
            st.rerun()
        except Exception:
            pass



# =========================================================
# v23.37.26 CHALLENGER RECENT FORM ENGINE
# Lee automáticamente datos/challenger/historicos/atp_matches_qual_chall_2026*.csv
# para confirmar mercados sin pegar fichas cada día.
# No toca el motor Over: solo añade señales y ajusta la acción final.
# =========================================================

_CH_FORM_CACHE_VERSION = "v23.37.26-challenger-form-2026"


def _ch_hist_paths():
    paths = []
    try:
        base = os.path.join("datos", "challenger", "historicos")
        patterns = [
            os.path.join(base, "atp_matches_qual_chall_2026*.csv"),
            os.path.join(base, "*.csv"),
        ]
        for pat in patterns:
            for p in glob.glob(pat):
                if p not in paths:
                    paths.append(p)
    except Exception:
        pass
    return paths


def _ch_parse_match_score(score):
    """Devuelve pares de juegos, sets winner/loser y flags básicos desde scores tipo '7-6(4) 3-6 6-4'."""
    pairs = _score_pairs_from_ta_score(score)
    if not pairs:
        return None
    sw = sl = 0
    total_games = 0
    tb_sets = 0
    long_sets = 0
    easy_sets = 0
    for a, b in pairs:
        total_games += a + b
        if a > b:
            sw += 1
        elif b > a:
            sl += 1
        if (a, b) in [(7, 6), (6, 7)]:
            tb_sets += 1
        if a + b >= 10 or sorted([a, b]) in ([6, 7], [5, 7]):
            long_sets += 1
        if sorted([a, b]) in ([0, 6], [1, 6], [2, 6]):
            easy_sets += 1
    return {
        "pairs": pairs,
        "sets_w": sw,
        "sets_l": sl,
        "total_games": total_games,
        "over18": total_games >= 19,
        "over19": total_games >= 20,
        "three_sets": len(pairs) >= 3 or (sw >= 2 and sl >= 1),
        "tb_sets": tb_sets,
        "long_sets": long_sets,
        "easy_sets": easy_sets,
        "short_straight": (len(pairs) == 2 and total_games <= 17),
    }


@st.cache_data(show_spinner=False)
def _load_challenger_recent_form_db(cache_version=_CH_FORM_CACHE_VERSION):
    rows = []
    for path in _ch_hist_paths():
        try:
            df = pd.read_csv(path)
        except Exception:
            try:
                df = pd.read_csv(path, encoding="latin-1")
            except Exception:
                continue
        if df is None or df.empty:
            continue
        cw = buscar_columna(df, ["winner_name", "Winner", "Ganador"])
        cl = buscar_columna(df, ["loser_name", "Loser", "Perdedor"])
        cs = buscar_columna(df, ["score", "Score", "Marcador"])
        cdate = buscar_columna(df, ["tourney_date", "Date", "fecha"])
        csurf = buscar_columna(df, ["surface", "Surface", "Superficie"])
        clevel = buscar_columna(df, ["tourney_level", "Level", "Series"])
        ctourney = buscar_columna(df, ["tourney_name", "Tournament", "Torneo"])
        if not cw or not cl or not cs:
            continue
        for _, r in df.iterrows():
            score = normalizar_texto(r.get(cs, ""))
            if not score or any(x in score.upper() for x in ["RET", "W/O", "DEF", "ABD", "WALKOVER"]):
                continue
            parsed = _ch_parse_match_score(score)
            if not parsed:
                continue
            winner = normalizar_texto(r.get(cw, ""))
            loser = normalizar_texto(r.get(cl, ""))
            if not winner or not loser:
                continue
            surface = normalizar_superficie_hist(r.get(csurf, "Hard") if csurf else "Hard", default="Hard")
            try:
                date_raw = str(r.get(cdate, "")) if cdate else ""
                # Jeff Sackmann: YYYYMMDD
                if re.match(r"^\d{8}$", date_raw):
                    dt = pd.to_datetime(date_raw, format="%Y%m%d", errors="coerce")
                else:
                    dt = pd.to_datetime(date_raw, errors="coerce")
            except Exception:
                dt = pd.NaT
            level = normalizar_texto(r.get(clevel, "")) if clevel else ""
            tourney = normalizar_texto(r.get(ctourney, "")) if ctourney else ""
            base = {
                "date": dt,
                "surface": surface,
                "score": score,
                "level": level,
                "tourney": tourney,
                **parsed,
            }
            rows.append({**base, "player": winner, "opponent": loser, "won": True, "set_won": True, "lost_02": False})
            rows.append({**base, "player": loser, "opponent": winner, "won": False, "set_won": parsed.get("sets_l", 0) >= 1, "lost_02": parsed.get("sets_l", 0) == 0 and parsed.get("sets_w", 0) >= 2})
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    try:
        out["player_clean"] = out["player"].apply(limpiar)
        out["surname"] = out["player"].apply(surname_key)
    except Exception:
        pass
    return out


def _ch_find_player_rows(player_name, surface=None, limit=12):
    db = _load_challenger_recent_form_db()
    if db is None or db.empty:
        return pd.DataFrame()
    name = normalizar_texto(player_name)
    ck = limpiar(name)
    sk = surname_key(name)
    candidates = db[db.get("player_clean", "").astype(str).eq(ck)].copy() if ck else pd.DataFrame()
    if candidates.empty and sk:
        # apellido exacto + mejor fuzzy por nombre completo
        tmp = db[db.get("surname", "").astype(str).eq(sk)].copy()
        if not tmp.empty:
            names = tmp["player"].dropna().astype(str).unique().tolist()
            best_name, best_score = None, 0.0
            for nm in names:
                sc = similitud_nombre(name, nm)
                if sc > best_score:
                    best_name, best_score = nm, sc
            if best_name and best_score >= 0.72:
                candidates = tmp[tmp["player"].astype(str).eq(best_name)].copy()
    if candidates.empty:
        # último intento fuzzy controlado
        unique = db["player"].dropna().astype(str).unique().tolist()
        best_name, best_score = None, 0.0
        for nm in unique:
            sc = similitud_nombre(name, nm)
            if sc > best_score:
                best_name, best_score = nm, sc
        if best_name and best_score >= 0.88:
            candidates = db[db["player"].astype(str).eq(best_name)].copy()
    if candidates.empty:
        return candidates
    if surface in ["Hard", "Clay", "Grass"]:
        surf_rows = candidates[candidates["surface"].astype(str).eq(surface)].copy()
        if len(surf_rows) >= 5:
            candidates = surf_rows
    if "date" in candidates.columns:
        candidates = candidates.sort_values("date", ascending=False, na_position="last")
    return candidates.head(limit).copy()


def _ch_recent_stats(player_name, surface=None, limit=12):
    rows = _ch_find_player_rows(player_name, surface=surface, limit=limit)
    if rows is None or rows.empty:
        return {
            "player": player_name, "n": 0, "found": False, "summary": "sin histórico Challenger reciente"
        }
    n = int(len(rows))
    def mean_bool(col):
        try:
            return float(rows[col].astype(bool).mean()) if col in rows.columns and n else 0.0
        except Exception:
            return 0.0
    wins = int(rows.get("won", pd.Series(dtype=bool)).astype(bool).sum()) if "won" in rows.columns else 0
    over18 = int(rows.get("over18", pd.Series(dtype=bool)).astype(bool).sum()) if "over18" in rows.columns else 0
    over19 = int(rows.get("over19", pd.Series(dtype=bool)).astype(bool).sum()) if "over19" in rows.columns else 0
    threes = int(rows.get("three_sets", pd.Series(dtype=bool)).astype(bool).sum()) if "three_sets" in rows.columns else 0
    setwon = int(rows.get("set_won", pd.Series(dtype=bool)).astype(bool).sum()) if "set_won" in rows.columns else 0
    lost02 = int(rows.get("lost_02", pd.Series(dtype=bool)).astype(bool).sum()) if "lost_02" in rows.columns else 0
    short = int(rows.get("short_straight", pd.Series(dtype=bool)).astype(bool).sum()) if "short_straight" in rows.columns else 0
    easy_sets = int(rows.get("easy_sets", pd.Series(dtype=int)).fillna(0).sum()) if "easy_sets" in rows.columns else 0
    avg_games = float(rows.get("total_games", pd.Series(dtype=float)).dropna().mean()) if "total_games" in rows.columns else 0.0
    matched = rows["player"].dropna().astype(str).iloc[0] if "player" in rows.columns and len(rows) else player_name
    return {
        "player": matched,
        "n": n,
        "found": True,
        "wins": wins,
        "win_rate": wins / n if n else 0,
        "over18": over18,
        "over18_rate": over18 / n if n else 0,
        "over19": over19,
        "over19_rate": over19 / n if n else 0,
        "three_sets": threes,
        "three_rate": threes / n if n else 0,
        "set_won": setwon,
        "set_won_rate": setwon / n if n else 0,
        "lost_02": lost02,
        "lost02_rate": lost02 / n if n else 0,
        "short": short,
        "short_rate": short / n if n else 0,
        "easy_sets": easy_sets,
        "avg_games": avg_games,
        "surface": surface or "All",
        "summary": f"CH {matched}: n={n}, Over18 {over18}/{n}, Set {setwon}/{n}, 3 sets {threes}/{n}, 0-2 {lost02}/{n}, media {avg_games:.1f}",
    }


def _parse_pct_cell(v, default=0.0):
    try:
        return float(leer_porcentaje(v, default))
    except Exception:
        return default


def _format_pct(v):
    try:
        return f"{float(v):.1%}"
    except Exception:
        return ""


def _ch_form_signal_for_row(row):
    partido = str(row.get("Partido", ""))
    if " vs " not in partido:
        return None
    p1, p2 = [x.strip() for x in partido.split(" vs ", 1)]
    surface = normalizar_superficie_hist(row.get("Superficie", "Hard"), default="Hard")
    s1 = _ch_recent_stats(p1, surface=surface, limit=12)
    s2 = _ch_recent_stats(p2, surface=surface, limit=12)
    n1, n2 = int(s1.get("n", 0)), int(s2.get("n", 0))
    combined_n = n1 + n2
    if combined_n <= 0:
        return {"usable": False, "s1": s1, "s2": s2, "motivo": "sin histórico Challenger útil"}

    def wavg(key):
        return ((s1.get(key, 0) * n1) + (s2.get(key, 0) * n2)) / max(combined_n, 1)

    over18_comb = wavg("over18_rate")
    over19_comb = wavg("over19_rate")
    three_comb = wavg("three_rate")
    avg_games = ((s1.get("avg_games", 0) * n1) + (s2.get("avg_games", 0) * n2)) / max(combined_n, 1)
    blowout = wavg("lost02_rate")
    short = wavg("short_rate")
    usable = n1 >= 5 and n2 >= 5

    mercado = str(row.get("🎯 Mercado más probable", ""))
    prob = _parse_pct_cell(row.get("🎯 Prob máxima", 0), 0)
    app_over = _parse_pct_cell(row.get("Over 18.5", 0), 0)
    accion = str(row.get("🎯 Acción final", ""))

    # Señales por mercado.
    estado = "neutral"
    nuevo_mercado = ""
    nueva_prob = None
    decision = ""
    motivo = []

    if usable:
        # Confirmar Over 18.5 si la app ya lo ve fuerte o el mercado candidato es over.
        if ("OVER 18.5" in mercado.upper() or app_over >= 0.80) and over18_comb >= 0.70 and avg_games >= 22.0 and blowout <= 0.28 and short <= 0.33:
            estado = "confirmar"
            nuevo_mercado = "Over 18.5"
            # prob conservadora: no inventa 95%; la limita a 88%.
            base = max(prob, app_over, 0.80)
            bonus = min(0.08, max(0.0, over18_comb - 0.68) * 0.25 + max(0.0, avg_games - 22.0) * 0.01)
            nueva_prob = min(0.88, base + bonus)
            decision = "🔥 Alto acierto + CH histórico"
            motivo.append(f"CH confirma Over18: combinado {over18_comb:.0%}, 3 sets {three_comb:.0%}, media {avg_games:.1f}, paliza {blowout:.0%}")
        # Gana set: muy estricto.
        if "GANA AL MENOS 1 SET" in mercado.upper() or "GANA SET" in mercado.upper():
            lado, chosen = _inferir_jugador_mercado(row)
            st_chosen = s1 if lado == "J1" else s2 if lado == "J2" else None
            if st_chosen and st_chosen.get("n", 0) >= 8:
                set_rate = st_chosen.get("set_won_rate", 0)
                lost02 = st_chosen.get("lost02_rate", 0)
                set_elite = (set_rate >= 0.90 and lost02 <= 0.10 and prob >= 0.94) or (set_rate >= 0.999 and lost02 <= 0.001 and st_chosen.get("n", 0) >= 10)
                if set_elite:
                    estado = "confirmar"
                    nuevo_mercado = mercado
                    nueva_prob = min(0.94, max(prob, 0.90 + (set_rate - 0.90) * 0.18))
                    decision = "🔥 Alto acierto + CH set élite"
                    motivo.append(f"CH confirma gana-set ÉLITE {chosen}: set {set_rate:.0%}, 0-2 {lost02:.0%}, n={st_chosen.get('n')}")
                elif set_rate < 0.82 or lost02 >= 0.18:
                    estado = "descartar"
                    motivo.append(f"CH NO confirma gana-set {chosen}: set {set_rate:.0%}, 0-2 {lost02:.0%}")
        # Si estaba JUGAR pero el histórico contradice, bajar.
        if "JUGAR" in accion.upper():
            if "OVER 18.5" in mercado.upper() and (over18_comb < 0.60 or blowout >= 0.38 or avg_games < 20.5):
                estado = "descartar"
                motivo.append(f"CH contradice Over: Over18 {over18_comb:.0%}, media {avg_games:.1f}, paliza {blowout:.0%}")
    else:
        motivo.append(f"CH parcial: J1 n={n1}, J2 n={n2}; no sube a JUGAR sin 5+ por jugador")

    if not motivo:
        motivo.append(f"CH neutral: Over18 {over18_comb:.0%}, 3 sets {three_comb:.0%}, media {avg_games:.1f}, paliza {blowout:.0%}")
    return {
        "usable": usable,
        "estado": estado,
        "nuevo_mercado": nuevo_mercado,
        "nueva_prob": nueva_prob,
        "decision": decision,
        "motivo": " | ".join(motivo),
        "s1": s1,
        "s2": s2,
        "over18_comb": over18_comb,
        "over19_comb": over19_comb,
        "three_comb": three_comb,
        "avg_games": avg_games,
        "blowout": blowout,
        "short": short,
    }


def aplicar_challenger_recent_form_engine(df):
    """Añade señales CH Recent Form y ajusta Acción final de forma conservadora.
    Se aplica automáticamente a tablas analizadas con partidos Challenger.
    """
    if df is None or df.empty:
        return df
    # Si no hay histórico, no tocar.
    try:
        hist = _load_challenger_recent_form_db()
        if hist is None or hist.empty:
            return df
    except Exception:
        return df
    out = df.copy()
    new_cols = [
        "🧠 CH Form J1", "🧠 CH Form J2", "🧠 CH Over18 combinado", "🧠 CH Set J1", "🧠 CH Set J2",
        "🧠 CH 3 sets", "🧠 CH Riesgo paliza", "🧠 CH Señal histórica"
    ]
    for c in new_cols:
        if c not in out.columns:
            out[c] = ""
    for idx, row in out.iterrows():
        try:
            sig = _ch_form_signal_for_row(row)
            if not sig:
                continue
            s1, s2 = sig.get("s1", {}), sig.get("s2", {})
            out.at[idx, "🧠 CH Form J1"] = s1.get("summary", "")
            out.at[idx, "🧠 CH Form J2"] = s2.get("summary", "")
            out.at[idx, "🧠 CH Over18 combinado"] = _format_pct(sig.get("over18_comb", 0))
            out.at[idx, "🧠 CH Set J1"] = _format_pct(s1.get("set_won_rate", 0)) if s1.get("n", 0) else ""
            out.at[idx, "🧠 CH Set J2"] = _format_pct(s2.get("set_won_rate", 0)) if s2.get("n", 0) else ""
            out.at[idx, "🧠 CH 3 sets"] = _format_pct(sig.get("three_comb", 0))
            out.at[idx, "🧠 CH Riesgo paliza"] = _format_pct(sig.get("blowout", 0))
            out.at[idx, "🧠 CH Señal histórica"] = sig.get("motivo", "")

            estado = sig.get("estado")
            if estado == "confirmar":
                if sig.get("nuevo_mercado") and "🎯 Mercado más probable" in out.columns:
                    out.at[idx, "🎯 Mercado más probable"] = sig.get("nuevo_mercado")
                if sig.get("nueva_prob") is not None and "🎯 Prob máxima" in out.columns:
                    out.at[idx, "🎯 Prob máxima"] = f"{float(sig.get('nueva_prob')):.1%}"
                if "🎯 Acción final" in out.columns:
                    out.at[idx, "🎯 Acción final"] = "✅ JUGAR"
                if "🎯 Decisión acierto" in out.columns:
                    out.at[idx, "🎯 Decisión acierto"] = sig.get("decision") or "🔥 Alto acierto + CH histórico"
                if "🎯 Aviso acierto" in out.columns:
                    out.at[idx, "🎯 Aviso acierto"] = "Histórico Challenger 2026 confirma el mercado; no usa cuotas."
            elif estado == "descartar":
                if "🎯 Acción final" in out.columns:
                    out.at[idx, "🎯 Acción final"] = "👀 OBSERVAR"
                if "🎯 Aviso acierto" in out.columns:
                    old = str(out.at[idx, "🎯 Aviso acierto"] or "")
                    out.at[idx, "🎯 Aviso acierto"] = (old + " | " if old else "") + "Histórico Challenger no confirma: bajado a OBSERVAR."
        except Exception:
            continue
    return out

# =========================================================
# v23.37.11 EXCEL LIMPIO 2 HOJAS
# =========================================================

PICKS_LIMPIOS_COLS = [
    "Hora", "Fecha", "Torneo", "Superficie", "Partido",
    "🎯 Acción final", "🎯 Mercado más probable", "🎯 Prob máxima", "🎯 Confianza acierto",
    "📥 Necesita Datos extra", "📥 Prioridad Datos extra", "📥 Estado Datos extra", "📥 Ajuste Datos extra", "📥 Qué mirar Datos extra",
    "🎯 Motivo acierto", "📥 Motivo Datos extra",
    "Mejor juegos jugador", "Prob juegos jugador", "Confianza juegos jugador",
]

DETALLE_TECNICO_FIRST_COLS = [
    "Versión app", "Fecha", "Hora", "Torneo", "Superficie", "Partido",
    "🎯 Acción final", "🎯 Decisión acierto", "🎯 Mercado más probable", "🎯 Prob máxima",
    "🎯 Confianza acierto", "🎯 Aviso acierto", "🎯 Motivo acierto",
    "📥 Necesita Datos extra", "📥 Prioridad Datos extra", "📥 Estado Datos extra", "📥 Ajuste Datos extra", "📥 Qué mirar Datos extra", "📥 Motivo Datos extra",
    "🧠 CH Form J1", "🧠 CH Form J2", "🧠 CH Over18 combinado", "🧠 CH Set J1", "🧠 CH Set J2", "🧠 CH 3 sets", "🧠 CH Riesgo paliza", "🧠 CH Señal histórica",
    "Recomendación", "Pick oficial", "Mercado recomendado", "Prob mercado recomendado",
    "Favorito modelo", "ML favorito", "Over 18.5", "Over 19.5", "Under 22.5",
    "Jugador gana set", "Prob gana set", "Favorito gana al menos 1 set",
    "J1 +10.5 juegos", "J1 +11.5 juegos", "J2 +10.5 juegos", "J2 +11.5 juegos",
    "Mejor juegos jugador", "Prob juegos jugador", "Confianza juegos jugador",
    "Signal Trust", "Confianza mínima", "Mín. partidos superficie", "Riesgos",
]


def _limpiar_texto_excel_cell(v):
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).replace("OBSERVAR ", "").replace("JUGAR ", "").strip()


# =========================================================
# v23.46.7 EXCEL VALUE RANKING
# Solo ordena y presenta picks por utilidad operativa para apostar.
# No toca fórmulas, probabilidades, simulaciones ni motor Over.
# =========================================================

def _prob_to_float_v23467(v):
    try:
        return float(leer_porcentaje(v, 0) or 0)
    except Exception:
        return 0.0


def _clasificar_estrategia_apuesta_v23467(row):
    mercado = str(row.get("Mercado", row.get("🎯 Mercado más probable", ""))).lower()
    accion = str(row.get("Acción", row.get("🎯 Acción final", ""))).upper()
    confianza = str(row.get("Confianza", row.get("🎯 Confianza acierto", ""))).upper()
    motivo = str(row.get("Motivo", row.get("🎯 Motivo acierto", ""))).lower()

    is_jugar = "JUGAR" in accion
    is_obs = "OBSERVAR" in accion
    is_gana_set = ("set" in mercado and ("gana" in mercado or "al menos" in mercado))
    is_over18 = ("over" in mercado and "18" in mercado)
    is_over19 = ("over" in mercado and "19" in mercado)
    is_ml = ("ganador" in mercado or "ml" in mercado or "gana partido" in mercado)

    # Priorización basada en el backtest que hemos visto: ganar 1 set > Over 18.5 > ML fuerte.
    if is_jugar and is_gana_set:
        return "🥇 Gana 1 set JUGAR", 100
    if is_jugar and is_over18:
        return "🥈 Over 18.5 JUGAR", 92
    if is_jugar and is_ml:
        return "🥉 ML favorito JUGAR", 84
    if is_jugar and is_over19:
        return "⚡ Over 19.5 JUGAR", 78
    if is_jugar:
        return "✅ Otro JUGAR", 70

    if is_obs and is_gana_set:
        return "👀 Gana 1 set OBSERVAR", 56
    if is_obs and is_over18:
        return "👀 Over 18.5 OBSERVAR", 52
    if is_obs and is_ml:
        return "👀 ML OBSERVAR", 46
    if is_obs:
        return "👀 Otro OBSERVAR", 35

    return "— Sin prioridad", 0


def _modificadores_score_apuesta_v23468(row, estrategia, base):
    """v23.46.8: desempate inteligente para que el TOP sea más parecido a una selección humana.
    No cambia predicciones: solo ordena el Excel.
    """
    mercado = str(row.get("Mercado", row.get("🎯 Mercado más probable", ""))).lower()
    accion = str(row.get("Acción", row.get("🎯 Acción final", ""))).upper()
    confianza = str(row.get("Confianza", row.get("🎯 Confianza acierto", ""))).upper()
    motivo = str(row.get("Motivo", row.get("🎯 Motivo acierto", ""))).lower()
    superficie = str(row.get("Superficie", "")).lower()
    torneo = str(row.get("Torneo", "")).lower()
    partido = str(row.get("Partido", "")).lower()
    txt = " ".join([mercado, accion, confianza, motivo, superficie, torneo, partido])

    mod = 0.0
    notas = []

    # 1) Confianza visual de la app.
    if "MUY" in confianza and "ALTA" in confianza:
        mod += 9; notas.append("confianza muy alta")
    elif "ALTA" in confianza:
        mod += 7; notas.append("confianza alta")
    elif "MEDIA" in confianza:
        mod += 3; notas.append("confianza media")
    elif "BAJA" in confianza or "DUD" in confianza:
        mod -= 10; notas.append("confianza baja")

    # 2) Señales del propio motivo.
    if "muy alto" in txt or "muy alta" in txt:
        mod += 7; notas.append("señal muy alta")
    elif "alto" in txt or "alta" in txt:
        mod += 4; notas.append("señal alta")

    # 3) Superficie: en estos días de hierba hemos visto Overs fuertes, pero sin tocar fórmula.
    if "over" in mercado and "18" in mercado:
        if "hierba" in superficie or "grass" in superficie:
            mod += 6; notas.append("over en hierba")
        elif "tierra" in superficie or "clay" in superficie:
            mod += 2; notas.append("over en tierra")
    if "set" in mercado and ("gana" in mercado or "al menos" in mercado):
        mod += 3; notas.append("mercado set estable")

    # 4) Penalización por palabras de peligro. Esto solo baja ranking, no elimina picks.
    danger_words = [
        "riesgo", "paliza", "dudoso", "baja", "datos bajos", "sin datos", "no encontrado",
        "inflada", "retirada", "retirado", "cancel", "inestable", "volatil", "varianza"
    ]
    if any(w in txt for w in danger_words):
        mod -= 12; notas.append("riesgo detectado")

    # 5) Preferencia práctica: si es OBSERVAR, que no se cuele por encima de JUGAR salvo score real muy alto.
    if "OBSERVAR" in accion:
        mod -= 18; notas.append("solo observar")

    # 6) Ligero ajuste por circuito/torneo para separar empates.
    if "challenger" in torneo and "over" in mercado:
        mod += 1; notas.append("challenger over")
    if "wta" in torneo and "over" in mercado and "17" not in mercado:
        mod -= 3; notas.append("wta over prudente")

    return round(mod, 2), "; ".join(notas)


def aplicar_ranking_valor_apuesta_v23467(picks):
    """Añade Prioridad/Score y ordena el Excel por valor práctico para apostar.
    v23.46.8: usa score compuesto: mercado + probabilidad + confianza + superficie + avisos.
    """
    if picks is None or picks.empty:
        return picks
    out = picks.copy()
    estrategias, bases, probs, mods, scores, razones = [], [], [], [], [], []
    for _, row in out.iterrows():
        estrategia, base = _clasificar_estrategia_apuesta_v23467(row)
        prob = _prob_to_float_v23467(row.get("Prob.", row.get("🎯 Prob máxima", 0)))
        mod, razon = _modificadores_score_apuesta_v23468(row, estrategia, base)
        # La probabilidad suma, pero no manda más que el tipo de mercado y la estabilidad.
        score = float(base) + float(prob * 25.0) + float(mod)
        estrategias.append(estrategia)
        bases.append(base)
        probs.append(prob)
        mods.append(mod)
        razones.append(razon)
        scores.append(round(score, 2))
    out["Prioridad apuesta"] = estrategias
    out["Score apuesta"] = scores
    out["Ajuste ranking"] = mods
    out["Motivo ranking"] = razones
    out["Ranking apuesta"] = range(1, len(out) + 1)
    try:
        sort_cols = ["Score apuesta"]
        if "Prob." in out.columns:
            out["__prob_sort"] = out["Prob."].apply(_prob_to_float_v23467)
            sort_cols.append("__prob_sort")
        out = out.sort_values(sort_cols, ascending=[False] * len(sort_cols)).reset_index(drop=True)
        out["Ranking apuesta"] = [f"#{i}" for i in range(1, len(out) + 1)]
        out = out.drop(columns=["__prob_sort"], errors="ignore")
    except Exception:
        pass

    # Columnas nuevas al principio.
    first = ["Ranking apuesta", "Prioridad apuesta", "Score apuesta", "Ajuste ranking", "Motivo ranking"]
    cols = first + [c for c in out.columns if c not in first]
    return out[cols]


def crear_top_apuestas_df_v23467(picks_df, n=10):
    if picks_df is None or picks_df.empty:
        return pd.DataFrame()
    df = picks_df.copy()
    if "Score apuesta" not in df.columns:
        df = aplicar_ranking_valor_apuesta_v23467(df)
    keep = [c for c in ["Ranking apuesta", "Prioridad apuesta", "Score apuesta", "Hora", "Fecha", "Torneo", "Superficie", "Partido", "Acción", "Mercado", "Prob.", "Confianza", "Motivo"] if c in df.columns]
    return df.head(int(n))[keep].copy()


def crear_combinada_recomendada_df_v23467(picks_df):
    """Propone una combinada de 2 siguiendo la jerarquía acordada.
    1) Dos gana-set JUGAR.
    2) Un gana-set JUGAR + mejor Over 18.5 JUGAR.
    3) Dos Over 18.5 JUGAR.
    Si no hay 2 selecciones fuertes, devuelve aviso de NO APOSTAR.
    """
    if picks_df is None or picks_df.empty:
        return pd.DataFrame([{"Estado": "NO APOSTAR", "Motivo": "No hay picks suficientes"}])
    df = picks_df.copy()
    if "Score apuesta" not in df.columns:
        df = aplicar_ranking_valor_apuesta_v23467(df)
    accion = df.get("Acción", pd.Series([""]*len(df))).astype(str).str.upper()
    mercado = df.get("Mercado", pd.Series([""]*len(df))).astype(str).str.lower()
    jugar = df[accion.str.contains("JUGAR", na=False)].copy()
    if jugar.empty:
        return pd.DataFrame([{"Estado": "NO APOSTAR", "Motivo": "No hay picks JUGAR"}])

    def _sin_repetir_partido(cands):
        rows, seen = [], set()
        for _, r in cands.sort_values("Score apuesta", ascending=False).iterrows():
            partido = str(r.get("Partido", ""))
            if partido and partido in seen:
                continue
            rows.append(r)
            if partido:
                seen.add(partido)
            if len(rows) >= 2:
                break
        return pd.DataFrame(rows)

    gana_set = jugar[mercado.loc[jugar.index].str.contains("set", na=False) & (mercado.loc[jugar.index].str.contains("gana|al menos", regex=True, na=False))]
    over18 = jugar[mercado.loc[jugar.index].str.contains("over", na=False) & mercado.loc[jugar.index].str.contains("18", na=False)]

    seleccion = _sin_repetir_partido(gana_set)
    estrategia = "A: 2 x Gana 1 set JUGAR"
    if len(seleccion) < 2:
        seleccion = _sin_repetir_partido(pd.concat([gana_set, over18], ignore_index=False))
        estrategia = "B: Gana 1 set JUGAR + mejor Over 18.5 JUGAR"
    if len(seleccion) < 2:
        seleccion = _sin_repetir_partido(over18)
        estrategia = "C: 2 x Over 18.5 JUGAR"

    if len(seleccion) < 2:
        return pd.DataFrame([{
            "Estado": "NO APOSTAR",
            "Estrategia": "Disciplina",
            "Motivo": "No hay 2 selecciones JUGAR suficientemente fuertes. Mejor no forzar.",
        }])

    keep = [c for c in ["Ranking apuesta", "Prioridad apuesta", "Score apuesta", "Hora", "Torneo", "Superficie", "Partido", "Acción", "Mercado", "Prob.", "Confianza"] if c in seleccion.columns]
    out = seleccion[keep].copy().reset_index(drop=True)
    out.insert(0, "Selección", ["Pick 1", "Pick 2"][:len(out)])
    out.insert(1, "Estrategia", estrategia)
    out.insert(2, "Estado", "COMBINADA RECOMENDADA")
    return out


def crear_picks_limpios_df(df):
    """Hoja simple para decidir: una fila por partido y solo columnas útiles."""
    if df is None or df.empty:
        return pd.DataFrame(columns=PICKS_LIMPIOS_COLS)
    out = df.copy()
    for c in PICKS_LIMPIOS_COLS:
        if c not in out.columns:
            out[c] = ""
    picks = out[[c for c in PICKS_LIMPIOS_COLS if c in out.columns]].copy()
    picks = picks.rename(columns={
        "🎯 Acción final": "Acción",
        "🎯 Mercado más probable": "Mercado",
        "🎯 Prob máxima": "Prob.",
        "🎯 Confianza acierto": "Confianza",
        "📥 Necesita Datos extra": "Datos extra",
        "📥 Prioridad Datos extra": "Prioridad Datos extra",
        "📥 Estado Datos extra": "Estado Datos extra",
        "📥 Ajuste Datos extra": "Ajuste Datos extra",
        "📥 Qué mirar Datos extra": "Qué mirar",
        "🎯 Motivo acierto": "Motivo",
        "📥 Motivo Datos extra": "Motivo Datos extra",
    })
    if "Mercado" in picks.columns:
        picks["Mercado"] = picks["Mercado"].apply(_limpiar_texto_excel_cell)

    def _score(row):
        accion = str(row.get("Acción", "")).upper()
        prio = str(row.get("Prioridad Datos extra", "")).upper()
        prob = leer_porcentaje(row.get("Prob.", 0), 0)
        if "JUGAR" in accion:
            base = 300
        elif "ALTA" in prio:
            base = 220
        elif "MEDIA" in prio:
            base = 160
        elif "OBSERVAR" in accion:
            base = 80
        else:
            base = 0
        return base + float(prob or 0)

    try:
        picks["_orden"] = picks.apply(_score, axis=1)
        picks = picks.sort_values("_orden", ascending=False).drop(columns=["_orden"], errors="ignore")
    except Exception:
        pass
    try:
        picks = aplicar_ranking_valor_apuesta_v23467(picks)
    except Exception:
        pass
    return picks


def crear_detalle_tecnico_df(df):
    """Hoja completa con todo lo técnico, pero con lo importante al principio."""
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    cols = [c for c in DETALLE_TECNICO_FIRST_COLS if c in out.columns] + [c for c in out.columns if c not in DETALLE_TECNICO_FIRST_COLS]
    return out[cols]


def aplicar_estilo_excel_limpio(writer):
    """Formato común para las hojas exportadas."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = str(ws.cell(row=1, column=col_idx).value or "")
            max_len = len(header)
            for r in range(2, min(ws.max_row, 80) + 1):
                max_len = max(max_len, len(str(ws.cell(row=r, column=col_idx).value or "")))
            width = min(max(max_len + 2, 10), 42)
            if header in ["Partido", "Mercado", "Qué mirar", "Motivo", "Motivo Datos extra", "🎯 Motivo acierto", "📥 Qué mirar Datos extra", "📥 Motivo Datos extra"]:
                width = min(max(width, 26), 50)
            ws.column_dimensions[letter].width = width
        headers = {str(ws.cell(row=1, column=i).value or ""): i for i in range(1, ws.max_column + 1)}
        accion_col = headers.get("Acción") or headers.get("🎯 Acción final")
        datos_extra_col = headers.get("Datos extra") or headers.get("📥 Necesita Datos extra")
        for r in range(2, ws.max_row + 1):
            fill = None
            accion = str(ws.cell(row=r, column=accion_col).value or "").upper() if accion_col else ""
            datos_extra = str(ws.cell(row=r, column=datos_extra_col).value or "").upper() if datos_extra_col else ""
            if "JUGAR" in accion:
                fill = PatternFill("solid", fgColor="C6EFCE")
            elif "EVITAR" in accion:
                fill = PatternFill("solid", fgColor="F4CCCC")
            elif "ALTA" in datos_extra:
                fill = PatternFill("solid", fgColor="FFF2CC")
            elif "OBSERVAR" in accion or "MEDIA" in datos_extra:
                fill = PatternFill("solid", fgColor="E2F0D9")
            if fill:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill

def crear_juegos_jugador_df_v23480(df):
    """Hoja nueva para validar mercados de jugador +10.5/+11.5 juegos.
    No crea picks oficiales: solo ordena señales para observación y estadística.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    rows = []
    for _, r in df.iterrows():
        partido = str(r.get("Partido", ""))
        if not partido or " vs " not in partido:
            continue
        try:
            p1, p2 = partido.split(" vs ", 1)
        except Exception:
            p1, p2 = "Jugador 1", "Jugador 2"
        specs = [
            (p1, "Jugador 1 +10.5 juegos", r.get("J1 +10.5 juegos", ""), 0.86, 0.80, 0.74),
            (p1, "Jugador 1 +11.5 juegos", r.get("J1 +11.5 juegos", ""), 0.80, 0.74, 0.68),
            (p2, "Jugador 2 +10.5 juegos", r.get("J2 +10.5 juegos", ""), 0.86, 0.80, 0.74),
            (p2, "Jugador 2 +11.5 juegos", r.get("J2 +11.5 juegos", ""), 0.80, 0.74, 0.68),
        ]
        for jugador, mercado, prob_txt, th_jugar, th_alta, th_obs in specs:
            prob = _prob_to_float_v23467(prob_txt)
            if prob <= 0:
                continue
            if prob >= th_jugar:
                accion = "✅ JUGAR"
                confianza = "🔥 Muy alta"
            elif prob >= th_alta:
                accion = "👀 OBSERVAR ALTO"
                confianza = "✅ Alta"
            elif prob >= th_obs:
                accion = "👀 OBSERVAR"
                confianza = "⚖️ Media-alta"
            else:
                continue
            rows.append({
                "Acción": accion,
                "Confianza": confianza,
                "Prob.": f"{prob:.1%}",
                "Jugador": jugador,
                "Mercado": mercado,
                "Partido": partido,
                "Hora": r.get("Hora", ""),
                "Fecha": r.get("Fecha", ""),
                "Torneo": r.get("Torneo", ""),
                "Superficie": r.get("Superficie", ""),
                "Circuito": r.get("Circuito cálculo", r.get("Circuito datos", "")),
                "Over 18.5": r.get("Over 18.5", ""),
                "Over 19.5": r.get("Over 19.5", ""),
                "Partido a 3 sets": r.get("Partido a 3 sets", ""),
                "Gana set jugador": (r.get(f"{jugador} gana al menos 1 set", "") if f"{jugador} gana al menos 1 set" in r.index else ""),
                "Total games estimado": r.get("Total games", ""),
                "Motivo": "Mercado intermedio entre Over total y gana set; validar unos días antes de usar fuerte.",
            })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["__prob_sort"] = out["Prob."].apply(_prob_to_float_v23467)
    out = out.sort_values(["__prob_sort", "Acción"], ascending=[False, True]).drop(columns=["__prob_sort"]).reset_index(drop=True)
    out.insert(0, "Ranking", [f"#{i}" for i in range(1, len(out)+1)])
    return out


def batch_excel_bytes(df):
    """Exporta Excel limpio: hoja diaria + detalle técnico."""
    output = io.BytesIO()
    picks_df = crear_picks_limpios_df(df)
    detalle_df = crear_detalle_tecnico_df(df)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        picks_df.to_excel(writer, index=False, sheet_name="PICKS LIMPIOS")
        try:
            crear_top_apuestas_df_v23467(picks_df, n=10).to_excel(writer, index=False, sheet_name="TOP APUESTAS")
            crear_combinada_recomendada_df_v23467(picks_df).to_excel(writer, index=False, sheet_name="COMBINADA RECOMENDADA")
            crear_juegos_jugador_df_v23480(detalle_df).to_excel(writer, index=False, sheet_name="JUEGOS_JUGADOR")
        except Exception:
            pass
        detalle_df.to_excel(writer, index=False, sheet_name="DETALLE TECNICO")
        try:
            wm_df = picks_df[picks_df.get("Datos extra", "").astype(str).str.upper().str.contains("SÍ", na=False)].copy()
            if not wm_df.empty:
                for _c in ["Resultado Datos extra", "Forma últimos 10", "Marcadores largos", "Nota manual"]:
                    wm_df[_c] = ""
                wm_df.to_excel(writer, index=False, sheet_name="DATOS EXTRA A REVISAR")
        except Exception:
            pass
        aplicar_estilo_excel_limpio(writer)
    output.seek(0)
    return output.getvalue()

def batch_excel_with_not_found_bytes(ok_df, ko_df, db):
    """Exporta Excel con 2 hojas útiles + no encontrados si existen."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ok_sheet = ok_df.copy() if ok_df is not None else pd.DataFrame()
        ko_sheet = enrich_not_found_with_suggestions(ko_df, db) if ko_df is not None and not ko_df.empty else pd.DataFrame()
        picks_df = crear_picks_limpios_df(ok_sheet)
        detalle_df = crear_detalle_tecnico_df(ok_sheet)
        picks_df.to_excel(writer, index=False, sheet_name="PICKS LIMPIOS")
        try:
            crear_top_apuestas_df_v23467(picks_df, n=10).to_excel(writer, index=False, sheet_name="TOP APUESTAS")
            crear_combinada_recomendada_df_v23467(picks_df).to_excel(writer, index=False, sheet_name="COMBINADA RECOMENDADA")
            crear_juegos_jugador_df_v23480(detalle_df).to_excel(writer, index=False, sheet_name="JUEGOS_JUGADOR")
        except Exception:
            pass
        detalle_df.to_excel(writer, index=False, sheet_name="DETALLE TECNICO")
        if not ko_sheet.empty:
            ko_sheet.to_excel(writer, index=False, sheet_name="NO ENCONTRADOS")
        aplicar_estilo_excel_limpio(writer)
    output.seek(0)
    return output.getvalue()

# =========================================================
# v23.33 Control Panel + Safe UX Helpers
# Solo añade visibilidad, control de estado y backups. No cambia cálculos.
# =========================================================

def _safe_exists(path):
    try:
        return os.path.exists(path)
    except Exception:
        return False


def _file_status_row(label, path):
    return {
        "Archivo": label,
        "Ruta": path,
        "Estado": "✅ OK" if _safe_exists(path) else "❌ Falta"
    }


def build_control_panel_files(circuito):
    r = rutas(circuito)
    rows = [
        _file_status_row("Elo", r.get("elo", "")),
        _file_status_row("Serve general", r.get("serve", "")),
        _file_status_row("Return general", r.get("return", "")),
        _file_status_row("Break general", r.get("break", "")),
        _file_status_row("Históricos", r.get("historicos", "")),
    ]
    for sf in ["Hard", "Clay", "Grass"]:
        rows.append(_file_status_row(f"Serve {sf}", ruta_stats_superficie(circuito, "serve", sf)))
        rows.append(_file_status_row(f"Return {sf}", ruta_stats_superficie(circuito, "return", sf)))
        rows.append(_file_status_row(f"Break {sf}", ruta_stats_superficie(circuito, "break", sf)))
    if str(circuito).upper() == "ATP":
        rows.append(_file_status_row("Challenger Elo", os.path.join("datos", "challenger", "challenger_elo.xlsx")))
    return pd.DataFrame(rows)


def render_control_panel(db, circuito):
    st.subheader("🧭 Panel de Control")
    st.caption("Panel seguro: solo comprueba estado, archivos y sesión. No toca predicciones ni mercados Over.")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Versión app", APP_VERSION)
    c2.metric("Jugadores cargados", f"{len(db):,}")
    c3.metric("Circuito", circuito)
    c4.metric("Over", "🔒 Blindado" if OVER_ENGINE_LOCKED else "Abierto")

    st.success(OVER_ENGINE_LOCK_NOTE)
    st.info("Fuente de partidos fijada: SofaScore. Datos extra/listas simples quedan fuera del flujo principal para evitar lecturas mezcladas.")

    st.divider()
    st.subheader("📁 Archivos detectados")
    files_df = build_control_panel_files(circuito)
    st.dataframe(files_df, width='stretch', hide_index=True)

    faltan = files_df[files_df["Estado"].astype(str).str.contains("Falta", na=False)]
    if not faltan.empty:
        st.warning(f"Faltan {len(faltan)} rutas/archivos. La app puede funcionar con fallbacks, pero conviene revisarlo.")
    else:
        st.success("Archivos principales encontrados.")

    st.divider()
    st.subheader("🧹 Herramientas seguras")
    col_a, col_b = st.columns(2)
    with col_a:
        if st.button("🧹 Limpiar caché ahora", key="control_clear_cache", width='stretch'):
            st.cache_data.clear()
            st.success("Caché limpiada. Recarga o vuelve a ejecutar el análisis.")
    with col_b:
        if st.button("🔄 Reiniciar resultados de lote", key="control_clear_batch", width='stretch'):
            for _k in ["batch_ok_df", "batch_ko_df", "batch_last_ready"]:
                st.session_state.pop(_k, None)
            st.success("Resultados de lote eliminados de la sesión.")

    if "batch_ok_df" in st.session_state:
        ok = st.session_state.get("batch_ok_df")
        ko = st.session_state.get("batch_ko_df")
        st.divider()
        st.subheader("📊 Último análisis en sesión")
        a, b, c = st.columns(3)
        a.metric("Partidos analizados", len(ok) if ok is not None else 0)
        b.metric("No encontrados", len(ko) if ko is not None else 0)
        if ok is not None and isinstance(ok, pd.DataFrame) and "Pick oficial" in ok.columns:
            c.metric("Picks oficiales", int(ok["Pick oficial"].astype(str).str.strip().ne("").sum()))
        else:
            c.metric("Picks oficiales", "-")


def backup_analysis_excel(ok_df, ko_df, db, prefix="analisis_lista_tennis_ia"):
    """Guarda una copia local del Excel en exports/. No modifica datos ni cálculos."""
    try:
        os.makedirs("exports", exist_ok=True)
        stamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"{prefix}_{stamp}.xlsx"
        path = os.path.join("exports", filename)
        with open(path, "wb") as f:
            f.write(batch_excel_with_not_found_bytes(ok_df, ko_df, db))
        return True, path
    except Exception as e:
        return False, str(e)


# =========================================================
# v23.35 SOFASCORE AUTO EXTRACTOR
# Extrae calendario de tenis desde SofaScore y lo convierte al mismo
# formato que ya entiende el parser manual. NO toca Over ni simulaciones.
# =========================================================

SOFASCORE_AUTO_EXTRACTOR_VERSION = "v23.35.1-sofascore-auto-anti403"
SOFASCORE_API_BASES = [
    # SofaScore usa estos endpoints internamente en la web.
    # Algunos servidores pueden recibir 403; por eso abajo usamos sesión, prewarm y fallback.
    "https://api.sofascore.com/api/v1",
    "https://www.sofascore.com/api/v1",
]


def _sofa_get_nested(d, path, default=""):
    cur = d or {}
    for part in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(part)
        if cur is None:
            return default
    return cur


def _sofa_headers(kind="api", fecha_iso=None):
    """Cabeceras tipo navegador. No toca análisis ni Over; solo ayuda a evitar 403."""
    referer_date = ""
    try:
        if fecha_iso:
            referer_date = "/" + str(fecha_iso)
    except Exception:
        referer_date = ""

    common = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": f"https://www.sofascore.com/tennis{referer_date}",
        "Origin": "https://www.sofascore.com",
        "Sec-Fetch-Site": "same-site",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
    }
    if kind == "html":
        common.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Dest": "document",
        })
    else:
        common.update({
            "Accept": "application/json, text/plain, */*",
            "X-Requested-With": "XMLHttpRequest",
        })
    return common


def _sofa_create_session():
    """Crea sesión normal u opcionalmente cloudscraper si está instalado."""
    try:
        import cloudscraper  # opcional: pip install cloudscraper
        return cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows", "mobile": False}), "cloudscraper"
    except Exception:
        return requests.Session(), "requests"


def _sofa_country(team):
    c = _sofa_get_nested(team, ["country", "name"], "")
    if c:
        return str(c)
    alpha = _sofa_get_nested(team, ["country", "alpha2"], "")
    # País seguro para que el parser no rompa si SofaScore no trae país.
    # No afecta al análisis: solo sirve para mantener el formato de pegado.
    return str(alpha).upper() if alpha else "USA"


def _sofa_team_name(team):
    for k in ["name", "shortName", "slug"]:
        v = (team or {}).get(k)
        if v:
            return normalizar_texto(v)
    return ""


def _sofa_is_doubles(event):
    txt = " ".join([
        _sofa_team_name(event.get("homeTeam", {})),
        _sofa_team_name(event.get("awayTeam", {})),
        str(_sofa_get_nested(event, ["tournament", "name"], "")),
        str(_sofa_get_nested(event, ["tournament", "uniqueTournament", "name"], "")),
    ]).lower()
    return "/" in txt or " doubles" in txt or "dobles" in txt


def _sofa_event_context(event):
    tournament = event.get("tournament", {}) or {}
    uniq = tournament.get("uniqueTournament", {}) or {}
    cat = tournament.get("category", {}) or uniq.get("category", {}) or {}
    parts = [
        tournament.get("name", ""),
        uniq.get("name", ""),
        cat.get("name", ""),
        tournament.get("slug", ""),
        uniq.get("slug", ""),
        cat.get("slug", ""),
    ]
    txt = " ".join(normalizar_texto(x) for x in parts if x)
    up = txt.upper()

    # Circuito fuente para que luego filtrar_matches_por_circuito_pegado funcione igual.
    if "WTA" in up or "WOMEN" in up:
        circuito = "WTA_125" if "125" in up else "WTA"
    elif "CHALLENGER" in up:
        circuito = "CHALLENGER_ATP"
    elif "ITF" in up:
        circuito = "ITF_WTA" if ("WOMEN" in up or "WTA" in up) else "ITF_ATP"
    elif "ATP" in up or "MEN" in up or "GRAND SLAM" in up:
        circuito = "ATP"
    else:
        # Fallback conservador: si no sabemos, lo dejamos desconocido para no mezclar cuadros.
        circuito = "DESCONOCIDO"

    surface = normalizar_superficie_pegada(txt, default="Clay")
    torneo = tournament.get("name") or uniq.get("name") or "SofaScore Tennis"
    return str(torneo), circuito, surface


def _sofa_time_hhmm(event):
    ts = event.get("startTimestamp")
    try:
        if ts:
            return pd.to_datetime(int(ts), unit="s", utc=True).tz_convert("Europe/Madrid").strftime("%H:%M")
    except Exception:
        pass
    return "00:00"


def _sofa_event_to_match(event, fecha_iso=""):
    home = event.get("homeTeam", {}) or {}
    away = event.get("awayTeam", {}) or {}
    p1 = _sofa_team_name(home)
    p2 = _sofa_team_name(away)
    if not p1 or not p2:
        return None
    if _sofa_is_doubles(event):
        return None
    if es_rival_pendiente_pegado(p1) or es_rival_pendiente_pegado(p2):
        return None

    torneo, circuito_detectado, surface = _sofa_event_context(event)
    return {
        "date": fecha_iso,
        "time": _sofa_time_hhmm(event),
        "p1_country": _sofa_country(home),
        "p2_country": _sofa_country(away),
        "p1_raw": p1,
        "p2_raw": p2,
        "surface": surface,
        "torneo": torneo,
        "circuito_detectado": circuito_detectado,
        "odd1": None,
        "odd2": None,
        "quoted_side": None,
        "quoted_odd": None,
        "quoted_text": None,
        "source": "SofaScore auto",
    }


@st.cache_data(show_spinner=False, ttl=900)
def fetch_sofascore_tennis_schedule(fecha_iso):
    """
    Devuelve partidos del día desde SofaScore.
    Best-effort anti-403: sesión persistente, pre-carga de la web y varias bases API.
    Si SofaScore bloquea el servidor, el flujo manual sigue intacto.
    """
    last_error = ""
    debug_attempts = []
    session, session_kind = _sofa_create_session()

    # 1) Prewarm: primero abrimos la página pública para recoger cookies si SofaScore las entrega.
    try:
        warm_url = f"https://www.sofascore.com/tennis/{fecha_iso}"
        warm = session.get(warm_url, headers=_sofa_headers("html", fecha_iso), timeout=12)
        debug_attempts.append(f"warm {warm.status_code}")
        time.sleep(0.25)
    except Exception as e:
        debug_attempts.append(f"warm error: {str(e)[:80]}")

    # 2) Endpoint interno usado por la web.
    for base in SOFASCORE_API_BASES:
        url = f"{base}/sport/tennis/scheduled-events/{fecha_iso}"
        try:
            r = session.get(url, headers=_sofa_headers("api", fecha_iso), timeout=15)
            debug_attempts.append(f"{base} -> {r.status_code}")
            if r.status_code != 200:
                body_hint = ""
                try:
                    body_hint = str(r.text[:140]).replace("\n", " ").strip()
                except Exception:
                    body_hint = ""
                last_error = f"{r.status_code} en {base}" + (f" · {body_hint}" if body_hint else "")
                continue
            data = r.json()
            events = data.get("events", []) if isinstance(data, dict) else []
            matches = []
            for ev in events:
                if not isinstance(ev, dict):
                    continue
                m = _sofa_event_to_match(ev, fecha_iso=fecha_iso)
                if m:
                    matches.append(m)
            matches = sorted(matches, key=lambda x: (str(x.get("time", "")), str(x.get("torneo", "")), str(x.get("p1_raw", ""))))
            return {"ok": True, "matches": matches, "url": url, "error": "", "debug": "; ".join(debug_attempts), "session": session_kind}
        except Exception as e:
            last_error = str(e)
            debug_attempts.append(f"{base} error: {str(e)[:80]}")

    # 3) Mensaje claro: no es problema del modelo, es bloqueo de acceso.
    if "403" in str(last_error):
        last_error = (
            "403: SofaScore ha bloqueado la extracción automática desde este servidor. "
            "La app conserva el pegado manual. Para intentar saltarlo en local puedes instalar: pip install cloudscraper"
        )
    return {"ok": False, "matches": [], "url": "", "error": last_error or "No se pudo conectar con SofaScore", "debug": "; ".join(debug_attempts), "session": session_kind}


def _sofa_circuit_label(circuito_detectado):
    c = str(circuito_detectado).upper()
    if c == "WTA_125":
        return "WTA 125"
    if c in ["WTA", "ATP"]:
        return c
    if c == "CHALLENGER_ATP":
        return "Challenger"
    if c == "CHALLENGER_WTA":
        return "Challenger Women"
    if c == "ITF_WTA":
        return "ITF Women"
    if c == "ITF_ATP":
        return "ITF"
    return "ATP/WTA"


def sofascore_matches_to_paste_text(matches):
    """Convierte los eventos auto al mismo formato de pegado que ya usa la app."""
    if not matches:
        return ""
    grouped = {}
    for m in matches:
        key = (m.get("torneo", "SofaScore Tennis"), m.get("circuito_detectado", "DESCONOCIDO"), m.get("surface", "Clay"))
        grouped.setdefault(key, []).append(m)

    lines = []
    for (torneo, circuito_detectado, surface), items in sorted(grouped.items(), key=lambda kv: str(kv[0])):
        lines.append(str(torneo))
        lines.append(_sofa_circuit_label(circuito_detectado))
        lines.append({"Clay": "Tierra batida", "Hard": "Dura", "Grass": "Hierba"}.get(surface, surface))
        lines.append(str(len(items)))
        for m in sorted(items, key=lambda x: str(x.get("time", ""))):
            lines.extend([
                str(m.get("time", "")),
                "-",
                str(m.get("p1_country", "USA")),
                str(m.get("p1_raw", "")),
                str(m.get("p2_country", "USA")),
                str(m.get("p2_raw", "")),
            ])
        lines.append("")
    return "\n".join(lines).strip()


def render_sofascore_auto_extractor(circuito):
    """UI segura para rellenar el cuadro de texto con partidos de SofaScore."""
    with st.expander("⚡ Extraer automáticamente desde SofaScore", expanded=False):
        st.caption("Carga el calendario de tenis del día y lo pega en el cuadro de abajo con el formato que ya entiende la app. Si SofaScore bloquea o cambia algo, usa el pegado manual como respaldo.")
        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            fecha_auto = st.date_input("Fecha SofaScore", value=pd.Timestamp.today(tz="Europe/Madrid").date(), key="sofa_auto_date")
        with c2:
            solo_circuito = st.toggle(f"Filtrar ahora por {circuito}", value=True, key="sofa_auto_filter_now")
        with c3:
            st.write("")
            st.write("")
            do_fetch = st.button("📥 Cargar partidos", width='stretch', key="btn_fetch_sofascore_auto")

        if do_fetch:
            fecha_iso = pd.to_datetime(fecha_auto).strftime("%Y-%m-%d")
            res = fetch_sofascore_tennis_schedule(fecha_iso)
            if not res.get("ok"):
                st.error(f"No se pudo extraer SofaScore automáticamente: {res.get('error','error desconocido')}")
                st.info("Puedes seguir usando el pegado manual de SofaScore; no se ha tocado ningún cálculo.")
                if mostrar_debug:
                    st.caption(f"Debug SofaScore: {res.get('debug','')} · sesión: {res.get('session','')}")
            else:
                matches = res.get("matches", [])
                total = len(matches)
                if solo_circuito:
                    matches = filtrar_matches_por_circuito_pegado(matches, circuito)
                texto = sofascore_matches_to_paste_text(matches)
                st.session_state["sofa_raw_batch"] = texto
                st.success(f"Cargados {len(matches)} partidos para {circuito} ({total} total tenis). Revisa el cuadro de texto y pulsa ANALIZAR LISTA.")
                if len(matches) == 0:
                    st.warning("SofaScore respondió, pero no quedaron partidos tras filtrar. Prueba desactivar el filtro o cambia ATP/WTA en la barra lateral.")
                st.rerun()


# =========================================================
# v23.36 MOBILE API TENNIS LOADER
# Fuente alternativa para usar la app desde móvil sin copiar/pegar SofaScore.
# Solo rellena el cuadro de texto en formato ya soportado. NO toca Over ni simulaciones.
# =========================================================

API_TENNIS_IMPORT_VERSION = "v23.36.2-api-tennis-grandslam-filter-fix"
API_TENNIS_BASE_URL = "https://api.api-tennis.com/tennis/"


def _secret_value(*names, default=""):
    """Lee claves desde st.secrets o variables de entorno sin romper la app."""
    for name in names:
        try:
            if name in st.secrets:
                v = st.secrets.get(name, "")
                if v:
                    return str(v).strip()
        except Exception:
            pass
        try:
            v = os.environ.get(name, "")
            if v:
                return str(v).strip()
        except Exception:
            pass
    return default


def _api_tennis_event_type_to_circuit(event_type, tournament_name="", preferred_circuit=""):
    """Clasifica eventos de API Tennis.

    v23.36.2 corrige un caso importante: algunos Grand Slam/qualy vienen con
    event_type_type genérico y el circuito aparece solo en el torneo o se debe
    heredar del selector lateral ATP/WTA. Si no lo hacemos, la app se queda con
    Challenger y parece que no hay picks ATP/WTA.
    """
    txt = normalizar_texto(f"{event_type} {tournament_name}").upper()
    pref = str(preferred_circuit or "").upper().strip()

    if "DOUBLES" in txt or "DOUBLES" in limpiar(txt):
        return "DOBLES"

    # ITF primero: no debe entrar como WTA/ATP normal.
    if "ITF" in txt:
        if "WOMEN" in txt or "WTA" in txt or pref == "WTA":
            return "ITF_WTA"
        return "ITF_ATP"

    if "WTA" in txt or "WOMEN" in txt or "WOMAN" in txt:
        if "CHALLENGER" in txt or "125" in txt:
            return "WTA_125"
        return "WTA"

    if "CHALLENGER" in txt:
        if "WOMEN" in txt or "WTA" in txt or pref == "WTA":
            return "CHALLENGER_WTA"
        return "CHALLENGER_ATP"

    if "ATP" in txt or "MEN" in txt or "MAN" in txt:
        return "ATP"

    # Grand Slam / Major / Qualy genéricos: heredan ATP/WTA del selector de la app.
    major_tokens = [
        "GRAND SLAM", "GRANDSLAM", "ROLAND GARROS", "FRENCH OPEN",
        "AUSTRALIAN OPEN", "WIMBLEDON", "US OPEN", "QUALIFYING", "QUALIFICATION"
    ]
    if any(tok in txt for tok in major_tokens):
        if pref in {"ATP", "WTA"}:
            return pref

    return "DESCONOCIDO"


def _api_tennis_is_doubles(item):
    txt = " ".join([
        str(item.get("event_type_type", "")),
        str(item.get("tournament_name", "")),
        str(item.get("event_first_player", "")),
        str(item.get("event_second_player", "")),
    ]).lower()
    return "doubles" in txt or "/" in txt or " dobles" in txt


def _api_tennis_surface(item, default_surface="Clay"):
    txt = " ".join([
        str(item.get("tournament_name", "")),
        str(item.get("tournament_round", "")),
        str(item.get("event_type_type", "")),
    ])
    return normalizar_superficie_pegada(txt, default=default_surface)


def _api_tennis_to_match(item, fecha_iso="", default_surface="Clay", preferred_circuit=""):
    p1 = normalizar_texto(item.get("event_first_player", ""))
    p2 = normalizar_texto(item.get("event_second_player", ""))
    if not p1 or not p2:
        return None
    if _api_tennis_is_doubles(item):
        return None
    if es_rival_pendiente_pegado(p1) or es_rival_pendiente_pegado(p2):
        return None

    circuito_detectado = _api_tennis_event_type_to_circuit(item.get("event_type_type", ""), item.get("tournament_name", ""), preferred_circuit=preferred_circuit)
    if circuito_detectado == "DOBLES":
        return None

    torneo = normalizar_texto(item.get("tournament_name", "")) or "API Tennis"
    ronda = normalizar_texto(item.get("tournament_round", ""))
    if ronda and ronda not in torneo:
        torneo = f"{torneo} · {ronda}"

    return {
        "date": fecha_iso or normalizar_texto(item.get("event_date", "")),
        "time": normalizar_texto(item.get("event_time", "")) or "00:00",
        # API Tennis no siempre devuelve país en fixtures; usamos marcador neutro para conservar formato.
        "p1_country": "USA",
        "p2_country": "USA",
        "p1_raw": p1,
        "p2_raw": p2,
        "surface": _api_tennis_surface(item, default_surface=default_surface),
        "torneo": torneo,
        "circuito_detectado": circuito_detectado,
        "odd1": None,
        "odd2": None,
        "quoted_side": None,
        "quoted_odd": None,
        "quoted_text": None,
        "source": "API Tennis auto",
        "api_tennis_event_key": item.get("event_key", ""),
    }


@st.cache_data(show_spinner=False, ttl=900)
def fetch_api_tennis_fixtures(fecha_iso, default_surface="Clay", preferred_circuit=""):
    """
    Carga fixtures desde API-Tennis. Requiere API_TENNIS_KEY en secrets.toml.
    Solo convierte datos a formato de pegado; no modifica análisis ni mercados.
    """
    api_key = _secret_value("API_TENNIS_KEY", "API_TENNIS_API_KEY", "APITENNIS_KEY")
    if not api_key:
        return {
            "ok": False,
            "matches": [],
            "error": "Falta API_TENNIS_KEY en .streamlit/secrets.toml",
            "debug": "sin_api_key",
        }

    params = {
        "method": "get_fixtures",
        "APIkey": api_key,
        "date_start": fecha_iso,
        "date_stop": fecha_iso,
        "timezone": "Europe/Madrid",
    }
    safe_debug = f"method=get_fixtures date={fecha_iso} timezone=Europe/Madrid"
    try:
        r = requests.get(API_TENNIS_BASE_URL, params=params, timeout=20)
        if r.status_code != 200:
            return {"ok": False, "matches": [], "error": f"HTTP {r.status_code} en API Tennis", "debug": safe_debug}
        data = r.json()
    except Exception as e:
        return {"ok": False, "matches": [], "error": f"Error conectando con API Tennis: {str(e)[:160]}", "debug": safe_debug}

    if not isinstance(data, dict):
        return {"ok": False, "matches": [], "error": "Respuesta API Tennis no válida", "debug": safe_debug}

    if str(data.get("success", "0")) not in ["1", "True", "true"]:
        err = data.get("error") or data.get("message") or data.get("result") or "API Tennis no devolvió success=1"
        return {"ok": False, "matches": [], "error": str(err)[:220], "debug": safe_debug}

    result = data.get("result", [])
    if not isinstance(result, list):
        return {"ok": False, "matches": [], "error": "API Tennis no devolvió lista de partidos", "debug": safe_debug}

    matches = []
    for item in result:
        if not isinstance(item, dict):
            continue
        m = _api_tennis_to_match(item, fecha_iso=fecha_iso, default_surface=default_surface, preferred_circuit=preferred_circuit)
        if m:
            matches.append(m)

    matches = sorted(matches, key=lambda x: (str(x.get("time", "")), str(x.get("torneo", "")), str(x.get("p1_raw", ""))))
    return {"ok": True, "matches": matches, "error": "", "debug": safe_debug, "raw_count": len(result)}


def render_api_tennis_mobile_loader(circuito):
    """UI móvil: carga partidos por API Tennis y rellena el cuadro de SofaScore."""
    with st.expander("📲 Modo móvil: cargar partidos con API Tennis", expanded=False):
        st.caption("Alternativa para fines de semana: desde el móvil pulsas un botón y la app rellena el cuadro. Requiere API_TENNIS_KEY en secrets.toml. No toca Over ni probabilidades.")

        api_key_present = bool(_secret_value("API_TENNIS_KEY", "API_TENNIS_API_KEY", "APITENNIS_KEY"))
        if api_key_present:
            st.success("API_TENNIS_KEY detectada")
        else:
            st.warning("Falta API_TENNIS_KEY. Añádela en .streamlit/secrets.toml para activar este modo.")
            st.code('API_TENNIS_KEY = "tu_clave_de_api_tennis"', language="toml")

        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            fecha_api = st.date_input("Fecha API Tennis", value=pd.Timestamp.today(tz="Europe/Madrid").date(), key="api_tennis_date")
        with c2:
            solo_circuito_api = st.toggle(f"Filtrar por {circuito}", value=True, key="api_tennis_filter_now")
        with c3:
            default_surface_label = st.selectbox("Superficie si falta", ["Clay", "Hard", "Grass"], index=0, key="api_tennis_default_surface")

        circuito_ui_api = str(circuito).upper().strip()
        if circuito_ui_api == "ATP":
            secondary_label = "Incluir Challenger ATP"
            secondary_on_msg = "Challenger ATP incluido. Se analizará con el motor Challenger si hay datos; si no hay muestra suficiente, la app lo bloqueará como contexto."
            secondary_off_msg = "Challenger ATP excluido: solo ATP principal / Grand Slam / qualy ATP."
        else:
            secondary_label = "Incluir WTA 125 / Challenger WTA"
            secondary_on_msg = "WTA 125 / Challenger WTA incluido. No se mezclará con ATP."
            secondary_off_msg = "WTA 125 / Challenger WTA excluido: solo WTA principal / Grand Slam / qualy WTA."

        incluir_challenger_api = st.toggle(secondary_label, value=True, key="api_tennis_include_challenger")
        incluir_itf_api = st.toggle(f"Incluir ITF {'ATP' if circuito_ui_api == 'ATP' else 'WTA'}", value=False, key="api_tennis_include_itf")
        if incluir_challenger_api:
            st.caption(secondary_on_msg)
        else:
            st.caption(secondary_off_msg)
        if incluir_itf_api:
            st.warning("ITF activado solo para observación: puede tener menos datos y más ruido. No recomendado para picks oficiales.")
        else:
            st.caption("ITF queda excluido por defecto para no ensuciar los pronósticos.")

        do_api_fetch = st.button("📥 Cargar desde API Tennis", width='stretch', key="btn_fetch_api_tennis_mobile")

        if do_api_fetch:
            fecha_iso = pd.to_datetime(fecha_api).strftime("%Y-%m-%d")
            res = fetch_api_tennis_fixtures(fecha_iso, default_surface=default_surface_label, preferred_circuit=(circuito if solo_circuito_api else ""))
            if not res.get("ok"):
                st.error(f"No se pudo cargar API Tennis: {res.get('error', 'error desconocido')}")
                if mostrar_debug:
                    st.caption(f"Debug API Tennis: {res.get('debug','')}")
            else:
                matches = res.get("matches", [])
                total = len(matches)
                raw_count = res.get("raw_count", total)
                from collections import Counter as _Counter
                circuit_counts_before = _Counter(str(m.get("circuito_detectado", "DESCONOCIDO")) for m in matches)
                itf_count = sum(1 for m in matches if "ITF" in str(m.get("circuito_detectado", "")).upper())

                if solo_circuito_api:
                    matches = filtrar_matches_por_circuito_pegado(matches, circuito, incluir_itf=incluir_itf_api, incluir_challenger=incluir_challenger_api)
                elif not incluir_itf_api:
                    matches = [m for m in matches if "ITF" not in str(m.get("circuito_detectado", "")).upper()]

                circuit_counts = _Counter(str(m.get("circuito_detectado", "DESCONOCIDO")) for m in matches)
                texto = sofascore_matches_to_paste_text(matches)
                st.session_state["sofa_raw_batch"] = texto
                st.success(f"API Tennis cargó {len(matches)} partidos para {circuito} ({total} singles útiles / {raw_count} eventos brutos). Revisa el cuadro y pulsa ANALIZAR LISTA.")

                if circuit_counts_before:
                    st.caption("API devolvió: " + " · ".join([f"{k}: {v}" for k, v in circuit_counts_before.items()]))
                if circuit_counts:
                    st.caption("Tras filtro del menú lateral: " + " · ".join([f"{k}: {v}" for k, v in circuit_counts.items()]))
                if itf_count and not incluir_itf_api:
                    st.info(f"Se han excluido {itf_count} partidos ITF por defecto para no ensuciar los pronósticos.")
                if len(matches) > 0 and all("CHALLENGER" in str(m.get("circuito_detectado", "")).upper() for m in matches):
                    st.warning("Solo han quedado partidos Challenger. El motor Challenger sigue activo, pero si los jugadores no tienen muestra/elo suficiente saldrán como contexto y no como oficiales. No es un fallo del Over.")
                if len(matches) == 0:
                    st.warning("La API respondió, pero no quedaron partidos tras filtrar. Prueba cambiar ATP/WTA, activar secundarios del circuito correcto o revisar la fecha.")
                st.rerun()

# =========================================================
# v23.44.0 PREDICTOR · TA FINE TUNE
# Permite pegar ficha completa TennisAbstract/Flashscore de los dos jugadores
# dentro del Predictor individual. No toca el motor Over ni la simulación:
# crea un ajuste de lectura final y muestra porcentajes afinados con trazabilidad.
# =========================================================

def _pred_ta_num(d, key, default=0.0):
    try:
        if not isinstance(d, dict):
            return default
        v = d.get(key, default)
        if v in [None, "", "nan"]:
            return default
        return float(v)
    except Exception:
        return default


def _pred_ta_pct_label(x):
    try:
        return f"{float(x):.1%}"
    except Exception:
        return "N/A"


def _pred_ta_confidence(ta1, ta2):
    n1 = _pred_ta_num(ta1, "matches", 0)
    n2 = _pred_ta_num(ta2, "matches", 0)
    n = min(n1, n2) if (n1 and n2) else max(n1, n2)
    if n >= 8:
        return 0.30, "alta"
    if n >= 5:
        return 0.22, "media"
    if n >= 3:
        return 0.14, "baja"
    return 0.08, "muy baja"


def _pred_blend(base, signal, weight, floor=0.02, ceil=0.98):
    try:
        base = float(base)
        if signal is None:
            return float(np.clip(base, floor, ceil))
        signal = float(signal)
        return float(np.clip(base * (1 - weight) + signal * weight, floor, ceil))
    except Exception:
        return base


def _pred_ta_compute_adjustments(base, ta1, ta2, p1_name="Jugador 1", p2_name="Jugador 2"):
    """Devuelve porcentajes ajustados por fichas TA recientes.
    Es una capa visual del Predictor; no modifica el motor Monte Carlo.
    """
    base = base or {}
    w, conf_label = _pred_ta_confidence(ta1, ta2)

    def avg_key(key, default=0.0):
        vals = []
        for d in [ta1, ta2]:
            if isinstance(d, dict) and d.get(key, "") not in [None, ""]:
                vals.append(_pred_ta_num(d, key, default))
        return float(np.mean(vals)) if vals else default

    over_rate = avg_key("over_rate", 0.0)
    over19_rate = avg_key("over19_rate", 0.0)
    three_rate = avg_key("three_rate", 0.0)
    easy_rate = avg_key("easy_rate", 0.0)
    avg_games = avg_key("avg_games", 0.0)
    set1 = _pred_ta_num(ta1, "set_rate", 0.0)
    set2 = _pred_ta_num(ta2, "set_rate", 0.0)
    lost1 = _pred_ta_num(ta1, "lost02_rate", 0.0)
    lost2 = _pred_ta_num(ta2, "lost02_rate", 0.0)
    elo1 = _pred_ta_num(ta1, "elo", 0.0)
    elo2 = _pred_ta_num(ta2, "elo", 0.0)
    wins1 = _pred_ta_num(ta1, "wins", 0.0)
    wins2 = _pred_ta_num(ta2, "wins", 0.0)
    n1 = max(1.0, _pred_ta_num(ta1, "matches", 0.0))
    n2 = max(1.0, _pred_ta_num(ta2, "matches", 0.0))
    form1 = wins1 / n1 if ta1 else 0.50
    form2 = wins2 / n2 if ta2 else 0.50

    # v23.46: Señales TA más prudentes.
    # La ficha TA NO debe convertir una señal aislada en conclusión fuerte.
    # El histórico Challenger mostró que three_rate por sí solo apenas predice 3 sets,
    # y que easy_rate alto debe frenar Overs altos y lecturas de partido largo.
    over18_signal = None
    if over_rate > 0 or avg_games > 0 or three_rate > 0:
        easy_cut = 0.18 * easy_rate
        if easy_rate >= 0.55:
            easy_cut += 0.08
        elif easy_rate >= 0.45:
            easy_cut += 0.04
        over18_signal = 0.50 + (over_rate - 0.50) * 0.50 + ((avg_games or 21.0) - 21.0) * 0.010 + (three_rate - 0.30) * 0.04 - easy_cut
        over18_signal = float(np.clip(over18_signal, 0.34, 0.84))
    over19_signal = None
    if over19_rate > 0 or three_rate > 0 or avg_games > 0:
        easy_cut = 0.22 * easy_rate
        if easy_rate >= 0.55:
            easy_cut += 0.08
        over19_signal = 0.45 + (over19_rate - 0.45) * 0.50 + ((avg_games or 21.0) - 21.5) * 0.008 + (three_rate - 0.30) * 0.035 - easy_cut
        over19_signal = float(np.clip(over19_signal, 0.28, 0.78))
    over20_signal = None
    if over19_rate > 0 or avg_games > 0:
        easy_cut = 0.24 * easy_rate
        if easy_rate >= 0.50:
            easy_cut += 0.08
        over20_signal = 0.40 + (over19_rate - 0.45) * 0.34 + ((avg_games or 21.0) - 22.0) * 0.010 - easy_cut
        over20_signal = float(np.clip(over20_signal, 0.18, 0.68))

    ml1_base = float(base.get("p1", 0.50) or 0.50)
    ml1_signal = None
    if elo1 and elo2:
        ml1_signal = elo_prob(elo1, elo2)
        # Pequeño ajuste por forma reciente, capado.
        ml1_signal += float(np.clip((form1 - form2) * 0.10, -0.035, 0.035))
        ml1_signal = float(np.clip(ml1_signal, 0.12, 0.88))

    p1_adj = _pred_blend(ml1_base, ml1_signal, min(w, 0.24), 0.05, 0.95)
    p2_adj = 1 - p1_adj
    over18_adj = _pred_blend(base.get("over18", 0.0), over18_signal, w, 0.05, 0.95)
    over19_adj = _pred_blend(base.get("over19", 0.0), over19_signal, w, 0.05, 0.90)
    over20_adj = _pred_blend(base.get("over20", 0.0), over20_signal, w * 0.85, 0.03, 0.85)

    # Gana al menos un set desde TA. v23.46: capado y con penalización fuerte si pierde 0-2 o easy_rate es alto.
    # En el backtest y en los ejemplos reales estaba inflado en zona media.
    set1_signal = None
    set2_signal = None
    if set1:
        set1_signal = float(np.clip(0.45 + set1 * 0.34 - lost1 * 0.25 - easy_rate * 0.08, 0.32, 0.88))
    if set2:
        set2_signal = float(np.clip(0.45 + set2 * 0.34 - lost2 * 0.25 - easy_rate * 0.08, 0.32, 0.88))

    notes = []
    if over_rate >= 0.70:
        notes.append(f"TA refuerza Over: Over18 reciente combinado {over_rate:.0%}")
    elif over_rate and over_rate <= 0.45:
        notes.append(f"TA frena Over: Over18 reciente bajo {over_rate:.0%}")
    if easy_rate >= 0.35:
        notes.append(f"Riesgo de paliza/partido corto: {easy_rate:.0%}")
    if three_rate >= 0.35:
        notes.append(f"Buena señal de partido largo: 3 sets {three_rate:.0%}")
    if elo1 and elo2:
        notes.append(f"Elo TA: {p1_name} {elo1:.0f} vs {p2_name} {elo2:.0f}")
    if not notes:
        notes.append("TA leído, pero sin señal fuerte; ajuste prudente.")

    markets = [
        (p1_adj, f"{p1_name} gana", "ML"),
        (p2_adj, f"{p2_name} gana", "ML"),
        (over18_adj, "Over 18.5", "OVER18"),
        (over19_adj, "Over 19.5", "OVER19"),
        (over20_adj, "Over 20.5", "OVER20"),
    ]
    if set1_signal is not None:
        markets.append((set1_signal, f"{p1_name} gana al menos 1 set", "SET"))
    if set2_signal is not None:
        markets.append((set2_signal, f"{p2_name} gana al menos 1 set", "SET"))
    markets = sorted(markets, key=lambda x: x[0], reverse=True)

    return {
        "confidence": conf_label,
        "weight": w,
        "p1_adj": p1_adj,
        "p2_adj": p2_adj,
        "over18_adj": over18_adj,
        "over19_adj": over19_adj,
        "over20_adj": over20_adj,
        "set1_signal": set1_signal,
        "set2_signal": set2_signal,
        "notes": notes,
        "markets": markets,
        "raw_signals": {
            "over_rate": over_rate, "over19_rate": over19_rate, "three_rate": three_rate,
            "easy_rate": easy_rate, "avg_games": avg_games, "set1": set1, "set2": set2,
            "lost1": lost1, "lost2": lost2, "elo1": elo1, "elo2": elo2,
        }
    }



# =========================================================
# v23.44.1 PREDICTOR · EXCEL DOWNLOAD
# Descarga un Excel del Predictor individual con datos base y, si existe,
# afinado TennisAbstract. No toca motor ni fórmulas; solo exporta lectura.
# =========================================================

def _predictor_excel_safe_pct(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _predictor_excel_bytes(payload, ta_adj=None):
    """Exporta el Predictor en UNA sola hoja, más cómodo para revisar/imprimir.
    No modifica cálculos: solo junta resumen, mercados base, ajuste TA, notas y señales.
    """
    output = io.BytesIO()
    payload = payload or {}
    base = payload.get("base", {}) if isinstance(payload.get("base", {}), dict) else {}

    rows = []

    def add_section(title):
        rows.append({"Sección": title, "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": ""})

    def pct(v):
        try:
            if v is None or v == "":
                return ""
            return f"{float(v):.1%}"
        except Exception:
            return ""

    def add_info(campo, valor):
        rows.append({"Sección": "RESUMEN", "Campo/Mercado": campo, "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": valor})

    add_section("RESUMEN")
    add_info("Versión app", APP_VERSION)
    add_info("Jugador 1", payload.get("p1_name", ""))
    add_info("Jugador 2", payload.get("p2_name", ""))
    add_info("Partido", f"{payload.get('p1_name','')} vs {payload.get('p2_name','')}")
    add_info("Circuito", payload.get("circuito", ""))
    add_info("Superficie", payload.get("surface", ""))
    add_info("Formato", "5 sets" if int(payload.get("best_of", 3) or 3) == 5 else "3 sets")
    add_info("Nota", "Exportación del Predictor individual en una sola hoja. No modifica cálculos.")

    add_section("MERCADOS BASE")
    mercados = [
        (f"{payload.get('p1_name','Jugador 1')} gana", base.get("p1"), "ML"),
        (f"{payload.get('p2_name','Jugador 2')} gana", base.get("p2"), "ML"),
        ("Over 17.5 WTA", base.get("over17"), "Over"),
        ("Over 18.5", base.get("over18"), "Over"),
        ("Over 19.5", base.get("over19"), "Over"),
        ("Over 20.5", base.get("over20"), "Over"),
        ("Over 22.5", base.get("over22"), "Over"),
    ]
    for mercado, prob, tipo in mercados:
        if _predictor_excel_safe_pct(prob) is not None:
            rows.append({"Sección": "MERCADOS BASE", "Campo/Mercado": mercado, "Tipo": tipo, "Base %": pct(prob), "TA ajustada %": "", "Valor/Nota": ""})

    # v23.45.1: exportar también el Deep Match Analyzer en esta misma hoja.
    deep = payload.get("deep", {}) if isinstance(payload.get("deep", {}), dict) else {}
    add_section("DEEP MATCH ANALYZER · RESUMEN")
    if deep:
        deep_info = [
            ("Media juegos esperada", deep.get("avg_games")),
            ("Mediana juegos", deep.get("med_games")),
            (f"Media juegos {payload.get('p1_name','Jugador 1')}", deep.get("avg_g1")),
            (f"Media juegos {payload.get('p2_name','Jugador 2')}", deep.get("avg_g2")),
            (f"Mediana juegos {payload.get('p1_name','Jugador 1')}", deep.get("med_g1")),
            (f"Mediana juegos {payload.get('p2_name','Jugador 2')}", deep.get("med_g2")),
            ("Diferencia media J1-J2", deep.get("game_diff")),
            (f"Rank {payload.get('p1_name','Jugador 1')}", deep.get("p1_rank")),
            (f"Rank {payload.get('p2_name','Jugador 2')}", deep.get("p2_rank")),
            (f"Elo superficie {payload.get('p1_name','Jugador 1')}", deep.get("p1_elo_surface")),
            (f"Elo superficie {payload.get('p2_name','Jugador 2')}", deep.get("p2_elo_surface")),
            (f"Hold {payload.get('p1_name','Jugador 1')}", deep.get("hold1")),
            (f"Hold {payload.get('p2_name','Jugador 2')}", deep.get("hold2")),
            (f"Return {payload.get('p1_name','Jugador 1')}", deep.get("ret1")),
            (f"Return {payload.get('p2_name','Jugador 2')}", deep.get("ret2")),
            ("Volatilidad", deep.get("vol")),
        ]
        for campo, valor in deep_info:
            if valor is not None and valor != "":
                rows.append({"Sección": "DEEP MATCH ANALYZER · RESUMEN", "Campo/Mercado": campo, "Tipo": "Dato profundo", "Base %": "", "TA ajustada %": "", "Valor/Nota": valor})
    else:
        rows.append({"Sección": "DEEP MATCH ANALYZER · RESUMEN", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": "Sin datos profundos calculados."})

    add_section("DEEP MATCH ANALYZER · INTELLIGENCE")
    try:
        intel = _deep_intelligence_report(payload, ta_adj if isinstance(ta_adj, dict) else None)
        rows.append({"Sección": "DEEP MATCH ANALYZER · INTELLIGENCE", "Campo/Mercado": "Historia final", "Tipo": "Lectura", "Base %": "", "TA ajustada %": "", "Valor/Nota": f"{intel.get('story','')} · {intel.get('story_detail','')}"})
        for r in intel.get("rows", []):
            rows.append({"Sección": "DEEP MATCH ANALYZER · INTELLIGENCE", "Campo/Mercado": r.get("Área", ""), "Tipo": r.get("Estado", ""), "Base %": "", "TA ajustada %": "", "Valor/Nota": f"{r.get('Lectura','')} · {r.get('Motivo','')}"})
    except Exception as e:
        rows.append({"Sección": "DEEP MATCH ANALYZER · INTELLIGENCE", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": f"No se pudo exportar intelligence: {type(e).__name__}: {e}"})

    add_section("DEEP MATCH ANALYZER · MERCADOS")
    try:
        df_deep = _deep_build_markets(payload, ta_adj if isinstance(ta_adj, dict) else None)
        if isinstance(df_deep, pd.DataFrame) and not df_deep.empty:
            for _, r in df_deep.iterrows():
                rows.append({
                    "Sección": "DEEP MATCH ANALYZER · MERCADOS",
                    "Campo/Mercado": r.get("Mercado / lectura", ""),
                    "Tipo": r.get("Grupo", ""),
                    "Base %": r.get("Probabilidad", ""),
                    "TA ajustada %": "",
                    "Valor/Nota": f"{r.get('Nivel','')} · {r.get('Lectura','')} {r.get('Nota','')}",
                })
        else:
            rows.append({"Sección": "DEEP MATCH ANALYZER · MERCADOS", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": "Sin mercados profundos disponibles."})
    except Exception as e:
        rows.append({"Sección": "DEEP MATCH ANALYZER · MERCADOS", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": f"No se pudo exportar deep markets: {type(e).__name__}: {e}"})

    add_section("DEEP MATCH ANALYZER · CONCLUSIONES")
    try:
        conclusiones = _deep_conclusions(payload, ta_adj if isinstance(ta_adj, dict) else None)
        for i, n in enumerate(conclusiones or [], 1):
            rows.append({"Sección": "DEEP MATCH ANALYZER · CONCLUSIONES", "Campo/Mercado": f"Conclusión {i}", "Tipo": "Lectura deportiva", "Base %": "", "TA ajustada %": "", "Valor/Nota": str(n)})
    except Exception as e:
        rows.append({"Sección": "DEEP MATCH ANALYZER · CONCLUSIONES", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": f"No se pudieron exportar conclusiones: {type(e).__name__}: {e}"})

    add_section("AJUSTE TENNIS ABSTRACT")
    if isinstance(ta_adj, dict):
        ta_markets = [
            (payload.get("p1_name", "Jugador 1") + " gana", base.get("p1"), ta_adj.get("p1_adj"), "ML"),
            (payload.get("p2_name", "Jugador 2") + " gana", base.get("p2"), ta_adj.get("p2_adj"), "ML"),
            ("Over 18.5", base.get("over18"), ta_adj.get("over18_adj"), "Over"),
            ("Over 19.5", base.get("over19"), ta_adj.get("over19_adj"), "Over"),
            ("Over 20.5", base.get("over20"), ta_adj.get("over20_adj"), "Over"),
        ]
        for mercado, base_prob, adj_prob, tipo in ta_markets:
            if _predictor_excel_safe_pct(adj_prob) is not None:
                rows.append({"Sección": "AJUSTE TENNIS ABSTRACT", "Campo/Mercado": mercado, "Tipo": tipo, "Base %": pct(base_prob), "TA ajustada %": pct(adj_prob), "Valor/Nota": ""})
        if ta_adj.get("set1_signal") is not None:
            rows.append({"Sección": "AJUSTE TENNIS ABSTRACT", "Campo/Mercado": payload.get("p1_name", "Jugador 1") + " gana al menos 1 set", "Tipo": "Gana set", "Base %": "", "TA ajustada %": pct(ta_adj.get("set1_signal")), "Valor/Nota": ""})
        if ta_adj.get("set2_signal") is not None:
            rows.append({"Sección": "AJUSTE TENNIS ABSTRACT", "Campo/Mercado": payload.get("p2_name", "Jugador 2") + " gana al menos 1 set", "Tipo": "Gana set", "Base %": "", "TA ajustada %": pct(ta_adj.get("set2_signal")), "Valor/Nota": ""})

        add_section("NOTAS TA")
        notes = ta_adj.get("notes", []) or []
        if notes:
            for i, n in enumerate(notes, 1):
                rows.append({"Sección": "NOTAS TA", "Campo/Mercado": f"Nota {i}", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": str(n)})
        else:
            rows.append({"Sección": "NOTAS TA", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": "Sin notas TA."})

        add_section("SEÑALES TA")
        raw = ta_adj.get("raw_signals", {}) or {}
        if raw:
            for k, v in raw.items():
                rows.append({"Sección": "SEÑALES TA", "Campo/Mercado": str(k), "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": v})
        else:
            rows.append({"Sección": "SEÑALES TA", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": "Sin señales TA."})
    else:
        rows.append({"Sección": "AJUSTE TENNIS ABSTRACT", "Campo/Mercado": "", "Tipo": "", "Base %": "", "TA ajustada %": "", "Valor/Nota": "Sin ajuste TA calculado todavía."})

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet = "PREDICTOR"
        df.to_excel(writer, index=False, sheet_name=sheet)
        wb = writer.book
        ws = writer.sheets[sheet]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # Resaltar filas de sección.
        try:
            from openpyxl.styles import PatternFill, Font
            fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
            bold = Font(bold=True)
            for row in range(2, ws.max_row + 1):
                campo = ws.cell(row=row, column=2).value
                tipo = ws.cell(row=row, column=3).value
                base_val = ws.cell(row=row, column=4).value
                ta_val = ws.cell(row=row, column=5).value
                nota = ws.cell(row=row, column=6).value
                if not any([campo, tipo, base_val, ta_val, nota]):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill
                        ws.cell(row=row, column=col).font = bold
        except Exception:
            pass

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

    output.seek(0)
    return output.getvalue()

def render_predictor_excel_download_panel():
    payload = st.session_state.get("predictor_last_payload_v23440")
    if not isinstance(payload, dict):
        return
    ta_adj = st.session_state.get("predictor_last_ta_adjustment_v23441")
    st.divider()
    with st.expander("📥 Descargar Excel del Predictor", expanded=False):
        st.caption("Descarga un Excel del Predictor en una sola hoja, incluyendo Deep Match Analyzer, Intelligence Challenger/TA, mercados derivados, conclusiones y ajuste TennisAbstract si lo has calculado. No modifica ningún cálculo.")
        try:
            data = _predictor_excel_bytes(payload, ta_adj if isinstance(ta_adj, dict) else None)
            p1 = limpiar(payload.get("p1_name", "jugador1"))[:18] or "jugador1"
            p2 = limpiar(payload.get("p2_name", "jugador2"))[:18] or "jugador2"
            st.download_button(
                "📊 Descargar Excel Predictor",
                data=data,
                file_name=f"predictor_{p1}_vs_{p2}_{APP_VERSION}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_predictor_excel_v23451"
            )
        except Exception as e:
            st.warning(f"No se pudo preparar el Excel del Predictor: {type(e).__name__}: {e}")

def render_predictor_ta_finetune_panel():
    payload = st.session_state.get("predictor_last_payload_v23440")
    if not isinstance(payload, dict):
        return
    p1_name = payload.get("p1_name", "Jugador 1")
    p2_name = payload.get("p2_name", "Jugador 2")
    partido = f"{p1_name} vs {p2_name}"

    st.divider()
    with st.expander("🧪 Afinar Predictor con fichas completas TennisAbstract", expanded=False):
        st.caption("Pega la ficha completa de TennisAbstract/Flashscore de cada jugador. La app lee resultados recientes, Over, 3 sets, palizas, Elo si aparece y te muestra porcentajes ajustados. No modifica el motor Over ni la simulación base.")
        c1, c2 = st.columns(2)
        cache1_txt, cache1 = _ta_profile_cache_status_text(p1_name)
        cache2_txt, cache2 = _ta_profile_cache_status_text(p2_name)
        with c1:
            if cache1_txt:
                st.success(cache1_txt)
            txt1 = st.text_area(f"Ficha completa TA / Flashscore · {p1_name}", value=str((cache1 or {}).get("raw_text", "") or ""), height=220, key=f"pred_ta_txt1_{limpiar(partido)}")
        with c2:
            if cache2_txt:
                st.success(cache2_txt)
            txt2 = st.text_area(f"Ficha completa TA / Flashscore · {p2_name}", value=str((cache2 or {}).get("raw_text", "") or ""), height=220, key=f"pred_ta_txt2_{limpiar(partido)}")

        save_cache = st.checkbox("💾 Guardar estas fichas en caché para próximos análisis", value=True, key=f"pred_ta_save_{limpiar(partido)}")
        if st.button("🔬 Analizar fichas y afinar porcentajes", key=f"pred_ta_btn_{limpiar(partido)}", use_container_width=True):
            d1 = _extraer_datos_extra_desde_texto(txt1 or "", {"Partido": partido})
            d2 = _extraer_datos_extra_desde_texto(txt2 or "", {"Partido": partido})
            d1 = _invalidar_lectura_si_no_coincide(d1, p1_name)
            d2 = _invalidar_lectura_si_no_coincide(d2, p2_name)
            ta1 = d1.get("tennisabstract") or d1.get("flashscore") if isinstance(d1, dict) else None
            ta2 = d2.get("tennisabstract") or d2.get("flashscore") if isinstance(d2, dict) else None

            if save_cache:
                if str(txt1 or "").strip():
                    ok, msg = _ta_profile_cache_upsert_from_raw(p1_name, txt1)
                    if ok: st.caption(f"✅ {msg}")
                    else: st.caption(f"⚠️ {p1_name}: {msg}")
                if str(txt2 or "").strip():
                    ok, msg = _ta_profile_cache_upsert_from_raw(p2_name, txt2)
                    if ok: st.caption(f"✅ {msg}")
                    else: st.caption(f"⚠️ {p2_name}: {msg}")

            st.subheader("📌 Lectura de fichas")
            r1a, r1b = _estado_lectura_ficha_datos_extra(d1)
            r2a, r2b = _estado_lectura_ficha_datos_extra(d2)
            cc1, cc2 = st.columns(2)
            with cc1:
                st.info(f"{r1a}\n\n{r1b}")
            with cc2:
                st.info(f"{r2a}\n\n{r2b}")

            if not ta1 and not ta2:
                st.warning("No he podido leer una ficha válida. Prueba copiando más contenido de TennisAbstract, especialmente la zona de resultados recientes/summary.")
                return

            adj = _pred_ta_compute_adjustments(payload.get("base", {}), ta1, ta2, p1_name, p2_name)
            st.session_state["predictor_last_ta_adjustment_v23441"] = adj
            st.subheader("🎯 Porcentajes afinados por TA")
            st.caption(f"Confianza del ajuste: {adj['confidence']} · peso aplicado sobre el motor base: {adj['weight']:.0%}")
            m1, m2, m3, m4, m5 = st.columns(5)
            base = payload.get("base", {})
            with m1: st.metric(f"{p1_name} ML", _pred_ta_pct_label(adj["p1_adj"]), f"base {_pred_ta_pct_label(base.get('p1',0.5))}")
            with m2: st.metric(f"{p2_name} ML", _pred_ta_pct_label(adj["p2_adj"]), f"base {_pred_ta_pct_label(base.get('p2',0.5))}")
            with m3: st.metric("Over 18.5", _pred_ta_pct_label(adj["over18_adj"]), f"base {_pred_ta_pct_label(base.get('over18',0))}")
            with m4: st.metric("Over 19.5", _pred_ta_pct_label(adj["over19_adj"]), f"base {_pred_ta_pct_label(base.get('over19',0))}")
            with m5: st.metric("Over 20.5", _pred_ta_pct_label(adj["over20_adj"]), f"base {_pred_ta_pct_label(base.get('over20',0))}")

            if adj.get("set1_signal") is not None or adj.get("set2_signal") is not None:
                s1, s2 = st.columns(2)
                with s1:
                    if adj.get("set1_signal") is not None:
                        st.metric(f"{p1_name} gana set", _pred_ta_pct_label(adj.get("set1_signal")))
                with s2:
                    if adj.get("set2_signal") is not None:
                        st.metric(f"{p2_name} gana set", _pred_ta_pct_label(adj.get("set2_signal")))

            best = adj["markets"][0] if adj.get("markets") else None
            if best:
                prob, label, typ = best
                if prob >= 0.80:
                    st.success(f"Mejor mercado TA: **{label}** · {prob:.1%}")
                elif prob >= 0.70:
                    st.warning(f"Mejor mercado TA: **{label}** · {prob:.1%} · mejor como simple/observación, no forzar combinada.")
                else:
                    st.info(f"TA no deja mercado premium claro. Mejor mercado: **{label}** · {prob:.1%}")

            with st.expander("Ver ranking de mercados TA", expanded=False):
                dfm = pd.DataFrame([{"Mercado": x[1], "Prob TA ajustada": f"{x[0]:.1%}", "Tipo": x[2]} for x in adj.get("markets", [])])
                st.dataframe(dfm, use_container_width=True)
            for n in adj.get("notes", []):
                st.caption("• " + str(n))



# =========================================================
# v23.45.0 PREDICTOR · DEEP MATCH ANALYZER
# Solo añade lectura profunda en Predictor individual. No toca Analista,
# motor Over, simulación ni selector por lista.
# =========================================================

def _deep_pct(v):
    try:
        return f"{float(v):.1%}"
    except Exception:
        return "N/A"


def _deep_float(payload, section, key, default=0.0):
    try:
        sec = payload.get(section, {}) if isinstance(payload, dict) else {}
        return float(sec.get(key, default) or default)
    except Exception:
        return default


def _deep_level(prob):
    try:
        p = float(prob)
    except Exception:
        return ""
    if p >= 0.82: return "🔥 Muy alta"
    if p >= 0.74: return "✅ Alta"
    if p >= 0.64: return "🟡 Media-alta"
    if p >= 0.55: return "⚖️ Ajustada"
    return "⚠️ Baja"


def _deep_build_markets(payload, ta_adj=None):
    base = payload.get("base", {}) if isinstance(payload, dict) else {}
    deep = payload.get("deep", {}) if isinstance(payload, dict) else {}
    p1 = payload.get("p1_name", "Jugador 1")
    p2 = payload.get("p2_name", "Jugador 2")
    dog = deep.get("model_dog_name", "Underdog")
    fav = deep.get("model_fav_name", "Favorito")
    rows = []
    def add(grupo, mercado, prob, lectura="", nota=""):
        try:
            probf = float(prob)
        except Exception:
            return
        rows.append({
            "Grupo": grupo,
            "Mercado / lectura": mercado,
            "Probabilidad": probf,
            "Nivel": _deep_level(probf),
            "Lectura": lectura,
            "Nota": nota,
        })

    add("Ganador", f"{p1} gana partido", base.get("p1", 0), "ML calibrado")
    add("Ganador", f"{p2} gana partido", base.get("p2", 0), "ML calibrado")
    add("Primer set", f"{p1} gana 1er set", base.get("p1_fs", 0), "Inicio de partido")
    add("Primer set", f"{p2} gana 1er set", base.get("p2_fs", 0), "Inicio de partido")
    if payload.get("circuito") == "WTA":
        add("Juegos totales", "Over 17.5 WTA", base.get("over17", 0), "Línea baja WTA")
    add("Juegos totales", "Over 18.5", base.get("over18", 0), "Línea base")
    add("Juegos totales", "Over 19.5", base.get("over19", 0), "Línea superior")
    add("Juegos totales", "Over 20.5", base.get("over20", 0), "Línea superior")
    add("Juegos totales", "Over 22.5", base.get("over22", 0), "Línea alta")
    add("Juegos totales", "Under 22.5", 1 - float(base.get("over22", 0) or 0), "Línea defensiva")
    add("Sets", "Más de 2.5 sets" if int(payload.get("best_of", 3) or 3) == 3 else "Partido largo / 4+ sets", base.get("set3", 0), "Resistencia / set extra")
    add("Sets", f"{dog} gana al menos 1 set", base.get("dog_wins_set", 0), "Mercado de resistencia")
    add("Sets", f"{fav} gana 2-0", base.get("fav_2_0", 0), "Dominio del favorito")
    add("Especiales", "Tie-break en el partido", base.get("tb", 0), "Saque/hold alto")
    add("Especiales", "Partido largo", base.get("long_match", 0), "Lectura global")
    add("Especiales", "Favorito + Under 22.5", base.get("fav_under22", 0), "Favorito controla sin guerra")
    add("Especiales", "Dog + Over 20.5", base.get("dog_over20", 0), "Underdog compite")

    # Juegos por jugador aproximados desde Monte Carlo.
    avg_g1 = float(deep.get("avg_g1", 0) or 0)
    avg_g2 = float(deep.get("avg_g2", 0) or 0)
    if avg_g1 > 0 and avg_g2 > 0:
        for player, avg in [(p1, avg_g1), (p2, avg_g2)]:
            # Probabilidades aproximadas visuales, derivadas de media esperada; no sustituyen al motor.
            add("Juegos jugador", f"{player} supera {max(6.5, np.floor(avg-1)+0.5):.1f} juegos", min(0.86, max(0.40, 0.58 + (avg-9.5)*0.055)), f"Media esperada {avg:.1f}")
            add("Juegos jugador", f"{player} supera {max(7.5, np.floor(avg)+0.5):.1f} juegos", min(0.80, max(0.30, 0.50 + (avg-10.0)*0.050)), f"Media esperada {avg:.1f}")

    # Si hay TA afinado, añadir mercados TA como bloque aparte.
    if isinstance(ta_adj, dict):
        add("TA ajustado", f"{p1} gana partido · TA", ta_adj.get("p1_adj"), "Capa TennisAbstract")
        add("TA ajustado", f"{p2} gana partido · TA", ta_adj.get("p2_adj"), "Capa TennisAbstract")
        add("TA ajustado", "Over 18.5 · TA", ta_adj.get("over18_adj"), "Capa TennisAbstract")
        add("TA ajustado", "Over 19.5 · TA", ta_adj.get("over19_adj"), "Capa TennisAbstract")
        add("TA ajustado", "Over 20.5 · TA", ta_adj.get("over20_adj"), "Capa TennisAbstract")
        if ta_adj.get("set1_signal") is not None:
            add("TA ajustado", f"{p1} gana al menos 1 set · TA", ta_adj.get("set1_signal"), "Capa TennisAbstract")
        if ta_adj.get("set2_signal") is not None:
            add("TA ajustado", f"{p2} gana al menos 1 set · TA", ta_adj.get("set2_signal"), "Capa TennisAbstract")

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Probabilidad"], ascending=False).reset_index(drop=True)
        df["Probabilidad"] = df["Probabilidad"].map(lambda x: f"{x:.1%}")
    return df




def _di_float(v, default=0.0):
    try:
        if v in [None, "", "nan"]:
            return default
        return float(v)
    except Exception:
        return default


def _di_pct(v):
    try:
        return f"{float(v):.1%}"
    except Exception:
        return "N/A"


def _deep_is_challenger_clay(payload):
    """Predictor profundo: aplicar inteligencia específica Challenger Clay.
    En la app a veces Challenger se calcula como ATP; por eso aceptamos ATP/CHALLENGER
    siempre que la superficie sea Clay y el formato sea BO3.
    """
    try:
        surface = str((payload or {}).get("surface", "")).strip().lower()
        circuito = str((payload or {}).get("circuito", "")).strip().upper()
        best_of = int((payload or {}).get("best_of", 3) or 3)
        return surface == "clay" and best_of == 3 and circuito in {"ATP", "CHALLENGER"}
    except Exception:
        return False


def _deep_ta_raw(ta_adj):
    if isinstance(ta_adj, dict):
        raw = ta_adj.get("raw_signals", {})
        return raw if isinstance(raw, dict) else {}
    return {}


def _deep_intelligence_report(payload, ta_adj=None):
    """v23.46 · Capa de razonamiento del Predictor.
    No cambia el motor. Interpreta conjuntamente Monte Carlo + TA + superficie.
    Basado en el backtest Challenger Clay:
      - Model3Sets apenas predice Real3Sets.
      - Over18 mejora cuando ModelOver18 alto + HoldAvg decente + favorito no ultra dominante.
      - easy_rate alto debe frenar Overs altos y lecturas de set extra.
      - dog gana set solo es fiable en rangos altos.
    """
    payload = payload or {}
    base = payload.get("base", {}) if isinstance(payload.get("base", {}), dict) else {}
    deep = payload.get("deep", {}) if isinstance(payload.get("deep", {}), dict) else {}
    raw = _deep_ta_raw(ta_adj)
    p1 = payload.get("p1_name", "Jugador 1")
    p2 = payload.get("p2_name", "Jugador 2")
    fav = deep.get("model_fav_name", "Favorito")
    dog = deep.get("model_dog_name", "Underdog")

    over18 = _di_float(base.get("over18"), 0.0)
    over19 = _di_float(base.get("over19"), 0.0)
    over20 = _di_float(base.get("over20"), 0.0)
    set3 = _di_float(base.get("set3"), 0.0)
    dog_set = _di_float(base.get("dog_wins_set"), 0.0)
    fav20 = _di_float(base.get("fav_2_0"), 0.0)
    mlmax = max(_di_float(base.get("p1"), 0.5), _di_float(base.get("p2"), 0.5))
    avg_games = _di_float(deep.get("avg_games"), 0.0)
    med_games = _di_float(deep.get("med_games"), 0.0)
    hold1 = _di_float(deep.get("hold1"), 0.0)
    hold2 = _di_float(deep.get("hold2"), 0.0)
    ret1 = _di_float(deep.get("ret1"), 0.0)
    ret2 = _di_float(deep.get("ret2"), 0.0)
    hold_avg = (hold1 + hold2) / 2 if hold1 and hold2 else 0.0
    hold_diff = abs(hold1 - hold2) if hold1 and hold2 else 0.0
    ret_avg = (ret1 + ret2) / 2 if ret1 and ret2 else 0.0
    elo1 = _di_float(deep.get("p1_elo_surface"), 0.0)
    elo2 = _di_float(deep.get("p2_elo_surface"), 0.0)
    elo_diff = abs(elo1 - elo2) if elo1 and elo2 else 0.0
    rank1 = _di_float(deep.get("p1_rank"), 999)
    rank2 = _di_float(deep.get("p2_rank"), 999)
    rank_gap = abs(rank1 - rank2) if rank1 < 900 and rank2 < 900 else 0.0

    ta_over = _di_float(raw.get("over_rate"), 0.0)
    ta_over19 = _di_float(raw.get("over19_rate"), 0.0)
    ta_three = _di_float(raw.get("three_rate"), 0.0)
    ta_easy = _di_float(raw.get("easy_rate"), 0.0)
    ta_avg_games = _di_float(raw.get("avg_games"), 0.0)
    ta_set1 = _di_float(raw.get("set1"), 0.0)
    ta_set2 = _di_float(raw.get("set2"), 0.0)
    ta_lost1 = _di_float(raw.get("lost1"), 0.0)
    ta_lost2 = _di_float(raw.get("lost2"), 0.0)

    surface = str(payload.get("surface", "")).strip()
    rows = []
    notes = []
    contradictions = []

    def add(area, estado, lectura, motivo):
        rows.append({"Área": area, "Estado": estado, "Lectura": lectura, "Motivo": motivo})

    is_ch_clay = _deep_is_challenger_clay(payload)

    # Perfil base de datos utilizados.
    add("Base datos", "Info", "Datos usados", f"Superficie {surface}; HoldAvg {hold_avg:.1%}; HoldDiff {hold_diff:.1%}; ML favorito {mlmax:.1%}; media {avg_games:.1f}; TA easy {ta_easy:.0%}; TA 3 sets {ta_three:.0%}; TA media {ta_avg_games:.1f}")

    # Contradicciones fuertes.
    if over18 >= 0.76 and ta_easy >= 0.50:
        contradictions.append("Over alto del motor pero TA muestra mucho partido corto/paliza.")
    if fav20 >= 0.60 and dog_set >= 0.65:
        contradictions.append("Favorito 2-0 alto pero el dog gana set también sale alto.")
    if set3 >= 0.45 and ta_easy >= 0.50:
        contradictions.append("Set extra alto pero TA muestra easy_rate alto.")
    if mlmax >= 0.80 and max(ta_set1, ta_set2) >= 0.75:
        contradictions.append("Favorito muy claro pero TA sugiere resistencia alta de al menos un jugador.")

    # Over 18.5 / 19.5.
    if is_ch_clay:
        if over18 >= 0.78 and hold_avg >= 0.60 and mlmax < 0.90 and ta_easy < 0.50:
            add("Over 18.5", "✅ Sólido", "Over bajo fuerte para Challenger Clay", f"Backtest: ModelOver18 >=78% y HoldAvg >=60% funcionan mejor; favorito no ultra dominante ({mlmax:.0%}) y easy_rate controlado ({ta_easy:.0%}).")
        elif ta_easy >= 0.55 or mlmax >= 0.90 or hold_avg < 0.60 or hold_diff >= 0.18:
            add("Over 18.5", "⚠️ Cautela", "No tratar como Over fuerte", f"Riesgo corto: easy_rate {ta_easy:.0%}, ML fav {mlmax:.0%}, HoldAvg {hold_avg:.0%}, HoldDiff {hold_diff:.0%}. En Clay esto debe frenar la conclusión.")
        elif over18 >= 0.74:
            add("Over 18.5", "🟡 Apto", "Over bajo posible, no subir línea sin más", f"Over18 {over18:.0%}; no hay confirmación suficiente de línea superior.")
        else:
            add("Over 18.5", "❌ Débil", "Over bajo no tiene señal clara", f"Over18 {over18:.0%}; por debajo del corte útil Challenger Clay.")

        if over19 >= 0.68 and ta_easy < 0.45 and hold_avg >= 0.62 and mlmax < 0.85:
            add("Over 19.5/20.5", "✅ Revisable", "Línea superior con lógica deportiva", f"Over19 {over19:.0%}, HoldAvg {hold_avg:.0%}, easy_rate {ta_easy:.0%}.")
        elif over20 >= 0.60 or over19 >= 0.65:
            add("Over 19.5/20.5", "⚠️ Fina", "Cuidado con subir línea", f"Over19 {over19:.0%}, Over20 {over20:.0%}; si easy_rate/ML/hold no acompañan, no convertirlo en conclusión fuerte.")
        else:
            add("Over 19.5/20.5", "❌ No prioritaria", "No hay fuerza para línea alta", f"Over19 {over19:.0%}, Over20 {over20:.0%}.")
    else:
        # Fallback general; no toca reglas específicas ATP/WTA.
        if over18 >= 0.78 and ta_easy < 0.50:
            add("Over 18.5", "✅ Bueno", "Over bajo coherente", f"Over18 {over18:.0%}, easy_rate {ta_easy:.0%}.")
        elif ta_easy >= 0.50:
            add("Over 18.5", "⚠️ Cautela", "TA avisa de partido corto", f"easy_rate {ta_easy:.0%}; no convertir Over en lectura fuerte sin más.")
        else:
            add("Over 18.5", "🟡 Neutral", "Lectura no concluyente", f"Over18 {over18:.0%}.")

    # 3 sets.
    if is_ch_clay:
        if set3 >= 0.50 and dog_set >= 0.65 and ta_easy < 0.40:
            add("3 sets", "🟡 Posible", "Set extra posible, no seguro", f"En el backtest Model3Sets no predice bien por sí solo; aquí solo se acepta por set3 {set3:.0%} + dog_set {dog_set:.0%} + easy bajo {ta_easy:.0%}.")
        else:
            add("3 sets", "⚠️ No fiable", "No usar como conclusión fuerte", f"Model3Sets {set3:.0%}; backtest Challenger Clay casi nulo. Solo considerar si otras señales lo confirman.")
    else:
        if set3 >= 0.50 and ta_easy < 0.45:
            add("3 sets", "🟡 Posible", "Set extra posible", f"Set3 {set3:.0%}, easy_rate {ta_easy:.0%}.")
        else:
            add("3 sets", "⚠️ Débil", "No convertir partido largo en 3 sets", f"Set3 {set3:.0%}, easy_rate {ta_easy:.0%}.")

    # Dog gana set / favorito 2-0.
    if dog_set >= 0.70 and ta_easy < 0.55:
        add("Dog gana set", "✅ Fuerte", f"{dog} puede ganar set", f"ModelDogWinsSet {dog_set:.0%}; en backtest solo es útil en rango alto.")
    elif dog_set >= 0.55:
        add("Dog gana set", "🟡 Media", "Resistencia posible, no concluyente", f"DogWinsSet {dog_set:.0%}; zona media históricamente ruidosa.")
    else:
        add("Dog gana set", "⚠️ Baja", "No apoyar mercado de set del dog", f"DogWinsSet {dog_set:.0%}.")

    if fav20 >= 0.70 and dog_set < 0.45 and (ta_easy >= 0.45 or hold_diff >= 0.16 or mlmax >= 0.82):
        add("Favorito 2-0", "✅ Coherente", f"{fav} 2-0 encaja", f"Fav2-0 {fav20:.0%}, dog_set {dog_set:.0%}, easy {ta_easy:.0%}, HoldDiff {hold_diff:.0%}.")
    elif fav20 >= 0.60 and dog_set >= 0.55:
        add("Favorito 2-0", "⚠️ Contradictorio", "No vender como dominio claro", f"Fav2-0 {fav20:.0%} pero dog_set {dog_set:.0%}. Esto debe frenar el guion dominante.")
    else:
        add("Favorito 2-0", "🟡 No central", "No es la lectura principal", f"Fav2-0 {fav20:.0%}.")

    # Historia final.
    if contradictions:
        story = "⚠️ SEÑALES CONTRADICTORIAS"
        story_detail = " | ".join(contradictions[:4])
    elif ta_easy >= 0.55 and (over18 >= 0.74 or set3 >= 0.40):
        story = "⚠️ RIESGO DE PARTIDO CORTO"
        story_detail = "TA easy_rate alto debe mandar sobre lecturas de Over/3 sets."
    elif is_ch_clay and over18 >= 0.78 and hold_avg >= 0.60 and ta_easy < 0.45 and mlmax < 0.85:
        story = "✅ CHALLENGER CLAY · OVER BAJO COHERENTE"
        story_detail = "Over18 alto, HoldAvg útil y sin favorito ultra dominante ni easy_rate alto."
    elif mlmax >= 0.80 and dog_set < 0.45:
        story = "✅ FAVORITO CON CONTROL"
        story_detail = "ML alto y dog_set bajo; revisar 2-0/Under antes que partido largo."
    elif dog_set >= 0.70:
        story = "✅ FAVORITO/DOG COMPETIDO"
        story_detail = "El dog tiene señal alta de ganar set; no vender 2-0 cómodo."
    elif over18 >= 0.74:
        story = "🟡 PARTIDO CON JUEGOS, NO NECESARIAMENTE 3 SETS"
        story_detail = "Over bajo posible; no convertir automáticamente en set extra."
    else:
        story = "🟡 SIN HISTORIA FUERTE"
        story_detail = "Mejor lectura prudente; no forzar mercados derivados."

    # Notas resumen para conclusiones.
    notes.append(f"🎭 Historia final: {story}. {story_detail}")
    if contradictions:
        for c in contradictions[:3]:
            notes.append(f"🔴 Contradicción: {c}")
    if is_ch_clay:
        notes.append("📌 Regla Challenger Clay activa: prioriza HoldAvg + Over18 alto + easy_rate bajo; rebaja 3 sets como señal fuerte.")

    return {"story": story, "story_detail": story_detail, "rows": rows, "notes": notes, "contradictions": contradictions}

def _deep_conclusions(payload, ta_adj=None):
    base = payload.get("base", {}) if isinstance(payload, dict) else {}
    deep = payload.get("deep", {}) if isinstance(payload, dict) else {}
    p1 = payload.get("p1_name", "Jugador 1")
    p2 = payload.get("p2_name", "Jugador 2")
    avg_games = float(deep.get("avg_games", 0) or 0)
    med_games = float(deep.get("med_games", 0) or 0)
    over18 = float(base.get("over18", 0) or 0)
    over19 = float(base.get("over19", 0) or 0)
    over20 = float(base.get("over20", 0) or 0)
    set3 = float(base.get("set3", 0) or 0)
    tb = float(base.get("tb", 0) or 0)
    dog_set = float(base.get("dog_wins_set", 0) or 0)
    fav20 = float(base.get("fav_2_0", 0) or 0)
    mlmax = max(float(base.get("p1", 0.5) or 0.5), float(base.get("p2", 0.5) or 0.5))
    hold1 = float(deep.get("hold1", 0) or 0)
    hold2 = float(deep.get("hold2", 0) or 0)
    game_diff = float(deep.get("game_diff", 0) or 0)
    notes = []

    # v23.46: conclusiones basadas primero en la capa de razonamiento conjunto.
    try:
        intel = _deep_intelligence_report(payload, ta_adj if isinstance(ta_adj, dict) else None)
        notes.extend(intel.get("notes", []))
    except Exception:
        intel = None

    # Conclusiones antiguas quedan como apoyo, pero ahora son más prudentes.
    if over18 >= 0.78 and over19 >= 0.68 and not (isinstance(intel, dict) and intel.get("contradictions")):
        notes.append("✅ Perfil claro de partido con juegos: Over 18.5 es base, y Over 19.5 merece revisión deportiva.")
    elif over18 >= 0.74 and over19 < 0.62:
        notes.append("🟡 Over bajo interesante, pero no hay fuerza suficiente para subir línea sin confirmación TA.")
    if over20 >= 0.62 and avg_games >= 22.3 and not (isinstance(intel, dict) and intel.get("contradictions")):
        notes.append("🟡 Línea superior de juegos revisable, pero solo si la historia del partido no muestra contradicción ni easy_rate alto.")
    if set3 >= 0.45:
        notes.append("⚠️ Set extra posible, pero v23.46 no lo trata como señal fuerte sin apoyo de dog_set/easy_rate/superficie.")
    elif fav20 >= 0.62 and mlmax >= 0.66:
        notes.append("✅ Guion de favorito dominante: vigilar 2-0 favorito y evitar forzar Overs altos si la media de juegos no acompaña.")
    if dog_set >= 0.74:
        notes.append("✅ El underdog compite: mercado 'gana al menos un set' o handicap de juegos puede tener sentido deportivo.")
    if tb >= 0.34 and min(hold1, hold2) >= 0.72:
        notes.append("🎯 Entorno de tie-break: ambos holds sostienen sets largos.")
    if avg_games:
        notes.append(f"📊 Total esperado: media {avg_games:.1f} juegos, mediana {med_games:.1f}; diferencia media J1-J2 {game_diff:+.1f} juegos.")
    if isinstance(ta_adj, dict) and ta_adj.get("notes"):
        notes.append("🧪 TA aporta lectura extra: " + " | ".join([str(x) for x in ta_adj.get("notes", [])[:3]]))
    if not notes:
        notes.append("Sin conclusión fuerte: partido mejor para observación o stake mínimo si no aparece mercado muy claro.")
    return notes


def render_predictor_deep_match_analyzer_panel():
    payload = st.session_state.get("predictor_last_payload_v23440")
    if not isinstance(payload, dict):
        return
    ta_adj = st.session_state.get("predictor_last_ta_adjustment_v23441")
    p1 = payload.get("p1_name", "Jugador 1")
    p2 = payload.get("p2_name", "Jugador 2")
    base = payload.get("base", {}) if isinstance(payload.get("base", {}), dict) else {}
    deep = payload.get("deep", {}) if isinstance(payload.get("deep", {}), dict) else {}

    st.divider()
    with st.expander("🔬 Deep Match Analyzer · análisis profundo del partido", expanded=True):
        st.caption("Laboratorio del Predictor individual. Usa la simulación base, Elo/superficie, hold/return, mercados derivados y, si existe, ficha TA afinada. No modifica el Analista por lista ni el motor Over.")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Media juegos", f"{float(deep.get('avg_games',0) or 0):.1f}", f"mediana {float(deep.get('med_games',0) or 0):.1f}")
        with c2:
            st.metric(f"Juegos {p1}", f"{float(deep.get('avg_g1',0) or 0):.1f}", f"med {float(deep.get('med_g1',0) or 0):.1f}")
        with c3:
            st.metric(f"Juegos {p2}", f"{float(deep.get('avg_g2',0) or 0):.1f}", f"med {float(deep.get('med_g2',0) or 0):.1f}")
        with c4:
            st.metric("3 sets / set extra", _deep_pct(base.get("set3", 0)), _deep_level(base.get("set3", 0)))

        c5, c6, c7, c8 = st.columns(4)
        with c5: st.metric("Primer set J1", _deep_pct(base.get("p1_fs", 0)))
        with c6: st.metric("Primer set J2", _deep_pct(base.get("p2_fs", 0)))
        with c7: st.metric("Tie-break", _deep_pct(base.get("tb", 0)), _deep_level(base.get("tb", 0)))
        with c8: st.metric("Dog gana set", _deep_pct(base.get("dog_wins_set", 0)), _deep_level(base.get("dog_wins_set", 0)))

        st.subheader("📌 Conclusiones deportivas")
        for n in _deep_conclusions(payload, ta_adj if isinstance(ta_adj, dict) else None):
            st.write("• " + str(n))

        # v23.46: lectura conjunta basada en backtest Challenger Clay + TA.
        try:
            intel = _deep_intelligence_report(payload, ta_adj if isinstance(ta_adj, dict) else None)
            st.subheader("🧠 Challenger / TA Intelligence")
            st.info(f"{intel.get('story','')} · {intel.get('story_detail','')}")
            dfi = pd.DataFrame(intel.get("rows", []))
            if not dfi.empty:
                st.dataframe(dfi, use_container_width=True, hide_index=True)
        except Exception as e:
            st.caption(f"No se pudo pintar Intelligence: {type(e).__name__}: {e}")

        st.subheader("📊 Todos los mercados derivados")
        dfm = _deep_build_markets(payload, ta_adj if isinstance(ta_adj, dict) else None)
        if not dfm.empty:
            st.dataframe(dfm, use_container_width=True, hide_index=True)
        else:
            st.info("No hay mercados derivados suficientes para mostrar.")

        with st.expander("🧾 Datos internos usados en esta lectura", expanded=False):
            st.write({
                "surface": payload.get("surface"),
                "circuito": payload.get("circuito"),
                "best_of": payload.get("best_of"),
                "rank": {p1: deep.get("p1_rank"), p2: deep.get("p2_rank")},
                "elo_surface": {p1: deep.get("p1_elo_surface"), p2: deep.get("p2_elo_surface")},
                "hold": {p1: deep.get("hold1"), p2: deep.get("hold2")},
                "return": {p1: deep.get("ret1"), p2: deep.get("ret2")},
                "volatilidad": deep.get("vol"),
                "notas_caps": deep.get("market_cap_notes", []),
            })


# =========================================================
# UI
# =========================================================

with st.sidebar:
    st.header("🎾 Tennis IA")
    st.caption(APP_VERSION)
    st.caption("ATP + Challenger ELO/Stats Engine · SofaScore workflow")
    st.caption("🔒 Over blindado")
    if st.button("🧹 Limpiar caché"):
        st.cache_data.clear()
        st.success("Caché limpiada")
    circuito = st.radio("Circuito", ["ATP", "Challenger", "WTA"])
    modo = st.radio("Modo", ["Panel de control", "Predictor", "Analizador por lista", "Validador histórico", "Analyzer"])
    mostrar_debug = st.toggle("🔧 Mostrar diagnóstico técnico", value=False)

# =========================================================
# UX Progress Engine: carga visible de datos
# =========================================================

load_status = st.status("🔄 Preparando datos del modelo...", expanded=False)
with load_status:
    st.caption("Fase 1/3 · Leyendo jugadores, Elo, stats y fatiga. Si es la primera carga puede tardar.")
    load_bar = st.progress(5, text="Fase 1/3 · Cargando base de jugadores...")
    db = cargar_datos_app(circuito)

    load_bar.progress(65, text="Fase 2/3 · Cargando Match Count / QualityMap...")
    st.caption("Fase 2/3 · Integrando ATP + Challenger/Qualy si existe datos/challenger.")
    try:
        qm_preview = crear_quality_map(circuito)
        q_players = qm_preview.get("_meta", {}).get("raw_player_count", "?") if isinstance(qm_preview, dict) else "?"
    except Exception:
        q_players = "?"

    load_bar.progress(100, text=f"Fase 3/3 · Datos listos · jugadores={len(db)} · quality players={q_players}")
    st.caption("✅ Datos preparados. Las siguientes cargas deberían ir más rápido por caché.")

load_status.update(label=f"✅ Datos listos · {len(db)} jugadores cargados", state="complete", expanded=False)

# v23.46.5: Challenger visible como circuito independiente.
# Para validar/simular se usa motor ATP, pero leyendo datos/históricos de Challenger.
circuito_sim_bt = "ATP" if str(circuito).upper().strip() == "CHALLENGER" else circuito


if not db:
    st.error("No se encontraron jugadores. Revisa carpetas y archivos.")
    st.stop()

# v23.25: evita que Streamlit conserve tablas/descargas antiguas al cambiar de versión.
if st.session_state.get("batch_app_version") != APP_VERSION:
    for _k in ["batch_ok_df", "batch_ko_df", "batch_last_ready"]:
        st.session_state.pop(_k, None)
    st.session_state["batch_app_version"] = APP_VERSION

if modo == "Panel de control":
    render_control_panel(db, circuito)

elif modo == "Predictor":
    with st.sidebar:
        players = sorted(db.keys())
        surface = st.selectbox("Superficie", ["Hard","Clay","Grass"])
        formato = st.radio("Formato", ["ATP Tour (3 sets)", "Grand Slam (5 sets)"])
        modo_velocidad = st.radio(
            "Velocidad",
            ["⚡ Rápido", "⚖️ Equilibrado", "🎯 Preciso"],
            index=1,
            help="Rápido usa menos simulaciones y responde antes. Preciso es más lento."
        )
        sims = st.select_slider(
            "Simulaciones",
            [1000, 2500, 5000, 10000, 20000],
            value=5000
        )

        if modo_velocidad.startswith("⚡"):
            sims = min(sims, 2500)
            st.caption("⚡ Modo rápido: máximo 2.500 sims para acelerar.")
        elif modo_velocidad.startswith("⚖️"):
            sims = min(sims, 5000)
            st.caption("⚖️ Modo equilibrado: máximo 5.000 sims.")
        else:
            st.caption("🎯 Modo preciso: usa el valor seleccionado.")

    c1, c2 = st.columns(2)
    with c1: p1_name = st.selectbox("Jugador 1", players)
    with c2: p2_name = st.selectbox("Jugador 2", players, index=min(1, len(players)-1))

    if st.button("🚀 ANALIZAR PARTIDO", width='stretch'):
        d1, d2 = db[p1_name], db[p2_name]
        best_of = 5 if "5" in formato else 3
        if sims >= 10000:
            st.warning("🎯 Modo preciso: puede tardar bastante en Streamlit Cloud. Para pruebas rápidas usa 2.500 o 5.000 sims.")
        sim_status = st.status(f"🎲 Preparando simulación Monte Carlo...", expanded=True)
        with sim_status:
            sim_bar = st.progress(1, text="Fase 1/2 · Preparando engines...")
            sim_msg = st.empty()
            sim_start = time.time()
            sim_msg.caption("Calculando hold/return, fatiga, rating sanity, clay engines y contexto...")

            last_ui_update = {"t": 0.0}

            def update_sim_progress(done, total):
                elapsed = max(0.001, time.time() - sim_start)

                if done <= 0:
                    sim_bar.progress(3, text=f"Fase 1/2 · Preparando engines · 0 / {total:,}")
                    sim_msg.caption("Aún no ha empezado el bucle Monte Carlo; esto es normal.")
                    return

                # Limitar repintados: suficiente para parecer vivo, sin ralentizar demasiado.
                now = time.time()
                if done < total and (now - last_ui_update["t"]) < 0.20:
                    return
                last_ui_update["t"] = now

                pct_float = min(1.0, max(0.03, done / total)) if total else 1.0
                pct_int = int(round(pct_float * 100))
                rate = done / elapsed if done else 0
                eta = (total - done) / rate if rate > 0 else 0

                sim_bar.progress(
                    pct_int,
                    text=f"Fase 2/2 · Monte Carlo {done:,} / {total:,} ({pct_int}%) · {elapsed:.1f}s · ETA {eta:.1f}s"
                )
                sim_msg.caption(f"Simulando partidos... último lote procesado: {done:,}")

            # Primer repintado antes de entrar al cálculo pesado.
            update_sim_progress(0, sims)

            sim = sim_match(
                d1, d2, surface, circuito, best_of, sims,
                context_row={},
                progress_callback=update_sim_progress
            )
            total_elapsed = time.time() - sim_start
            sim_bar.progress(100, text=f"Simulación completada · {sims:,} partidos · {total_elapsed:.1f}s")
            sim_msg.caption("✅ Resultados listos para pintar en pantalla.")

        sim_status.update(label=f"✅ Simulación completada · {sims:,} partidos", state="complete", expanded=False)

        st.caption(
            f"📁 Stats usadas: {stats_filename_label(circuito, 'serve', surface)} · "
            f"{stats_filename_label(circuito, 'return', surface)} · "
            f"{stats_filename_label(circuito, 'break', surface)}"
        )

        p1, p2, p1c, p2c = sim["p1"], sim["p2"], sim["p1_cal"], sim["p2_cal"]
        games = sim["games"]
        avg_games, med_games = np.mean(games), np.median(games)
        games_p1 = sim.get("games_p1", [])
        games_p2 = sim.get("games_p2", [])
        avg_g1 = float(np.mean(games_p1)) if len(games_p1) else 0.0
        avg_g2 = float(np.mean(games_p2)) if len(games_p2) else 0.0
        med_g1 = float(np.median(games_p1)) if len(games_p1) else 0.0
        med_g2 = float(np.median(games_p2)) if len(games_p2) else 0.0
        game_diff = avg_g1 - avg_g2
        over17_raw = sum(x > 17.5 for x in games)/sims
        over18_raw = sum(x > 18.5 for x in games)/sims
        over19_raw = sum(x > 19.5 for x in games)/sims
        over20_raw = sum(x > 20.5 for x in games)/sims
        over22_raw = sum(x > 22.5 for x in games)/sims

        market_caps = aplicar_market_sanity_caps(sim, circuito, surface, over18_raw, over19_raw, over20_raw, over22_raw)
        over18 = market_caps["over18"]
        over19 = market_caps["over19"]
        over20 = market_caps["over20"]
        over22 = market_caps["over22"]
        under22 = 1-over22
        over17 = over17_raw if circuito == "WTA" else 0.0

        sim["market_over17"] = over17
        sim["market_over18"] = over18
        sim["market_over19"] = over19
        sim["market_over20"] = over20
        sim["market_over22"] = over22
        sim["market_cap_notes"] = market_caps.get("notes", [])

        # v23.44.0: guardar último Predictor para poder afinar con fichas TA sin repetir simulación.
        st.session_state["predictor_last_payload_v23440"] = {
            "p1_name": d1.get("Player", p1_name),
            "p2_name": d2.get("Player", p2_name),
            "surface": surface,
            "circuito": circuito,
            "best_of": best_of,
            "base": {
                "p1": float(sim.get("p1_cal", sim.get("p1", 0.5))),
                "p2": float(sim.get("p2_cal", sim.get("p2", 0.5))),
                "over17": float(over17),
                "over18": float(over18),
                "over19": float(over19),
                "over20": float(over20),
                "over22": float(over22),
                "p1_fs": float(sim.get("p1_fs", 0.5)),
                "p2_fs": float(sim.get("p2_fs", 0.5)),
                "set3": float(sim.get("set3", 0.0)),
                "tb": float(sim.get("tb", 0.0)),
                "dog_wins_set": float(sim.get("dog_wins_set", 0.0)),
                "fav_2_0": float(sim.get("fav_2_0", 0.0)),
                "long_match": float(sim.get("long_match", 0.0)),
                "fav_under22": float(sim.get("fav_under22", 0.0)),
                "dog_over20": float(sim.get("dog_over20", 0.0)),
            },
            "deep": {
                "avg_games": float(avg_games),
                "med_games": float(med_games),
                "avg_g1": float(avg_g1),
                "avg_g2": float(avg_g2),
                "med_g1": float(med_g1),
                "med_g2": float(med_g2),
                "game_diff": float(game_diff),
                "hold1": float(sim.get("hold1", 0.0)),
                "hold2": float(sim.get("hold2", 0.0)),
                "ret1": float(sim.get("ret1", 0.0)),
                "ret2": float(sim.get("ret2", 0.0)),
                "vol": float(sim.get("vol", 0.0)),
                "model_fav_name": sim.get("model_fav_name", "Favorito"),
                "model_dog_name": sim.get("model_dog_name", "Underdog"),
                "p1_rank": int(d1.get("Rank", 999) or 999),
                "p2_rank": int(d2.get("Rank", 999) or 999),
                "p1_elo_surface": float(d1.get(surface, 1500) or 1500),
                "p2_elo_surface": float(d2.get(surface, 1500) or 1500),
                "p1_profile": sim.get("p1_profile", "normal"),
                "p2_profile": sim.get("p2_profile", "normal"),
                "rating_sanity": sim.get("rating_sanity", {}),
                "market_cap_notes": sim.get("market_cap_notes", []),
            }
        }

        elo_ref = elo_prob(d1[surface], d2[surface])

        risk = "🟢 Riesgo bajo"
        if max(p1c,p2c) < 0.56: risk = "🔴 Riesgo alto"
        elif max(p1c,p2c) < 0.63: risk = "🟡 Riesgo medio"

        st.divider()
        st.subheader("🏆 Ganador del Partido")
        a,b = st.columns(2)
        with a:
            st.metric(d1["Player"], f"{p1c:.1%}", f"{nivel(p1c)} · bruta {p1:.1%}")
            st.caption(f"Rank #{d1['Rank']} · Elo {surface}: {d1[surface]:.0f} · {edge_calibracion(p1,p1c)}")
        with b:
            st.metric(d2["Player"], f"{p2c:.1%}", f"{nivel(p2c)} · bruta {p2:.1%}")
            st.caption(f"Rank #{d2['Rank']} · Elo {surface}: {d2[surface]:.0f} · {edge_calibracion(p2,p2c)}")
        st.caption(f"Referencia Elo puro: {elo_ref:.1%} / {1-elo_ref:.1%} · {risk}")

        st.divider()
        st.subheader("🎾 Primer Set")
        fs1, fs2 = st.columns(2)
        with fs1: st.metric(f"{d1['Player']} gana", f"{sim['p1_fs']:.1%}", nivel(sim["p1_fs"]))
        with fs2: st.metric(f"{d2['Player']} gana", f"{sim['p2_fs']:.1%}", nivel(sim["p2_fs"]))

        st.divider()
        st.subheader("📊 Mercados principales")
        if circuito == "WTA":
            m0,m1,m2,m3,m4,m5 = st.columns(6)
            with m0: st.metric("Over 17.5 WTA", f"{over17:.1%}", nivel(over17))
        else:
            m1,m2,m3,m4,m5 = st.columns(5)
        with m1: st.metric("Over 18.5", f"{over18:.1%}", nivel(over18))
        with m2: st.metric("Over 19.5", f"{over19:.1%}", nivel(over19))
        with m3: st.metric("Over 20.5", f"{over20:.1%}", nivel(over20))
        with m4: st.metric("Over 22.5", f"{over22:.1%}", nivel(over22))
        with m5: st.metric("Under 22.5", f"{under22:.1%}", nivel(under22))

        if best_of == 5 and circuito == "ATP":
            gs5 = calcular_grand_slam_5set_markets(sim, d1["Player"], d2["Player"])
            st.warning("🏆 Grand Slam 5 sets activo: ignora Over 18.5/19.5 como mercado oficial. Usa las líneas GS de abajo.")
            st.subheader("🏆 Mercados Grand Slam 5 sets")
            gm = gs5.get("markets", {})
            cgs1, cgs2, cgs3, cgs4, cgs5 = st.columns(5)
            with cgs1: st.metric("Over 30.5 GS", _fmt_pct_gs(gm, "Over 30.5 GS"), nivel(gm.get("Over 30.5 GS", 0)))
            with cgs2: st.metric("Over 32.5 GS", _fmt_pct_gs(gm, "Over 32.5 GS"), nivel(gm.get("Over 32.5 GS", 0)))
            with cgs3: st.metric("Over 35.5 GS", _fmt_pct_gs(gm, "Over 35.5 GS"), nivel(gm.get("Over 35.5 GS", 0)))
            with cgs4: st.metric("Over 37.5 GS", _fmt_pct_gs(gm, "Over 37.5 GS"), nivel(gm.get("Over 37.5 GS", 0)))
            with cgs5: st.metric("Over 39.5 GS", _fmt_pct_gs(gm, "Over 39.5 GS"), nivel(gm.get("Over 39.5 GS", 0)))
            sg1, sg2, sg3, sg4 = st.columns(4)
            with sg1: st.metric("Partido 4+ sets", f"{gs5.get('four_plus',0):.1%}", nivel(gs5.get("four_plus",0)))
            with sg2: st.metric("Partido 5 sets", f"{gs5.get('set5',0):.1%}", nivel(gs5.get("set5",0)))
            with sg3: st.metric("Favorito NO 3-0", f"{gs5.get('fav_no_3_0',0):.1%}", nivel(gs5.get("fav_no_3_0",0)))
            with sg4: st.metric("Underdog gana set", f"{gs5.get('dog_set',0):.1%}", nivel(gs5.get("dog_set",0)))
            if gs5.get("best_label") and gs5.get("best_label") != "Sin mercado GS claro":
                st.success(f"{gs5.get('action')} · {gs5.get('best_label')} → {gs5.get('best_prob',0):.1%}")
            else:
                st.info("Sin mercado Grand Slam 5 sets claro con los umbrales actuales.")

        model_dog_name = sim.get("model_dog_name", "Underdog")
        e1,e2,e3 = st.columns(3)
        with e1: st.metric("3 sets", f"{sim['set3']:.1%}", nivel(sim["set3"]))
        with e2: st.metric("Tie-break", f"{sim['tb']:.1%}", nivel(sim["tb"]))
        with e3: st.metric(f"{model_dog_name} gana al menos 1 set", f"{sim['dog_wins_set']:.1%}", nivel(sim["dog_wins_set"]))

        st.subheader("🎮 Juegos esperados")
        gcol1, gcol2, gcol3, gcol4 = st.columns(4)
        with gcol1:
            st.metric(d1["Player"], f"{avg_g1:.1f}", f"Mediana {med_g1:.0f}")
        with gcol2:
            st.metric(d2["Player"], f"{avg_g2:.1f}", f"Mediana {med_g2:.0f}")
        with gcol3:
            st.metric("Total esperado", f"{avg_games:.1f}", f"Mediana {med_games:.0f}")
        with gcol4:
            diff_label = d1["Player"] if game_diff >= 0 else d2["Player"]
            st.metric("Diferencia juegos", f"{abs(game_diff):.1f}", f"Ventaja {diff_label}")

        st.caption(
            f"Proyección: {d1['Player']} {avg_g1:.1f} - {avg_g2:.1f} {d2['Player']} · "
            f"Total medio {avg_games:.1f}"
        )
        if sim.get("market_cap_notes"):
            st.caption("🧯 Market sanity: " + " · ".join(sim.get("market_cap_notes", [])))

        with st.expander("🔧 Ver diagnóstico técnico", expanded=mostrar_debug):
            st.divider()
            st.subheader("📋 Mercados derivados")
            dm1, dm2, dm3 = st.columns(3)
            with dm1:
                st.metric(f"{sim.get('model_fav_name','Favorito')} 2-0", f"{sim['fav_2_0']:.1%}", nivel(sim["fav_2_0"]))
            with dm2:
                st.metric("Partido largo", f"{sim['long_match']:.1%}", nivel(sim["long_match"]))
            with dm3:
                st.metric("Fav + Under 22.5", f"{sim['fav_under22']:.1%}", nivel(sim["fav_under22"]))

            st.divider()
            st.subheader("🎾 Hold / Return Engine")
            h1,h2 = st.columns(2)
            with h1:
                st.metric(d1["Player"], f"{sim['hold1']:.1%}", f"Raw hold {sim['raw_hold1']:.1%}")
                st.caption(f"{perfil_legible(sim['p1_profile'])} · Return strength {sim['ret1']:.1%}")
            with h2:
                st.metric(d2["Player"], f"{sim['hold2']:.1%}", f"Raw hold {sim['raw_hold2']:.1%}")
                st.caption(f"{perfil_legible(sim['p2_profile'])} · Return strength {sim['ret2']:.1%}")



            if circuito == "ATP" and surface == "Clay" and sim.get("elite_clay", {}).get("active", False):
                st.divider()
                st.subheader("👑 Elite ATP Clay Protection")

                ec = sim.get("elite_clay", {})
                e1,e2,e3 = st.columns(3)

                with e1:
                    st.metric("Activo", "Sí")

                with e2:
                    st.metric("Favorito elite", ec.get("fav","").upper())

                with e3:
                    st.metric("Vol mult", f"{ec.get('vol_mult',1.0):.2f}")

            if circuito == "ATP" and surface == "Clay":
                st.divider()
                st.subheader("🧱 Rating Sanity Engine")

                ce = sim.get("clay_engine", {})
                c1,c2,c3 = st.columns(3)

                with c1:
                    st.metric("Activo", "Sí" if ce.get("active", False) else "No")

                with c2:
                    st.metric("Perfil", ce.get("profile", "neutral"))

                with c3:
                    st.metric("Vol mult", f"{ce.get('vol_mult',1.0):.2f}")

                if ce.get("notes"):
                    st.caption(" · ".join(ce.get("notes", [])))

            st.divider()
            st.subheader("🎯 Tie-break Intelligence")

            tb1, tb2, tb3 = st.columns(3)
            with tb1:
                st.metric("TB boost", f"{sim.get('tb_intel_boost', 0):+.1%}")
            with tb2:
                st.metric(f"Presión {d1['Player']}", f"{sim.get('pressure_skill1', 0):+.1%}")
            with tb3:
                st.metric(f"Presión {d2['Player']}", f"{sim.get('pressure_skill2', 0):+.1%}")


            st.divider()
            st.subheader("🏟️ Tournament Engine")

            tc = sim.get("tournament_ctx", {})
            t1,t2,t3 = st.columns(3)

            with t1:
                st.metric("Series", tc.get("series","ATP"))

            with t2:
                st.metric("Court", tc.get("court","Outdoor"))

            with t3:
                st.metric("Round", tc.get("round","Main"))

            st.divider()
            st.subheader("🔋 Fatigue Engine")

            fcol1, fcol2 = st.columns(2)
            with fcol1:
                f = sim.get("fatigue1", {})
                st.metric(d1["Player"], f"{f.get('fatigue_score',0):.1%}", f"Ajuste {sim.get('fatigue_adj1',0):+.1%}")
                st.caption(f"7 días: {f.get('matches7',0)} partidos · {f.get('games7',0)} games · descanso {f.get('rest_days',7)} días")
            with fcol2:
                f = sim.get("fatigue2", {})
                st.metric(d2["Player"], f"{f.get('fatigue_score',0):.1%}", f"Ajuste {sim.get('fatigue_adj2',0):+.1%}")
                st.caption(f"7 días: {f.get('matches7',0)} partidos · {f.get('games7',0)} games · descanso {f.get('rest_days',7)} días")


            st.divider()
            st.subheader("🎾 Rating Sanity Engine")

            sd = sim.get("set_dynamics", {})

            s1,s2,s3 = st.columns(3)

            with s1:
                st.metric("2-0 boost", f"{sd.get('fav20_boost',0):+.1%}")

            with s2:
                st.metric("Dog suppression", f"{sd.get('dog_set_suppress',0):+.1%}")

            with s3:
                st.metric("Long match adj", f"{sd.get('long_match_adj',0):+.1%}")

            ssg = sim.get("straight_sets_guard", {}) or {}
            if ssg.get("active", False):
                st.info(
                    f"🧭 Straight Sets Guard activo · 2-0 {ssg.get('fav20_boost',0):+.1%} · "
                    f"Dog set {ssg.get('dogset_cut',0):-.1%} · Largo {ssg.get('long_cut',0):-.1%}"
                )
                notes = ssg.get("notes", [])
                if notes:
                    st.caption(" · ".join(notes))



            if sim.get("wta_engine_active", False):
                st.divider()
                st.subheader("🎭 WTA Match Script")

                ws2 = sim.get("wta_script", {})
                m1,m2,m3,m4 = st.columns(4)

                with m1:
                    st.metric("Activo", "Sí" if ws2.get("active", False) else "No")

                with m2:
                    st.metric("Script", ws2.get("script","neutral"))

                with m3:
                    st.metric("Fav2-0", f"{ws2.get('fav20_mult',1.0):.2f}x")

                with m4:
                    st.metric("Long", f"{ws2.get('long_mult',1.0):.2f}x")

            if sim.get("wta_engine_active", False):
                st.divider()
                st.subheader("🎾 Elite WTA Separation")

                ws = sim.get("wta_separation", {})
                w1,w2,w3 = st.columns(3)

                with w1:
                    st.metric("Activo", "Sí" if ws.get("active", False) else "No")

                with w2:
                    st.metric("Perfil", ws.get("edge_label", "normal"))

                with w3:
                    st.metric("Vol mult", f"{ws.get('vol_mult',1.0):.2f}")


            st.divider()
            st.subheader("🧠 Rating Sanity Engine")

            rs = sim.get("rating_sanity", {})
            r1c, r2c, r3c = st.columns(3)

            with r1c:
                p = rs.get("p1", {})
                st.metric(d1["Player"], f"{p.get('confidence',1):.0%}", delta=f"{p.get('matches_surface',0)} partidos {surface}")
                st.caption(f"Quality {p.get('tour_quality',0):.0%} · Stability {p.get('stability',0):.0%} · Elo eff {p.get('elo_effective', d1[surface]):.0f}")
                if p.get("flags"):
                    st.caption(" · ".join(p.get("flags", [])))

            with r2c:
                p = rs.get("p2", {})
                st.metric(d2["Player"], f"{p.get('confidence',1):.0%}", delta=f"{p.get('matches_surface',0)} partidos {surface}")
                st.caption(f"Quality {p.get('tour_quality',0):.0%} · Stability {p.get('stability',0):.0%} · Elo eff {p.get('elo_effective', d2[surface]):.0f}")
                if p.get("flags"):
                    st.caption(" · ".join(p.get("flags", [])))

            with r3c:
                st.metric("Vol mult", f"{rs.get('vol_mult',1.0):.2f}")
                st.caption(f"Engine {rs.get('version','v22')}")

            st.caption("🔎 Match Count Debug")
            hist_diag = historicos_diagnostics(circuito)
            st.caption(
                f"Históricos: files={hist_diag.get('files_count',0)} · rows={hist_diag.get('rows',0)} · "
                f"Winner={hist_diag.get('winner_col')} · Loser={hist_diag.get('loser_col')} · Surface={hist_diag.get('surface_col')} · "
                f"Dates={hist_diag.get('date_min','N/A')}→{hist_diag.get('date_max','N/A')}"
            )
            if hist_diag.get("folder_counts"):
                parts = []
                for fd, cnt in (hist_diag.get("folder_counts", {}) or {}).items():
                    parts.append(f"{fd}={cnt}")
                st.caption("Archivos por carpeta: " + " · ".join(parts))
                for fd, samples in (hist_diag.get("folder_samples", {}) or {}).items():
                    if samples:
                        st.caption(f"Muestras {fd}: " + ", ".join([str(x) for x in samples[:5]]))
            if hist_diag.get("sample_players"):
                st.caption("Ejemplos histórico: " + ", ".join(hist_diag.get("sample_players", [])[:8]))
            if hist_diag.get("level_cols"):
                st.caption("Columnas nivel: " + ", ".join([str(x) for x in hist_diag.get("level_cols", [])[:8]]))
                sample_bits = []
                for col, vals in (hist_diag.get("level_samples", {}) or {}).items():
                    if vals:
                        sample_bits.append(f"{col}={', '.join([str(v) for v in vals[:4]])}")
                if sample_bits:
                    st.caption("Ejemplos nivel: " + " · ".join(sample_bits[:4]))
            # Debug interno del Quality Map/cache. Si esto sale a 0, Streamlit está usando un mapa vacío o cache viejo.
            qm1 = rs.get("p1", {}).get("quality_meta", {}) or rs.get("p2", {}).get("quality_meta", {}) or {}
            if qm1:
                st.caption(
                    f"QualityMap: version={qm1.get('version','N/A')} · players={qm1.get('raw_player_count','?')} · keys={qm1.get('quality_keys','?')}"
                )
                if qm1.get("sample_names"):
                    st.caption("QualityMap ejemplos: " + ", ".join([str(x) for x in qm1.get("sample_names", [])[:8]]))
                st.caption("Tour Quality v22.27: amplía Upset Risk Guard para favoritos clay con dog de muestra corta; limpia etiquetas contradictorias.")
            else:
                st.caption("QualityMap: sin meta visible — posible cache viejo o quality_map vacío")
            if hist_diag.get("files_count", 0) == 0 or hist_diag.get("rows", 0) == 0 or not hist_diag.get("winner_col") or not hist_diag.get("loser_col"):
                st.warning(
                    "Históricos no detectados bien: "
                    f"folder={hist_diag.get('folder')} · existe={hist_diag.get('folder_exists')} · "
                    f"files={hist_diag.get('files_count')} · rows={hist_diag.get('rows')} · "
                    f"Winner={hist_diag.get('winner_col')} · Loser={hist_diag.get('loser_col')} · Surface={hist_diag.get('surface_col')}"
                )
                if hist_diag.get("sample_files"):
                    st.caption("Files detectados: " + ", ".join(hist_diag.get("sample_files", [])))
                if hist_diag.get("columns"):
                    st.caption("Columnas detectadas: " + ", ".join([str(x) for x in hist_diag.get("columns", [])[:12]]))
                if hist_diag.get("sample_players"):
                    st.caption("Jugadores ejemplo: " + ", ".join(hist_diag.get("sample_players", [])[:6]))
            dbg1 = rs.get("p1", {})
            dbg2 = rs.get("p2", {})
            dc1, dc2 = st.columns(2)
            with dc1:
                st.caption(
                    f"{d1['Player']} → match: {dbg1.get('matched_name','N/A')} "
                    f"({dbg1.get('match_score',0):.0%}) · total {dbg1.get('matches_total',0)} · "
                    f"H/C/G {dbg1.get('matches_by_surface',{}).get('Hard',0)}/"
                    f"{dbg1.get('matches_by_surface',{}).get('Clay',0)}/"
                    f"{dbg1.get('matches_by_surface',{}).get('Grass',0)}"
                )
                lc = dbg1.get('level_counts', {})
                st.caption(f"Tour/Ch/ITF/Q/Unk {lc.get('tour',0)}/{lc.get('challenger',0)}/{lc.get('itf',0)}/{lc.get('qualy',0)}/{lc.get('unknown',0)}")
                if dbg1.get('source_files'):
                    st.caption("Files: " + ", ".join(dbg1.get('source_files', [])[:4]))
                radar = []
                for c in dbg1.get('candidate_matches', [])[:5]:
                    sf = c.get('surface', {}) if isinstance(c.get('surface', {}), dict) else {}
                    radar.append(f"{c.get('name','?')} {c.get('score',0):.0%} {c.get('reason','')} ({c.get('matches_total',0)} · C{sf.get('Clay',0)})")
                st.caption("Radar nombres: " + (" | ".join(radar) if radar else "sin candidatos"))
            with dc2:
                st.caption(
                    f"{d2['Player']} → match: {dbg2.get('matched_name','N/A')} "
                    f"({dbg2.get('match_score',0):.0%}) · total {dbg2.get('matches_total',0)} · "
                    f"H/C/G {dbg2.get('matches_by_surface',{}).get('Hard',0)}/"
                    f"{dbg2.get('matches_by_surface',{}).get('Clay',0)}/"
                    f"{dbg2.get('matches_by_surface',{}).get('Grass',0)}"
                )
                lc = dbg2.get('level_counts', {})
                st.caption(f"Tour/Ch/ITF/Q/Unk {lc.get('tour',0)}/{lc.get('challenger',0)}/{lc.get('itf',0)}/{lc.get('qualy',0)}/{lc.get('unknown',0)}")
                if dbg2.get('source_files'):
                    st.caption("Files: " + ", ".join(dbg2.get('source_files', [])[:4]))
                radar = []
                for c in dbg2.get('candidate_matches', [])[:5]:
                    sf = c.get('surface', {}) if isinstance(c.get('surface', {}), dict) else {}
                    radar.append(f"{c.get('name','?')} {c.get('score',0):.0%} {c.get('reason','')} ({c.get('matches_total',0)} · C{sf.get('Clay',0)})")
                st.caption("Radar nombres: " + (" | ".join(radar) if radar else "sin candidatos"))

        st.divider()
        st.subheader("🧠 Perfil del Partido")
        tags = []
        if sim["p1_profile"] in ["big_server","elite_server"]: tags.append(f"🚀 {d1['Player']} gran sacador")
        if sim["p2_profile"] in ["big_server","elite_server"]: tags.append(f"🚀 {d2['Player']} gran sacador")
        if sim["set3"] > 0.45: tags.append("⚠️ Partido volátil")
        if sim["tb"] > 0.32: tags.append("🎯 Tie-break probable")
        if avg_games > 24: tags.append("📈 Partido largo")
        if avg_games < 21: tags.append("📉 Partido corto")
        if max(p1c,p2c) < 0.55: tags.append("⚠️ Favorito débil")
        if sim["vol"] > 0.06: tags.append("🌪️ Alta volatilidad por diferencia Elo/superficie")
        if surface == "Hard" and sim["fav_raw_est"] >= 0.66: tags.append("🧊 Hard compression activada")
        if sim["hold1"] > 0.76 and sim["hold2"] > 0.76 and surface == "Hard": tags.append("🎯 Smart tie-break boost")
        if sim.get("tb_intel_boost", 0) >= 0.08: tags.append("🧠 TB Intelligence alto")
        if sim.get("fatigue1", {}).get("fatigue_score", 0) >= 0.030: tags.append(f"🔋 Fatiga {d1['Player']}")
        if sim.get("fatigue2", {}).get("fatigue_score", 0) >= 0.030: tags.append(f"🔋 Fatiga {d2['Player']}")
        if sim.get("long_match", 0) >= 0.65: tags.append("📈 Partido largo probable")
        if sim.get("rating_sanity", {}).get("active", False): tags.append("🧠 Rating sanity activo")
        if circuito == "ATP" and surface == "Clay" and sim.get("clay_engine", {}).get("active", False):
            tags.append(f"🧱 Clay engine: {sim.get('clay_engine', {}).get('profile','neutral')}")
        upset_active = (sim.get("upset_risk_guard", {}) or {}).get("active", False)
        fav20_tag_value = sim.get("fav_2_0", 0)
        elite_opp_active = (sim.get("elite_clay_opponent_guard", {}) or {}).get("active", False)
        high_gap_label = abs(max(p1c, p2c) - sim.get("fav_raw_est", max(p1c, p2c))) >= 0.10
        if fav20_tag_value >= 0.55:
            if upset_active:
                tags.append("⚠️ Favorito vulnerable")
            elif elite_opp_active or high_gap_label:
                if fav20_tag_value >= 0.65:
                    tags.append("✅ Favorito 2-0 viable")
                else:
                    tags.append("⚖️ Favorito 2-0 moderado")
            elif fav20_tag_value >= 0.78:
                tags.append("🔥 Spot favorito 2-0")
            elif fav20_tag_value >= 0.65:
                tags.append("✅ Favorito 2-0 viable")
            else:
                tags.append("⚖️ Favorito 2-0 moderado")
        if upset_active: tags.append("⚠️ Riesgo upset")
        if (sim.get("low_over_split_guard", {}) or {}).get("active", False):
            tags.append("🎚️ Over bajo ≠ partido largo")
        tc = sim.get("tournament_ctx", {})
        if "Indoor" in tc.get("court",""): tags.append("🏟️ Indoor boost")
        if "Final" in tc.get("round",""): tags.append("🎯 Final pressure")
        if "Grand Slam" in tc.get("series",""): tags.append("🏆 Grand Slam intensity")
        if sim.get("wta_engine_active", False): tags.append("🎾 WTA Chaos control")
        if sim.get("wta_separation", {}).get("active", False): tags.append("👑 Elite WTA separation")
        if sim.get("wta_script", {}).get("active", False): tags.append(f"🎭 {sim.get('wta_script',{}).get('script','neutral')}")
        st.info(" · ".join(tags) if tags else "Sin perfil extremo detectado.")

        st.divider()
        st.subheader("🎯 Señal principal del modelo")
        markets = {
            "ML favorito calibrado": max(p1c,p2c), "Over 18.5": over18,
            "Over 19.5": over19, "Over 20.5": over20, "Over 22.5": over22, "Under 22.5": under22,
            "Tie-break": sim["tb"], "Fav + Under 22.5": sim["fav_under22"],
            "Dog + Over 20.5": sim["dog_over20"],
            f"{sim.get('model_dog_name','Underdog')} gana set": sim["dog_wins_set"],
            f"{sim.get('model_fav_name','Favorito')} 2-0": sim["fav_2_0"],
            "Partido largo": sim["long_match"]
        }
        if circuito == "WTA":
            markets["Over 17.5 WTA"] = over17
        best = max(markets.items(), key=lambda x: x[1])
        st.success(f"{best[0]} → {best[1]:.1%}")
        filters = betting_filter_engine(circuito, surface, sim, d1["Player"], d2["Player"])
        render_betting_filters(filters)

        st.caption(f"Tennis IA v23.25 Fix Países + Watchlist Label · {sims:,} simulaciones Monte Carlo")


    # Mostrar afinador TA aunque el botón de simulación ya no esté pulsado.
    render_predictor_ta_finetune_panel()
    render_predictor_deep_match_analyzer_panel()
    render_predictor_excel_download_panel()

elif modo == "Analizador por lista":
    st.subheader("📋 Analizador por lista pegada")
    st.caption("Pega lista diaria o resultados de SofaScore. Se ignoran dobles/cancelados/retirados.")
    st.success("Fuente fijada: SofaScore. Esta pantalla ya no usa Datos extra ni listas simples como flujo principal.")
    st.info("Nota: en Sofascore resultados, el ML se valida con ganador real. Si el pegado incluye juegos por set, también valida Over 18.5 y, solo en WTA, Over 17.5. En tie-breaks puede contar puntos extra si Sofascore los copia como números separados.")

    with st.sidebar:
        surface = st.selectbox("Superficie lista", ["Hard","Clay","Grass"], index=1)
        formato = st.radio("Formato lista", ["ATP Tour (3 sets)", "Grand Slam (5 sets)"])
        sims_batch = st.select_slider(
            "Simulaciones por partido",
            [500, 1000, 2500, 5000],
            value=1000,
            help="Para lotes largos usa 500/1000. Para confirmar señales concretas usa el Predictor individual."
        )
        max_batch = st.number_input(
            "Máx partidos a analizar",
            1, 150, 60, 1,
            help="Puedes subirlo para analizar el día completo en una sola hoja. Si va lento, baja simulaciones a 500/1000."
        )
        formato_pegado = st.radio(
            "Formato SofaScore",
            ["Sofascore día auto ATP/WTA/Challenger", "Sofascore resultados auto ATP/WTA/Challenger"],
            index=0,
            help="Pega todo el día junto y la app filtra según el circuito elegido en la barra lateral. Ignora dobles/cancelados/retirados."
        )
        usar_cuotas = False
        vista_resultados_simple = st.toggle(
            "Vista simple resultados",
            value=True,
            help="En Sofascore resultados muestra ML, Over 18.5 y, solo en WTA, Over 17.5 para revisar aciertos."
        )

    ejemplo = """16:20
-
Russia
A. Rublev
Georgia
N. Basilashvili

17:30
-
Serbia
H. Medjedović
Spain
M. Landaluce

Sebastián Baez - Roberto Carballés Baena"""

    render_sofascore_auto_extractor(circuito)
    render_api_tennis_mobile_loader(circuito)

    raw_batch = st.text_area(
        "Pega aquí los partidos/resultados de SofaScore",
        height=260,
        placeholder=ejemplo,
        key="sofa_raw_batch"
    )

    # v23.37.27 FIX: si cambia la lista/formato/circuito, no mostrar ni reutilizar
    # resultados guardados de un análisis anterior. Esto evita que aparezca un día
    # viejo, por ejemplo 20/05, cuando ya se pegó otra lista.
    try:
        _batch_input_key = f"{formato_pegado}|{circuito}|{str(raw_batch).strip()}"
        _prev_input_key = st.session_state.get("batch_input_key_v233727")
        if _prev_input_key is not None and _prev_input_key != _batch_input_key:
            for _k in [
                "batch_ok_df", "batch_ko_df", "batch_last_ready",
                "datos_extra_ajustes", "ta_extra_ajustes",
            ]:
                if _k in st.session_state:
                    del st.session_state[_k]
        st.session_state["batch_input_key_v233727"] = _batch_input_key
    except Exception:
        pass

    cprev, crun = st.columns([1,1])
    with cprev:
        if raw_batch.strip():
            if formato_pegado == "Sofascore día":
                preview = parse_sofascore_paste(raw_batch)
            elif formato_pegado == "Sofascore día auto ATP/WTA/Challenger":
                preview_all = parse_sofascore_day_grouped_paste(raw_batch)
                st.caption(f"Auto detectado total: {len(preview_all)} · Se analizarán solo {circuito} según el selector lateral.")
                preview = filtrar_matches_por_circuito_pegado(preview_all, circuito)
            elif formato_pegado == "Sofascore resultados":
                preview = parse_sofascore_results_paste(raw_batch)
            elif formato_pegado == "Sofascore resultados auto ATP/WTA/Challenger":
                preview_all = parse_sofascore_results_grouped_paste(raw_batch)
                st.caption(f"Auto resultados detectado total: {len(preview_all)} · Se analizarán solo {circuito} según el selector lateral.")
                preview = filtrar_matches_por_circuito_pegado(preview_all, circuito)
            else:
                preview = parse_datos_extra_paste(raw_batch) if usar_cuotas else parse_simple_match_list(raw_batch)
                # autodetección si se pegó formato Sofascore pero quedó seleccionado Casa/Datos extra.
                if not preview and any(is_time_line_sofa(x.strip()) for x in raw_batch.splitlines()):
                    preview = parse_sofascore_paste(raw_batch)
                if not preview and any(is_date_line_sofa_result(x.strip()) for x in raw_batch.splitlines()):
                    preview = parse_sofascore_results_paste(raw_batch)
        else:
            preview = []
        st.metric("Partidos detectados", len(preview))
    with crun:
        total_est_sims = min(len(preview), int(max_batch)) * int(sims_batch)
        st.metric("Simulaciones estimadas", f"{total_est_sims:,}")
        if total_est_sims > 30000:
            st.warning("Lote pesado para Streamlit Cloud. Mejor baja a 500/1000 sims o analiza menos partidos.")

    if preview:
        with st.expander("👀 Vista previa detectada", expanded=True):
            prev_df = pd.DataFrame(preview[:int(max_batch)])

            # v23.25.4 Debug Games Preview:
            # Solo añade visibilidad. No cambia cálculos, simulaciones ni filtros.
            if "actual_total_games" in prev_df.columns:
                prev_df["games_leidos"] = prev_df["actual_total_games"].apply(lambda x: "✅ Sí" if pd.notna(x) and str(x) != "" else "⚠️ No")
            if "p1_sets_real" in prev_df.columns and "p2_sets_real" in prev_df.columns:
                prev_df["sets_real"] = prev_df.apply(
                    lambda r: f"{int(r['p1_sets_real'])}-{int(r['p2_sets_real'])}"
                    if pd.notna(r.get('p1_sets_real')) and pd.notna(r.get('p2_sets_real')) else "",
                    axis=1
                )

            show_cols = [c for c in [
                "date", "time", "circuito_detectado", "torneo", "surface",
                "p1_raw", "p2_raw", "sets_real", "score_games",
                "actual_total_games", "games_leidos", "p1_sets_real", "p2_sets_real"
            ] if c in prev_df.columns]
            st.dataframe(prev_df[show_cols] if show_cols else prev_df, width='stretch', hide_index=True)

            if "actual_total_games" in prev_df.columns:
                con_games = int(prev_df["actual_total_games"].notna().sum())
                sin_games = int(len(prev_df) - con_games)
                st.caption(f"Juegos reales detectados en {con_games}/{len(prev_df)} partidos. Sin juegos reales: {sin_games}.")

    if st.button("🚀 ANALIZAR LISTA", width='stretch'):
        # v23.10/v23.37.27: liberar resultado anterior antes de un lote nuevo.
        for _k in ["batch_ok_df", "batch_ko_df", "batch_last_ready", "datos_extra_ajustes", "ta_extra_ajustes"]:
            if _k in st.session_state:
                del st.session_state[_k]
        try:
            st.session_state["batch_input_key_v233727"] = f"{formato_pegado}|{circuito}|{str(raw_batch).strip()}"
        except Exception:
            pass
        gc.collect()

        if formato_pegado == "Sofascore día":
            parsed = parse_sofascore_paste(raw_batch)[:int(max_batch)]
        elif formato_pegado == "Sofascore día auto ATP/WTA/Challenger":
            parsed_all = parse_sofascore_day_grouped_paste(raw_batch)
            parsed = filtrar_matches_por_circuito_pegado(parsed_all, circuito)[:int(max_batch)]
        elif formato_pegado == "Sofascore resultados":
            parsed = parse_sofascore_results_paste(raw_batch)[:int(max_batch)]
        elif formato_pegado == "Sofascore resultados auto ATP/WTA/Challenger":
            parsed_all = parse_sofascore_results_grouped_paste(raw_batch)
            parsed = filtrar_matches_por_circuito_pegado(parsed_all, circuito)[:int(max_batch)]
        else:
            parsed = (parse_datos_extra_paste(raw_batch) if usar_cuotas else parse_simple_match_list(raw_batch))
            # autodetección si se pegó formato Sofascore pero quedó seleccionado Casa/Datos extra.
            if not parsed and any(is_time_line_sofa(x.strip()) for x in raw_batch.splitlines()):
                parsed = parse_sofascore_paste(raw_batch)
            if not parsed and any(is_date_line_sofa_result(x.strip()) for x in raw_batch.splitlines()):
                parsed = parse_sofascore_results_paste(raw_batch)
            parsed = parsed[:int(max_batch)]

        # v23.10: seguridad para evitar superar memoria en Cloud.
        if len(parsed) * int(sims_batch) > 50000:
            st.error("Lote demasiado pesado para Streamlit Cloud. Baja simulaciones o reduce Máx partidos. Recomendado: 10-15 partidos a 500/1000 sims.")
            st.stop()

        if not parsed:
            st.error("No he detectado partidos para el circuito seleccionado. Revisa que hayas elegido el formato correcto: SofaScore día auto o SofaScore resultados auto, y que arriba esté seleccionado ATP, WTA o Challenger correctamente.")
            st.stop()

        best_of = 5 if "5" in formato else 3
        if best_of == 5:
            st.success("🏆 Modo Grand Slam 5 sets activo: se usarán mercados GS 30.5/32.5/35.5/37.5/39.5, 4+ sets, 5 sets y 3-0/no 3-0. El Over BO3 sigue blindado.")

        status = st.status(f"📋 Analizando {len(parsed)} partidos...", expanded=True)
        with status:
            bar = st.progress(1, text="Preparando análisis por lote...")
            msg = st.empty()
            start = time.time()

            def update_batch(done, total, label=""):
                pct = int(round((done / total) * 100)) if total else 100
                pct = max(1, min(100, pct))
                elapsed = time.time() - start
                bar.progress(pct, text=f"Analizando lista {done}/{total} · {elapsed:.1f}s")
                if label:
                    msg.caption(label)

            df_batch = analyze_batch_matches(
                parsed, db, circuito, surface, best_of, int(sims_batch),
                progress_callback=update_batch
            )
            # v23.26.1: eliminar duplicados reales antes de separar OK/KO y antes de exportar.
            if isinstance(df_batch, pd.DataFrame) and not df_batch.empty and "Partido" in df_batch.columns:
                dedupe_cols = [c for c in ["Fecha", "Hora", "Partido"] if c in df_batch.columns]
                if dedupe_cols:
                    before_dedup = len(df_batch)
                    df_batch = df_batch.drop_duplicates(subset=dedupe_cols, keep="last").reset_index(drop=True)
                    if before_dedup != len(df_batch):
                        msg.caption(f"✅ Lista analizada. Duplicados eliminados: {before_dedup - len(df_batch)}")
            bar.progress(100, text=f"Análisis completado · {len(parsed)} partidos")
            msg.caption("✅ Lista analizada.")

        status.update(label=f"✅ Lista analizada · {len(parsed)} partidos", state="complete", expanded=False)

        st.divider()
        st.subheader("🔥 Resumen ordenado")
        if "Estado" in df_batch.columns:
            # v23.25.9: los partidos con fallback estimado SÍ están analizados.
            # Antes caían en "No encontrados" porque el split usaba Estado == "OK" estricto.
            # Ahora la tabla principal muestra OK + OK con jugador estimado;
            # la pestaña "No encontrados" queda solo para no analizables reales.
            estados_ok = ["OK", "OK con jugador estimado"]
            ok = df_batch[df_batch["Estado"].isin(estados_ok)].copy()
            ko = df_batch[~df_batch["Estado"].isin(estados_ok)].copy()
        else:
            ok, ko = df_batch, pd.DataFrame()

        if not ok.empty:
            ok = prepare_batch_display_table(ok)

            # Orden simple: recomendación/trust/edge.
            def _edge_num(x):
                try:
                    s = str(x).replace("%","").replace(",", ".").strip()
                    return float(s) if s else -999
                except Exception:
                    return -999

            rec_order = {
                "APTA + VALUE": 5,
                "APTA": 4,
                "DUDOSA CON VALUE": 3,
                "DUDOSA": 2,
                "VALUE NUMÉRICO PERO RIESGO": 1,
                "MIRAR FAVORITO 2-0": 4,
                "WATCH FAVORITO 2-0": 1,
                "WATCHLIST OVER": 1,
                "NO BET": 0,
            }

            ok["_edge_sort"] = ok["Edge"].apply(_edge_num) if "Edge" in ok.columns else -999
            ok["_rec_sort"] = ok["Recomendación"].astype(str).apply(lambda x: max([v for k,v in rec_order.items() if k in x] or [0]))
            ok = ok.sort_values(["_rec_sort","_edge_sort"], ascending=[False, False]).drop(columns=["_edge_sort","_rec_sort"], errors="ignore")

            # v23.11: vista simple para backtest de resultados.
            if formato_pegado in ["Sofascore resultados", "Sofascore resultados auto ATP/WTA/Challenger"] and vista_resultados_simple:
                # v23.37.4: la vista simple de backtest NO debe ocultar el Selector Maestro.
                # Antes esta lista recortaba las columnas 🎯 y por eso el Excel parecía igual
                # aunque el motor nuevo ya estuviera calculándolas.
                simple_cols = [
                    "Versión app",
                    "Fecha",
                    "Partido",
                    "🎯 Acción final",
                    "🎯 Decisión acierto",
                    "🎯 Mercado más probable",
                    "🎯 Prob máxima",
                    "🎯 Confianza acierto",
                    "🎯 Aviso acierto",
                    "🎯 Motivo acierto",
                    "🏆 GS 5 sets activo",
                    "🏆 Mercado GS 5 sets",
                    "🏆 Prob GS 5 sets",
                    "🏆 Acción GS",
                    "🏆 Confianza GS",
                    "GS Over 30.5",
                    "GS Over 32.5",
                    "GS Over 35.5",
                    "GS Over 37.5",
                    "GS Over 39.5",
                    "GS Partido 4+ sets",
                    "GS Partido 5 sets",
                    "GS Favorito 3-0",
                    "GS Favorito NO 3-0",
                    "GS Underdog gana set",
                    "GS Juegos reales",
                    "GS Sets reales totales",
                    "GS Over 30.5 real",
                    "GS Over 32.5 real",
                    "GS Over 35.5 real",
                    "GS Over 37.5 real",
                    "GS Over 39.5 real",
                    "GS Partido 4+ sets real",
                    "GS Partido 5 sets real",
                    "GS Favorito 3-0 real",
                    "GS Favorito NO 3-0 real",
                    "GS Underdog gana set real",
                    "Acierta mercado GS",
                    "📥 Necesita Datos extra",
                    "📥 Prioridad Datos extra",
                    "📥 Qué mirar Datos extra",
                    "📥 Motivo Datos extra",
                    "Favorito modelo",
                    "ML favorito",
                    "Ganador real",
                    "Resultado sets",
                    "Acierta ML modelo",
                    "Mejor mercado WTA",
                    "WTA Over17 Priority",
                    "Over 17.5",
                    "Over 17.5 real",
                    "Acierta Over 17.5",
                    "Over 18.5",
                    "Over 18.5 real",
                    "Acierta Over 18.5",
                    "WTA Watchlist",
                    "Signal Trust",
                    "Recomendación",
                    "Mercado recomendado",
                    "Prob mercado recomendado",
                    "Motivo Market Selector",
                    "Favorito 2-0",
                    "Riesgos",
                    "Pick oficial"
                ]
                ok = ok[[c for c in simple_cols if c in ok.columns]]

            # v23.37.26: aplicar Challenger Recent Form Engine antes de guardar.
            try:
                ok = aplicar_challenger_recent_form_engine(ok)
                # v23.38.4: candado FINAL después del CH engine, porque CH podía recuperar Over 18.5.
                ok = aplicar_grand_slam_bo5_selector_final(ok)
            except Exception:
                try:
                    ok = aplicar_grand_slam_bo5_selector_final(ok)
                except Exception:
                    pass

            # v23.5: guardar resultado para que descargar Excel/CSV no limpie pantalla tras rerun.
            st.session_state["batch_ok_df"] = ok
            st.session_state["batch_ko_df"] = ko
            st.session_state["batch_last_ready"] = True

        else:
            st.warning("No se pudo analizar ningún partido.")
            st.session_state["batch_ok_df"] = pd.DataFrame()
            st.session_state["batch_ko_df"] = ko
            st.session_state["batch_last_ready"] = True

    # v23.5: render persistente fuera del botón. Así los downloads no borran la tabla.
    if st.session_state.get("batch_last_ready", False):
        ok_saved = st.session_state.get("batch_ok_df", pd.DataFrame())
        ko_saved = st.session_state.get("batch_ko_df", pd.DataFrame())
        try:
            ok_saved = aplicar_challenger_recent_form_engine(ok_saved)
            # v23.38.4: candado FINAL persistente después del CH engine.
            ok_saved = aplicar_grand_slam_bo5_selector_final(ok_saved)
            st.session_state["batch_ok_df"] = ok_saved
        except Exception:
            try:
                ok_saved = aplicar_grand_slam_bo5_selector_final(ok_saved)
                st.session_state["batch_ok_df"] = ok_saved
            except Exception:
                pass

        if ok_saved is not None and not ok_saved.empty:
            st.divider()

            # v23.30.2: panel separado, claro y SIEMPRE visible solo con picks oficiales.
            # Si la sesión viene de una tabla anterior sin la columna, la reconstruimos al vuelo.
            if "Pick oficial" not in ok_saved.columns:
                try:
                    ok_saved = ok_saved.copy()
                    ok_saved["Pick oficial"] = ok_saved.apply(pick_oficial_v23301, axis=1)
                    st.session_state["batch_ok_df"] = ok_saved
                except Exception:
                    pass

            st.subheader("🎯 Picks oficiales")

            if "Pick oficial" in ok_saved.columns:
                _oficiales = ok_saved[ok_saved["Pick oficial"].astype(str).str.strip() != ""].copy()

                if not _oficiales.empty:
                    _n_over = int(_oficiales["Pick oficial"].astype(str).str.upper().str.contains("OVER", na=False).sum())
                    _n_ml = int(_oficiales["Pick oficial"].astype(str).str.upper().str.contains("ML", na=False).sum())

                    cpo1, cpo2, cpo3 = st.columns(3)
                    cpo1.metric("Picks oficiales", len(_oficiales))
                    cpo2.metric("Overs oficiales", _n_over)
                    cpo3.metric("ML oficiales", _n_ml)

                    _cols_oficiales_base = [
                        "Hora",
                        "Partido",
                        "Pick oficial",
                        "Mercado recomendado",
                        "Prob mercado recomendado",
                        "Resultado sets",
                        "Ganador real",
                        "Over 18.5 real",
                        "Acierta Over 18.5",
                        "Acierta ML modelo",
                        "Confianza mínima",
                        "Mín. partidos superficie",
                        "Over Quality Guard",
                        "ML Quality Guard",
                        "WTA Over17 Official Guard",
                        "Motivo Market Selector",
                    ]
                    _cols_oficiales = [c for c in _cols_oficiales_base if c in _oficiales.columns]

                    st.dataframe(
                        _oficiales[_cols_oficiales],
                        width='stretch',
                        hide_index=True
                    )

                    with st.expander("Ver solo motivos de picks oficiales"):
                        _motivo_cols = [c for c in ["Partido", "Pick oficial", "Motivo Market Selector", "Motivos Over Guard", "Motivos ML Guard", "Riesgos"] if c in _oficiales.columns]
                        if _motivo_cols:
                            st.dataframe(_oficiales[_motivo_cols], width='stretch', hide_index=True)
                else:
                    st.warning("🎯 Picks oficiales detectados: 0. No forzar combinada si solo hay WATCH/NO BET.")
            else:
                st.warning("No se ha podido crear la columna 🎯 Pick oficial en esta ejecución.")

            with st.expander("📲 Enviar picks a Telegram", expanded=False):
                render_telegram_sender_panel(ok_saved)

            render_datos_extra_reanalysis_panel(ok_saved)

            try:
                _jg_view = crear_juegos_jugador_df_v23480(ok_saved)
                if _jg_view is not None and not _jg_view.empty:
                    st.subheader("🎯 Juegos jugador +10.5 / +11.5")
                    st.caption("Mercado nuevo en observación: no toca motor Over. Úsalo unos días para validar estadísticas.")
                    st.dataframe(_jg_view.head(20), width='stretch', hide_index=True)
            except Exception as _e_jg:
                st.caption(f"Juegos jugador no disponible: {type(_e_jg).__name__}: {_e_jg}")

            st.subheader("🔥 Resumen ordenado completo")
            _hide_cols_v23480 = [c for c in ["Cuota pegada", "Jugador cuota", "Cuota justa", "Edge"] if c in ok_saved.columns]
            st.dataframe(ok_saved.drop(columns=_hide_cols_v23480, errors="ignore"), width='stretch', hide_index=True)

            # v23.48.0: pantalla principal limpia. Se ocultan cuotas, stake y value.
            # La estrategia se decide con rankings, Over, gana set y juegos jugador.

            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    "⬇️ Descargar CSV",
                    data=ok_saved.to_csv(index=False).encode("utf-8-sig"),
                    file_name="analisis_lista_tennis_ia_v23_26_0.csv",
                    mime="text/csv",
                    key="download_batch_csv"
                )
            with dl2:
                st.download_button(
                    "📊 Descargar Excel",
                    data=batch_excel_with_not_found_bytes(ok_saved, ko_saved, db),
                    file_name="analisis_lista_tennis_ia_v23_26_0.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_batch_excel"
                )

            if st.button("💾 Guardar copia automática en carpeta exports", key="backup_batch_excel"):
                _ok_bk, _msg_bk = backup_analysis_excel(ok_saved, ko_saved, db)
                if _ok_bk:
                    st.success(f"Copia guardada: {_msg_bk}")
                else:
                    st.warning(f"No se pudo guardar la copia: {_msg_bk}")

        if ko_saved is not None and not ko_saved.empty:
            st.divider()
            st.subheader("⚠️ No encontrados / revisar nombres")
            ko_view = enrich_not_found_with_suggestions(ko_saved, db)
            st.dataframe(ko_view, width='stretch', hide_index=True)

            nf1, nf2 = st.columns(2)
            with nf1:
                st.download_button(
                    "⬇️ Descargar no encontrados CSV",
                    data=ko_view.to_csv(index=False).encode("utf-8-sig"),
                    file_name="no_encontrados_tennis_ia.csv",
                    mime="text/csv",
                    key="download_not_found_csv"
                )
            with nf2:
                st.download_button(
                    "📊 Descargar no encontrados Excel",
                    data=batch_excel_bytes(ko_view),
                    file_name="no_encontrados_tennis_ia.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_not_found_excel"
                )


elif modo == "Validador histórico":
    st.subheader("📚 Validador histórico")
    hist_df = cargar_historicos(circuito)
    if hist_df.empty:
        st.error("No se encontraron históricos.")
        st.stop()
    with st.sidebar:
        surface_filter = st.selectbox("Superficie histórica", ["Todas","Hard","Clay","Grass"])
        max_matches = st.number_input("Máx partidos a validar", 10, 5000, 500, 50)
        sims_bt = st.select_slider("Simulaciones por partido", [300,500,1000,2000], value=500)
    st.info(f"Históricos cargados ({circuito}): {len(hist_df):,} partidos.")
    if st.button("🚀 EJECUTAR VALIDACIÓN", width='stretch'):
        with st.spinner("Validando partidos históricos..."):
            val = validar_historico(db,hist_df,circuito_sim_bt,surface_filter,int(max_matches),int(sims_bt))
        if val.empty:
            st.error("No se pudieron emparejar partidos.")
            st.stop()
        ml_acc = val["CalFavWasWinner"].mean()
        over17_acc = ((val["ModelOver17"] >= 0.50) == val["RealOver17"]).mean() if "ModelOver17" in val.columns else 0
        over18_acc = ((val["ModelOver18"] >= 0.50) == val["RealOver18"]).mean()
        over19_acc = ((val["ModelOver19"] >= 0.50) == val["RealOver19"]).mean() if "ModelOver19" in val.columns else 0
        over20_acc = ((val["ModelOver20"] >= 0.50) == val["RealOver20"]).mean()
        over22_acc = ((val["ModelOver22"] >= 0.50) == val["RealOver22"]).mean()
        set3_acc = ((val["Model3Sets"] >= 0.50) == val["Real3Sets"]).mean()
        tb_acc = ((val["ModelTB"] >= 0.50) == val["RealTB"]).mean()
        games_error = np.mean(np.abs(val["ModelAvgGames"] - val["RealGames"]))
        st.divider()
        st.subheader("📊 Resumen validación")
        c1,c2,c3,c4 = st.columns(4)
        with c1: st.metric("ML accuracy", f"{ml_acc:.1%}")
        with c2: st.metric("Over 20.5 accuracy", f"{over20_acc:.1%}")
        with c3: st.metric("3 sets accuracy", f"{set3_acc:.1%}")
        with c4: st.metric("Error medio games", f"{games_error:.2f}")
        d1,d2,d3 = st.columns(3)
        with d1: st.metric("Over 17.5 accuracy", f"{over17_acc:.1%}")
        with d2: st.metric("Over 18.5 / 19.5", f"{over18_acc:.1%} / {over19_acc:.1%}")
        with d3: st.metric("Tie-break accuracy", f"{tb_acc:.1%}")
        st.caption(f"Over 20.5 accuracy: {over20_acc:.1%} · Over 22.5 accuracy: {over22_acc:.1%}")
        st.dataframe(val, width='stretch')
        st.download_button("⬇️ Descargar CSV", data=val.to_csv(index=False).encode("utf-8"), file_name="validacion_tennis_ia_v22.csv", mime="text/csv")

else:
    st.subheader("📊 Analyzer Engine")
    hist_df = cargar_historicos(circuito)
    if hist_df.empty:
        st.error("No se encontraron históricos.")
        st.stop()
    with st.sidebar:
        surface_filter = st.selectbox("Superficie analyzer", ["Todas","Hard","Clay","Grass"])
        max_matches = st.number_input("Máx partidos", 50, 5000, 1000, 100)
        sims_bt = st.select_slider("Simulaciones por partido", [300,500,1000], value=500)
        min_casos = st.number_input("Mínimo casos por segmento", 5, 200, 25, 5)
    st.info(f"Este modo busca patrones históricos por mercado, superficie, ranking gap y perfiles. Circuito activo: {circuito}")
    if st.button("🚀 EJECUTAR ANALYZER", width='stretch'):
        with st.spinner("Generando validación base para Analyzer..."):
            val = validar_historico(db,hist_df,circuito_sim_bt,surface_filter,int(max_matches),int(sims_bt))
        if val.empty:
            st.error("No se pudieron emparejar partidos.")
            st.stop()
        tables = crear_analyzer_tables(val, int(min_casos))
        bt_summary, bt_detail = crear_backtest_oficial_observar(val, circuito, int(min_casos))
        st.divider()
        st.subheader("🎯 Backtest OFICIAL vs OBSERVAR")
        if bt_summary.empty:
            st.warning("No hay suficientes señales OFICIAL/OBSERVAR con el mínimo de casos seleccionado.")
        else:
            bt_show = bt_summary.copy()
            for _c in ["Porcentaje_acierto", "Prob_media_modelo"]:
                if _c in bt_show.columns:
                    bt_show[_c] = bt_show[_c].apply(lambda x: f"{x:.1%}" if pd.notna(x) else "")
            if "Games_media_real" in bt_show.columns:
                bt_show["Games_media_real"] = bt_show["Games_media_real"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")
            st.dataframe(bt_show, width='stretch', hide_index=True)
            st.download_button(
                "⬇️ Descargar backtest Oficial/Observar CSV",
                data=bt_detail.to_csv(index=False).encode("utf-8"),
                file_name="backtest_oficial_observar_tennis_ia.csv",
                mime="text/csv"
            )
        st.divider()
        st.subheader("🏆 Resumen general")
        g1,g2,g3,g4 = st.columns(4)
        with g1: st.metric("Partidos analizados", f"{len(val):,}")
        with g2: st.metric("ML accuracy", f"{val['CalFavWasWinner'].mean():.1%}")
        with g3: st.metric("Games error", f"{np.mean(np.abs(val['ModelAvgGames'] - val['RealGames'])):.2f}")
        with g4: st.metric("TB real/modelo", f"{val['RealTB'].mean():.1%} / {val['ModelTB'].mean():.1%}")
        st.divider()
        st.subheader("📈 Mercados por umbral")
        mt = tables.get("Mercados por umbral", pd.DataFrame())
        if not mt.empty:
            st.dataframe(mt.sort_values(["Acierto real","Casos"], ascending=[False,False]), width='stretch')
        else:
            st.warning("No hay suficientes casos para mercados por umbral.")
        st.divider()
        st.subheader("🎯 ML por tramos")
        st.dataframe(tables.get("ML por tramos", pd.DataFrame()), width='stretch')
        st.divider()
        st.subheader("🌍 Por superficie")
        st.dataframe(tables.get("Por superficie", pd.DataFrame()), width='stretch')
        st.divider()
        st.subheader("🚀 Big server")
        st.dataframe(tables.get("Big server", pd.DataFrame()), width='stretch')
        st.divider()
        st.subheader("📊 Ranking gap")
        st.dataframe(tables.get("Ranking gap", pd.DataFrame()), width='stretch')
        st.divider()
        st.subheader("🧾 Detalle base")
        st.dataframe(val, width='stretch')
        st.download_button("⬇️ Descargar Analyzer CSV", data=val.to_csv(index=False).encode("utf-8"), file_name="analyzer_tennis_ia_v22.csv", mime="text/csv")